"""
Test B1 Dashboard build (v1.4.2).

Проверяет, что `scripts/build_dashboard.py`:
1. Успешно загружает артефакты 4 MC-движков.
2. Генерирует функции engine_summary и stage_gate_funnel без ошибок.
3. Содержит все 4 движка с согласованными числами.
4. Stage-gate воронка монотонно убывает.
5. Tornado от matrix_27 даёт корректные экстремумы.
6. Значения в dashboard совпадают с оригинальными JSON (no drift).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from build_dashboard import (  # noqa: E402
    ANCHOR_BASE,
    ANCHOR_LOWER,
    ANCHOR_UPPER,
    engine_summary,
    load_artifacts,
    stage_gate_funnel,
    tornado_from_matrix27,
)


@pytest.fixture(scope="module")
def data() -> dict:
    return load_artifacts()


def test_load_artifacts_all_present(data):
    for key in ("mc", "boot", "gate", "lhs", "m27"):
        assert key in data, f"Missing artifact: {key}"
        assert isinstance(data[key], dict)


def test_engine_summary_four_engines(data):
    es = engine_summary(data)
    assert len(es) == 4
    names = [e["engine"] for e in es]
    assert names == ["Naive MC", "Block Bootstrap", "Stage-Gate", "LHS+Copula"]


def test_engine_summary_fields_numeric(data):
    es = engine_summary(data)
    for e in es:
        for key in ("mean", "std", "p5", "p50", "p95", "var95", "breach_p"):
            assert key in e
            assert isinstance(e[key], (int, float))
            assert e[key] >= 0 or key == "mean"  # mean может быть любой


def test_lhs_mean_close_to_anchor(data):
    es = engine_summary(data)
    lhs = next(e for e in es if e["engine"] == "LHS+Copula")
    # LHS mean должен быть в [2800, 3050] — в пределах нормы для якоря
    assert 2800 <= lhs["mean"] <= 3050, f"LHS mean out of range: {lhs['mean']}"


def test_lhs_variance_less_than_naive_mc(data):
    """LHS+copula должен давать меньшую (или равную) std vs naive MC."""
    es = engine_summary(data)
    mc = next(e for e in es if e["engine"] == "Naive MC")
    lhs = next(e for e in es if e["engine"] == "LHS+Copula")
    assert lhs["std"] <= mc["std"] + 5, (
        f"LHS std {lhs['std']} should be <= MC std {mc['std']} (+tolerance)"
    )


def test_lhs_breach_p_less_than_naive_mc(data):
    """LHS+copula breach должен быть не больше naive MC (вариационная редукция rank-bias)."""
    es = engine_summary(data)
    mc = next(e for e in es if e["engine"] == "Naive MC")
    lhs = next(e for e in es if e["engine"] == "LHS+Copula")
    assert lhs["breach_p"] <= mc["breach_p"] + 0.005


def test_stage_gate_funnel_monotone_decreasing(data):
    funnel = stage_gate_funnel(data["gate"])
    assert len(funnel) == 5
    counts = [f[1] for f in funnel]
    # Монотонно убывает (нестрого)
    for i in range(len(counts) - 1):
        assert counts[i] >= counts[i + 1], f"Funnel not monotone at {i}: {counts}"
    # Начинается с n_films
    assert counts[0] == data["gate"]["n_films"]
    # Последний равен ≈ n * P(reach_release)
    expected_last = data["gate"]["n_films"] * data["gate"]["p_reach_release"]
    assert abs(counts[-1] - expected_last) < 0.01


def test_tornado_four_drivers(data):
    tornado = tornado_from_matrix27(data["m27"])
    assert len(tornado) == 4
    for name, low, high in tornado:
        assert low <= high, f"tornado {name}: low {low} > high {high}"


def test_tornado_fx_is_widest(data):
    """FX 0→20% должен быть в топ-2 по ширине среди одно-драйверных."""
    tornado = tornado_from_matrix27(data["m27"])
    single_drivers = [(n, lo, hi) for n, lo, hi in tornado if not n.startswith("Комбо")]
    widths = {n: hi - lo for n, lo, hi in single_drivers}
    fx_width = widths["FX 0→20%"]
    sorted_widths = sorted(widths.values(), reverse=True)
    # FX должен быть среди топ-2
    assert fx_width in sorted_widths[:2], (
        f"FX width {fx_width} not in top-2 of single drivers: {widths}"
    )


def test_tornado_combo_widest(data):
    """Комбинированный сценарий должен иметь максимальную ширину."""
    tornado = tornado_from_matrix27(data["m27"])
    widths = {name: high - low for name, low, high in tornado}
    combo_width = max(w for n, w in widths.items() if n.startswith("Комбо"))
    single_widths = [w for n, w in widths.items() if not n.startswith("Комбо")]
    assert combo_width >= max(single_widths), (
        f"Combo width {combo_width} should be >= single max {max(single_widths)}"
    )


def test_values_match_source_json(data):
    """Значения в engine_summary совпадают с исходными JSON (no drift)."""
    es = engine_summary(data)
    lhs = next(e for e in es if e["engine"] == "LHS+Copula")
    assert lhs["mean"] == data["lhs"]["mean_ebitda"]
    assert lhs["var95"] == data["lhs"]["var_95_mln_rub"]
    assert lhs["breach_p"] == data["lhs"]["breach_probability"]


def test_anchor_constants():
    assert ANCHOR_BASE == 3000.0
    assert ANCHOR_LOWER == 2970.0
    assert ANCHOR_UPPER == 3030.0
    assert ANCHOR_UPPER - ANCHOR_LOWER == 60.0  # ±1%


def test_html_dashboard_exists_after_build(data, tmp_path):
    """Smoke-тест: build_html_dashboard создаёт файл и он содержит plotly."""
    from build_dashboard import build_html_dashboard
    out = tmp_path / "dash.html"
    build_html_dashboard(data, out)
    assert out.exists()
    content = out.read_text()
    assert "plotly" in content.lower()
    # Plotly кодирует кириллицу в unicode-escape, ищем по escape-последовательности
    assert "\\u0424\\u0438\\u043d\\u043c\\u043e\\u0434\\u0435\\u043b\\u044c" in content, (
        "HTML должен содержать 'Финмодель' (unicode-escaped)"
    )
    assert "3000" in content
    assert "LHS+Copula" in content


def test_png_charts_produced(data, tmp_path):
    from build_dashboard import build_png_charts
    produced = build_png_charts(data, tmp_path)
    assert len(produced) == 5
    for p in produced:
        assert p.exists()
        assert p.stat().st_size > 1000  # не пустые


def test_xlsx_dashboard_structure(data, tmp_path):
    from build_dashboard import build_xlsx_dashboard
    from openpyxl import load_workbook
    out = tmp_path / "dash.xlsx"
    build_xlsx_dashboard(data, out)
    wb = load_workbook(out)
    expected_sheets = ["Summary", "Matrix27", "StageGate", "LHSCopula"]
    assert wb.sheetnames == expected_sheets
    # Summary должен содержать 4 движка (строки 5-8)
    ws = wb["Summary"]
    engine_names_in_sheet = [ws.cell(row=r, column=1).value for r in range(5, 9)]
    assert engine_names_in_sheet == [
        "Naive MC",
        "Block Bootstrap",
        "Stage-Gate",
        "LHS+Copula",
    ]
    # Matrix27 должен иметь 27 сценариев + header + footer
    ws2 = wb["Matrix27"]
    assert ws2.max_row >= 28  # header + 27 scenarios + footer rows
