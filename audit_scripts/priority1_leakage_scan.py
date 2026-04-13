#!/usr/bin/env python3
"""
PRIORITY 1 SECURITY AUDIT: Internal Data Leakage Scan
Target: investor_model_v1.0_Public.xlsx
Checks for any traces of internal/confidential data leaking into the public model.
"""

import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────
TARGET = "/home/user/TrendStudio-Holding/Investor_Package/investor_model_v1.0_Public.xlsx"

# Patterns to scan in raw XML
SENSITIVE_PATTERNS = {
    "Internal keyword": re.compile(r"[Ii]nternal", re.IGNORECASE),
    "W5 reference": re.compile(r"\bW5\b"),
    "W5 unicode subscript": re.compile(r"W\u2085"),
    "Waterfall 5": re.compile(r"[Ww]aterfall\s*5", re.IGNORECASE),
    "Internal IRR 24.75": re.compile(r"24[\.\,]75"),
    "Internal IRR 2475": re.compile(r"\b2475\b"),
    "Internal MC IRR 13.95": re.compile(r"13[\.\,]95"),
    "Email address": re.compile(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"),
    "File path (Windows)": re.compile(r"[A-Z]:\\[^\s<>\"]+"),
    "File path (UNC)": re.compile(r"\\\\[^\s<>\"]+"),
    "File path (Unix)": re.compile(r"/(?:home|Users|tmp|var|etc)/[^\s<>\"]+"),
    "Hidden sheet attr": re.compile(r"(?:hidden|veryHidden)", re.IGNORECASE),
}

findings = []

def add_finding(severity, category, location, detail):
    findings.append({
        "severity": severity,
        "category": category,
        "location": location,
        "detail": detail,
    })

def severity_order(s):
    return {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}.get(s, 99)

# ─────────────────────────────────────────────────────────────────────
# 1. openpyxl-based checks (data_only=False to preserve formulas)
# ─────────────────────────────────────────────────────────────────────
print("=" * 90)
print("PRIORITY 1 — INTERNAL DATA LEAKAGE SCAN")
print(f"Target: {TARGET}")
print(f"File size: {os.path.getsize(TARGET):,} bytes")
print("=" * 90)

wb = openpyxl.load_workbook(TARGET, data_only=False)

# ── 1a. Sheet inventory (including hidden / very hidden) ──
print("\n" + "─" * 90)
print("SECTION 1: SHEET INVENTORY")
print("─" * 90)
all_sheets = wb.sheetnames
print(f"Total sheets: {len(all_sheets)}")
for sname in all_sheets:
    ws = wb[sname]
    vis = ws.sheet_state  # 'visible', 'hidden', 'veryHidden'
    tag = ""
    if vis == "hidden":
        tag = "  ** HIDDEN **"
        add_finding("HIGH", "Hidden Sheet", sname, f"Sheet '{sname}' is HIDDEN (state={vis})")
    elif vis == "veryHidden":
        tag = "  ** VERY HIDDEN **"
        add_finding("CRITICAL", "Very Hidden Sheet", sname, f"Sheet '{sname}' is VERY HIDDEN (state={vis})")
    print(f"  [{vis:>10}] {sname}{tag}")

# ── 1b. Cell-by-cell scan across ALL sheets ──
print("\n" + "─" * 90)
print("SECTION 2: CELL-BY-CELL SCAN (formulas, values, comments)")
print("─" * 90)

internal_re = re.compile(r"internal", re.IGNORECASE)
ref_error_re = re.compile(r"#REF!", re.IGNORECASE)
w5_re = re.compile(r"\bW5\b|W\u2085|[Ww]aterfall\s*5", re.IGNORECASE)
irr_re = re.compile(r"24[\.\,]75|13[\.\,]95|\b2475\b")

cells_scanned = 0
for sname in all_sheets:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            cells_scanned += 1
            cell_ref = f"'{sname}'!{cell.coordinate}"
            val = cell.value
            val_str = str(val) if val is not None else ""

            # Check value / formula
            if val_str:
                if internal_re.search(val_str):
                    sev = "CRITICAL" if val_str.startswith("=") else "HIGH"
                    add_finding(sev, "Internal reference in cell", cell_ref,
                                f"Value/formula contains 'Internal': {val_str[:200]}")
                if ref_error_re.search(val_str):
                    add_finding("HIGH", "#REF! error", cell_ref,
                                f"Broken reference: {val_str[:200]}")
                if w5_re.search(val_str):
                    add_finding("CRITICAL", "W5/Waterfall 5 reference", cell_ref,
                                f"Internal waterfall ref: {val_str[:200]}")
                if irr_re.search(val_str):
                    add_finding("MEDIUM", "Suspicious IRR value", cell_ref,
                                f"Potential internal IRR: {val_str[:200]}")

            # Check comments
            if cell.comment:
                comment_text = str(cell.comment.text) if cell.comment.text else ""
                comment_author = str(cell.comment.author) if cell.comment.author else ""
                if comment_text or comment_author:
                    # Any comment is worth flagging in a public model
                    add_finding("MEDIUM", "Comment found", cell_ref,
                                f"Author: {comment_author} | Text: {comment_text[:300]}")
                    if internal_re.search(comment_text) or internal_re.search(comment_author):
                        add_finding("HIGH", "Internal ref in comment", cell_ref,
                                    f"Author: {comment_author} | Text: {comment_text[:300]}")

print(f"  Cells scanned: {cells_scanned:,}")
cell_findings = [f for f in findings if f["category"] in (
    "Internal reference in cell", "#REF! error", "W5/Waterfall 5 reference",
    "Suspicious IRR value", "Comment found", "Internal ref in comment")]
print(f"  Cell-level findings so far: {len(cell_findings)}")

# ── 1c. Defined names ──
print("\n" + "─" * 90)
print("SECTION 3: DEFINED NAMES / NAMED RANGES")
print("─" * 90)
if wb.defined_names:
    for dn in wb.defined_names.definedName:
        name = dn.name
        value = dn.attr_text  # the formula / range reference
        hidden = dn.hidden
        scope = dn.localSheetId
        print(f"  Name: {name}  |  Value: {value}  |  Hidden: {hidden}  |  Scope: {scope}")
        if internal_re.search(str(name)) or internal_re.search(str(value)):
            add_finding("CRITICAL", "Internal ref in defined name", name,
                        f"Name='{name}' Value='{value}' Hidden={hidden}")
        if ref_error_re.search(str(value)):
            add_finding("HIGH", "#REF! in defined name", name,
                        f"Broken ref in name: {value}")
        if hidden:
            add_finding("HIGH", "Hidden defined name", name,
                        f"Hidden name '{name}' = {value}")
        # Check for orphan references (sheet doesn't exist)
        # Extract sheet names from the value
        sheet_refs = re.findall(r"'?([^'!]+)'?!", str(value))
        for sr in sheet_refs:
            sr_clean = sr.strip("'")
            if sr_clean not in all_sheets:
                add_finding("CRITICAL", "Orphan defined name", name,
                            f"References non-existent sheet '{sr_clean}': {value}")
else:
    print("  No defined names found.")

# ── 1d. Document properties ──
print("\n" + "─" * 90)
print("SECTION 4: DOCUMENT PROPERTIES / METADATA")
print("─" * 90)
props = wb.properties
prop_fields = {
    "creator": props.creator,
    "title": props.title,
    "subject": props.subject,
    "description": props.description,
    "keywords": props.keywords,
    "category": props.category,
    "lastModifiedBy": props.lastModifiedBy,
    "created": props.created,
    "modified": props.modified,
    "company": getattr(props, "company", None),
    "manager": getattr(props, "manager", None),
    "version": props.version,
    "revision": props.revision,
    "identifier": props.identifier,
    "language": props.language,
    "contentStatus": props.contentStatus,
}
for k, v in prop_fields.items():
    if v:
        print(f"  {k}: {v}")
        v_str = str(v)
        if internal_re.search(v_str):
            add_finding("HIGH", "Internal ref in metadata", k, f"{k} = {v_str}")
        # Check for emails
        if re.search(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", v_str):
            add_finding("MEDIUM", "Email in metadata", k, f"{k} = {v_str}")
        # Any personal name in creator/lastModifiedBy is worth flagging
        if k in ("creator", "lastModifiedBy") and v_str.strip():
            add_finding("MEDIUM", "Personal info in metadata", k, f"{k} = {v_str}")

# Also check custom properties
try:
    custom = wb.custom_doc_props
    if custom:
        print("\n  Custom document properties:")
        for cp in custom:
            print(f"    {cp.name} = {cp.value}")
            if internal_re.search(str(cp.name)) or internal_re.search(str(cp.value)):
                add_finding("HIGH", "Internal ref in custom prop", cp.name,
                            f"{cp.name} = {cp.value}")
except Exception as e:
    print(f"  (Custom properties access: {e})")

wb.close()

# ─────────────────────────────────────────────────────────────────────
# 2. ZIP-level XML deep scan
# ─────────────────────────────────────────────────────────────────────
print("\n" + "─" * 90)
print("SECTION 5: RAW XML / ZIP DEEP SCAN")
print("─" * 90)

with zipfile.ZipFile(TARGET, "r") as zf:
    xml_files = zf.namelist()
    print(f"  Files inside ZIP: {len(xml_files)}")
    for fn in sorted(xml_files):
        print(f"    {fn}")

    print()

    # ── 5a. Scan each file for sensitive patterns ──
    for fn in xml_files:
        try:
            raw = zf.read(fn)
            # Try to decode as text; skip binary images etc.
            try:
                text = raw.decode("utf-8")
            except UnicodeDecodeError:
                try:
                    text = raw.decode("latin-1")
                except Exception:
                    continue

            for pattern_name, pattern_re in SENSITIVE_PATTERNS.items():
                for m in pattern_re.finditer(text):
                    start = max(0, m.start() - 60)
                    end = min(len(text), m.end() + 60)
                    context = text[start:end].replace("\n", " ").replace("\r", "")

                    # Determine severity
                    if pattern_name in ("Internal keyword",):
                        sev = "CRITICAL"
                    elif pattern_name in ("W5 reference", "W5 unicode subscript", "Waterfall 5"):
                        sev = "CRITICAL"
                    elif pattern_name in ("Internal IRR 24.75", "Internal IRR 2475", "Internal MC IRR 13.95"):
                        sev = "HIGH"
                    elif pattern_name in ("Email address", "File path (Windows)", "File path (UNC)", "File path (Unix)"):
                        sev = "MEDIUM"
                    elif pattern_name == "Hidden sheet attr":
                        sev = "HIGH"
                    else:
                        sev = "MEDIUM"

                    add_finding(sev, f"XML: {pattern_name}", fn,
                                f"Match: '{m.group()}' | Context: ...{context}...")
        except Exception as e:
            print(f"  Error reading {fn}: {e}")

    # ── 5b. Full content of docProps/core.xml and docProps/app.xml ──
    for docprop in ("docProps/core.xml", "docProps/app.xml"):
        if docprop in xml_files:
            print(f"\n  ── Full content: {docprop} ──")
            try:
                content = zf.read(docprop).decode("utf-8")
                print(content)
            except Exception as e:
                print(f"  Error: {e}")
        else:
            print(f"\n  {docprop}: NOT FOUND")

    # ── 5c. calcChain.xml analysis ──
    print(f"\n  ── calcChain.xml analysis ──")
    if "xl/calcChain.xml" in xml_files:
        cc_raw = zf.read("xl/calcChain.xml").decode("utf-8")
        print(f"  calcChain.xml size: {len(cc_raw):,} bytes")
        # Parse and look for sheet references
        try:
            cc_tree = ET.fromstring(cc_raw)
            ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            entries = cc_tree.findall(".//x:c", ns)
            if not entries:
                # try without namespace
                entries = cc_tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c")
            sheet_ids_in_calc = set()
            for entry in entries:
                si = entry.get("i")
                if si:
                    sheet_ids_in_calc.add(si)
            print(f"  Unique sheet IDs referenced in calcChain: {sorted(sheet_ids_in_calc)}")
            print(f"  Total calc entries: {len(entries)}")

            # Cross-reference with workbook.xml to see if any sheet IDs are orphaned
            if "xl/workbook.xml" in xml_files:
                wb_raw = zf.read("xl/workbook.xml").decode("utf-8")
                wb_tree = ET.fromstring(wb_raw)
                # Get all sheet elements
                ns2 = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
                sheets_el = wb_tree.findall(".//x:sheet", ns2)
                if not sheets_el:
                    sheets_el = wb_tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet")
                wb_sheet_ids = set()
                for s in sheets_el:
                    sid = s.get("sheetId")
                    if sid:
                        wb_sheet_ids.add(sid)
                    sname_xml = s.get("name")
                    state = s.get("state", "visible")
                    print(f"    Workbook sheet: name='{sname_xml}' sheetId={sid} state={state}")
                    if state in ("hidden", "veryHidden"):
                        add_finding("HIGH" if state == "hidden" else "CRITICAL",
                                    "Hidden sheet in workbook XML", sname_xml,
                                    f"Sheet '{sname_xml}' state={state} sheetId={sid}")

                orphan_ids = sheet_ids_in_calc - wb_sheet_ids
                if orphan_ids:
                    print(f"\n  *** ORPHAN SHEET IDs in calcChain: {sorted(orphan_ids)} ***")
                    add_finding("CRITICAL", "Orphan calcChain references", "xl/calcChain.xml",
                                f"calcChain references sheet IDs not in workbook: {sorted(orphan_ids)}")
                else:
                    print(f"  No orphan sheet IDs in calcChain.")
        except Exception as e:
            print(f"  Error parsing calcChain: {e}")
    else:
        print("  calcChain.xml: NOT FOUND")

    # ── 5d. workbook.xml — defined names at XML level ──
    print(f"\n  ── workbook.xml defined names (XML level) ──")
    if "xl/workbook.xml" in xml_files:
        wb_raw = zf.read("xl/workbook.xml").decode("utf-8")
        # Print definedNames section
        dn_start = wb_raw.find("<definedNames")
        if dn_start >= 0:
            dn_end = wb_raw.find("</definedNames>", dn_start)
            if dn_end >= 0:
                dn_section = wb_raw[dn_start:dn_end + len("</definedNames>")]
                print(f"  definedNames section ({len(dn_section)} chars):")
                print(dn_section[:3000])
                if len(dn_section) > 3000:
                    print(f"  ... (truncated, total {len(dn_section)} chars)")
            else:
                print("  definedNames opening tag found but no closing tag")
        else:
            print("  No definedNames section found in workbook.xml")

    # ── 5e. Check for externalLinks ──
    print(f"\n  ── External Links ──")
    ext_links = [f for f in xml_files if "externalLink" in f.lower() or "externalReference" in f.lower()]
    if ext_links:
        for el in ext_links:
            print(f"  FOUND external link file: {el}")
            try:
                el_content = zf.read(el).decode("utf-8")
                print(f"  Content ({len(el_content)} chars):")
                print(el_content[:2000])
                add_finding("HIGH", "External link", el,
                            f"External link file found: {el}")
            except Exception as e:
                print(f"  Error: {e}")
    else:
        print("  No external link files found.")

    # ── 5f. Check for printerSettings, vbaProject, or other unusual files ──
    print(f"\n  ── Unusual / Noteworthy Files ──")
    suspicious_patterns = ["vbaProject", "macro", "activeX", "externalLink", "connections", "customXml"]
    for fn in xml_files:
        for sp in suspicious_patterns:
            if sp.lower() in fn.lower():
                print(f"  SUSPICIOUS FILE: {fn}")
                add_finding("MEDIUM", "Suspicious file in ZIP", fn,
                            f"File '{fn}' matches pattern '{sp}'")


# ─────────────────────────────────────────────────────────────────────
# 3. Hidden sheets content dump
# ─────────────────────────────────────────────────────────────────────
print("\n" + "─" * 90)
print("SECTION 6: HIDDEN SHEET CONTENT DUMP")
print("─" * 90)

wb2 = openpyxl.load_workbook(TARGET, data_only=False)
for sname in wb2.sheetnames:
    ws = wb2[sname]
    if ws.sheet_state in ("hidden", "veryHidden"):
        print(f"\n  ── Hidden sheet: '{sname}' (state={ws.sheet_state}) ──")
        row_count = 0
        for row in ws.iter_rows(values_only=False):
            non_empty = [(c.coordinate, c.value) for c in row if c.value is not None]
            if non_empty:
                for coord, val in non_empty:
                    print(f"    {coord}: {str(val)[:300]}")
                row_count += 1
        if row_count == 0:
            print("    (empty sheet)")
wb2.close()


# ─────────────────────────────────────────────────────────────────────
# FINAL REPORT
# ─────────────────────────────────────────────────────────────────────
print("\n" + "=" * 90)
print("FINAL AUDIT REPORT — PRIORITY 1: INTERNAL DATA LEAKAGE")
print("=" * 90)

# Deduplicate findings
seen = set()
unique_findings = []
for f in findings:
    key = (f["severity"], f["category"], f["location"], f["detail"][:150])
    if key not in seen:
        seen.add(key)
        unique_findings.append(f)

unique_findings.sort(key=lambda f: (severity_order(f["severity"]), f["category"], f["location"]))

# Summary counts
counts = defaultdict(int)
for f in unique_findings:
    counts[f["severity"]] += 1

print(f"\nTotal unique findings: {len(unique_findings)}")
print(f"  CRITICAL: {counts.get('CRITICAL', 0)}")
print(f"  HIGH:     {counts.get('HIGH', 0)}")
print(f"  MEDIUM:   {counts.get('MEDIUM', 0)}")
print(f"  LOW:      {counts.get('LOW', 0)}")

print()
for f in unique_findings:
    sev = f["severity"]
    marker = {"CRITICAL": "!!! ", "HIGH": "!!  ", "MEDIUM": "!   ", "LOW": "    "}.get(sev, "    ")
    print(f"[{sev:>8}] {marker}{f['category']}")
    print(f"           Location: {f['location']}")
    print(f"           Detail:   {f['detail'][:500]}")
    print()

if counts.get("CRITICAL", 0) > 0:
    print("*** VERDICT: CRITICAL LEAKAGE DETECTED — DO NOT DISTRIBUTE PUBLIC MODEL ***")
elif counts.get("HIGH", 0) > 0:
    print("*** VERDICT: HIGH-RISK ISSUES FOUND — REVIEW BEFORE DISTRIBUTION ***")
elif counts.get("MEDIUM", 0) > 0:
    print("*** VERDICT: MEDIUM-RISK ISSUES — RECOMMEND CLEANUP ***")
else:
    print("*** VERDICT: NO SIGNIFICANT LEAKAGE DETECTED ***")

print("\n" + "=" * 90)
print("END OF PRIORITY 1 AUDIT")
print("=" * 90)
