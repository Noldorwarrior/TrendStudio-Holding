"""
tests/test_30_sterilize.py — Phase 0 tests for xlsx sterilization.

Verifies:
  - No Internal leakage (F-001, F-002, F-009, F-010, F-011)
  - No absPath in workbook.xml (F-003)
  - Clean metadata in core.xml (F-020, F-021, F-022)
  - No dangling = sign (F-036)
  - No backup/bak files in Investor_Package (F-039)
  - verify.py uses dynamic INPUT_FILES count (F-033)

R-items tested: R-001, R-002, R-003, R-004, R-024
"""
from __future__ import annotations

import re
import sys
import zipfile
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).parent.parent.parent
PIPELINE_ROOT = Path(__file__).parent.parent

# Both Public xlsx locations (sterilized in-place)
PUBLIC_XLSX_FILES = [
    REPO_ROOT / "investor_model_v1.0_Public.xlsx",
    REPO_ROOT / "Investor_Package" / "investor_model_v1.0_Public.xlsx",
]


def _get_sterilized_files() -> list[Path]:
    """Return existing sterilized Public xlsx files."""
    return [p for p in PUBLIC_XLSX_FILES if p.exists()]


def _read_zip_entry(xlsx_path: Path, entry: str) -> str | None:
    """Read a specific entry from xlsx zip as UTF-8 string."""
    with zipfile.ZipFile(str(xlsx_path), "r") as z:
        if entry in z.namelist():
            return z.read(entry).decode("utf-8", errors="replace")
    return None


def _count_internal_in_xlsx(xlsx_path: Path, *, allow_irr: bool = True) -> int:
    """Count 'Internal' occurrences across all XML entries, excluding IRR."""
    count = 0
    with zipfile.ZipFile(str(xlsx_path), "r") as z:
        for name in z.namelist():
            data = z.read(name).decode("utf-8", errors="replace")
            for m in re.finditer(r"[Ii]nternal", data):
                if allow_irr:
                    end_far = min(len(data), m.end() + 200)
                    context = data[m.start() : end_far]
                    if "Rate of Return" in context:
                        continue
                count += 1
    return count


# ─── F-002 / R-002: No Internal leakage ─────────────────────────────────


@pytest.mark.parametrize("xlsx_path", _get_sterilized_files(), ids=lambda p: p.name)
class TestNoInternalLeak:
    """F-002 / R-002: grep -i 'internal' = 0 (except IRR)."""

    def test_zero_internal_occurrences(self, xlsx_path: Path):
        count = _count_internal_in_xlsx(xlsx_path)
        assert count == 0, (
            f"{xlsx_path.name}: found {count} 'Internal' occurrence(s) "
            f"(excluding 'Internal Rate of Return')"
        )

    def test_irr_preserved(self, xlsx_path: Path):
        """Ensure 'Internal Rate of Return' in Glossary is preserved."""
        import openpyxl

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        if "37_Glossary" in wb.sheetnames:
            ws = wb["37_Glossary"]
            d19 = ws["D19"].value
            assert d19 is not None, "37_Glossary!D19 should not be empty"
            assert "Rate of Return" in str(d19), (
                "37_Glossary!D19 should contain 'Internal Rate of Return'"
            )
        wb.close()


# ─── F-003 / R-003: No absPath ──────────────────────────────────────────


@pytest.mark.parametrize("xlsx_path", _get_sterilized_files(), ids=lambda p: p.name)
class TestNoAbsPath:
    """F-003 / R-003: No x15ac:absPath in workbook.xml."""

    def test_no_abspath(self, xlsx_path: Path):
        wb_xml = _read_zip_entry(xlsx_path, "xl/workbook.xml")
        assert wb_xml is not None, "xl/workbook.xml not found"
        assert "absPath" not in wb_xml, "absPath still present in workbook.xml"

    def test_no_x15ac_namespace(self, xlsx_path: Path):
        wb_xml = _read_zip_entry(xlsx_path, "xl/workbook.xml")
        assert wb_xml is not None
        assert "x15ac" not in wb_xml, "x15ac namespace still present"


# ─── F-020/021/022 / R-004: Clean metadata ──────────────────────────────


@pytest.mark.parametrize("xlsx_path", _get_sterilized_files(), ids=lambda p: p.name)
class TestCleanMetadata:
    """F-020/021/022 / R-004: docProps/core.xml is clean."""

    def test_author_is_trendstudio(self, xlsx_path: Path):
        core = _read_zip_entry(xlsx_path, "docProps/core.xml")
        assert core is not None
        assert "<dc:creator" in core
        m = re.search(r"<dc:creator[^>]*>(.*?)</dc:creator>", core)
        assert m, "dc:creator not found"
        assert m.group(1) == "TrendStudio", f"Author is '{m.group(1)}', expected 'TrendStudio'"

    def test_description_empty(self, xlsx_path: Path):
        """F-021: description should not leak anchor value."""
        core = _read_zip_entry(xlsx_path, "docProps/core.xml")
        assert core is not None
        m = re.search(r"<dc:description[^>]*>(.*?)</dc:description>", core)
        if m:
            assert m.group(1).strip() == "", (
                f"Description should be empty, got: '{m.group(1)}'"
            )

    def test_no_l3_in_keywords(self, xlsx_path: Path):
        """F-022: keywords should not contain 'L3'."""
        core = _read_zip_entry(xlsx_path, "docProps/core.xml")
        assert core is not None
        assert "L3" not in core, "Keywords still contain 'L3'"

    def test_last_modified_by_not_a(self, xlsx_path: Path):
        """F-020: lastModifiedBy should not be 'a'."""
        core = _read_zip_entry(xlsx_path, "docProps/core.xml")
        assert core is not None
        m = re.search(r"<cp:lastModifiedBy>(.*?)</cp:lastModifiedBy>", core)
        if m:
            assert m.group(1) != "a", "lastModifiedBy is still 'a'"
            assert m.group(1) == "TrendStudio", (
                f"lastModifiedBy is '{m.group(1)}', expected 'TrendStudio'"
            )

    def test_no_anchor_in_metadata(self, xlsx_path: Path):
        """Ensure anchor value 3000 is not in any metadata field."""
        core = _read_zip_entry(xlsx_path, "docProps/core.xml")
        assert core is not None
        assert "3 000" not in core, "Anchor '3 000' leaked in core.xml"
        assert "3000" not in core, "Anchor '3000' leaked in core.xml"


# ─── F-001: Specific cell checks ────────────────────────────────────────


@pytest.mark.parametrize("xlsx_path", _get_sterilized_files(), ids=lambda p: p.name)
class TestSpecificCellsScrubbed:
    """Verify specific cells are scrubbed of Internal references."""

    def test_f001_investor_returns_b49(self, xlsx_path: Path):
        """F-001: 24_Investor_Returns B49 should not reference Internal W₅."""
        import openpyxl

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["24_Investor_Returns"]
        val = str(ws["B49"].value or "")
        wb.close()
        assert "Internal" not in val, f"B49 still contains 'Internal': {val[:80]}"
        assert "W₅" not in val, f"B49 still contains 'W₅': {val[:80]}"

    def test_f009_changelog_c30(self, xlsx_path: Path):
        """F-009: 03_Change_Log C30 should not mention Internal."""
        import openpyxl

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["03_Change_Log"]
        val = str(ws["C30"].value or "")
        wb.close()
        assert "Internal" not in val, f"C30 still contains 'Internal': {val}"

    def test_f010_valuation_b46(self, xlsx_path: Path):
        """F-010: 23_Valuation_Multiples B46 should not have (internal)."""
        import openpyxl

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["23_Valuation_Multiples"]
        val = str(ws["B46"].value or "")
        wb.close()
        assert "internal" not in val.lower(), f"B46 still contains 'internal': {val}"

    def test_f011_comps_no_internal(self, xlsx_path: Path):
        """F-011: 32_Comparable_Transactions E12, E13, J12 clean."""
        import openpyxl

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["32_Comparable_Transactions"]
        for coord in ["E12", "E13", "J12"]:
            val = str(ws[coord].value or "")
            assert "internal" not in val.lower(), (
                f"{coord} still contains 'internal': {val}"
            )
        wb.close()

    def test_f036_no_dangling_equals(self, xlsx_path: Path):
        """F-036: 17_Deal_Structures B31 should not be bare '='."""
        import openpyxl

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["17_Deal_Structures"]
        val = str(ws["B31"].value or "")
        wb.close()
        assert val.strip() != "=", f"B31 is still a bare '='"


# ─── F-039: No backup files ─────────────────────────────────────────────


class TestNoBackupFiles:
    """F-039: No .bak or _backup files in Investor_Package."""

    def test_no_bak_files(self):
        ip = REPO_ROOT / "Investor_Package"
        if not ip.exists():
            pytest.skip("Investor_Package not found")
        bak_files = list(ip.glob("*.bak*"))
        assert len(bak_files) == 0, (
            f"Found {len(bak_files)} .bak file(s): {[f.name for f in bak_files]}"
        )

    def test_no_backup_files(self):
        ip = REPO_ROOT / "Investor_Package"
        if not ip.exists():
            pytest.skip("Investor_Package not found")
        backup_files = list(ip.glob("*backup*"))
        assert len(backup_files) == 0, (
            f"Found {len(backup_files)} backup file(s): {[f.name for f in backup_files]}"
        )

    def test_no_fuse_hidden(self):
        ip = REPO_ROOT / "Investor_Package"
        if not ip.exists():
            pytest.skip("Investor_Package not found")
        fuse = list(ip.glob(".fuse_hidden*"))
        assert len(fuse) == 0, (
            f"Found {len(fuse)} FUSE hidden file(s)"
        )

    def test_no_lock_files(self):
        ip = REPO_ROOT / "Investor_Package"
        if not ip.exists():
            pytest.skip("Investor_Package not found")
        locks = list(ip.glob(".~lock.*"))
        assert len(locks) == 0, (
            f"Found {len(locks)} lock file(s)"
        )


# ─── F-033 / R-024: verify.py dynamic count ─────────────────────────────


class TestVerifyDynamic:
    """F-033 / R-024: verify.py should use len(INPUT_FILES), not hardcoded 14."""

    def test_no_hardcoded_14_in_verify(self):
        verify_path = PIPELINE_ROOT / "scripts" / "verify.py"
        if not verify_path.exists():
            pytest.skip("verify.py not found")
        content = verify_path.read_text(encoding="utf-8")
        # Should not have == 14 or "/14" in completeness check context
        assert "== 14" not in content, (
            "verify.py still has hardcoded '== 14' — should use len(INPUT_FILES)"
        )

    def test_uses_len_input_files(self):
        verify_path = PIPELINE_ROOT / "scripts" / "verify.py"
        if not verify_path.exists():
            pytest.skip("verify.py not found")
        content = verify_path.read_text(encoding="utf-8")
        assert "len(INPUT_FILES)" in content or "len(used)" in content, (
            "verify.py should use dynamic length from INPUT_FILES"
        )
