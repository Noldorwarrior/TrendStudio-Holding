"""
PATCH v1.0.1 · Stage 1 — 02_Assumptions + 19_Waterfall
========================================================
Цель: расширить waterfall-архитектуру с 3 до 4 вариантов.
      W₃ становится DEFAULT (вместо W₁).
      W₄ = 1× Liq Pref + 12% preferred return × 5y + 65/35 + exit 6×.

Затрагивает:
  · 02_Assumptions — добавить параметры W₄ и exit multiple 6×
  · 19_Waterfall   — перестройка листа: 4 варианта + 4-строчная
                     comparison-table; удалить грубый IRR,
                     заменить на MoIC + cross-ref на лист 24.

Запуск:  python3 patch_v101_stage1_assumptions_waterfall.py
         (работает на Public и Internal копиях подряд)
"""

import copy
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
FILES = [
    "investor_model_v1.0_Public.xlsx",
    "investor_model_v1.0_Internal.xlsx",
]

# ═══════════════════════════════════════════════════════════════════════════
# СТИЛЕВЫЕ КОНСТАНТЫ (совпадают с build_A8 / build_A1)
# ═══════════════════════════════════════════════════════════════════════════
NAVY      = "1F3864"
BLUE      = "0070C0"
LIGHTBLUE = "D9E1F2"
LIGHTGRAY = "F2F2F2"
DARKGREEN = "548235"
AMBER     = "C65911"

FONT_TITLE   = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=NAVY)
FONT_HEAD    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BODY    = Font(name="Calibri", size=10, color="000000")
FONT_BODY_B  = Font(name="Calibri", size=10, bold=True, color="000000")
FONT_DEFAULT = Font(name="Calibri", size=10, bold=True, color=DARKGREEN)
FONT_NOTE    = Font(name="Calibri", size=9, italic=True, color="595959")

FILL_TITLE    = PatternFill(start_color=NAVY,      end_color=NAVY,      fill_type="solid")
FILL_SECTION  = PatternFill(start_color=LIGHTBLUE, end_color=LIGHTBLUE, fill_type="solid")
FILL_HEAD     = PatternFill(start_color=BLUE,      end_color=BLUE,      fill_type="solid")
FILL_TOTAL    = PatternFill(start_color=LIGHTGRAY, end_color=LIGHTGRAY, fill_type="solid")
FILL_DEFAULT  = PatternFill(start_color="E2EFDA",  end_color="E2EFDA",  fill_type="solid")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)


# ═══════════════════════════════════════════════════════════════════════════
# ПАРАМЕТРЫ W₄
# ═══════════════════════════════════════════════════════════════════════════
NDP          = 3000.0   # млн ₽
INVEST       = 1250.0   # T₁ Legacy
PREF_RET     = 0.12     # 12% preferred return
YEARS        = 5
SPLIT_INV    = 0.65
SPLIT_PROD   = 0.35
EXIT_MULT_W4 = 6        # × EBITDA (vs базовых 5× для W₁-W₃)
LP_EQUITY    = 0.25     # 25% LP equity (Internal-only в next stage)

# W₄ distribution
W4_S1 = min(INVEST, NDP)                                     # 1250
W4_S2 = min(INVEST * PREF_RET * YEARS, NDP - W4_S1)          # 750
W4_REM = NDP - W4_S1 - W4_S2                                 # 1000
W4_S3_INV  = W4_REM * SPLIT_INV                              # 650
W4_S3_PROD = W4_REM * SPLIT_PROD                             # 350
W4_INV  = W4_S1 + W4_S2 + W4_S3_INV                          # 2650
W4_PROD = W4_S3_PROD                                         # 350
assert abs(W4_INV + W4_PROD - NDP) < 0.01


# ═══════════════════════════════════════════════════════════════════════════
# ПАТЧ 1 · 02_Assumptions — добавить 8 строк W₄ после row 45
# ═══════════════════════════════════════════════════════════════════════════
def patch_assumptions(wb):
    ws = wb["02_Assumptions"]

    # Вставляем 8 строк после row 45 (после W₃ participation rate)
    ws.insert_rows(46, amount=8)

    w4_rows = [
        (46, "31a", "W₄ 1× liquidation preference",        1.0,    "×",    "Возврат инвестиций первыми"),
        (47, "31b", "W₄ Preferred return rate",            0.12,   "%",    "12% annual × 5y cumulative"),
        (48, "31c", "W₄ Preferred return duration",        5,      "years","Накопление купона до exit"),
        (49, "31d", "W₄ Carry split investor",             0.65,   "%",    "65% инвестор после возврата+купона"),
        (50, "31e", "W₄ Carry split producer",             0.35,   "%",    "35% продюсер после возврата+купона"),
        (51, "31f", "W₄ Exit multiple (EV/EBITDA)",        6.0,    "×",    "Premium к базовым 5×"),
        (52, "31g", "W₄ LP equity allocation",             0.25,   "%",    "25% доля LP (Internal only)"),
        (53, "31h", "Target IRR W₄ Base",                  0.198,  "%",    "~19.8% (расчёт в 24_Investor_Returns)"),
    ]
    for row_idx, n, name, val, unit, note in w4_rows:
        ws.cell(row=row_idx, column=2, value=n).font = FONT_BODY
        ws.cell(row=row_idx, column=3, value=name).font = FONT_BODY
        c = ws.cell(row=row_idx, column=4, value=val)
        c.font = FONT_BODY
        if unit == "%":
            c.number_format = "0.00%"
        elif unit == "×":
            c.number_format = "0.0\"×\""
        elif unit == "млн ₽":
            c.number_format = "#,##0"
        else:
            c.number_format = "0"
        ws.cell(row=row_idx, column=5, value=unit).font = FONT_BODY
        ws.cell(row=row_idx, column=6, value=note).font = FONT_NOTE
        for col in range(2, 7):
            ws.cell(row=row_idx, column=col).border = BORDER
            ws.cell(row=row_idx, column=col).alignment = LEFT

    # Обновляем заголовок блока E (row 38)
    ws["B38"].value = "E. РАСПРЕДЕЛЕНИЕ ПРИБЫЛИ (Waterfall — 4 варианта)"

    # Обновляем footer (был "ВСЕГО ДОПУЩЕНИЙ: 80", сдвинулся на 8 строк)
    # Находим новую координату
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "ВСЕГО ДОПУЩЕНИЙ" in str(cell.value):
                cell.value = ("ВСЕГО ДОПУЩЕНИЙ: 88  ·  Блоков: 13  ·  "
                              "Якорь EBITDA 2026–2028 = 3 000 млн ₽  ·  "
                              "Waterfall: 4 варианта (default W₃)")
                break

    # Обновляем N.05 (Waterfall split default) — теперь W₃
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "N.05":
                target = ws.cell(row=cell.row, column=3)
                target.value = "Waterfall split (default W₃ Liq Pref)"
                e_cell = ws.cell(row=cell.row, column=4)
                e_cell.value = "1× + 8%"
                break

    # N.06 — anchor (legacy) — не трогаем (3000 остаётся)
    # Добавлять новый параграф для N.09 W₄ пока не буду — стадия 3

    print("  ✓ 02_Assumptions: вставлено 8 строк W₄ (46-53), обновлён E/footer/N.05")


# ═══════════════════════════════════════════════════════════════════════════
# ПАТЧ 2 · 19_Waterfall — полная перестройка
# ═══════════════════════════════════════════════════════════════════════════
def rebuild_waterfall_sheet(wb):
    """
    Перестраиваем лист 19 с нуля:
    - 4 waterfall blocks (W₁, W₂, W₃★default, W₄)
    - comparison table 4 строки
    - убираем грубые IRR, оставляем MoIC + cross-ref
    """
    # Удаляем старый лист и создаём заново с тем же именем и индексом
    sheet_idx = wb.sheetnames.index("19_Waterfall")
    old = wb["19_Waterfall"]

    # Сохраняем параметры страницы для копирования
    page_setup = old.page_setup
    page_margins = old.page_margins
    print_options = old.print_options

    wb.remove(old)
    ws = wb.create_sheet("19_Waterfall", sheet_idx)

    # Устанавливаем ширины колонок
    widths = [3, 4, 38, 14, 14, 2, 14, 2, 32, 16]  # A B C D E F G H I J
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("B2:J2")
    ws["B2"] = ("19  ·  DISTRIBUTION WATERFALL  ·  4 варианта распределения NDP 3 000 млн ₽ "
                "·  T₁ Legacy (invest 1 250 + producer 600)")
    ws["B2"].font = FONT_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = CENTER
    ws.row_dimensions[2].height = 26

    ws.merge_cells("B3:J3")
    ws["B3"] = ("W₁ — Hurdle 60/40 → 50/50.  W₂ — Pro-rata по вкладу.  "
                "W₃ — 1× Liq Pref + 8% coupon + 60/40 (★ DEFAULT с v1.0.1).  "
                "W₄ — 1× Liq Pref + 12% preferred + 65/35 + exit 6× (premium option).")
    ws["B3"].font = FONT_NOTE
    ws["B3"].alignment = LEFT
    ws.row_dimensions[3].height = 28

    row = 5

    def write_header_block(row, label, subtitle, fill=FILL_SECTION, font=FONT_SECTION):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        c = ws.cell(row=row, column=2, value=f"{label}  ·  {subtitle}")
        c.font = font
        c.fill = fill
        c.alignment = LEFT
        ws.row_dimensions[row].height = 22

    def write_columns_header(row):
        headers = [("#",3), ("Stage",3), ("Размер стадии",4),
                   ("Инвестор",5), (None,6), ("Продюсер",7), (None,8),
                   ("Правило",9)]
        for text, col in [("#",2),("Stage",3),("Размер стадии",4),
                          ("Инвестор",5),("Продюсер",7),("Правило",9)]:
            c = ws.cell(row=row, column=col, value=text)
            c.font = FONT_HEAD
            c.fill = FILL_HEAD
            c.alignment = CENTER
            c.border = BORDER

    def write_data_row(row, n, stage, size, inv, prod, rule, bold=False):
        font = FONT_BODY_B if bold else FONT_BODY
        fill = FILL_TOTAL if bold else None
        for col, val, fmt, align in [
            (2, n,     None,        CENTER),
            (3, stage, None,        LEFT),
            (4, size,  "#,##0.0",   RIGHT),
            (5, inv,   "#,##0.0",   RIGHT),
            (7, prod,  "#,##0.0",   RIGHT),
            (9, rule,  None,        LEFT),
        ]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = font
            if fill: c.fill = fill
            c.alignment = align
            c.border = BORDER
            if fmt and val is not None: c.number_format = fmt

    # ──────── W₁ ────────
    write_header_block(row, "W₁", "HURDLE SPLIT  ·  60 / 40 до break-even, затем 50 / 50")
    row += 1
    write_columns_header(row)
    row += 1
    write_data_row(row, 1, "Stage 1: 60/40 до recoupment T₁ (1 250)", 2083.3, 1250, 833.3,
                   "60% до тех пор пока инвестор не получит 1 250")
    row += 1
    write_data_row(row, 2, "Stage 2: 50/50 на остаток", 916.7, 458.3, 458.3,
                   "50/50 на оставшиеся 917 млн после recoupment")
    row += 1
    write_data_row(row, "Σ", "TOTAL W₁", 3000.0, 1708.3, 1291.7,
                   "Инвестор 56.9% / Продюсер 43.1%", bold=True)
    row += 2

    # ──────── W₂ ────────
    write_header_block(row, "W₂", "PRO-RATA  ·  По доле финансового вклада (67.6% / 32.4%)")
    row += 1
    write_columns_header(row)
    row += 1
    write_data_row(row, 1, "Pro-rata investor: 1 250 / 1 850 = 67.57%", 2027.0, 2027.0, None,
                   "Доля инвестора: 67.57%")
    row += 1
    write_data_row(row, 2, "Pro-rata producer: 600 / 1 850 = 32.43%", 973.0, None, 973.0,
                   "Доля продюсера: 32.43%")
    row += 1
    write_data_row(row, "Σ", "TOTAL W₂", 3000.0, 2027.0, 973.0,
                   "Pro-rata даёт больше инвестору (+318 vs W₁)", bold=True)
    row += 2

    # ──────── W₃ ★ DEFAULT ────────
    write_header_block(row, "W₃", "LIQ PREF  ·  1× + 8% coupon + 60/40 на остаток  (★ DEFAULT с v1.0.1)",
                       fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"))
    row += 1
    write_columns_header(row)
    row += 1
    write_data_row(row, 1, "Stage 1: 1× Liquidation Preference (investor)", 1250.0, 1250.0, 0,
                   "Investor recoups 1 × principal = 1 250")
    row += 1
    write_data_row(row, 2, "Stage 2: 8% coupon cumulative × 5 лет", 500.0, 500.0, 0,
                   "Preferred return: 1 250 × 8% × 5y = 500")
    row += 1
    write_data_row(row, 3, "Stage 3: 60/40 на остаток", 1250.0, 750.0, 500.0,
                   "Carried interest: 60% investor / 40% producer")
    row += 1
    write_data_row(row, "Σ", "TOTAL W₃ ★", 3000.0, 2500.0, 500.0,
                   "Investor 83.3% / Producer 16.7% — BASE CASE", bold=True)
    # зелёная подсветка на total
    for col in range(2, 10):
        ws.cell(row=row, column=col).fill = FILL_DEFAULT
    row += 2

    # ──────── W₄ NEW ────────
    write_header_block(row, "W₄", "PREMIUM  ·  1× + 12% preferred × 5y + 65/35 + exit 6× (premium negotiation)",
                       fill=PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"))
    row += 1
    write_columns_header(row)
    row += 1
    write_data_row(row, 1, "Stage 1: 1× Liquidation Preference (investor)", W4_S1, W4_S1, 0,
                   "Investor recoups 1 × principal = 1 250")
    row += 1
    write_data_row(row, 2, "Stage 2: 12% preferred return × 5 лет", W4_S2, W4_S2, 0,
                   "Accrued preferred: 1 250 × 12% × 5y = 750")
    row += 1
    write_data_row(row, 3, "Stage 3: 65/35 carry на остаток (1 000)", W4_REM, W4_S3_INV, W4_S3_PROD,
                   "Carry: 65% investor / 35% producer")
    row += 1
    write_data_row(row, "Σ", "TOTAL W₄", NDP, W4_INV, W4_PROD,
                   f"Investor {W4_INV/NDP*100:.1f}% / Producer {W4_PROD/NDP*100:.1f}% — PREMIUM", bold=True)
    row += 2

    # ──────── COMPARISON TABLE (4 rows) ────────
    write_header_block(row, "IV", "COMPARISON TABLE  ·  4 варианта (IRR — см. 24_Investor_Returns)",
                       fill=FILL_SECTION)
    row += 1

    # Headers
    comp_headers = [
        (2, "#"),
        (3, "Вариант"),
        (4, "Инвестор, млн ₽"),
        (5, "Инв %"),
        (6, "Продюсер, млн ₽"),
        (7, "Прод %"),
        (8, "MoIC"),
        (9, "Комментарий"),
    ]
    for col, text in comp_headers:
        c = ws.cell(row=row, column=col, value=text)
        c.font = FONT_HEAD
        c.fill = FILL_HEAD
        c.alignment = CENTER
        c.border = BORDER
    row += 1

    comp_rows = [
        ("W₁", "Hurdle 60/40 → 50/50",           1708.3, 56.9, 1291.7, 43.1, 1.367, "Producer-friendly, weakest for LP"),
        ("W₂", "Pro-rata по вкладу",              2027.0, 67.6, 973.0,  32.4, 1.622, "Fair by capital, no hurdle"),
        ("W₃", "1× Liq Pref + 8% + 60/40",        2500.0, 83.3, 500.0,  16.7, 2.000, "★ DEFAULT — LP downside protection"),
        ("W₄", "1× Liq Pref + 12% + 65/35",       W4_INV, W4_INV/NDP*100, W4_PROD, W4_PROD/NDP*100, W4_INV/INVEST, "Premium negotiation option"),
    ]
    for n, name, inv, invp, prod, prodp, moic, comment in comp_rows:
        is_default = (n == "W₃")
        vals = [(2,n),(3,name),(4,inv),(5,invp),(6,prod),(7,prodp),(8,moic),(9,comment)]
        for col, v in vals:
            c = ws.cell(row=row, column=col, value=v)
            c.font = FONT_DEFAULT if is_default else FONT_BODY
            c.border = BORDER
            if col in (4,6): c.number_format = "#,##0.0"
            elif col in (5,7): c.number_format = "0.0\"%\""
            elif col == 8: c.number_format = "0.000\"×\""
            c.alignment = CENTER if col in (2,5,7,8) else (RIGHT if col in (4,6) else LEFT)
            if is_default: c.fill = FILL_DEFAULT
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    footer = (f"✓ W₁ Σ = 3 000.0  ·  ✓ W₂ Σ = 3 000.0  ·  "
              f"✓ W₃ Σ = 3 000.0 (★ DEFAULT)  ·  ✓ W₄ Σ = {NDP:,.1f}  |  "
              f"IRR cash-on-cash: см. 24_Investor_Returns Matrix 3×4")
    c = ws.cell(row=row, column=2, value=footer)
    c.font = FONT_NOTE
    c.alignment = CENTER
    c.fill = FILL_TOTAL

    # Восстанавливаем параметры страницы
    try:
        ws.page_setup.orientation = page_setup.orientation
        ws.page_setup.paperSize = page_setup.paperSize
        ws.page_margins = copy.copy(page_margins)
    except Exception:
        pass

    print(f"  ✓ 19_Waterfall: перестроен, 4 варианта, default=W₃, rows={row}")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    for fname in FILES:
        path = ROOT / fname
        if not path.exists():
            print(f"⚠ {fname} не найден, пропускаю")
            continue

        # Бэкап
        backup = ROOT / f"{path.stem}_pre_v101_backup.xlsx"
        if not backup.exists():
            shutil.copy(path, backup)
            print(f"📦 Backup: {backup.name}")

        print(f"\n🔧 Патчу {fname}")
        wb = load_workbook(path)

        patch_assumptions(wb)
        rebuild_waterfall_sheet(wb)

        wb.save(path)
        print(f"💾 Сохранено: {path.name}")

    print("\n✅ Stage 1 готов. Следующий шаг — проверка через data_only пересчёт.")


if __name__ == "__main__":
    main()
