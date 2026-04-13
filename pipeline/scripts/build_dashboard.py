#!/usr/bin/env python3
"""
B1 Dashboard Builder — v1.4.2

Читает артефакты MC-движков и строит:
1. `artifacts/dashboard.html` — интерактивный Plotly-дашборд (standalone).
2. `artifacts/dashboard_*.png` — статичные matplotlib-графики (для memo).
3. `artifacts/dashboard.xlsx` — многолистовой Excel с данными и формулами.

Запуск:
    python scripts/build_dashboard.py

Зависимости: plotly, matplotlib, openpyxl, python-docx (для memo).
"""
from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")  # noqa: E402
import matplotlib.pyplot as plt  # noqa: E402
import plotly.graph_objects as go  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.chart import BarChart, LineChart, Reference  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from plotly.subplots import make_subplots  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
ARTIFACTS = ROOT / "artifacts"
STRESS = ARTIFACTS / "stress_matrix"

ANCHOR_BASE = 3000.0
ANCHOR_LOWER = 2970.0  # −1%
ANCHOR_UPPER = 3030.0  # +1%

# Корпоративная палитра (ТрендСтудио / government-like)
COLOR_ANCHOR = "#0070C0"
COLOR_OK = "#2E8540"
COLOR_WARN = "#E08B00"
COLOR_RISK = "#C62828"
COLOR_GREY = "#6B7280"

ENGINE_COLORS = {
    "Naive MC": "#4C78A8",
    "Block Bootstrap": "#F58518",
    "Stage-Gate": "#C62828",
    "LHS+Copula": "#2E8540",
}


# ─────────────────────────────────────────────────────────────────────────────
# 1. Загрузка данных
# ─────────────────────────────────────────────────────────────────────────────
def load_artifacts() -> dict[str, Any]:
    """Загружает все четыре MC-артефакта + matrix_27."""
    files = {
        "mc": "monte_carlo.json",
        "boot": "market_bootstrap.json",
        "gate": "stage_gate.json",
        "lhs": "lhs_copula.json",
        "m27": "matrix_27.json",
    }
    return {k: json.loads((STRESS / f).read_text()) for k, f in files.items()}


def engine_summary(data: dict) -> list[dict]:
    """Возвращает таблицу сравнения 4 движков."""
    return [
        {
            "engine": "Naive MC",
            "mean": data["mc"]["mean_ebitda"],
            "std": data["mc"]["std_ebitda"],
            "p5": data["mc"]["p5_ebitda"],
            "p50": data["mc"]["p50_ebitda"],
            "p95": data["mc"]["p95_ebitda"],
            "var95": data["mc"]["var_95_mln_rub"],
            "breach_p": data["mc"]["breach_probability"],
        },
        {
            "engine": "Block Bootstrap",
            "mean": data["boot"]["mean_ebitda"],
            "std": data["boot"]["std_ebitda"],
            "p5": data["boot"]["p5_ebitda"],
            "p50": data["boot"]["p50_ebitda"],
            "p95": data["boot"]["p95_ebitda"],
            "var95": data["boot"]["var_95_mln_rub"],
            "breach_p": data["boot"]["breach_probability"],
        },
        {
            "engine": "Stage-Gate",
            "mean": data["gate"]["mean_ebitda"],
            "std": data["gate"]["std_ebitda"],
            "p5": data["gate"]["p5_ebitda"],
            "p50": data["gate"]["p50_ebitda"],
            "p95": data["gate"]["p95_ebitda"],
            "var95": data["gate"]["var_95_mln_rub"],
            "breach_p": data["gate"]["breach_probability"],
        },
        {
            "engine": "LHS+Copula",
            "mean": data["lhs"]["mean_ebitda"],
            "std": data["lhs"]["std_ebitda"],
            "p5": data["lhs"]["p5_ebitda"],
            "p50": data["lhs"]["p50_ebitda"],
            "p95": data["lhs"]["p95_ebitda"],
            "var95": data["lhs"]["var_95_mln_rub"],
            "breach_p": data["lhs"]["breach_probability"],
        },
    ]


def tornado_from_matrix27(m27: dict) -> list[tuple[str, float, float]]:
    """Возвращает список (driver, low_ebitda, high_ebitda) для tornado.

    Для каждого драйвера фиксирует остальные на нуле, варьирует один,
    находит min/max EBITDA. low_ebitda ≤ high_ebitda всегда.
    """
    base = m27["base_ebitda"]
    sc = {s["scenario_id"]: s for s in m27["scenarios"]}

    def extremes(ids: list[str]) -> tuple[float, float]:
        vals = [sc[i]["cumulative_ebitda"] for i in ids if i in sc]
        return (min(vals), max(vals)) if vals else (base, base)

    drivers: list[tuple[str, float, float]] = []

    # FX shock 0% → +20% (остальные нули)
    fx_lo, fx_hi = extremes(["FX0_I0_D0", "FX10_I0_D0", "FX20_I0_D0"])
    drivers.append(("FX 0→20%", fx_lo, fx_hi))

    # Инфляция 0% → +6%
    i_lo, i_hi = extremes(["FX0_I0_D0", "FX0_I3_D0", "FX0_I6_D0"])
    drivers.append(("Инфляция 0→6%", i_lo, i_hi))

    # Задержка 0 → 6 мес
    d_lo, d_hi = extremes(["FX0_I0_D0", "FX0_I0_D3", "FX0_I0_D6"])
    drivers.append(("Задержка 0→6 мес", d_lo, d_hi))

    # Комбинированный худший сценарий (one-sided: worst vs base)
    worst_id = m27["worst_scenario_id"]
    worst_val = m27["worst_ebitda"]
    lo = min(worst_val, base)
    hi = max(worst_val, base)
    drivers.append((f"Комбо {worst_id}", lo, hi))

    return drivers


def stage_gate_funnel(gate: dict) -> list[tuple[str, float]]:
    """Воронка stage-gate для 12 фильмов."""
    n = gate["n_films"]
    p1 = gate["p_dev_to_green"]
    p2 = gate["p_green_to_prod"]
    p3 = gate["p_prod_to_post"]
    p4 = gate["p_post_to_release"]
    stages = [
        ("Dev", n),
        ("Greenlight", n * p1),
        ("Production", n * p1 * p2),
        ("Post-prod", n * p1 * p2 * p3),
        ("Release", n * p1 * p2 * p3 * p4),
    ]
    return stages


# ─────────────────────────────────────────────────────────────────────────────
# 2. HTML dashboard (Plotly)
# ─────────────────────────────────────────────────────────────────────────────
def build_html_dashboard(data: dict, out_path: Path) -> None:
    engines = engine_summary(data)
    lhs = data["lhs"]
    mc = data["mc"]

    # 2.1 Anchor gauge
    gauge = go.Indicator(
        mode="gauge+number+delta",
        value=lhs["mean_ebitda"],
        number={"suffix": " млн ₽", "font": {"size": 36}},
        delta={
            "reference": ANCHOR_BASE,
            "increasing": {"color": COLOR_OK},
            "decreasing": {"color": COLOR_RISK},
        },
        title={"text": "Якорь: cumulative EBITDA 2026–2028 (LHS+Copula mean)"},
        gauge={
            "axis": {"range": [2500, 3500], "tickwidth": 1},
            "bar": {"color": COLOR_ANCHOR, "thickness": 0.6},
            "steps": [
                {"range": [2500, 2700], "color": "#fee2e2"},
                {"range": [2700, ANCHOR_LOWER], "color": "#fef3c7"},
                {"range": [ANCHOR_LOWER, ANCHOR_UPPER], "color": "#d1fae5"},
                {"range": [ANCHOR_UPPER, 3300], "color": "#dbeafe"},
                {"range": [3300, 3500], "color": "#e0e7ff"},
            ],
            "threshold": {
                "line": {"color": COLOR_RISK, "width": 3},
                "thickness": 0.85,
                "value": ANCHOR_LOWER,
            },
        },
    )

    # 2.2 MC comparison (mean ± std, 4 engines)
    engine_names = [e["engine"] for e in engines]
    means = [e["mean"] for e in engines]
    p5s = [e["p5"] for e in engines]
    p95s = [e["p95"] for e in engines]
    colors = [ENGINE_COLORS[n] for n in engine_names]

    mc_trace = go.Bar(
        x=engine_names,
        y=means,
        marker={"color": colors},
        error_y={
            "type": "data",
            "symmetric": False,
            "array": [p - m for p, m in zip(p95s, means)],
            "arrayminus": [m - p for m, p in zip(means, p5s)],
            "color": COLOR_GREY,
        },
        text=[f"{m:.0f}" for m in means],
        textposition="auto",
        name="Mean ± p5/p95",
        hovertemplate="%{x}<br>Mean: %{y:.0f} млн ₽<extra></extra>",
    )

    # 2.3 Tornado
    tornado = tornado_from_matrix27(data["m27"])
    t_names = [t[0] for t in tornado]
    t_lows = [t[1] - ANCHOR_BASE for t in tornado]
    t_highs = [t[2] - ANCHOR_BASE for t in tornado]
    tornado_trace_low = go.Bar(
        y=t_names,
        x=t_lows,
        orientation="h",
        marker={"color": COLOR_RISK},
        name="Low",
        hovertemplate="%{y}: base+%{x:.0f} млн ₽<extra></extra>",
    )
    tornado_trace_high = go.Bar(
        y=t_names,
        x=t_highs,
        orientation="h",
        marker={"color": COLOR_OK},
        name="High",
        hovertemplate="%{y}: base+%{x:.0f} млн ₽<extra></extra>",
    )

    # 2.4 Stage-gate funnel
    funnel = stage_gate_funnel(data["gate"])
    funnel_trace = go.Funnel(
        y=[f[0] for f in funnel],
        x=[f[1] for f in funnel],
        textinfo="value+percent initial",
        marker={
            "color": [
                "#1E3A8A",
                "#2563EB",
                "#3B82F6",
                "#60A5FA",
                "#93C5FD",
            ]
        },
    )

    # 2.5 VaR comparison table
    var_headers = ["Engine", "Mean", "Std", "p5", "p95", "VaR95", "Breach %"]
    var_cells = [
        engine_names,
        [f"{e['mean']:.0f}" for e in engines],
        [f"{e['std']:.0f}" for e in engines],
        [f"{e['p5']:.0f}" for e in engines],
        [f"{e['p95']:.0f}" for e in engines],
        [f"{e['var95']:.0f}" for e in engines],
        [f"{e['breach_p']*100:.2f}%" for e in engines],
    ]
    var_table = go.Table(
        header={
            "values": var_headers,
            "fill_color": COLOR_ANCHOR,
            "font": {"color": "white", "size": 13},
            "align": "center",
        },
        cells={
            "values": var_cells,
            "fill_color": [["#f8fafc", "#e2e8f0"] * 4],
            "align": "center",
            "font": {"size": 12},
        },
    )

    # 2.6 Собираем grid 3×2
    fig = make_subplots(
        rows=3,
        cols=2,
        specs=[
            [{"type": "indicator"}, {"type": "bar"}],
            [{"type": "bar"}, {"type": "funnel"}],
            [{"type": "table", "colspan": 2}, None],
        ],
        subplot_titles=(
            "Anchor Gauge (LHS+Copula)",
            "Сравнение 4 MC-движков: mean ± p5/p95",
            "Tornado: Δ EBITDA от драйверов",
            "Stage-Gate воронка (12 фильмов)",
            "Сводная таблица метрик",
        ),
        vertical_spacing=0.10,
        horizontal_spacing=0.08,
    )
    fig.add_trace(gauge, row=1, col=1)
    fig.add_trace(mc_trace, row=1, col=2)
    fig.add_trace(tornado_trace_low, row=2, col=1)
    fig.add_trace(tornado_trace_high, row=2, col=1)
    fig.add_trace(funnel_trace, row=2, col=2)
    fig.add_trace(var_table, row=3, col=1)

    fig.update_layout(
        title={
            "text": (
                "<b>Финмодель «ТрендСтудио» — B1 Dashboard v1.4.2</b><br>"
                f"<sub>Якорь: 3000 млн ₽ ± 1% · "
                f"LHS+Copula mean: {lhs['mean_ebitda']:.0f} · "
                f"VaR95: {lhs['var_95_mln_rub']:.0f} · "
                f"Breach prob: {lhs['breach_probability']*100:.2f}%</sub>"
            ),
            "x": 0.5,
            "xanchor": "center",
        },
        showlegend=False,
        height=1400,
        barmode="relative",
        paper_bgcolor="white",
        plot_bgcolor="#f8fafc",
        font={"family": "Inter, -apple-system, sans-serif", "size": 13},
        margin={"t": 110, "b": 40, "l": 60, "r": 40},
    )
    fig.update_yaxes(title_text="EBITDA, млн ₽", row=1, col=2)
    fig.update_xaxes(title_text="Δ от базы, млн ₽", row=2, col=1)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.write_html(
        out_path,
        include_plotlyjs="cdn",
        full_html=True,
        config={"displaylogo": False, "responsive": True},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 3. PNG charts (matplotlib)
# ─────────────────────────────────────────────────────────────────────────────
def build_png_charts(data: dict, out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    produced: list[Path] = []
    engines = engine_summary(data)
    lhs = data["lhs"]

    # 3.1 Anchor bar
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.axhspan(ANCHOR_LOWER, ANCHOR_UPPER, color="#d1fae5", label="Anchor ±1%")
    ax.axhline(ANCHOR_BASE, color=COLOR_ANCHOR, linestyle="--", linewidth=1.5, label="Base 3000")
    names = [e["engine"] for e in engines]
    values = [e["mean"] for e in engines]
    colors = [ENGINE_COLORS[n] for n in names]
    bars = ax.bar(names, values, color=colors, edgecolor="#334155", linewidth=0.8)
    for bar, v in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, v + 30, f"{v:.0f}", ha="center", fontsize=10)
    ax.set_ylabel("Cumulative EBITDA, млн ₽", fontsize=11)
    ax.set_title("Сравнение 4 MC-движков vs Anchor 3000 ± 1%", fontsize=13, fontweight="bold")
    ax.set_ylim(1500, 4500)
    ax.grid(axis="y", alpha=0.3)
    ax.legend(loc="upper right", fontsize=9)
    fig.tight_layout()
    p = out_dir / "dashboard_mc_comparison.png"
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    produced.append(p)

    # 3.2 Tornado
    fig, ax = plt.subplots(figsize=(8, 4.5))
    tornado = tornado_from_matrix27(data["m27"])
    t_names = [t[0] for t in tornado]
    deltas_low = [t[1] - ANCHOR_BASE for t in tornado]
    deltas_high = [t[2] - ANCHOR_BASE for t in tornado]
    y = list(range(len(t_names)))
    ax.barh(y, deltas_low, color=COLOR_RISK, label="Low (−)")
    ax.barh(y, deltas_high, color=COLOR_OK, label="High (+)")
    ax.axvline(0, color="#334155", linewidth=0.8)
    ax.set_yticks(y)
    ax.set_yticklabels(t_names)
    ax.set_xlabel("Δ от базы, млн ₽", fontsize=11)
    ax.set_title("Tornado: чувствительность к драйверам (matrix 27)", fontsize=13, fontweight="bold")
    ax.grid(axis="x", alpha=0.3)
    ax.legend(loc="lower right", fontsize=9)
    fig.tight_layout()
    p = out_dir / "dashboard_tornado.png"
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    produced.append(p)

    # 3.3 Stage-gate funnel
    fig, ax = plt.subplots(figsize=(8, 4.5))
    funnel = stage_gate_funnel(data["gate"])
    stages = [f[0] for f in funnel]
    counts = [f[1] for f in funnel]
    colors_f = ["#1E3A8A", "#2563EB", "#3B82F6", "#60A5FA", "#93C5FD"]
    ax.barh(stages[::-1], counts[::-1], color=colors_f[::-1], edgecolor="#0f172a")
    for i, (s, c) in enumerate(zip(stages[::-1], counts[::-1])):
        ax.text(c + 0.15, i, f"{c:.2f} ({c / counts[0] * 100:.1f}%)", va="center", fontsize=10)
    ax.set_xlabel("Количество фильмов (из 12)", fontsize=11)
    ax.set_title(
        f"Stage-Gate воронка: P(reach)={data['gate']['p_reach_release']*100:.1f}%, "
        f"mean={data['gate']['mean_released_count']:.2f}",
        fontsize=13,
        fontweight="bold",
    )
    ax.set_xlim(0, 14)
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout()
    p = out_dir / "dashboard_stage_gate_funnel.png"
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    produced.append(p)

    # 3.4 VaR comparison (LHS vs MC)
    fig, ax = plt.subplots(figsize=(8, 4.5))
    percentiles = ["p5", "p25", "p50", "p75", "p95"]
    mc_vals = [data["mc"][f"{p}_ebitda"] for p in percentiles]
    lhs_vals = [data["lhs"][f"{p}_ebitda"] for p in percentiles]
    x = range(len(percentiles))
    width = 0.35
    ax.bar([i - width / 2 for i in x], mc_vals, width, color=ENGINE_COLORS["Naive MC"], label="Naive MC")
    ax.bar([i + width / 2 for i in x], lhs_vals, width, color=ENGINE_COLORS["LHS+Copula"], label="LHS+Copula")
    ax.axhline(ANCHOR_BASE, color=COLOR_ANCHOR, linestyle="--", linewidth=1, label="Anchor 3000")
    ax.axhline(ANCHOR_LOWER, color=COLOR_RISK, linestyle=":", linewidth=1, label="Lower −1%")
    ax.set_xticks(list(x))
    ax.set_xticklabels(percentiles)
    ax.set_ylabel("EBITDA, млн ₽", fontsize=11)
    ax.set_title("Распределение EBITDA: Naive MC vs LHS+Copula", fontsize=13, fontweight="bold")
    ax.legend(fontsize=9)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    p = out_dir / "dashboard_var_comparison.png"
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    produced.append(p)

    # 3.5 Anchor gauge (отдельный PNG)
    fig, ax = plt.subplots(figsize=(6, 4))
    theta = [i * math.pi / 50 for i in range(51)]
    # Полукруговой фон (упрощённая версия)
    r_out, r_in = 1.0, 0.55
    zones = [
        (0, 0.2, "#fee2e2"),
        (0.2, 0.35, "#fef3c7"),
        (0.35, 0.65, "#d1fae5"),
        (0.65, 0.8, "#dbeafe"),
        (0.8, 1.0, "#e0e7ff"),
    ]
    for z_start, z_end, col in zones:
        ts = [math.pi * (1 - z_start - (z_end - z_start) * i / 20) for i in range(21)]
        xs_o = [r_out * math.cos(t) for t in ts]
        ys_o = [r_out * math.sin(t) for t in ts]
        xs_i = [r_in * math.cos(t) for t in reversed(ts)]
        ys_i = [r_in * math.sin(t) for t in reversed(ts)]
        ax.fill(xs_o + xs_i, ys_o + ys_i, color=col, edgecolor="white")
    # Стрелка
    mean_val = lhs["mean_ebitda"]
    pos = max(0.0, min(1.0, (mean_val - 2500) / 1000))
    angle = math.pi * (1 - pos)
    ax.plot([0, 0.9 * math.cos(angle)], [0, 0.9 * math.sin(angle)], color=COLOR_ANCHOR, linewidth=3)
    ax.scatter([0], [0], color=COLOR_ANCHOR, s=80, zorder=5)
    ax.text(0, -0.15, f"{mean_val:.0f}", ha="center", fontsize=20, fontweight="bold", color=COLOR_ANCHOR)
    ax.text(0, -0.3, "млн ₽ (LHS mean)", ha="center", fontsize=10, color=COLOR_GREY)
    ax.text(-1, 0.05, "2500", ha="center", fontsize=8)
    ax.text(1, 0.05, "3500", ha="center", fontsize=8)
    ax.text(0, 1.05, "3000 ± 1%", ha="center", fontsize=10, fontweight="bold", color=COLOR_ANCHOR)
    ax.set_xlim(-1.2, 1.2)
    ax.set_ylim(-0.5, 1.25)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("Якорный показатель", fontsize=13, fontweight="bold")
    fig.tight_layout()
    p = out_dir / "dashboard_anchor_gauge.png"
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    produced.append(p)

    return produced


# ─────────────────────────────────────────────────────────────────────────────
# 4. XLSX dashboard
# ─────────────────────────────────────────────────────────────────────────────
def build_xlsx_dashboard(data: dict, out_path: Path) -> None:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0070C0")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(color="0070C0", bold=True, size=14)
    border = Border(*([Side(style="thin", color="D1D5DB")] * 4))
    center = Alignment(horizontal="center", vertical="center")

    # Лист 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "B1 Dashboard — Финмодель «ТрендСтудио»"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")
    ws["A2"] = "Версия v1.4.2  ·  11 апреля 2026  ·  Anchor: 3000 млн ₽ ± 1%"
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:G2")

    # Заголовки сравнения
    headers = ["Engine", "Mean", "Std", "p5", "p50", "p95", "VaR95", "Breach %"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    engines = engine_summary(data)
    for row_idx, e in enumerate(engines, start=5):
        ws.cell(row=row_idx, column=1, value=e["engine"]).border = border
        for col_idx, key in enumerate(
            ["mean", "std", "p5", "p50", "p95", "var95"], start=2
        ):
            c = ws.cell(row=row_idx, column=col_idx, value=round(e[key], 2))
            c.number_format = "#,##0"
            c.border = border
            c.alignment = center
        c = ws.cell(row=row_idx, column=8, value=e["breach_p"])
        c.number_format = "0.00%"
        c.border = border
        c.alignment = center

    # Формулы: отклонение от якоря
    ws.cell(row=9, column=1, value="Δ от Anchor").font = Font(bold=True)
    for col_idx in range(2, 9):
        if col_idx == 8:
            continue
        ws.cell(row=10, column=1, value=f"Delta (mean − 3000):")
    for row_idx, e in enumerate(engines, start=5):
        ws.cell(row=row_idx, column=9, value=f"=B{row_idx}-3000").number_format = "#,##0;-#,##0"
    ws.cell(row=4, column=9, value="Δ Anchor").fill = header_fill
    ws.cell(row=4, column=9).font = header_font
    ws.cell(row=4, column=9).alignment = center

    # Column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 18

    # Mini bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Mean EBITDA по движкам"
    chart.y_axis.title = "млн ₽"
    chart.x_axis.title = "Engine"
    chart.height = 8
    chart.width = 16
    data_ref = Reference(ws, min_col=2, min_row=4, max_row=8, max_col=2)
    cats = Reference(ws, min_col=1, min_row=5, max_row=8)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "K4")

    # Лист 2: Matrix27 scenarios
    ws2 = wb.create_sheet("Matrix27")
    ws2.append(["scenario_id", "fx_shock_pct", "inflation_pct", "delay_months", "ebitda", "Δ от base", "breach?"])
    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
    for s in data["m27"]["scenarios"]:
        ws2.append(
            [
                s["scenario_id"],
                s["fx_shock_pct"],
                s["inflation_pct"],
                s["delay_months"],
                round(s["cumulative_ebitda"], 2),
                round(s["cumulative_ebitda"] - data["m27"]["base_ebitda"], 2),
                "Y" if s["cumulative_ebitda"] < 2700 else "",
            ]
        )
    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    # Формула: процент scenarios with breach
    last_row = ws2.max_row
    ws2.cell(row=last_row + 2, column=1, value="Breach count:").font = Font(bold=True)
    ws2.cell(row=last_row + 2, column=2, value=f'=COUNTIF(G2:G{last_row},"Y")')
    ws2.cell(row=last_row + 3, column=1, value="Total:").font = Font(bold=True)
    ws2.cell(row=last_row + 3, column=2, value=last_row - 1)
    ws2.cell(row=last_row + 4, column=1, value="Breach %:").font = Font(bold=True)
    ws2.cell(row=last_row + 4, column=2, value=f"=B{last_row + 2}/B{last_row + 3}").number_format = "0.00%"

    # Лист 3: Stage-Gate
    ws3 = wb.create_sheet("StageGate")
    ws3.append(["Stage", "Expected count", "% of cohort"])
    for c in ws3[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
    funnel = stage_gate_funnel(data["gate"])
    for stage, count in funnel:
        ws3.append([stage, round(count, 3), f"=B{ws3.max_row+1 if False else ws3.max_row}/{funnel[0][1]}"])
    # Исправим формулы
    for r in range(2, ws3.max_row + 1):
        ws3.cell(row=r, column=3, value=f"=B{r}/B2").number_format = "0.00%"
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 15

    # Доп метрики
    gate = data["gate"]
    ws3.append([])
    ws3.append(["P(reach release)", round(gate["p_reach_release"], 4)])
    ws3.append(["Mean released", round(gate["mean_released_count"], 3)])
    ws3.append(["Std released", round(gate["std_released_count"], 3)])
    ws3.append(["Mean sunk cost", round(gate["mean_sunk_cost_mln_rub"], 1)])
    ws3.append(["p95 sunk cost", round(gate["p95_sunk_cost_mln_rub"], 1)])

    # Лист 4: LHS+Copula
    ws4 = wb.create_sheet("LHSCopula")
    lhs = data["lhs"]
    ws4.append(["Percentile", "EBITDA, млн ₽"])
    for c in ws4[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
    for p_key in ["p1", "p5", "p25", "p50", "p75", "p95", "p99"]:
        ws4.append([p_key, round(lhs[f"{p_key}_ebitda"], 2)])
    ws4.append([])
    ws4.append(["Mean", round(lhs["mean_ebitda"], 2)])
    ws4.append(["Std", round(lhs["std_ebitda"], 2)])
    ws4.append(["VaR95", round(lhs["var_95_mln_rub"], 2)])
    ws4.append(["VaR99", round(lhs["var_99_mln_rub"], 2)])
    ws4.append(["Breach p", lhs["breach_probability"]])
    ws4.cell(row=ws4.max_row, column=2).number_format = "0.00%"
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    data = load_artifacts()

    html_path = ARTIFACTS / "dashboard.html"
    build_html_dashboard(data, html_path)
    print(f"[OK] HTML: {html_path}")

    png_paths = build_png_charts(data, ARTIFACTS)
    for p in png_paths:
        print(f"[OK] PNG: {p}")

    xlsx_path = ARTIFACTS / "dashboard.xlsx"
    build_xlsx_dashboard(data, xlsx_path)
    print(f"[OK] XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
