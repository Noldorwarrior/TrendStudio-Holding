#!/usr/bin/env python3
"""
DD-Grade Audit: Step 0 - Map Internal XLSX Structure
=====================================================
Opens investor_model_v1.0_Internal.xlsx and produces a comprehensive
structural map for due-diligence purposes.
"""

import os
import re
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = "/home/user/TrendStudio-Holding/Investor_Package"
INTERNAL = os.path.join(BASE, "investor_model_v1.0_Internal.xlsx")
PUBLIC   = os.path.join(BASE, "investor_model_v1.0_Public.xlsx")

SEP = "=" * 80
SUB = "-" * 60

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
ERROR_TOKENS = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#GETTING_DATA", "#SPILL!"}

def visibility_label(state):
    """Return human-readable visibility."""
    if state == "visible" or state is None:
        return "visible"
    return str(state)  # 'hidden' or 'veryHidden'

def is_formula(cell):
    """True when a cell contains a formula string."""
    v = cell.value
    return isinstance(v, str) and v.startswith("=")

def is_error(value):
    """True when a cached value is an Excel error token."""
    if isinstance(value, str):
        return value.strip() in ERROR_TOKENS
    return False

def cell_ref(cell):
    return f"{get_column_letter(cell.column)}{cell.row}"

# ---------------------------------------------------------------------------
# 1. Load workbooks
# ---------------------------------------------------------------------------
print(SEP)
print("STEP 0 -- STRUCTURAL MAP OF INTERNAL MODEL")
print(f"File : {INTERNAL}")
print(f"Run  : {datetime.now().isoformat()}")
print(SEP)

print("\nLoading workbook (formulas) ...")
wb_f = openpyxl.load_workbook(INTERNAL, data_only=False, read_only=False)
print("Loading workbook (cached values) ...")
wb_v = openpyxl.load_workbook(INTERNAL, data_only=True, read_only=False)

# Also load Public for diff
print("Loading Public workbook for sheet diff ...")
pub_exists = os.path.exists(PUBLIC)
if pub_exists:
    wb_pub = openpyxl.load_workbook(PUBLIC, read_only=True, data_only=False)
    public_sheets = set(wb_pub.sheetnames)
    wb_pub.close()
else:
    public_sheets = set()
    print("  [WARNING] Public file not found -- skipping diff")

# ---------------------------------------------------------------------------
# 5. Document Properties
# ---------------------------------------------------------------------------
print(f"\n{SEP}")
print("SECTION A -- DOCUMENT PROPERTIES")
print(SEP)
props = wb_f.properties
prop_fields = [
    ("Title",            props.title),
    ("Subject",          props.subject),
    ("Creator",          props.creator),
    ("Last Modified By", props.lastModifiedBy),
    ("Created",          props.created),
    ("Modified",         props.modified),
    ("Company",          getattr(props, "company", None)),
    ("Keywords",         props.keywords),
    ("Description",      props.description),
    ("Category",         props.category),
    ("Version",          props.version),
    ("Revision",         props.revision),
]
for label, val in prop_fields:
    print(f"  {label:20s}: {val}")

# ---------------------------------------------------------------------------
# 3. Global Defined Names
# ---------------------------------------------------------------------------
print(f"\n{SEP}")
print("SECTION B -- GLOBAL DEFINED NAMES")
print(SEP)
dn_count = 0
# openpyxl 3.1+ uses DefinedNameDict (dict-like)
for name, dn in wb_f.defined_names.items():
    destinations = []
    try:
        for title, coord in dn.destinations:
            destinations.append(f"{title}!{coord}")
    except Exception:
        destinations.append(str(dn.attr_text) if hasattr(dn, 'attr_text') else str(dn.value))
    print(f"  {name:40s} | {', '.join(destinations)}")
    dn_count += 1

# Also check sheet-scoped defined names
for idx, sheet_name in enumerate(wb_f.sheetnames):
    ws = wb_f[sheet_name]
    if hasattr(ws, 'defined_names') and ws.defined_names:
        for sname, sdn in ws.defined_names.items():
            dests = []
            try:
                for title, coord in sdn.destinations:
                    dests.append(f"{title}!{coord}")
            except Exception:
                dests.append(str(sdn))
            print(f"  {sname:40s} | Sheet-scoped({sheet_name}) | {', '.join(dests)}")
            dn_count += 1

if dn_count == 0:
    print("  (none)")
print(f"\n  Total defined names: {dn_count}")

# ---------------------------------------------------------------------------
# 8. External Links
# ---------------------------------------------------------------------------
print(f"\n{SEP}")
print("SECTION C -- EXTERNAL LINKS")
print(SEP)
ext_links = []
# Check via workbook _external_links attribute
if hasattr(wb_f, '_external_links'):
    for el in wb_f._external_links:
        target = getattr(el, 'file_link', None) or getattr(el, 'Target', None) or str(el)
        ext_links.append(target)

if ext_links:
    for i, link in enumerate(ext_links, 1):
        print(f"  [{i}] {link}")
else:
    # Fallback: scan formulas for [*] external references
    print("  (No _external_links found; will scan formulas for external refs in per-sheet section)")

# ---------------------------------------------------------------------------
# Per-sheet analysis
# ---------------------------------------------------------------------------
print(f"\n{SEP}")
print("SECTION D -- PER-SHEET ANALYSIS")
print(SEP)

internal_sheets = set(wb_f.sheetnames)
total_formulas_all = 0
total_values_all = 0
total_errors_all = 0
external_refs_found = []

for idx, sheet_name in enumerate(wb_f.sheetnames):
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    print(f"\n{SUB}")
    print(f"SHEET [{idx}]: \"{sheet_name}\"")
    print(SUB)

    # 2. Visibility, dimensions
    vis = visibility_label(ws_f.sheet_state)
    print(f"  Visibility     : {vis}")
    print(f"  Dimensions     : {ws_f.dimensions}")
    print(f"  Min row/col    : ({ws_f.min_row}, {ws_f.min_column})")
    print(f"  Max row/col    : ({ws_f.max_row}, {ws_f.max_column})")

    # 9. Sheet protection
    prot = ws_f.protection
    is_protected = prot.sheet
    print(f"  Protected      : {is_protected}")
    if is_protected:
        prot_details = []
        for attr in ["password", "formatColumns", "formatRows", "formatCells",
                      "insertColumns", "insertRows", "insertHyperlinks",
                      "deleteColumns", "deleteRows", "selectLockedCells",
                      "sort", "autoFilter", "pivotTables", "selectUnlockedCells",
                      "objects", "scenarios"]:
            v = getattr(prot, attr, None)
            if v is not None and v is not False:
                prot_details.append(f"{attr}={v}")
        if prot_details:
            print(f"  Protection det.: {', '.join(prot_details)}")

    # Defined names scoped to this sheet
    ws_dns = getattr(ws_f, 'defined_names', {})
    if ws_dns:
        print(f"  Scoped names   : {len(ws_dns)}")
        for dn_name, dn_obj in list(ws_dns.items())[:5]:
            try:
                dests = [f"{t}!{c}" for t, c in dn_obj.destinations]
            except Exception:
                dests = [str(dn_obj)]
            print(f"    - {dn_name}: {', '.join(dests)}")

    # 6. Count formulas vs values  &  7. Error cells  &  4. Sample formulas
    formula_count = 0
    value_count = 0
    error_cells = []
    formula_samples = []
    ext_ref_cells = []

    for row in ws_f.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if is_formula(cell):
                formula_count += 1
                if len(formula_samples) < 10:
                    # Get cached value from value-only workbook
                    cached = ws_v[cell_ref(cell)].value
                    formula_samples.append((cell_ref(cell), cell.value, cached))
                # Check for external references in formula
                if re.search(r"\[.*?\]", str(cell.value)):
                    ext_ref_cells.append((cell_ref(cell), cell.value))
            else:
                value_count += 1

            # Check cached value for errors (in the value workbook)
            cached_val = ws_v[cell_ref(cell)].value
            if is_error(cached_val):
                error_cells.append((cell_ref(cell), cached_val, cell.value if is_formula(cell) else None))

    total_formulas_all += formula_count
    total_values_all += value_count
    total_errors_all += len(error_cells)

    print(f"  Formulas       : {formula_count:,}")
    print(f"  Values (non-empty): {value_count:,}")
    print(f"  Error cells    : {len(error_cells)}")

    if error_cells:
        print(f"  Error details  :")
        for ref, err_val, formula in error_cells[:20]:
            formula_str = f"  formula={formula}" if formula else ""
            print(f"    {ref:12s} -> {err_val}{formula_str}")
        if len(error_cells) > 20:
            print(f"    ... and {len(error_cells) - 20} more")

    # 4. Sample formula cells
    if formula_samples:
        print(f"  Formula samples:")
        for ref, formula, cached in formula_samples:
            cached_str = repr(cached) if cached is not None else "<None>"
            # Truncate long formulas
            f_display = formula if len(formula) <= 100 else formula[:97] + "..."
            print(f"    {ref:12s} | {f_display}")
            print(f"    {'':12s} | cached => {cached_str}")

    # External references in formulas
    if ext_ref_cells:
        external_refs_found.extend([(sheet_name, r, f) for r, f in ext_ref_cells])
        print(f"  External refs in formulas: {len(ext_ref_cells)}")
        for ref, formula in ext_ref_cells[:5]:
            print(f"    {ref}: {formula[:120]}")

# ---------------------------------------------------------------------------
# External references summary
# ---------------------------------------------------------------------------
print(f"\n{SEP}")
print("SECTION E -- EXTERNAL REFERENCES SUMMARY")
print(SEP)
if external_refs_found:
    # Extract unique external file names
    ext_files = set()
    for sn, ref, formula in external_refs_found:
        matches = re.findall(r"\[([^\]]+)\]", formula)
        ext_files.update(matches)
    print(f"  Total cells with external refs: {len(external_refs_found)}")
    print(f"  Unique external files referenced:")
    for f in sorted(ext_files):
        print(f"    - {f}")
    print(f"\n  First 15 external-ref cells:")
    for sn, ref, formula in external_refs_found[:15]:
        print(f"    Sheet=\"{sn}\" {ref}: {formula[:120]}")
else:
    print("  No external references found in formulas.")

# ---------------------------------------------------------------------------
# 10. Internal vs Public sheet diff
# ---------------------------------------------------------------------------
print(f"\n{SEP}")
print("SECTION F -- SHEET DIFF: INTERNAL vs PUBLIC")
print(SEP)
if pub_exists:
    only_internal = sorted(internal_sheets - public_sheets)
    only_public   = sorted(public_sheets - internal_sheets)
    common        = sorted(internal_sheets & public_sheets)

    print(f"  Internal sheet count : {len(internal_sheets)}")
    print(f"  Public sheet count   : {len(public_sheets)}")
    print(f"  Common sheets        : {len(common)}")

    if only_internal:
        print(f"\n  Sheets in INTERNAL but NOT in Public ({len(only_internal)}):")
        for s in only_internal:
            vis = visibility_label(wb_f[s].sheet_state)
            print(f"    - \"{s}\"  (visibility: {vis})")
    else:
        print(f"\n  No sheets unique to Internal.")

    if only_public:
        print(f"\n  Sheets in PUBLIC but NOT in Internal ({len(only_public)}):")
        for s in only_public:
            print(f"    - \"{s}\"")
    else:
        print(f"\n  No sheets unique to Public.")
else:
    print("  [SKIPPED] Public file not available.")

# ---------------------------------------------------------------------------
# Grand summary
# ---------------------------------------------------------------------------
print(f"\n{SEP}")
print("SECTION G -- GRAND SUMMARY")
print(SEP)
print(f"  Total sheets           : {len(wb_f.sheetnames)}")
print(f"  Total formulas         : {total_formulas_all:,}")
print(f"  Total non-empty values : {total_values_all:,}")
print(f"  Total error cells      : {total_errors_all:,}")
print(f"  Total defined names    : {dn_count}")
print(f"  External ref cells     : {len(external_refs_found)}")

hidden_sheets = [s for s in wb_f.sheetnames if wb_f[s].sheet_state != "visible" and wb_f[s].sheet_state is not None]
if hidden_sheets:
    print(f"  Hidden/VeryHidden sheets ({len(hidden_sheets)}):")
    for s in hidden_sheets:
        print(f"    - \"{s}\" ({visibility_label(wb_f[s].sheet_state)})")

protected_sheets = [s for s in wb_f.sheetnames if wb_f[s].protection.sheet]
if protected_sheets:
    print(f"  Protected sheets ({len(protected_sheets)}):")
    for s in protected_sheets:
        print(f"    - \"{s}\"")
else:
    print(f"  Protected sheets       : 0")

print(f"\n  Sheet index:")
for i, s in enumerate(wb_f.sheetnames):
    vis = visibility_label(wb_f[s].sheet_state)
    dim = wb_f[s].dimensions
    print(f"    [{i:2d}] {vis:12s} | {dim:20s} | \"{s}\"")

print(f"\n{SEP}")
print("END OF STRUCTURAL MAP")
print(SEP)

wb_f.close()
wb_v.close()
