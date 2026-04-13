#!/usr/bin/env python3
"""
Block B — Balance-Sheet Integrity & Cash Reconciliation Audit
=============================================================
DD-grade auditor script.  Opens the Public investor model (data_only)
and checks:
  1. Assets == Liabilities + Equity for every period
  2. Ending Cash (CF) == Cash (BS) for every period
  3. The model's own "Control" row (C1) == 0 for every period
  4. Summarises all findings with PASS / FAIL verdicts

File : investor_model_v1.0_Public.xlsx
Sheets: 11_Balance_Sheet, 10_Cash_Flow
"""

import sys, os
from pathlib import Path

# ── paths ──────────────────────────────────────────────────────────
MODEL = Path("/home/user/TrendStudio-Holding/Investor_Package/"
             "investor_model_v1.0_Public.xlsx")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed – pip install openpyxl")

TOLERANCE = 0.01          # absolute tolerance (MLN RUB)

# ── helpers ────────────────────────────────────────────────────────
def safe_float(v, default=None):
    """Return float or *default* when a cell is empty / text."""
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def find_sheet(wb, candidates):
    """Return the first matching sheet or None."""
    names_lower = {n.lower(): n for n in wb.sheetnames}
    for c in candidates:
        if c.lower() in names_lower:
            return wb[names_lower[c.lower()]]
    # partial match
    for c in candidates:
        for nl, name in names_lower.items():
            if c.lower() in nl:
                return wb[name]
    return None


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    print("=" * 78)
    print("BLOCK B  ·  BALANCE-SHEET INTEGRITY AUDIT")
    print(f"File: {MODEL.name}")
    print("=" * 78)

    wb = openpyxl.load_workbook(str(MODEL), data_only=True)
    print(f"\nLoaded workbook — {len(wb.sheetnames)} sheets.")

    # ── locate sheets ──────────────────────────────────────────────
    bs = find_sheet(wb, ["11_Balance_Sheet", "Balance_Sheet", "BS"])
    cf = find_sheet(wb, ["10_Cash_Flow", "Cash_Flow", "CF"])

    if bs is None:
        sys.exit("FATAL: could not locate Balance Sheet tab.")
    if cf is None:
        sys.exit("FATAL: could not locate Cash Flow tab.")

    print(f"Balance Sheet → '{bs.title}'  (rows={bs.max_row}, cols={bs.max_column})")
    print(f"Cash Flow     → '{cf.title}'  (rows={cf.max_row}, cols={cf.max_column})")

    # ── identify period columns (row 5 is header row) ─────────────
    HEADER_ROW = 5
    DATA_COL_START = 4          # col D onward
    # Read headers from BS row 5
    period_cols = {}             # col_idx → header label
    for cell in bs[HEADER_ROW]:
        if cell.column >= DATA_COL_START and cell.value is not None:
            label = str(cell.value).strip()
            # Skip comment columns
            if label.lower().startswith("комм") or label.lower().startswith("comment"):
                continue
            period_cols[cell.column] = label

    # Also check max period column in CF headers
    cf_period_cols = {}
    for cell in cf[HEADER_ROW]:
        if cell.column >= DATA_COL_START and cell.value is not None:
            label = str(cell.value).strip()
            if label.lower().startswith("комм") or label.lower().startswith("comment"):
                continue
            if label.startswith("Σ"):
                continue
            cf_period_cols[cell.column] = label

    # Use BS period columns (cols 4-19 = Q1'26..2032)
    # Filter: keep only cols present in both BS and CF if needed
    print(f"\nIdentified {len(period_cols)} period columns in BS:")
    for col, lbl in sorted(period_cols.items()):
        print(f"  Col {col:2d}: {lbl}")

    # ── Row mapping (from exploration) ─────────────────────────────
    # BS rows:
    #   7  A1  Cash
    #   8  A2  Content library
    #   9  A3  PP&E
    #  10  A4  TOTAL ASSETS
    #  13  B1  T₁ loan balance
    #  14  B2  Producer equity
    #  15  B3  Retained earnings
    #  16  B4  TOTAL L+E
    #  19  C1  Control (Assets − L+E), should be 0

    ROW_TOTAL_ASSETS = 10
    ROW_TOTAL_LE     = 16
    ROW_CONTROL      = 19
    ROW_CASH_BS      = 7

    # CF rows:
    #  23  4.3  Cash — end of period
    ROW_CASH_CF      = 23

    # ══════════════════════════════════════════════════════════════
    #  TEST 1: Assets == Liabilities + Equity
    # ══════════════════════════════════════════════════════════════
    print("\n" + "─" * 78)
    print("TEST 1  ·  Assets == Liabilities + Equity  (every period)")
    print("─" * 78)
    print(f"{'Period':<12} {'Total Assets':>14} {'Total L+E':>14} {'Delta':>12} {'Status':>8}")
    print("-" * 62)

    t1_pass = True
    t1_flags = []
    for col, lbl in sorted(period_cols.items()):
        assets = safe_float(bs.cell(row=ROW_TOTAL_ASSETS, column=col).value)
        le     = safe_float(bs.cell(row=ROW_TOTAL_LE,     column=col).value)
        if assets is None or le is None:
            print(f"{lbl:<12} {'N/A':>14} {'N/A':>14} {'—':>12} {'SKIP':>8}")
            continue
        delta = assets - le
        ok = abs(delta) <= TOLERANCE
        status = "OK" if ok else "FAIL"
        if not ok:
            t1_pass = False
            t1_flags.append((lbl, delta))
        print(f"{lbl:<12} {assets:>14,.2f} {le:>14,.2f} {delta:>12.4f} {status:>8}")

    if t1_pass:
        print("\n  ✓ TEST 1 PASSED — balance sheet balances for ALL periods.")
    else:
        print(f"\n  ✗ TEST 1 FAILED — {len(t1_flags)} period(s) out of balance:")
        for lbl, d in t1_flags:
            print(f"    {lbl}: delta = {d:.4f}")

    # ══════════════════════════════════════════════════════════════
    #  TEST 2: Model's own Control row (C1) == 0
    # ══════════════════════════════════════════════════════════════
    print("\n" + "─" * 78)
    print("TEST 2  ·  Model Control Row C1 == 0  (every period)")
    print("─" * 78)
    print(f"{'Period':<12} {'Control Value':>14} {'Status':>8}")
    print("-" * 36)

    t2_pass = True
    t2_flags = []
    for col, lbl in sorted(period_cols.items()):
        ctrl = safe_float(bs.cell(row=ROW_CONTROL, column=col).value)
        if ctrl is None:
            print(f"{lbl:<12} {'N/A':>14} {'SKIP':>8}")
            continue
        ok = abs(ctrl) <= TOLERANCE
        status = "OK" if ok else "FAIL"
        if not ok:
            t2_pass = False
            t2_flags.append((lbl, ctrl))
        print(f"{lbl:<12} {ctrl:>14.4f} {status:>8}")

    if t2_pass:
        print("\n  ✓ TEST 2 PASSED — Control row = 0 for ALL periods.")
    else:
        print(f"\n  ✗ TEST 2 FAILED — {len(t2_flags)} period(s) with non-zero control:")
        for lbl, d in t2_flags:
            print(f"    {lbl}: control = {d:.4f}")

    # ══════════════════════════════════════════════════════════════
    #  TEST 3: Cash (BS) == Ending Cash (CF)
    # ══════════════════════════════════════════════════════════════
    print("\n" + "─" * 78)
    print("TEST 3  ·  Cash on BS == Ending Cash on CF  (every period)")
    print("─" * 78)
    print(f"{'Period':<12} {'Cash (BS)':>14} {'Cash (CF)':>14} {'Delta':>12} {'Status':>8}")
    print("-" * 62)

    # Determine which columns to compare: intersection of BS & CF period cols
    common_cols = sorted(set(period_cols.keys()) & set(cf_period_cols.keys()))

    t3_pass = True
    t3_flags = []
    for col in common_cols:
        lbl = period_cols.get(col, cf_period_cols.get(col, f"Col{col}"))
        cash_bs = safe_float(bs.cell(row=ROW_CASH_BS, column=col).value)
        cash_cf = safe_float(cf.cell(row=ROW_CASH_CF, column=col).value)
        if cash_bs is None or cash_cf is None:
            print(f"{lbl:<12} {'N/A':>14} {'N/A':>14} {'—':>12} {'SKIP':>8}")
            continue
        delta = cash_bs - cash_cf
        ok = abs(delta) <= TOLERANCE
        status = "OK" if ok else "FAIL"
        if not ok:
            t3_pass = False
            t3_flags.append((lbl, delta, cash_bs, cash_cf))
        print(f"{lbl:<12} {cash_bs:>14,.2f} {cash_cf:>14,.2f} {delta:>12.4f} {status:>8}")

    if t3_pass:
        print("\n  ✓ TEST 3 PASSED — Cash reconciles for ALL periods.")
    else:
        print(f"\n  ✗ TEST 3 FAILED — {len(t3_flags)} period(s) with cash mismatch:")
        for lbl, d, bsv, cfv in t3_flags:
            print(f"    {lbl}: BS={bsv:.2f}  CF={cfv:.2f}  delta={d:.4f}")

    # ══════════════════════════════════════════════════════════════
    #  TEST 4: Cross-check individual BS components sum to totals
    # ══════════════════════════════════════════════════════════════
    print("\n" + "─" * 78)
    print("TEST 4  ·  Component-level verification (Assets=A1+A2+A3, L+E=B1+B2+B3)")
    print("─" * 78)
    print(f"{'Period':<12} {'A1+A2+A3':>14} {'Tot.Assets':>14} {'ΔA':>10}  "
          f"{'B1+B2+B3':>14} {'Tot.L+E':>14} {'ΔLE':>10} {'Status':>8}")
    print("-" * 100)

    ROW_A1, ROW_A2, ROW_A3 = 7, 8, 9
    ROW_B1, ROW_B2, ROW_B3 = 13, 14, 15

    t4_pass = True
    t4_flags = []
    for col, lbl in sorted(period_cols.items()):
        a1 = safe_float(bs.cell(row=ROW_A1, column=col).value, 0)
        a2 = safe_float(bs.cell(row=ROW_A2, column=col).value, 0)
        a3 = safe_float(bs.cell(row=ROW_A3, column=col).value, 0)
        total_a = safe_float(bs.cell(row=ROW_TOTAL_ASSETS, column=col).value)

        b1 = safe_float(bs.cell(row=ROW_B1, column=col).value, 0)
        b2 = safe_float(bs.cell(row=ROW_B2, column=col).value, 0)
        b3 = safe_float(bs.cell(row=ROW_B3, column=col).value, 0)
        total_le = safe_float(bs.cell(row=ROW_TOTAL_LE, column=col).value)

        if total_a is None or total_le is None:
            print(f"{lbl:<12} {'N/A':>14} {'N/A':>14} {'—':>10}  "
                  f"{'N/A':>14} {'N/A':>14} {'—':>10} {'SKIP':>8}")
            continue

        sum_a = a1 + a2 + a3
        sum_le = b1 + b2 + b3
        da = sum_a - total_a
        dle = sum_le - total_le

        ok_a = abs(da) <= TOLERANCE
        ok_le = abs(dle) <= TOLERANCE
        ok = ok_a and ok_le
        status = "OK" if ok else "FAIL"
        if not ok:
            t4_pass = False
            t4_flags.append((lbl, da, dle))

        print(f"{lbl:<12} {sum_a:>14,.2f} {total_a:>14,.2f} {da:>10.4f}  "
              f"{sum_le:>14,.2f} {total_le:>14,.2f} {dle:>10.4f} {status:>8}")

    if t4_pass:
        print("\n  ✓ TEST 4 PASSED — component sums match totals for ALL periods.")
    else:
        print(f"\n  ✗ TEST 4 FAILED — {len(t4_flags)} period(s) with component mismatch:")
        for lbl, da, dle in t4_flags:
            print(f"    {lbl}: ΔA={da:.4f}  ΔLE={dle:.4f}")

    # ══════════════════════════════════════════════════════════════
    #  TEST 5: CF consistency — Beginning Cash + Net Change == Ending Cash
    # ══════════════════════════════════════════════════════════════
    print("\n" + "─" * 78)
    print("TEST 5  ·  CF internal: Begin + NetChange == Ending Cash")
    print("─" * 78)
    # CF rows: 22=beginning, 21=net change, 23=ending
    ROW_CF_BEGIN = 22
    ROW_CF_CHANGE = 21
    ROW_CF_END = 23

    print(f"{'Period':<12} {'Begin':>12} {'Change':>12} {'Sum':>12} {'Ending':>12} {'Delta':>10} {'Status':>8}")
    print("-" * 80)

    t5_pass = True
    t5_flags = []
    for col, lbl in sorted(cf_period_cols.items()):
        begin  = safe_float(cf.cell(row=ROW_CF_BEGIN,  column=col).value)
        change = safe_float(cf.cell(row=ROW_CF_CHANGE, column=col).value)
        ending = safe_float(cf.cell(row=ROW_CF_END,    column=col).value)
        if begin is None or change is None or ending is None:
            print(f"{lbl:<12} {'N/A':>12} {'N/A':>12} {'N/A':>12} {'N/A':>12} {'—':>10} {'SKIP':>8}")
            continue
        computed = begin + change
        delta = computed - ending
        ok = abs(delta) <= TOLERANCE
        status = "OK" if ok else "FAIL"
        if not ok:
            t5_pass = False
            t5_flags.append((lbl, delta))
        print(f"{lbl:<12} {begin:>12,.2f} {change:>12,.2f} {computed:>12,.2f} "
              f"{ending:>12,.2f} {delta:>10.4f} {status:>8}")

    if t5_pass:
        print("\n  ✓ TEST 5 PASSED — CF beginning + change = ending for ALL periods.")
    else:
        print(f"\n  ✗ TEST 5 FAILED — {len(t5_flags)} period(s):")
        for lbl, d in t5_flags:
            print(f"    {lbl}: delta = {d:.4f}")

    # ══════════════════════════════════════════════════════════════
    #  TEST 6: CF section sums — OCF + ICF + FCF == Net Change
    # ══════════════════════════════════════════════════════════════
    print("\n" + "─" * 78)
    print("TEST 6  ·  CF section sums: OCF + ICF + FCF == Net Change")
    print("─" * 78)
    ROW_OCF = 10    # 1.4 Net Cash from Operations
    ROW_ICF = 14    # 2.3 Net Cash from Investing
    ROW_FCF = 19    # 3.4 Net Cash from Financing

    print(f"{'Period':<12} {'OCF':>10} {'ICF':>10} {'FCF':>10} {'Sum':>12} {'NetChg':>12} {'Delta':>10} {'Status':>8}")
    print("-" * 90)

    t6_pass = True
    t6_flags = []
    for col, lbl in sorted(cf_period_cols.items()):
        ocf = safe_float(cf.cell(row=ROW_OCF, column=col).value, 0)
        icf = safe_float(cf.cell(row=ROW_ICF, column=col).value, 0)
        fcf = safe_float(cf.cell(row=ROW_FCF, column=col).value, 0)
        net_chg = safe_float(cf.cell(row=ROW_CF_CHANGE, column=col).value)
        if net_chg is None:
            print(f"{lbl:<12} {'N/A':>10} {'N/A':>10} {'N/A':>10} {'N/A':>12} {'N/A':>12} {'—':>10} {'SKIP':>8}")
            continue
        computed = ocf + icf + fcf
        delta = computed - net_chg
        ok = abs(delta) <= TOLERANCE
        status = "OK" if ok else "FAIL"
        if not ok:
            t6_pass = False
            t6_flags.append((lbl, delta))
        print(f"{lbl:<12} {ocf:>10,.2f} {icf:>10,.2f} {fcf:>10,.2f} "
              f"{computed:>12,.2f} {net_chg:>12,.2f} {delta:>10.4f} {status:>8}")

    if t6_pass:
        print("\n  ✓ TEST 6 PASSED — OCF+ICF+FCF = Net Change for ALL periods.")
    else:
        print(f"\n  ✗ TEST 6 FAILED — {len(t6_flags)} period(s):")
        for lbl, d in t6_flags:
            print(f"    {lbl}: delta = {d:.4f}")

    # ══════════════════════════════════════════════════════════════
    #  SUMMARY
    # ══════════════════════════════════════════════════════════════
    print("\n" + "=" * 78)
    print("AUDIT SUMMARY — Balance Sheet Integrity")
    print("=" * 78)
    all_tests = [
        ("TEST 1", "Assets == L+E (every period)",            t1_pass),
        ("TEST 2", "Model Control Row C1 == 0",               t2_pass),
        ("TEST 3", "Cash BS == Ending Cash CF",               t3_pass),
        ("TEST 4", "Component sums == Totals",                t4_pass),
        ("TEST 5", "CF: Begin + Change == Ending",            t5_pass),
        ("TEST 6", "CF: OCF + ICF + FCF == Net Change",       t6_pass),
    ]
    n_pass = sum(1 for _, _, p in all_tests if p)
    n_total = len(all_tests)
    for tag, desc, passed in all_tests:
        mark = "PASS" if passed else "FAIL"
        sym  = "✓" if passed else "✗"
        print(f"  {sym} {tag}  {desc:<45s}  [{mark}]")

    print(f"\nOverall: {n_pass}/{n_total} tests passed.")
    overall = "PASS" if n_pass == n_total else "FAIL"
    print(f"BLOCK B BALANCE-SHEET AUDIT VERDICT: ** {overall} **")
    print("=" * 78)

    wb.close()
    return 0 if n_pass == n_total else 1


if __name__ == "__main__":
    sys.exit(main())
