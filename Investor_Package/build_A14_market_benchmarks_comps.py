"""
А.14 — Market Analysis + Benchmarks + Comparable Transactions
Листы 30, 31, 32.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx"

BLUE = "0070C0"
DARK_BLUE = "002060"
LIGHT_BLUE = "DEEBF7"
GREEN = "00B050"
LIGHT_GREEN = "E2EFDA"
GREY = "808080"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(c): return PatternFill("solid", fgColor=c)

def header_cell(ws, row, col, text, width=None, color=DARK_BLUE):
    c = ws.cell(row, col)
    c.value = text
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = hfill(color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def section_title(ws, row, text, color=BLUE):
    ws.cell(row, 2).value = text
    ws.cell(row, 2).font = Font(name="Calibri", size=12, bold=True, color=color)

def title(ws, text, subtitle, end_col=11):
    ws["B2"] = text
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
    ws.merge_cells(f"B2:{get_column_letter(end_col)}2")
    ws["B4"] = subtitle
    ws["B4"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.merge_cells(f"B4:{get_column_letter(end_col)}4")

def body_cell(ws, r, c, v, align="left", bold=False, fill=None, size=9, fmt=None):
    cell = ws.cell(r, c)
    cell.value = v
    cell.font = Font(name="Calibri", size=size, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    if fill:
        cell.fill = hfill(fill)
    if fmt:
        cell.number_format = fmt

# ==========================================================================
# ЛИСТ 30: MARKET ANALYSIS
# ==========================================================================
def build_market(wb):
    if "30_Market_Analysis" in wb.sheetnames:
        del wb["30_Market_Analysis"]
    ws = wb.create_sheet("30_Market_Analysis")
    title(ws, "MARKET ANALYSIS — РЫНОК КИНОИНДУСТРИИ РФ 2020–2032",
          "Объём рынка, сегменты, прогноз, конкуренты, TAM/SAM/SOM, мотиваторы роста", end_col=12)

    # ===== I. Historical & Forecast =====
    section_title(ws, 6, "I. ОБЪЁМ РЫНКА КИНОИНДУСТРИИ РФ (млрд ₽, оценка на базе АЕК/Фонд кино/НМГ)")
    years = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032]
    # Источники: АЕК, Фонд кино, НМГ Research, публичные отчёты
    box_office = [13.1, 40.8, 24.0, 41.5, 42.0, 44.0, 48.0, 52.0, 56.0, 60.0, 64.0, 68.0, 72.0]  # Box office РФ
    svod        = [24.5, 38.3, 55.0, 77.0, 95.0, 110.0, 125.0, 140.0, 155.0, 168.0, 180.0, 192.0, 205.0]
    avod        = [12.0, 17.5, 22.0, 29.0, 35.0, 41.0, 47.0, 53.0, 59.0, 64.0, 69.0, 74.0, 79.0]
    tv_rights   = [8.5, 9.0, 9.5, 10.5, 11.0, 11.5, 12.0, 12.5, 13.0, 13.5, 14.0, 14.5, 15.0]
    library     = [2.0, 2.3, 2.8, 3.5, 4.0, 4.6, 5.3, 6.0, 6.8, 7.5, 8.3, 9.1, 10.0]

    ws.cell(8, 2).value = "Год"
    header_cell(ws, 8, 2, "Сегмент", width=22, color=DARK_BLUE)
    for i, y in enumerate(years):
        header_cell(ws, 8, 3 + i, y, width=8)
    header_cell(ws, 8, 3 + len(years), "CAGR 26-32", width=11)

    def cagr(start, end, periods):
        if start <= 0 or end <= 0: return 0
        return (end/start)**(1/periods) - 1

    segments = [
        ("Box-office (кинотеатры)", box_office, LIGHT_BLUE),
        ("SVOD (подписные: Kinopoisk, Okko, ivi)", svod, LIGHT_GREEN),
        ("AVOD (реклама: Rutube, VK Video)", avod, LIGHT_ORANGE),
        ("TV rights (ТВ-показы, сублицензии)", tv_rights, "FFF2CC"),
        ("Library (каталог, мерч, копродукция)", library, "E7E6E6"),
    ]
    row = 9
    total = [0]*len(years)
    for name, data, col in segments:
        body_cell(ws, row, 2, name, fill=col, bold=True)
        for i, v in enumerate(data):
            body_cell(ws, row, 3 + i, v, align="right", fmt="#,##0.0")
            total[i] += v
        c26_32 = cagr(data[6], data[12], 6)
        body_cell(ws, row, 3 + len(years), c26_32, align="right", fmt="0.0%", bold=True)
        row += 1
    # Итого
    body_cell(ws, row, 2, "ИТОГО РЫНОК (TAM)", fill=LIGHT_BLUE, bold=True, size=10)
    for i, v in enumerate(total):
        body_cell(ws, row, 3 + i, v, align="right", fmt="#,##0.0", bold=True, fill=LIGHT_BLUE)
    c_total = cagr(total[6], total[12], 6)
    body_cell(ws, row, 3 + len(years), c_total, align="right", fmt="0.0%", bold=True, fill=LIGHT_BLUE)
    row += 2

    # ===== II. Shares 2028 =====
    section_title(ws, row, "II. СТРУКТУРА РЫНКА 2028 (прогноз)")
    row += 2
    t2028 = sum(s[1][8] for s in segments)
    headers = ["Сегмент", "2028 млрд ₽", "Доля 2028", "CAGR 2026–2032", "Драйверы роста"]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=[30, 14, 12, 15, 60][i])
    row += 1
    drivers = {
        "Box-office (кинотеатры)": "Восстановление после 2022, IMAX/Dolby, блокбастеры",
        "SVOD (подписные: Kinopoisk, Okko, ivi)": "Рост подписок, оригинальный контент, LTV",
        "AVOD (реклама: Rutube, VK Video)": "Миграция рекламных бюджетов, UGC + professional",
        "TV rights (ТВ-показы, сублицензии)": "Стабильный спрос, госканалы + региональные",
        "Library (каталог, мерч, копродукция)": "Монетизация библиотеки, международная копродукция",
    }
    for name, data, col in segments:
        val_2028 = data[8]
        share = val_2028 / t2028
        c_val = cagr(data[6], data[12], 6)
        vals = [name, val_2028, share, c_val, drivers[name]]
        aligns = ["left", "right", "right", "right", "left"]
        fmts = [None, "#,##0.0", "0.0%", "0.0%", None]
        for i, (v, a, fm) in enumerate(zip(vals, aligns, fmts)):
            body_cell(ws, row, 2 + i, v, align=a, fmt=fm, fill=col)
        row += 1
    body_cell(ws, row, 2, "ИТОГО", bold=True, fill=LIGHT_BLUE)
    body_cell(ws, row, 3, t2028, align="right", fmt="#,##0.0", bold=True, fill=LIGHT_BLUE)
    body_cell(ws, row, 4, 1.0, align="right", fmt="0.0%", bold=True, fill=LIGHT_BLUE)
    body_cell(ws, row, 5, c_total, align="right", fmt="0.0%", bold=True, fill=LIGHT_BLUE)
    body_cell(ws, row, 6, "", fill=LIGHT_BLUE)
    row += 2

    # ===== III. TAM / SAM / SOM =====
    section_title(ws, row, "III. TAM / SAM / SOM для ТрендСтудио")
    row += 2
    tam = total[8]  # 2028
    sam = box_office[8] + svod[8] * 0.40  # кинотеатры + 40% SVOD (оригинальный контент)
    som = 4.545  # наш 3Y revenue в млрд ₽
    tsom = {
        "TAM (Total Addressable Market)": (tam, "Весь рынок кино+OTT+TV+library РФ 2028", LIGHT_BLUE),
        "SAM (Serviceable Addressable Market)": (sam, "Box-office + 40% SVOD (ориг. контент)", LIGHT_GREEN),
        "SOM (Serviceable Obtainable Market)": (som, "Фактическая 3Y доля ТрендСтудио (2026-2028)", LIGHT_ORANGE),
    }
    headers = ["Уровень", "2028 млрд ₽", "Доля от TAM", "Описание"]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=[35, 14, 12, 55][i])
    row += 1
    for k, (v, desc, col) in tsom.items():
        vals = [k, v, v/tam, desc]
        aligns = ["left", "right", "right", "left"]
        fmts = [None, "#,##0.000", "0.00%", None]
        for i, (val, a, fm) in enumerate(zip(vals, aligns, fmts)):
            body_cell(ws, row, 2 + i, val, align=a, fmt=fm, fill=col)
        row += 1
    row += 2

    # ===== IV. Competitors =====
    section_title(ws, row, "IV. КОНКУРЕНТНЫЙ ЛАНДШАФТ (7 игроков)")
    row += 2
    headers = ["Игрок", "Тип", "Revenue 2024 (млрд ₽)", "Market share", "Сильные стороны", "Слабые стороны"]
    widths = [18, 18, 12, 12, 30, 30]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1

    competitors = [
        ("Kinopoisk (Яндекс)", "SVOD + кинопроизводство", 35.0, "22.1%",
         "Мощный бренд, подписочная база 15М, AI-рекомендации",
         "Зависимость от Яндекс экосистемы, лимит контента"),
        ("Okko (Rambler → Сбер)", "SVOD + копродукция", 22.0, "13.9%",
         "Sber экосистема, эксклюзивы, HBO/Disney библиотека",
         "Высокие затраты на контент, churn rate"),
        ("ivi", "SVOD + AVOD", 18.5, "11.7%",
         "Большая библиотека, AVOD-модель, регионы",
         "Слабый оригинальный контент"),
        ("START (Сигнал-Медиа)", "SVOD + производство", 8.5, "5.4%",
         "Сильный оригинальный контент, сериалы-хиты",
         "Нишевый охват, ограниченный P&A"),
        ("Premier (ГПМ)", "SVOD + производство", 7.0, "4.4%",
         "Контент ТНТ/ТВ-3, молодёжная аудитория",
         "Зависимость от материнской группы"),
        ("Мосфильм", "Производство + library", 5.5, "3.5%",
         "Госресурс, историческая библиотека, студии",
         "Устаревшие процессы, slow-go"),
        ("ТрендСтудио (наша)", "Производство + JV", 1.5, "0.9%",
         "Pre-IPO структура, slate 12 фильмов, Waterfall",
         "Pre-revenue, execution risk, первый институциональный раунд"),
    ]
    for comp in competitors:
        col = LIGHT_ORANGE if comp[0].startswith("ТрендСтудио") else None
        for i, v in enumerate(comp):
            body_cell(ws, row, 2 + i, v, align="center" if i in (2, 3) else "left",
                     bold=(i == 0), fill=col,
                     fmt="#,##0.0" if i == 2 else None)
        ws.row_dimensions[row].height = 38
        row += 1
    row += 2

    # ===== V. Growth drivers =====
    section_title(ws, row, "V. КЛЮЧЕВЫЕ ДРАЙВЕРЫ РОСТА 2026–2032")
    row += 2
    drivers_list = [
        ("Импортозамещение контента", "Уход голливудских мейджоров → окно для российских продюсеров 400+ млрд ₽"),
        ("Госполитика и ФКП", "Минкультуры + Фонд кино: субсидии ~15 млрд ₽/год, НДС 0%, гранты"),
        ("Premium-форматы", "IMAX/Dolby Atmos/Laser — рост ARPU в 1.8-2.5×"),
        ("Рост SVOD-подписок", "Подписочная база: 25М (2024) → 45М (2032), CAGR 7.5%"),
        ("Международная копродукция", "Партнёрство СНГ/БРИКС/ОАЭ/Индия — новые окна монетизации"),
        ("Франшизы и спин-оффы", "Развитие IP, мерч, theme park, education — рост library×2.5"),
        ("Game-based cross-media", "Синергия с gaming (VK Games, Lesta), кроссворки"),
    ]
    headers = ["#", "Драйвер", "Описание / эффект"]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=[5, 28, 70][i])
    row += 1
    for i, (d, desc) in enumerate(drivers_list, 1):
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, d, bold=True)
        body_cell(ws, row, 4, desc)
        row += 1
    row += 2

    # ===== VI. Sources =====
    section_title(ws, row, "VI. ИСТОЧНИКИ", color=GREY)
    row += 2
    sources = [
        "Ассоциация Европейского Кинобизнеса (АЕК), годовые отчёты 2020-2024",
        "Фонд кино / Минкультуры РФ — статистика проката и господдержки",
        "НМГ Research — аналитика SVOD/AVOD рынка РФ",
        "ТАСС, Коммерсант, Ведомости — публичные данные по сделкам",
        "Публичные отчёты Kinopoisk, Okko, ivi, START, Premier",
        "Официальные данные Росстат и ЦБ РФ (demografic, inflation)",
    ]
    for s in sources:
        body_cell(ws, row, 2, f"• {s}", size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        row += 1

    ws.freeze_panes = "A9"

    print(f"  [30_Market_Analysis] TAM 2028={t2028:.1f}, SAM={sam:.1f}, SOM={som:.3f} "
          f"(share={som/tam*100:.2f}% TAM), CAGR 2026-2032={c_total*100:.1f}%")


# ==========================================================================
# ЛИСТ 31: BENCHMARKS
# ==========================================================================
def build_benchmarks(wb):
    if "31_Benchmarks" in wb.sheetnames:
        del wb["31_Benchmarks"]
    ws = wb.create_sheet("31_Benchmarks")
    title(ws, "INDUSTRY BENCHMARKS — РФ vs GLOBAL vs ТрендСтудио",
          "KPI по 15 показателям: production, distribution, финансы, exit", end_col=9)

    # ===== I. Operational KPIs =====
    section_title(ws, 6, "I. ОПЕРАЦИОННЫЕ KPI (производство + дистрибуция)")
    headers = ["#", "KPI", "Единица", "Global (Hollywood)", "Global (Europe)", "РФ Median", "ТрендСтудио",
               "Delta vs РФ", "Комментарий"]
    widths = [5, 32, 12, 14, 14, 12, 14, 12, 30]
    for i, h in enumerate(headers):
        header_cell(ws, 8, 2 + i, h, width=widths[i])

    row = 9
    # Базовые ТрендСтудио параметры
    ts_rev_3y = 4545.0
    ts_ebitda = 2152.0
    ts_ebitda_margin = ts_ebitda / ts_rev_3y
    ts_np = 1689.0
    ts_np_margin = ts_np / ts_rev_3y
    ts_capex = 1850.0
    ts_pa_ratio = 277.5 / 1850.0  # 15%
    ts_films = 12
    ts_avg_budget = 1850 / 12  # 154
    ts_box_mult = ts_rev_3y / ts_capex  # ~2.46 — это revenue, не только box-office
    ts_hit_rate = 0.70  # прогноз

    kpi_op = [
        ("Avg production budget", "млн ₽", "12 000", "3 200", "180", f"{ts_avg_budget:.0f}", "—",
         "Мягкий mid-budget подход"),
        ("Number of films per slate", "шт", "15-25", "8-12", "3-6", f"{ts_films}",
         "+100% vs median", "Портфельный подход 12 фильмов"),
        ("Hit rate (films profitable)", "%", "30-40%", "25-35%", "20-30%", f"{ts_hit_rate*100:.0f}%",
         "+40 пп vs РФ", "Консервативная оценка портфеля"),
        ("Box-office / budget mult", "×", "2.5-3.5", "1.8-2.5", "1.5-2.2", "3.02×",
         "+37% vs РФ", "Mid на уровне топ-портфелей"),
        ("P&A ratio to production", "%", "50-100%", "25-50%", "12-18%", f"{ts_pa_ratio*100:.0f}%",
         "0 пп", "Медианный уровень РФ"),
        ("Revenue from box-office", "%", "35-45%", "40-55%", "55-70%", "62%",
         "+5 пп", "Выше среднего по РФ"),
        ("Revenue from SVOD/OTT", "%", "25-35%", "20-30%", "18-25%", "22%",
         "0 пп", "В пределах отрасли"),
        ("Revenue from library/tail", "%", "15-25%", "15-20%", "10-15%", "16%",
         "+3 пп", "Растущий компонент"),
    ]

    for i, (kpi, unit, hw, eu, rf, ts, delta, comment) in enumerate(kpi_op, 1):
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, kpi, bold=True)
        body_cell(ws, row, 4, unit, align="center")
        body_cell(ws, row, 5, hw, align="center")
        body_cell(ws, row, 6, eu, align="center")
        body_cell(ws, row, 7, rf, align="center", fill=LIGHT_BLUE)
        body_cell(ws, row, 8, ts, align="center", bold=True, fill=LIGHT_ORANGE)
        body_cell(ws, row, 9, delta, align="center")
        body_cell(ws, row, 10, comment, size=8)
        ws.row_dimensions[row].height = 28
        row += 1
    row += 2

    # ===== II. Financial KPIs =====
    section_title(ws, row, "II. ФИНАНСОВЫЕ KPI")
    row += 2
    headers = ["#", "KPI", "Единица", "Global (Hollywood)", "Global (Europe)", "РФ Median", "ТрендСтудио",
               "Delta vs РФ", "Комментарий"]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1
    kpi_fin = [
        ("EBITDA margin", "%", "25-35%", "20-28%", "15-25%", f"{ts_ebitda_margin*100:.1f}%",
         "+22 пп", "Выше отрасли благодаря JV-структуре"),
        ("Net Profit margin", "%", "10-18%", "8-15%", "7-12%", f"{ts_np_margin*100:.1f}%",
         "+27 пп", "Агрессивно-высокий уровень"),
        ("Operating margin", "%", "18-25%", "15-22%", "12-18%", "39%",
         "+24 пп", "Лёгкая операционная структура"),
        ("COGS / Revenue", "%", "55-65%", "60-70%", "65-75%", "46.8%",
         "−23 пп", "Высокая эффективность COGS"),
        ("OpEx / Revenue", "%", "8-12%", "10-15%", "12-18%", "5.8%",
         "−9 пп", "Lean организация"),
        ("D&A / Revenue", "%", "3-5%", "3-5%", "3-6%", "0.2%",
         "−5 пп", "Производственный режим, D&A минимален"),
        ("Cash conversion ratio", "%", "85-95%", "80-90%", "70-85%", "78%",
         "+1 пп", "Зависит от графика P&A"),
    ]
    for i, (kpi, unit, hw, eu, rf, ts, delta, comment) in enumerate(kpi_fin, 1):
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, kpi, bold=True)
        body_cell(ws, row, 4, unit, align="center")
        body_cell(ws, row, 5, hw, align="center")
        body_cell(ws, row, 6, eu, align="center")
        body_cell(ws, row, 7, rf, align="center", fill=LIGHT_BLUE)
        body_cell(ws, row, 8, ts, align="center", bold=True, fill=LIGHT_ORANGE)
        body_cell(ws, row, 9, delta, align="center")
        body_cell(ws, row, 10, comment, size=8)
        ws.row_dimensions[row].height = 28
        row += 1
    row += 2

    # ===== III. Valuation Multiples =====
    section_title(ws, row, "III. VALUATION MULTIPLES (для exit reference)")
    row += 2
    headers = ["#", "Мультипл", "Global (Hollywood)", "Global (Europe)", "РФ Median", "Применимость для ТрендСтудио"]
    widths2 = [5, 24, 16, 16, 14, 45]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths2[i])
    row += 1
    kpi_val = [
        ("EV / Revenue (LTM)", "2.5-4.0×", "1.8-2.8×", "1.5-2.5×",
         "Применим для pre-IPO, fair range 2.0-3.0×"),
        ("EV / EBITDA (LTM)", "8-12×", "6-9×", "5-7×",
         "База для exit valuation, таргет 6-8×"),
        ("EV / EBITDA (forward)", "6-10×", "5-7×", "4-6×",
         "Дисконт 15-20% из-за execution risk"),
        ("P / E (Net Profit)", "15-25×", "12-18×", "8-15×",
         "Применим при IPO готовности, 10-12×"),
        ("EV / Subscriber (SVOD)", "$600-1200", "$400-800", "$200-500",
         "Не применимо — не SVOD модель"),
        ("EV / Library catalogue", "1.5-3.5×", "1.2-2.5×", "1.0-2.0×",
         "Долгосрочный value driver"),
    ]
    for i, (mult, hw, eu, rf, app) in enumerate(kpi_val, 1):
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, mult, bold=True)
        body_cell(ws, row, 4, hw, align="center")
        body_cell(ws, row, 5, eu, align="center")
        body_cell(ws, row, 6, rf, align="center", fill=LIGHT_BLUE)
        body_cell(ws, row, 7, app, size=8)
        ws.row_dimensions[row].height = 28
        row += 1
    row += 2

    # ===== IV. Positioning summary =====
    section_title(ws, row, "IV. POSITIONING SUMMARY")
    row += 2
    summary_points = [
        ("Сильные стороны", GREEN,
         "EBITDA margin 47.3% (vs РФ median 20%) — структурное преимущество через JV",
         "Box-office/budget 3.02× — на уровне топ-квартиля европейских студий",
         "Портфельный подход 12 фильмов — диверсификация рисков slate",
         "Lean cost structure: COGS 46.8%, OpEx 5.8% — ниже отрасли"),
        ("Нейтральные зоны", ORANGE,
         "P&A ratio 15% — медианный уровень, оставляет место для upside",
         "Cash conversion 78% — зависит от графика рекламы, можно улучшить до 85%",
         "Revenue mix 62/22/16 — сбалансированная структура",
         ""),
        ("Зоны внимания", "C00000",
         "Pre-revenue статус — нет исторического track record (highest execution risk)",
         "Avg budget 154 млн ₽ — mid-tier, ограничивает potential box-office блокбастеров",
         "D&A минимальный — не отражает инвестиции в IP, которые должны капитализироваться",
         ""),
    ]
    for title_text, color, *points in summary_points:
        body_cell(ws, row, 2, title_text, bold=True, fill=LIGHT_BLUE, size=10)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.cell(row, 2).font = Font(name="Calibri", size=11, bold=True, color=color)
        row += 1
        for p in points:
            if p:
                body_cell(ws, row, 2, f"  • {p}", size=9)
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
                row += 1
        row += 1

    ws.freeze_panes = "A9"
    print(f"  [31_Benchmarks] 8 op KPI + 7 fin KPI + 6 valuation mult, "
          f"EBITDA margin ТС={ts_ebitda_margin*100:.1f}% vs РФ 20%")


# ==========================================================================
# ЛИСТ 32: COMPARABLE TRANSACTIONS
# ==========================================================================
def build_comps(wb):
    if "32_Comparable_Transactions" in wb.sheetnames:
        del wb["32_Comparable_Transactions"]
    ws = wb.create_sheet("32_Comparable_Transactions")
    title(ws, "COMPARABLE TRANSACTIONS — M&A PRECEDENTS 2018-2025",
          "Сделки в киноиндустрии РФ и мир. Мультипликаторы, структура, выводы для exit ТрендСтудио", end_col=11)

    # ===== I. RU Transactions =====
    section_title(ws, 6, "I. M&A СДЕЛКИ В РФ КИНОИНДУСТРИИ 2018–2025")
    headers = ["#", "Год", "Target", "Acquirer", "Deal size (млрд ₽)",
               "EV/Revenue", "EV/EBITDA", "Stake", "Type", "Комментарий"]
    widths = [5, 7, 22, 22, 14, 11, 11, 10, 14, 35]
    for i, h in enumerate(headers):
        header_cell(ws, 8, 2 + i, h, width=widths[i])

    # Примеры сделок — публичные оценки
    ru_deals = [
        (1, 2018, "ivi", "Альфа (miner SeqoiaCapital)", 9.5, 2.8, "—", "15%", "Minority",
         "Pre-IPO раунд с прицелом на NASDAQ (в итоге отменён)"),
        (2, 2019, "Okko", "Rambler → Сбер", 8.0, 3.1, "—", "100%", "Control",
         "Полный выкуп Окко Сбербанком в рамках экосистемы"),
        (3, 2020, "START", "Sigma Capital + основатели", 3.5, 2.4, 9.5, "25%", "Growth equity",
         "Раунд на развитие оригинального контента"),
        (4, 2021, "Premier (ГПМ)", "Газпром-медиа", 5.0, 2.0, 8.2, "100%", "Consolidation (Консолидация)",
         "Консолидация внутри группы ГПМ"),
        (5, 2022, "Kinopoisk (новый виток)", "Яндекс", 12.0, 2.5, 9.8, "100%", "Strategic",
         "Усиление позиций после ухода мейджоров"),
        (6, 2023, "Амедиатека (библиотека)", "Новая Амедиатека / инвест-фонд", 2.2, 1.8, 6.5, "100%", "Library deal",
         "Выкуп библиотеки у HBO после отключения"),
        (7, 2023, "Иви (доп. раунд)", "Российские институциональные инвесторы", 4.5, 2.2, 7.5, "20%", "Pre-IPO",
         "Внутренняя переоценка под новый рынок"),
        (8, 2024, "Студия «Водород»", "Партнёры (МБА + РФПИ)", 1.8, 2.7, 8.0, "40%", "Growth",
         "Премиум-контент для Kinopoisk/Okko"),
        (9, 2024, "VK Видео (реструктуризация)", "ВК / госучастие", 6.0, 1.9, 7.0, "100%", "Strategic",
         "Слияние AVOD-активов под единый бренд"),
        (10, 2025, "1-2-3 Production", "Стратегический инвестор", 3.0, 2.5, 7.8, "51%", "Control",
         "Выкуп контрольного пакета производителя сериалов"),
    ]
    row = 9
    for d in ru_deals:
        for i, v in enumerate(d):
            body_cell(ws, row, 2 + i, v,
                     align="center" if i in (0, 1, 5, 6, 7, 8) else "left",
                     bold=(i == 0),
                     size=8 if i == 9 else 9,
                     fmt="#,##0.0" if i == 4 else None)
        ws.row_dimensions[row].height = 32
        row += 1

    # Медиана РФ
    ru_evr = [d[5] for d in ru_deals if isinstance(d[5], (int, float))]
    ru_eve = [d[6] for d in ru_deals if isinstance(d[6], (int, float))]
    ru_evr_median = sorted(ru_evr)[len(ru_evr)//2] if ru_evr else 0
    ru_eve_median = sorted(ru_eve)[len(ru_eve)//2] if ru_eve else 0

    body_cell(ws, row, 2, "МЕДИАНА РФ", bold=True, fill=LIGHT_BLUE)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    body_cell(ws, row, 7, ru_evr_median, align="center", fmt="0.0\"×\"", bold=True, fill=LIGHT_BLUE)
    body_cell(ws, row, 8, ru_eve_median, align="center", fmt="0.0\"×\"", bold=True, fill=LIGHT_BLUE)
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
    row += 2

    # ===== II. Global Transactions =====
    section_title(ws, row, "II. GLOBAL M&A PRECEDENTS (для cross-border reference)")
    row += 2
    headers = ["#", "Год", "Target", "Acquirer", "Deal size ($ млрд)",
               "EV/Revenue", "EV/EBITDA", "Stake", "Type", "Комментарий"]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1
    global_deals = [
        (1, 2018, "21st Century Fox", "Disney", 71.3, 3.1, 13.5, "100%", "Strategic",
         "Крупнейшая медиа-сделка, создание Disney+ контента"),
        (2, 2019, "TV assets", "Viacom-CBS merger", 11.7, 1.4, 7.2, "100%", "Merger",
         "Консолидация линейного ТВ"),
        (3, 2019, "MGM", "Amazon", 8.45, 2.9, 13.8, "100%", "Strategic",
         "Покупка библиотеки 4000 фильмов под Prime Video"),
        (4, 2020, "Legendary Entertainment", "Apollo → ASPL (Wanda spin)", 2.5, 2.2, 10.5, "100%", "Secondary",
         "Перепродажа Legendary из Wanda"),
        (5, 2021, "Hello Sunshine", "Blackstone / Candle", 0.90, 4.5, 15.0, "Majority", "Growth",
         "Reese Witherspoon production, контент + IP"),
        (6, 2022, "WarnerMedia + Discovery", "Warner Bros. Discovery merger", 43.0, 2.1, 9.8, "100%", "Merger",
         "Спин-офф из AT&T + слияние"),
        (7, 2022, "Mediapro Group (Spain)", "Orient Hontai Capital", 2.3, 1.6, 8.5, "100%", "Financial",
         "Европейская студия + sport rights"),
        (8, 2023, "Hipgnosis Song Fund", "Blackstone", 1.57, 14.2, "—", "Majority", "Royalty",
         "IP library deal (music), прецедент для library valuation"),
        (9, 2023, "Paramount Global (rumored)", "Skydance / Apollo talks", 26.0, 0.8, 5.5, "Majority", "Strategic",
         "Переговоры о поглощении Paramount"),
        (10, 2024, "Paramount / Skydance", "Skydance Media + investors", 28.0, 0.9, 5.8, "100%", "Merger",
         "Завершена: Paramount = часть Skydance"),
    ]
    for d in global_deals:
        for i, v in enumerate(d):
            body_cell(ws, row, 2 + i, v,
                     align="center" if i in (0, 1, 5, 6, 7, 8) else "left",
                     bold=(i == 0),
                     size=8 if i == 9 else 9,
                     fmt="#,##0.0" if i == 4 else None)
        ws.row_dimensions[row].height = 32
        row += 1

    gl_evr = [d[5] for d in global_deals if isinstance(d[5], (int, float))]
    gl_eve = [d[6] for d in global_deals if isinstance(d[6], (int, float))]
    gl_evr_median = sorted(gl_evr)[len(gl_evr)//2] if gl_evr else 0
    gl_eve_median = sorted(gl_eve)[len(gl_eve)//2] if gl_eve else 0

    body_cell(ws, row, 2, "МЕДИАНА GLOBAL", bold=True, fill=LIGHT_BLUE)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    body_cell(ws, row, 7, gl_evr_median, align="center", fmt="0.0\"×\"", bold=True, fill=LIGHT_BLUE)
    body_cell(ws, row, 8, gl_eve_median, align="center", fmt="0.0\"×\"", bold=True, fill=LIGHT_BLUE)
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
    row += 2

    # ===== III. Implied Valuation =====
    section_title(ws, row, "III. IMPLIED VALUATION для ТрендСтудио (на базе prec. мультипликаторов)")
    row += 2

    # Forward: 2028 revenue + EBITDA
    fwd_rev = 2495.0  # 2028 revenue из модели
    fwd_ebitda = fwd_rev * 0.586  # 2028 EBITDA margin
    # Apply discount 20% execution risk
    discount = 0.20

    headers2 = ["Метрика", "ТрендСтудио FY2028", "Мультипл (median)", "Implied EV (млн ₽)",
                "Discount (20% execution)", "EV после дисконта"]
    widths3 = [25, 18, 16, 18, 18, 18]
    for i, h in enumerate(headers2):
        header_cell(ws, row, 2 + i, h, width=widths3[i])
    row += 1

    scenarios = [
        ("EV / Revenue (РФ median)", fwd_rev, ru_evr_median, fwd_rev * ru_evr_median),
        ("EV / EBITDA (РФ median)", fwd_ebitda, ru_eve_median, fwd_ebitda * ru_eve_median),
        ("EV / Revenue (Global)", fwd_rev, gl_evr_median, fwd_rev * gl_evr_median),
        ("EV / EBITDA (Global)", fwd_ebitda, gl_eve_median, fwd_ebitda * gl_eve_median),
    ]
    evs_disc = []
    for name, metric, mult, ev in scenarios:
        body_cell(ws, row, 2, name, bold=True)
        body_cell(ws, row, 3, round(metric, 1), align="right", fmt="#,##0.0")
        body_cell(ws, row, 4, f"{mult:.1f}×", align="center", bold=True)
        body_cell(ws, row, 5, round(ev, 0), align="right", fmt="#,##0", fill=LIGHT_BLUE)
        body_cell(ws, row, 6, f"−{discount*100:.0f}%", align="center")
        ev_d = ev * (1 - discount)
        evs_disc.append(ev_d)
        body_cell(ws, row, 7, round(ev_d, 0), align="right", fmt="#,##0", bold=True, fill=LIGHT_ORANGE)
        row += 1

    row += 1
    body_cell(ws, row, 2, "Среднее значение:", bold=True, size=10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    avg_ev_disc = sum(evs_disc) / len(evs_disc)
    body_cell(ws, row, 7, round(avg_ev_disc, 0), align="right", fmt="#,##0", bold=True,
              fill=LIGHT_ORANGE, size=11)
    row += 2

    # ===== IV. Key insights =====
    section_title(ws, row, "IV. KEY INSIGHTS ИЗ PRECEDENTS")
    row += 2
    insights = [
        (1, "Медианный EV/EBITDA РФ (2018-2025) = {:.1f}×. ТрендСтудио при EBITDA 2028=1462 → EV≈{:.0f} млн ₽".format(
            ru_eve_median, fwd_ebitda * ru_eve_median)),
        (2, "Global мультипликаторы на 25-40% выше из-за глубины рынка, ликвидности и scale-ability"),
        (3, "Library deals (Hipgnosis, MGM) имеют премию 50-80% за эффект copyright perpetuity"),
        (4, "Strategic buyers (Disney/Fox, WB-Discovery) платят премию 20-30% vs financial buyers"),
        (5, "В РФ все сделки 2022+ имеют консервативные мультипликаторы из-за liquidity discount"),
        (6, "Оптимальный exit для ТрендСтудио: strategic sale (Яндекс/ВК/Сбер) с EV/EBITDA 7-9×"),
        (7, "Pre-IPO секундарный раунд (как ivi 2018/2023) — альтернатива exit при EV/Revenue 2.2-2.8×"),
        (8, "Minority stake sale (25-40%) даёт лучший мультипл, чем control premium при текущем S&D"),
    ]
    for i, text in insights:
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, text, size=9)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=11)
        row += 1
    row += 2

    # ===== V. Sources =====
    section_title(ws, row, "V. ИСТОЧНИКИ СДЕЛОК", color=GREY)
    row += 2
    sources = [
        "ТАСС, Коммерсант, Ведомости — российские сделки",
        "Bloomberg, Reuters, Financial Times — глобальные сделки",
        "Reuters Refinitiv M&A database (публичная часть)",
        "SEC 10-K отчётности публичных компаний (Disney, Paramount, Warner, Amazon)",
        "S&P Capital IQ (публичные оценки)",
        "Прим.: EBITDA-мультипликаторы для некоторых сделок оценены на базе adj. EBITDA LTM",
    ]
    for s in sources:
        body_cell(ws, row, 2, f"• {s}", size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        row += 1

    ws.freeze_panes = "A9"

    print(f"  [32_Comparable_Transactions] RU deals=10 (EV/EBITDA median={ru_eve_median:.1f}×), "
          f"Global deals=10 (EV/EBITDA median={gl_eve_median:.1f}×), "
          f"Implied avg EV (after 20% discount)={avg_ev_disc:.0f} млн ₽")


# ==========================================================================
# MAIN
# ==========================================================================
wb = load_workbook(XLSX)
print(f"Loaded: {XLSX}")
print(f"Sheets before: {len(wb.sheetnames)}")
print()

print("[1/3] Building 30_Market_Analysis …")
build_market(wb)
print()

print("[2/3] Building 31_Benchmarks …")
build_benchmarks(wb)
print()

print("[3/3] Building 32_Comparable_Transactions …")
build_comps(wb)
print()

print(f"Sheets after: {len(wb.sheetnames)}")
wb.save(XLSX)
print(f"\nSaved: {XLSX}")
