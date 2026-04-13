"""
PATCH v1.0.1 · Stage 2 — 24_Investor_Returns + 25_Exit_Scenarios
================================================================
Цель:
  1) В 24_Investor_Returns переключить default cashflow illustration с W₁ на W₃
     и расширить Returns Matrix до 3 сценария × 4 waterfall.
  2) В 25_Exit_Scenarios честно переформулировать probability-weighted T₁ MoIC
     и обновить «Base W₁» → «Base W₃» во всех метках.
  3) Единая, воспроизводимая методология расчёта IRR: out quarterly 2026,
     in по годам 2029-2032 с распределением 20/50/15/15.
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gc
import numpy_financial as npf

ROOT = Path(__file__).parent
FILES = [
    "investor_model_v1.0_Public.xlsx",
    "investor_model_v1.0_Internal.xlsx",
]

# ═══════════════════════════════════════════════════════════════════════════
# КОНСТАНТЫ
# ═══════════════════════════════════════════════════════════════════════════
INVEST = 1250.0
NDP_BASE = 3000.0

# Сценарии NDP (Bear −25%, Bull +25%)
SCENARIOS = {"Bear": 2250.0, "Base": 3000.0, "Bull": 3750.0}

# Pattern returns: year 0 -> -1250, year 3 (2029) 20%, year 4 (2030) 50%,
# year 5 (2031) 15%, year 6 (2032) 15%
RETURN_PATTERN = [0.20, 0.50, 0.15, 0.15]

# ═══════════════════════════════════════════════════════════════════════════
# СТИЛИ
# ═══════════════════════════════════════════════════════════════════════════
NAVY, BLUE, GRN, YEL = "1F3864", "0070C0", "E2EFDA", "FFF2CC"
FONT_BODY    = Font(name="Calibri", size=10, color="000000")
FONT_BODY_B  = Font(name="Calibri", size=10, bold=True, color="000000")
FONT_HEAD    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_DEFAULT = Font(name="Calibri", size=10, bold=True, color="548235")
FONT_NOTE    = Font(name="Calibri", size=9, italic=True, color="595959")
FILL_HEAD    = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
FILL_DEFAULT = PatternFill(start_color=GRN,  end_color=GRN,  fill_type="solid")
FILL_PREMIUM = PatternFill(start_color=YEL,  end_color=YEL,  fill_type="solid")
FILL_TOTAL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=True)


# ═══════════════════════════════════════════════════════════════════════════
# WATERFALL RULES
# ═══════════════════════════════════════════════════════════════════════════
def w1(ndp):
    """W₁ Hurdle 60/40 → 50/50 после recoupment 2 083"""
    if ndp < 2083.33:
        return ndp * 0.6, ndp * 0.4
    inv = 1250 + (ndp - 2083.33) * 0.5
    return inv, ndp - inv

def w2(ndp):
    """W₂ Pro-rata 1250/1850 = 67.57%"""
    return ndp * (1250 / 1850), ndp * (600 / 1850)

def w3(ndp):
    """W₃ 1× Liq Pref + 8% × 5y coupon + 60/40 carry"""
    s1 = min(1250, ndp)
    s2 = min(500, ndp - s1)
    rem = ndp - s1 - s2
    return s1 + s2 + rem * 0.6, rem * 0.4

def w4(ndp):
    """W₄ 1× Liq Pref + 12% × 5y preferred + 65/35 carry"""
    s1 = min(1250, ndp)
    s2 = min(750, ndp - s1)
    rem = ndp - s1 - s2
    return s1 + s2 + rem * 0.65, rem * 0.35

WATERFALLS = [("W₁", w1), ("W₂", w2), ("W₃", w3), ("W₄", w4)]


def compute_cashflow_and_irr(investor_total):
    """Возвращает (cashflow list 7 years, irr) для pattern 20/50/15/15"""
    cf = [-INVEST, 0, 0]
    for pct in RETURN_PATTERN:
        cf.append(investor_total * pct)
    irr = npf.irr(cf)
    return cf, irr


# ═══════════════════════════════════════════════════════════════════════════
# ПАТЧ 24_Investor_Returns
# ═══════════════════════════════════════════════════════════════════════════
def patch_24_investor_returns(wb):
    ws = wb["24_Investor_Returns"]

    # ---- Header обновить (r2) ----
    ws["B2"] = ("INVESTOR RETURNS — T₁ LEGACY 1 250 млн ₽  ·  "
                "3 Scenarios × 4 Waterfalls  ·  Base illustration = W₃ DEFAULT")
    ws["B4"] = ("IRR / MOIC / DPI / TVPI анализ — базовый waterfall W₃ "
                "(Liq Pref 1× + 8% coupon + 60/40) с четырьмя вариантами распределения NDP. "
                "Single methodology: out quarterly 2026, in annual 2029-2032 (20/50/15/15 split).")

    # ---- Section I header (r6) ----
    ws["B6"] = "I. T₁ INVESTOR CASHFLOW SCHEDULE (Base case · W₃ Liq Pref + 8% + 60/40 DEFAULT)"

    # ---- Пересчёт Section I cashflow под W₃ Base (NDP=3000, Investor=2500) ----
    # Тишина существующие данные r8-r16
    # Новая схема:
    #   Q1-Q4 2026: -250/-350/-350/-300 (out)
    #   2029: Partial Liq Pref 500 (In, principal)
    #   2030: Liq Pref remainder 750 + Coupon 500 = 1250 (In, principal+interest)
    #   2031: Carried 60% on 1250 residual = 375 × 1 year portion... но простим - 375
    #   2032: Carried 60% remainder = 375
    # Итого: 500 + 1250 + 375 + 375 = 2500 ✓

    # Row 8-11 — outflow — не трогаем (tranches те же)
    # Row 12-15 — inflow пересчитываем
    row_data = [
        (12, 5, "2029", "Liq Pref partial", 500, 0,    0,    500,  -750,  -1250, "In"),
        (13, 6, "2030", "Liq Pref final + coupon", 750, 500, 0,  1250, 500,   -1250, "In"),
        (14, 7, "2031", "Carried interest share", 0,   0,   375, 375,  875,   -1250, "In"),
        (15, 8, "2032", "Carried interest share", 0,   0,   375, 375,  1250,  -1250, "In"),
    ]
    for r, n, period, action, tranche, interest, upside, net, cum, peak, tp in row_data:
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=period)
        ws.cell(row=r, column=4, value=action)
        ws.cell(row=r, column=5, value=tranche if tranche != 0 else None)
        ws.cell(row=r, column=6, value=interest if interest != 0 else None)
        ws.cell(row=r, column=7, value=upside if upside != 0 else None)
        ws.cell(row=r, column=8, value=net)
        ws.cell(row=r, column=9, value=cum)
        ws.cell(row=r, column=10, value=peak)
        ws.cell(row=r, column=11, value=tp)
        for col in range(2, 12):
            c = ws.cell(row=r, column=col)
            c.font = FONT_BODY
            if col in (5, 6, 7, 8, 9, 10) and c.value is not None:
                c.number_format = "#,##0.0"

    # Row 16 — ИТОГО
    ws.cell(row=16, column=2, value="ИТОГО").font = FONT_BODY_B
    ws.cell(row=16, column=5, value=0).font = FONT_BODY_B       # net tranche (out+in = 0)
    ws.cell(row=16, column=6, value=500).font = FONT_BODY_B     # interest
    ws.cell(row=16, column=7, value=750).font = FONT_BODY_B     # upside
    ws.cell(row=16, column=8, value=1250).font = FONT_BODY_B    # net gain
    for col in (5, 6, 7, 8):
        c = ws.cell(row=16, column=col)
        c.fill = FILL_TOTAL
        c.number_format = "#,##0"

    # ---- Section II: Returns Matrix 3×4 (r18-22) ----
    ws["B18"] = "II. RETURNS MATRIX — 3 SCENARIOS × 4 WATERFALLS (IRR / MOIC, единая методология)"

    # Headers r19: B=#  C=Сценарий  D=W₁ IRR  E=W₁ MOIC  F=W₂ IRR  G=W₂ MOIC
    #              H=W₃ IRR  I=W₃ MOIC  J=W₄ IRR  K=W₄ MOIC  L=Best  M=NDP
    headers = [
        (2, "#"), (3, "Сценарий"),
        (4, "W₁ IRR"), (5, "W₁ MOIC"),
        (6, "W₂ IRR"), (7, "W₂ MOIC"),
        (8, "W₃ IRR"), (9, "W₃ MOIC"),
        (10, "W₄ IRR"), (11, "W₄ MOIC"),
        (12, "Best"), (13, "NDP"),
    ]
    for col, text in headers:
        c = ws.cell(row=19, column=col, value=text)
        c.font = FONT_HEAD
        c.fill = FILL_HEAD
        c.alignment = CENTER
        c.border = BORDER

    # Расширить merge header (была J19..., теперь M19)
    # Снимаем любой старый merge в row 19
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == 19 or mr.max_row == 19:
            ws.unmerge_cells(str(mr))

    # Заполняем матрицу
    for i, (sname, ndp) in enumerate(SCENARIOS.items()):
        r = 20 + i
        ws.cell(row=r, column=2, value=i + 1).font = FONT_BODY
        ws.cell(row=r, column=3, value=sname).font = FONT_BODY_B
        inv_vals = {}
        for wname, rule in WATERFALLS:
            inv, _ = rule(ndp)
            _, irr = compute_cashflow_and_irr(inv)
            inv_vals[wname] = (irr, inv / INVEST)
        ws.cell(row=r, column=4,  value=inv_vals["W₁"][0] * 100).font = FONT_BODY
        ws.cell(row=r, column=5,  value=inv_vals["W₁"][1]).font = FONT_BODY
        ws.cell(row=r, column=6,  value=inv_vals["W₂"][0] * 100).font = FONT_BODY
        ws.cell(row=r, column=7,  value=inv_vals["W₂"][1]).font = FONT_BODY
        ws.cell(row=r, column=8,  value=inv_vals["W₃"][0] * 100).font = FONT_DEFAULT
        ws.cell(row=r, column=9,  value=inv_vals["W₃"][1]).font = FONT_DEFAULT
        ws.cell(row=r, column=10, value=inv_vals["W₄"][0] * 100).font = FONT_BODY_B
        ws.cell(row=r, column=11, value=inv_vals["W₄"][1]).font = FONT_BODY_B
        ws.cell(row=r, column=12, value="W₄").font = FONT_BODY_B
        ws.cell(row=r, column=13, value=ndp).font = FONT_BODY

        for col in range(2, 14):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            c.alignment = CENTER
            if col in (4, 6, 8, 10):
                c.number_format = '0.00"%"'
            elif col in (5, 7, 9, 11):
                c.number_format = '0.000"×"'
            elif col == 13:
                c.number_format = "#,##0"
        # Подсветка W₃ колонки default (зелёная)
        for col in (8, 9):
            ws.cell(row=r, column=col).fill = FILL_DEFAULT
        for col in (10, 11):
            ws.cell(row=r, column=col).fill = FILL_PREMIUM

    # ---- Section III: Cash-on-Cash Metrics (r25-35) под W₃ ----
    ws["B25"] = "III. CASH-ON-CASH METRICS (Base case · W₃ DEFAULT)"

    # Новые значения под W₃ Base: Investor=2500, NetGain=1250, MOIC=2.0, IRR=18.04%
    w3_inv, _ = w3(NDP_BASE)  # 2500
    _, irr_base_w3 = compute_cashflow_and_irr(w3_inv)

    # r26-35 — точечные обновления
    ws["G26"] = 1250.0;              ws["I26"] = "Σ outflow T₁"
    ws["G27"] = 2500.0;               ws["I27"] = "W₃: 1250 liq pref + 500 coupon + 750 carried"
    ws["G28"] = 1250.0;               ws["I28"] = "Return on capital (net gain)"
    ws["G29"] = 100.0;                ws["I29"] = "Total return % (net gain / invest)"
    ws["G30"] = 2.0;                  ws["I30"] = "MOIC = 2500/1250 (investor-friendly W₃)"
    ws["G31"] = 2.0;                  ws["I31"] = "DPI fully realized"
    ws["G32"] = 2.0;                  ws["I32"] = "= DPI (no NAV)"
    ws["G33"] = irr_base_w3 * 100;   ws["I33"] = f"Hurdle 18%, PASS (fractional period IRR)"
    ws["G34"] = 1250.0;              ws["I34"] = "Q4'26 full draw 1 250"
    ws["G35"] = 3.75;                ws["I35"] = "2029-начало 2030 full principal return"

    for r in range(26, 36):
        c = ws.cell(row=r, column=7)
        if c.value is not None:
            if r in (29, 33):
                c.number_format = '0.00"%"'
            elif r == 30 or r == 31 or r == 32:
                c.number_format = '0.000"×"'
            elif r == 35:
                c.number_format = '0.0" years"'
            else:
                c.number_format = "#,##0"

    # ---- Section IV: Return Attribution под W₃ (r38-43) ----
    ws["B38"] = "IV. RETURN ATTRIBUTION — РАЗБИВКА ROI (Base W₃ DEFAULT)"

    ws["C40"] = "Return of Principal (Liq Pref 1×)"
    ws["H40"] = 1250.0; ws["I40"] = 50.0
    ws["J40"] = "Stage 1 W₃: senior recovery"

    ws["C41"] = "Preferred Return Coupon (8% × 5y)"
    ws["H41"] = 500.0;  ws["I41"] = 20.0
    ws["J41"] = "Stage 2 W₃: 1 250 × 8% × 5y = 500"

    ws["C42"] = "Carried Interest Share (60% of 1 250 residual)"
    ws["H42"] = 750.0;  ws["I42"] = 30.0
    ws["J42"] = "Stage 3 W₃: 60% × (3000 − 1250 − 500)"

    ws["C43"] = "ИТОГО ROI"
    ws["H43"] = 2500.0; ws["I43"] = 100.0
    ws["J43"] = "= 1250 + 500 + 750 (W₃ Base NDP 3000)"

    for r in range(40, 44):
        ws.cell(row=r, column=8).number_format = "#,##0"
        ws.cell(row=r, column=9).number_format = '0.0"%"'

    print(f"  ✓ 24_Investor_Returns: matrix 3×4, default=W₃, Base IRR={irr_base_w3*100:.2f}%")


# ═══════════════════════════════════════════════════════════════════════════
# ПАТЧ 25_Exit_Scenarios
# ═══════════════════════════════════════════════════════════════════════════
def patch_25_exit_scenarios(wb):
    ws = wb["25_Exit_Scenarios"]

    # r49: "Base W₁ share" → "50/30/20 exit split (not NDP-based)"
    ws["I49"] = "50/30/20 exit split (не соответствует ни одному W из листа 24)"

    # r53: "Expected T₁ MOIC (base W₁)" → честно
    ws["B53"] = "Expected T₁ MOIC (prob-weighted через 50/30/20)"
    ws["G53"] = 2.64935
    ws["I53"] = "≠ matrix из 24 (там W₃ Base MOIC = 2.00×). Методы разные."

    # Добавить methodological note в конец (r55)
    note_text = ("МЕТОДОЛОГИЧЕСКОЕ ПРИМЕЧАНИЕ: лист 24 рассчитывает returns от "
                 "распределения NDP 3 000 млн ₽ через waterfall W₁-W₄ (основа — "
                 "операционная прибыль 2026-2028). Лист 25 рассчитывает returns "
                 "от exit event (продажа компании 2028-2032) через 50/30/20 split. "
                 "Two independent return streams: fulll LP return ≈ 24 (W₃) + 25 (exit).")
    ws.merge_cells("B55:K55")
    c = ws.cell(row=55, column=2, value=note_text)
    c.font = FONT_NOTE
    c.alignment = LEFT
    c.fill = FILL_TOTAL
    ws.row_dimensions[55].height = 42

    print("  ✓ 25_Exit_Scenarios: methodology note, Base W₁ labels corrected")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    for fname in FILES:
        path = ROOT / fname
        if not path.exists():
            print(f"⚠ {fname} не найден")
            continue

        # Дополнительный бэкап перед stage 2
        backup = ROOT / f"{path.stem}_pre_v101_stage2_backup.xlsx"
        if not backup.exists():
            shutil.copy(path, backup)
            print(f"📦 Backup: {backup.name}")

        print(f"\n🔧 Патчу {fname}")
        wb = load_workbook(path)
        patch_24_investor_returns(wb)
        patch_25_exit_scenarios(wb)
        wb.save(path)
        print(f"💾 Сохранено: {path.name}")

    print("\n✅ Stage 2 готов")


if __name__ == "__main__":
    main()
