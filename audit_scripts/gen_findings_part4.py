import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/home/user/TrendStudio-Holding/audit_scripts/audit_public_v1_findings.xlsx'
wb = openpyxl.load_workbook(OUT)
thin = Side(style='thin')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="003366")
wrap = Alignment(wrap_text=True, vertical="top")

def add_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = wrap; c.border = border
    ws.freeze_panes = "A2"

def add_row(ws, row_num, data):
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.alignment = wrap; c.border = border

# === SHEET 15: Numerical & Symbolic ===
ws15 = wb.create_sheet("Numerical & Symbolic")
add_header(ws15, ["Check","Result","Details","Status"])
numerical = [
["Float precision (IRR 20.09%)","Machine epsilon","IRR stored as constant, not computed in xlsx","N/A — static model"],
["BS Balance check","0.00 all 16 periods","Assets = Liabilities + Equity exactly","PASS"],
["Cash chain CF→BS","Match all periods","Ending Cash (CF Row 23) = Cash (BS Row 7)","PASS"],
["Revenue identity","4545 = sum of segments","300 + 1750 + 2495 = 4545","PASS"],
["EBITDA identity","4545 - 2008.8 - 368.8 = 2167.4","Top-down verified","PASS"],
["NDP bridge","2167 + 600 + 233 = 3000","Three-component bridge","PASS"],
["Cross-sheet Revenue","P&L = Revenue_Breakdown = KPI = ES","All show 4545","PASS"],
["Cross-sheet EBITDA","P&L = KPI = ES = 2167.4","Margin 47.7% consistent","PASS"],
["IRR cross-method","20.09% (model) vs 20.09% (anchor)","Consistent but static","PASS"],
["MC IRR distribution","Mean 11.44%, Median 12.00%","P5=-0.46%, P95=23.40%","VERIFIED"],
["Waterfall arithmetic","W3: 1250 + 500 + 750 = 2500","LP total verified","PASS"],
["MOIC check","2500/1250 = 2.00x","Consistent with model","PASS"],
["","","",""],
["Symbolic (SymPy)","Not executable","Model is static — no formulas to verify symbolically","N/A"],
["Condition numbers","Not applicable","No matrix operations in xlsx (all constants)","N/A"],
["Interval arithmetic","Not applicable","Would require live formulas","N/A"],
]
for i, n in enumerate(numerical, 2):
    add_row(ws15, i, n)
    c = ws15.cell(row=i, column=4)
    colors = {"PASS":"00AA00","VERIFIED":"0070C0","N/A":"808080"}
    for k,v in colors.items():
        if k in str(n[3]):
            c.font = Font(bold=True, color=v)
for col in range(1,5):
    ws15.column_dimensions[get_column_letter(col)].width = [25,25,50,18][col-1]

# === SHEET 16: UQ & Convergence ===
ws16 = wb.create_sheet("UQ & Convergence")
add_header(ws16, ["Parameter","Value","Source","Notes"])
uq = [
["MC Simulations","5,000","28_Monte_Carlo_Summary","seed=42"],
["MC Engine","numpy vectorized","manifest v1.0.1","Upgraded from random module"],
["","","",""],
["Stochastic Variables","","",""],
["rev_mult","Triangular(0.6, 1.0, 1.4)","manifest","Revenue multiplier"],
["ebitda_shock","Normal(0, 0.04)","manifest","EBITDA additive shock"],
["capex_over","LogNormal(0, 0.10)","manifest","CAPEX overrun"],
["exit_mult","Triangular(3, 5, 7)","manifest","Exit multiple"],
["hit_rate","Binomial(12, 0.7)/12","manifest","Film success rate"],
["","","",""],
["Results (Model)","","",""],
["Mean NDP","2,104 mln RUB","28_MC","vs base 3,000"],
["Mean IRR","11.44%","28_MC","vs deterministic 20.09%"],
["Median IRR","12.00%","28_MC",""],
["P5 IRR","-0.46%","28_MC","5th percentile"],
["P95 IRR","23.40%","28_MC","95th percentile"],
["P(IRR > 18%)","13.6%","28_MC","Below hurdle"],
["P(Loss)","5.5%","28_MC","Negative IRR"],
["Mean MOIC","1.542x","28_MC","vs deterministic 2.0x"],
["VaR 95%","1,770.7 mln RUB","28_MC",""],
["","","",""],
["Pipeline MC (4 engines)","","",""],
["Parametric MC","n=2000, Cholesky corr","pipeline/generators/monte_carlo.py","Triangular shocks"],
["Market Bootstrap","Block bootstrap YoY","pipeline/generators/market_bootstrap.py","EAIS historical"],
["Stage-Gate","12 films x 4 gates","pipeline/generators/stage_gate.py","Binomial tree"],
["LHS + Copula","n=2000, Gaussian copula","pipeline/generators/lhs_copula.py","Variance reduction 1.5-3x"],
["","","",""],
["Convergence","","",""],
["CLT check","Not run independently","Would need MC re-execution","DEFERRED"],
["Sobol indices","Not computed","Requires SALib installation","DEFERRED"],
["Bootstrap CI","Not computed independently","Available in pipeline","DEFERRED"],
]
for i, u in enumerate(uq, 2):
    add_row(ws16, i, u)
for col in range(1,5):
    ws16.column_dimensions[get_column_letter(col)].width = [22,28,35,30][col-1]

# === SHEET 17: DAG & Contracts ===
ws17 = wb.create_sheet("DAG & Contracts")
add_header(ws17, ["Component","Type","Description","Dependencies","Status"])
dag = [
["Pipeline Architecture","","","",""],
["inputs/ (18 YAML)","L1 - Data","Single source of truth","None","VERIFIED"],
["schemas/ (15 .py)","L2 - Contracts","Pydantic StrictModel extra=forbid","inputs/","VERIFIED"],
["generators/ (22 .py)","L3 - Logic","Pure functions, no I/O","schemas/","VERIFIED"],
["artifacts/ (xlsx,docx)","L4 - Output","Build artifacts","generators/","VERIFIED"],
["logs/","N3 - Provenance","Append-only audit trail","generators/","VERIFIED"],
["navigation/","N3 - Docs","Auto-generated documentation","all","VERIFIED"],
["","","","",""],
["Orchestration DAG","","","",""],
["Phase 1: load_inputs","Entry","Pydantic validation 18 YAML","None","PASS"],
["Phase 2: run_all","Core","3 scenarios + anchor check","Phase 1","PASS (3000.7)"],
["Phase 3: sensitivity","Analysis","WACC x growth grid","Phase 2","PASS"],
["Phase 4: stress_tests","Analysis","6 shock scenarios","Phase 2","PASS"],
["Phase 4+5: combined","Analysis","3x3x3 matrix + 4 MC engines","Phase 2","PASS"],
["Phase 6: provenance","Audit","Source ID registry","Phase 1","PASS"],
["Phase 7: hash_manifest","Audit","SHA-256 of all inputs","All","PASS"],
["Phase 8: xlsx_builder","Output","21-sheet model.xlsx","Phases 2-5","PASS"],
["Phase 9: docx_builder","Output","model_report.docx","Phases 2-5","PASS"],
["","","","",""],
["Contract Verification","","","",""],
["Pydantic StrictModel","Schema","extra='forbid', validate_assignment","All schemas","PASS"],
["ScenarioValues ordering","Invariant","cons <= base <= opt","All scenario fields","PASS"],
["Anchor invariant","Gate","3000 ±1% (actual 3000.7)","Phase 2","PASS"],
["Seed determinism","Property","random.Random(seed), PYTHONHASHSEED=0","All MC engines","PASS"],
["Hash manifest","Integrity","SHA-256 of inputs+schemas+generators","Phase 7","PASS"],
["Provenance tracking","Traceability","source_id per assumption","Phase 6","PASS"],
["","","","",""],
["Resilience","","","",""],
["Pipeline runtime","Performance","6.29 seconds end-to-end","","PASS"],
["Test suite","Quality","323/328 pass, 0 fail","","PASS"],
["Reproducibility","Determinism","Identical outputs on re-run","","PASS"],
["Error handling","Robustness","Anchor breach = exit code 1","","PASS"],
]
for i, d in enumerate(dag, 2):
    add_row(ws17, i, d)
    c = ws17.cell(row=i, column=5)
    if "PASS" in str(d[4]) or "VERIFIED" in str(d[4]):
        c.font = Font(bold=True, color="00AA00")
for col in range(1,6):
    ws17.column_dimensions[get_column_letter(col)].width = [25,14,40,25,12][col-1]

wb.save(OUT)
print(f"Part 4 DONE: All {len(wb.sheetnames)} sheets complete!")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {name}")
import os
size = os.path.getsize(OUT)
print(f"\nFile size: {size:,} bytes ({size/1024:.1f} KB)")
