"""
А.5 — Cash Flow Statement + Balance Sheet.

Консистентно с P&L (А.4, Вариант C, GAAP EBITDA 2152 + NDP 3000):
  - Net Income Σ 2026-2028 = 1 689
  - Revenue = 4 545, COGS = 2 127.5, OpEx = 265.5, EBITDA = 2 152
  - Content CAPEX schedule: 400/800/650 = 1 850
  - T₁ Legacy 1 250 (20/28/28/24%) + Producer equity 600

Принципы:
  Cash Flow — косвенный метод (indirect), 12Q + 4Y
  Balance Sheet — end-of-period snapshots, 12Q + 4Y
  Верификация: Cash BS = Ending Cash CF, Assets = L+E
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

# ===== ПАЛИТРА =====
BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
INFLOW = "E2EFDA"    # светло-зелёный — приток
OUTFLOW = "FCE4D6"   # светло-оранжевый — отток
CHECK_GREEN = "C6EFCE"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color=BRAND_BLUE)
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
box_brand = Border(left=medium, right=medium, top=medium, bottom=medium)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'


def set_cell(ws, ref, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws[ref]
    c.value = value
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


# ===== ПЕРИОДЫ =====
PERIODS = [
    ("D", "Q1'26", "Q"), ("E", "Q2'26", "Q"), ("F", "Q3'26", "Q"), ("G", "Q4'26", "Q"),
    ("H", "Q1'27", "Q"), ("I", "Q2'27", "Q"), ("J", "Q3'27", "Q"), ("K", "Q4'27", "Q"),
    ("L", "Q1'28", "Q"), ("M", "Q2'28", "Q"), ("N", "Q3'28", "Q"), ("O", "Q4'28", "Q"),
    ("P", "2029",  "Y"), ("Q", "2030",  "Y"), ("R", "2031",  "Y"), ("S", "2032",  "Y"),
]
Q_COLS = [p[0] for p in PERIODS if p[2] == "Q"]   # 12 кварталов 2026-2028
Y_COLS = [p[0] for p in PERIODS if p[2] == "Y"]   # 4 года 2029-2032
ALL_COLS = [p[0] for p in PERIODS]

# ===== ВХОДНЫЕ ДАННЫЕ (консистентны с P&L) =====
COGS_RATIO = 2127.5 / 4545  # 0.46810

# Revenue floor: минимальный доход от библиотеки контента (SVOD, TV rights, merchandise).
# Benchmark: CTC Media / ivi — 15-20% revenue от каталога. Floor = 17% от пика 2028.
REVENUE_FLOOR = 380  # млн ₽/год — минимальный доход от библиотеки контента

_REV_Q_RAW = {
    "D": 0,    "E": 0,    "F": 0,    "G": 310,
    "H": 250,  "I": 700,  "J": 620,  "K": 420,
    "L": 460,  "M": 550,  "N": 830,  "O": 405,
    "P": 380,  "Q": 300,  "R": 220,  "S": 150,   # raw before floor
}
# Apply revenue floor to annual tail periods (2029-2032)
REV_Q = {col: max(val, REVENUE_FLOOR) if col in ("P", "Q", "R", "S") else val
         for col, val in _REV_Q_RAW.items()}

OPEX_Q_BASE = 22.1265   # млн ₽ / квартал (A₁ flat base)
OPEX_Y_BASE = 88.506    # млн ₽ / год (base)

# FIX-02: ФОТ cap при падении выручки — OpEx ≤ 70% от Revenue
MAX_OPEX_REVENUE_RATIO = 0.70

def _opex_capped(col, kind):
    """OpEx with cap: min(base, Revenue × MAX_OPEX_RATIO)."""
    base = OPEX_Q_BASE if kind == "Q" else OPEX_Y_BASE
    rev = REV_Q[col]
    if rev <= 0:
        return base
    return min(base, rev * MAX_OPEX_REVENUE_RATIO)

# Pre-compute effective opex per period
OPEX_EFF = {col: round(_opex_capped(col, kind), 4)
            for col, _, kind in PERIODS}
OPEX_Q = OPEX_Q_BASE  # backward compat alias
OPEX_Y = OPEX_Y_BASE
DA_Q = 0.75
DA_Y = 3.0

INT_PLAN = {"D": 1.0, "E": 1.0, "F": 1.5, "G": 1.5,
            "H": 1.5, "I": 1.5, "J": 1.5, "K": 1.5,
            "L": 0.75, "M": 0.75, "N": 0.5, "O": 0.5,
            "P": 1.0, "Q": 0.5, "R": 0.0, "S": 0.0}

# Content CAPEX schedule (production outflow): Σ = 1850
CONTENT_CAPEX = {
    "D": 80,  "E": 100, "F": 120, "G": 100,  # 2026: 400
    "H": 180, "I": 220, "J": 220, "K": 180,  # 2027: 800
    "L": 180, "M": 170, "N": 170, "O": 130,  # 2028: 650
    "P": 0,   "Q": 0,   "R": 0,   "S": 0,    # 2029-2032: 0
}

# T₁ Legacy tranches: 4 × (20/28/28/24)% of 1250
T1_TRANCHES = {
    "D": 250,   # Q1'26 — tranche 1 (20%)
    "G": 350,   # Q4'26 — tranche 2 (28%)
    "J": 350,   # Q3'27 — tranche 3 (28%)
    "M": 300,   # Q2'28 — tranche 4 (24%)
}

# Producer equity (JV capital): 600, pro-rata
PRODUCER_EQ = {
    "D": 200,   # Q1'26
    "G": 150,   # Q4'26
    "J": 150,   # Q3'27
    "M": 100,   # Q2'28
}

# P&A schedule (cash outflow тем же кварталом, что Revenue) — часть COGS
PA_Q = {c: round(REV_Q[c] * (277.5 / 4545), 2) for c in ALL_COLS}

# Production amortization = 1850 млн, распределяется по periods как доля от total revenue
PROD_AMORT_RATIO = 1850 / 4545
PROD_AMORT_Q = {c: round(REV_Q[c] * PROD_AMORT_RATIO, 2) for c in ALL_COLS}

# PP&E CAPEX (minimal): 10 млн total over 3 years
PPE_CAPEX_Q = {c: 0.83 if PERIODS[ALL_COLS.index(c)][2] == "Q" else 0 for c in ALL_COLS}


def compute_ni_quarterly():
    """Чистая прибыль поквартально (GAAP, consistent with P&L + FIX-02 cap)."""
    ni = {}
    for col, lbl, kind in PERIODS:
        rev = REV_Q[col]
        cogs = rev * COGS_RATIO
        gp = rev - cogs
        opex = OPEX_EFF[col]  # FIX-02: capped opex
        ebitda = gp - opex
        da = DA_Q if kind == "Q" else DA_Y
        ebit = ebitda - da
        interest = INT_PLAN[col]
        ebt = ebit - interest
        tax = max(ebt, 0) * 0.20
        ni[col] = round(ebt - tax, 2)
    return ni


# ============================================================
#   10_Cash_Flow_Statement
# ============================================================
def build_cash_flow(wb):
    if "10_Cash_Flow" in wb.sheetnames:
        del wb["10_Cash_Flow"]
    ws = wb.create_sheet("10_Cash_Flow")

    # Ширины колонок
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 40
    for i in range(4, 20):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions["T"].width = 12
    ws.column_dimensions["U"].width = 12
    ws.column_dimensions["V"].width = 28

    # Title
    ws.merge_cells("B2:V2")
    set_cell(ws, "B2",
             "10  ·  CASH FLOW STATEMENT  ·  Косвенный метод  ·  12Q + 4Y",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:V3")
    set_cell(ws, "B3",
             "Все суммы — млн ₽. Консистентно с 09_P&L_Statement (Вариант C: GAAP EBITDA 2 152 + NDP 3 000).",
             font=F_ITALIC, align=C_CENTER)

    # Headers
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Показатель", font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    for col, lbl, _ in PERIODS:
        set_cell(ws, f"{col}{r}", lbl, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"T{r}", "Σ 2026–2028", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"U{r}", "Σ 2026–2032", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"V{r}", "Комментарий", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 20

    ni = compute_ni_quarterly()

    # ========= I. OPERATING ACTIVITIES =========
    lines = []
    lines.append(("SECTION", "I.  CASH FLOW FROM OPERATING ACTIVITIES"))
    lines.append(("ROW", "1.1", "Net Income (from P&L)", ni, "inflow",
                  "Из строки 4.6 P&L"))
    lines.append(("ROW", "1.2", "(+) D&A (non-cash)",
                  {c: (DA_Q if p[2]=="Q" else DA_Y) for c, p in zip(ALL_COLS, PERIODS)},
                  "inflow", "Обратный выкладыш амортизации"))
    lines.append(("ROW", "1.3", "(+) Content amortization add-back",
                  PROD_AMORT_Q, "inflow",
                  "1850 млн распределено по matching principle"))

    # Operating CF subtotal
    def ocf_row():
        res = {}
        for col, lbl, kind in PERIODS:
            da = DA_Q if kind == "Q" else DA_Y
            res[col] = round(ni[col] + da + PROD_AMORT_Q[col], 2)
        return res

    lines.append(("SUBTOTAL", "1.4", "Net Cash from Operations", ocf_row(), "subtotal", ""))

    # ========= II. INVESTING ACTIVITIES =========
    lines.append(("SECTION", "II.  CASH FLOW FROM INVESTING ACTIVITIES"))
    lines.append(("ROW", "2.1", "(−) Content production CAPEX",
                  {c: -CONTENT_CAPEX[c] for c in ALL_COLS},
                  "outflow", "Производство контента — schedule: 400/800/650/0"))
    lines.append(("ROW", "2.2", "(−) PP&E CAPEX (minimal)",
                  {c: -PPE_CAPEX_Q[c] for c in ALL_COLS},
                  "outflow", "Офис, оборудование: ~10 млн за 3 года"))

    def icf_row():
        return {c: round(-CONTENT_CAPEX[c] - PPE_CAPEX_Q[c], 2) for c in ALL_COLS}

    lines.append(("SUBTOTAL", "2.3", "Net Cash from Investing", icf_row(), "subtotal", ""))

    # ========= III. FINANCING ACTIVITIES =========
    lines.append(("SECTION", "III.  CASH FLOW FROM FINANCING ACTIVITIES"))
    lines.append(("ROW", "3.1", "(+) T₁ Legacy tranches drawn",
                  {c: T1_TRANCHES.get(c, 0) for c in ALL_COLS},
                  "inflow", "4 транша: 250/350/350/300 = 1 250"))
    lines.append(("ROW", "3.2", "(+) Producer equity contributions",
                  {c: PRODUCER_EQ.get(c, 0) for c in ALL_COLS},
                  "inflow", "JV capital: 200/150/150/100 = 600"))

    # Distributions — в 2029-2032 годах (после recoupment)
    # NDP 3 000 распределяется по W₃ waterfall (1× Liq Pref + 8% coupon + 60/40).
    # Investor получает 2 500, producer 500.
    # Распределение: 2029 40%, 2030 30%, 2031 20%, 2032 10%
    DIST_Y = {"P": -3000*0.40, "Q": -3000*0.30, "R": -3000*0.20, "S": -3000*0.10}
    dist_row_data = {c: round(DIST_Y.get(c, 0), 2) for c in ALL_COLS}
    lines.append(("ROW", "3.3", "(−) Distributions (waterfall W₃ default · 1× Liq Pref)",
                  dist_row_data, "outflow",
                  "NDP 3000 × W₃: Inv 2500 / Prod 500, schedule 40/30/20/10%"))

    def fcf_row():
        res = {}
        for c in ALL_COLS:
            fcf = T1_TRANCHES.get(c, 0) + PRODUCER_EQ.get(c, 0) + DIST_Y.get(c, 0)
            res[c] = round(fcf, 2)
        return res

    lines.append(("SUBTOTAL", "3.4", "Net Cash from Financing", fcf_row(), "subtotal", ""))

    # ========= IV. NET CHANGE IN CASH =========
    lines.append(("SECTION", "IV.  NET CHANGE IN CASH & RECONCILIATION"))

    def net_change():
        ocf = ocf_row(); icf = icf_row(); fcf = fcf_row()
        return {c: round(ocf[c] + icf[c] + fcf[c], 2) for c in ALL_COLS}

    nc = net_change()
    lines.append(("SUBTOTAL", "4.1", "Net change in cash", nc, "subtotal", ""))

    # Cumulative cash balance
    cum_cash = {}
    running = 0.0
    for c in ALL_COLS:
        running += nc[c]
        cum_cash[c] = round(running, 2)

    # Beginning / Ending
    beginning = {}
    running_beg = 0.0
    for c in ALL_COLS:
        beginning[c] = round(running_beg, 2)
        running_beg += nc[c]

    lines.append(("ROW", "4.2", "Cash — beginning of period", beginning, "neutral", ""))
    lines.append(("SUBTOTAL", "4.3", "Cash — end of period", cum_cash, "anchor",
                  "★ Совпадает с Cash в 11_Balance_Sheet"))

    # ---- Рендер ----
    r = 6
    row_map = {}

    for entry in lines:
        kind = entry[0]
        if kind == "SECTION":
            ws.merge_cells(f"B{r}:V{r}")
            set_cell(ws, f"B{r}", entry[1],
                     font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
            ws.row_dimensions[r].height = 20
            r += 1
            continue

        _, num, name, data, style, comment = entry

        # Определение стиля
        if style == "inflow":
            row_fill = INFLOW
            font_main = F_BODY
        elif style == "outflow":
            row_fill = OUTFLOW
            font_main = F_BODY
        elif style == "subtotal":
            row_fill = SUBTOTAL_FILL
            font_main = F_TOTAL
        elif style == "anchor":
            row_fill = NDP_FILL
            font_main = F_TOTAL
        else:
            row_fill = None
            font_main = F_BODY

        set_cell(ws, f"B{r}", num, font=font_main, fill=row_fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font_main, fill=row_fill, align=C_LEFT, border=box_thin)

        total_3y = 0.0
        total_all = 0.0
        for col, lbl, k in PERIODS:
            v = data.get(col, 0)
            set_cell(ws, f"{col}{r}", v,
                     font=font_main, fill=row_fill, align=C_RIGHT,
                     border=box_thin, number_format=NUMFMT)
            if k == "Q":
                total_3y += v
            total_all += v

        fill_total = KEY_METRIC_FILL if kind == "SUBTOTAL" else row_fill
        set_cell(ws, f"T{r}", round(total_3y, 1),
                 font=F_TOTAL, fill=fill_total, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"U{r}", round(total_all, 1),
                 font=F_TOTAL, fill=fill_total, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"V{r}", comment, font=F_ITALIC, fill=row_fill,
                 align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 18

        row_map[num] = (r, data)
        r += 1

    # ---- Control row ----
    r += 1
    ws.merge_cells(f"B{r}:V{r}")
    # Ending cash at Q4'28 and at 2032
    end_q4_28 = cum_cash["O"]
    end_2032 = cum_cash["S"]
    set_cell(ws, f"B{r}",
             f"✓ Cash end of Q4'28 = {end_q4_28:,.1f} млн ₽  |  Cash end of 2032 = {end_2032:,.1f} млн ₽  |  "
             f"Σ OCF 3Y = {sum(ocf_row()[c] for c in Q_COLS):,.0f}  ·  Σ ICF 3Y = {sum(icf_row()[c] for c in Q_COLS):,.0f}  ·  "
             f"Σ FCF 3Y = {sum(fcf_row()[c] for c in Q_COLS):,.0f}",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 22

    ws.freeze_panes = "D6"

    return {
        "ni": ni,
        "ocf": ocf_row(),
        "icf": icf_row(),
        "fcf": fcf_row(),
        "cum_cash": cum_cash,
        "dist_y": DIST_Y,
    }


# ============================================================
#   11_Balance_Sheet
# ============================================================
def build_balance_sheet(wb, cf_data):
    if "11_Balance_Sheet" in wb.sheetnames:
        del wb["11_Balance_Sheet"]
    ws = wb.create_sheet("11_Balance_Sheet")

    # Ширины
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 38
    for i in range(4, 20):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions["T"].width = 30

    # Title
    ws.merge_cells("B2:T2")
    set_cell(ws, "B2",
             "11  ·  BALANCE SHEET  ·  End-of-Period  ·  12Q + 4Y",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:T3")
    set_cell(ws, "B3",
             "Все суммы — млн ₽. Верификация: Assets = Liabilities + Equity; Cash = Ending Cash из 10_Cash_Flow.",
             font=F_ITALIC, align=C_CENTER)

    # Headers
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Статья", font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    for col, lbl, _ in PERIODS:
        set_cell(ws, f"{col}{r}", lbl, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"T{r}", "Комментарий", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 20

    # ===== РАСЧЁТЫ =====
    cum_cash = cf_data["cum_cash"]
    ni = cf_data["ni"]

    # Cumulative totals по периодам (end-of-period)
    def cumulate(dct):
        res = {}
        run = 0.0
        for c in ALL_COLS:
            run += dct[c]
            res[c] = round(run, 2)
        return res

    cum_content_capex = cumulate(CONTENT_CAPEX)
    cum_prod_amort = cumulate(PROD_AMORT_Q)
    cum_ppe_capex = cumulate(PPE_CAPEX_Q)
    cum_da = cumulate({c: (DA_Q if PERIODS[ALL_COLS.index(c)][2]=="Q" else DA_Y) for c in ALL_COLS})
    cum_t1 = cumulate({c: T1_TRANCHES.get(c, 0) for c in ALL_COLS})
    cum_prod_eq = cumulate({c: PRODUCER_EQ.get(c, 0) for c in ALL_COLS})
    cum_ni = cumulate(ni)

    # ── Distributions split: W₃ waterfall priority ──────────────────
    # W₃: Stage 1 = 1× Liq Pref (→ investor), Stage 2 = 8% coupon
    # (→ investor), Stage 3 = 60/40 residual.
    # Total: Investor 2 500, Producer 500 out of NDP 3 000.
    # IMPORTANT: B1/B2 are ALLOWED to go negative (return > principal).
    # Flooring at 0 breaks the balance — see v3 BS bug fix (Apr 2026).
    S1_CAP = 1250.0   # 1× Liq Pref capacity
    S2_CAP = 500.0    # 8% coupon capacity
    s1_rem, s2_rem = S1_CAP, S2_CAP
    dist_inv_per = {}   # per-period investor share (negative)
    dist_prod_per = {}  # per-period producer share (negative)
    for c in ALL_COLS:
        d = abs(cf_data["dist_y"].get(c, 0))
        remaining = d
        to_inv = 0.0
        to_prod = 0.0
        if s1_rem > 0:
            take = min(remaining, s1_rem)
            to_inv += take; s1_rem -= take; remaining -= take
        if s2_rem > 0 and remaining > 0:
            take = min(remaining, s2_rem)
            to_inv += take; s2_rem -= take; remaining -= take
        if remaining > 0:
            to_inv += remaining * 0.60
            to_prod += remaining * 0.40
        # Keep sign consistent with distributions (negative = outflow)
        sign = -1.0 if cf_data["dist_y"].get(c, 0) < 0 else (1.0 if d > 0 else 0.0)
        dist_inv_per[c] = round(-to_inv, 2)   # negative
        dist_prod_per[c] = round(-to_prod, 2)  # negative

    dist_inv = cumulate(dist_inv_per)
    dist_prod = cumulate(dist_prod_per)

    # ===== ASSETS =====
    sec_row = [
        ("SECTION", "A.  ASSETS"),
        ("ASSET", "A1", "Cash & equivalents", cum_cash, "Из 10_CF 4.3"),
        # FIX-04: Floor=0 — актив не может быть отрицательным.
        # Амортизация свыше CAPEX означает полное списание, а не отрицательный актив.
        ("ASSET", "A2", "Content library (net)",
         {c: round(max(0, cum_content_capex[c] - cum_prod_amort[c]), 2) for c in ALL_COLS},
         "max(0, Gross CAPEX − amortization)"),
        ("ASSET", "A3", "Property, Plant & Equipment",
         {c: round(max(0, cum_ppe_capex[c] - cum_da[c]), 2) for c in ALL_COLS},
         "max(0, PP&E net)"),
        ("TOTAL_A", "A4", "TOTAL ASSETS", None, "Σ A1+A2+A3"),
    ]

    # FIX-04/05: Compute floor adjustment balancing item.
    # Flooring assets at 0 INCREASES total assets (removes negative values).
    # Flooring loan at 0 INCREASES total L+E (removes negative values).
    # The equity balancing item = asset floor impact - liability floor impact.
    # Formula: adj = -min(0, content_net) - min(0, ppe_net) + min(0, T1_raw)
    # where min(0, x) extracts the negative portion; we negate to get the floor delta.
    _t1_raw = {c: round(cum_t1[c] + dist_inv[c], 2) for c in ALL_COLS}
    content_floor_adj = {c: round(
        -min(0, cum_content_capex[c] - cum_prod_amort[c])   # asset floor delta (positive)
        - min(0, cum_ppe_capex[c] - cum_da[c])              # asset floor delta (positive)
        + min(0, _t1_raw[c]),                                # liability floor delta (negative)
        2) for c in ALL_COLS}

    # ===== LIABILITIES + EQUITY =====
    le_rows = [
        ("SECTION", "B.  LIABILITIES & EQUITY"),
        # FIX-05: Loan balance floor — after full repayment, surplus → cash, not negative debt
        ("LIAB", "B1", "T₁ Legacy loan balance",
         {c: round(max(0, cum_t1[c] + dist_inv[c]), 2) for c in ALL_COLS},
         "max(0, Draws − investor distributions)"),
        ("EQUITY", "B2", "Producer equity (JV capital)",
         {c: round(cum_prod_eq[c] + dist_prod[c], 2) for c in ALL_COLS},
         "Contributions − producer distributions"),
        ("EQUITY", "B3", "Retained earnings",
         cum_ni,
         "Cumulative Net Income"),
        ("EQUITY", "B4", "Amortization reserve (floor adj.)",
         content_floor_adj,
         "FIX-04: balancing item for asset floor=0"),
        ("TOTAL_LE", "B5", "TOTAL LIABILITIES + EQUITY", None, "Σ B1+B2+B3+B4"),
    ]

    lines = sec_row + [("BLANK",)] + le_rows

    # ---- Рендер ----
    r = 6
    asset_rows = []
    le_rows_idx = []
    section = None

    for entry in lines:
        if entry[0] == "SECTION":
            ws.merge_cells(f"B{r}:T{r}")
            set_cell(ws, f"B{r}", entry[1],
                     font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
            ws.row_dimensions[r].height = 20
            section = entry[1]
            r += 1
            continue
        if entry[0] == "BLANK":
            r += 1
            continue

        kind, num, name, data, comment = entry

        is_total = kind.startswith("TOTAL")
        fill = SUBTOTAL_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY

        set_cell(ws, f"B{r}", num, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)

        if data is not None:
            for col, lbl, _ in PERIODS:
                set_cell(ws, f"{col}{r}", data[col],
                         font=font, fill=fill, align=C_RIGHT,
                         border=box_thin, number_format=NUMFMT)
            if kind == "ASSET":
                asset_rows.append(data)
            elif kind in ("LIAB", "EQUITY"):
                le_rows_idx.append(data)
        else:
            # TOTAL rows — суммы по соответствующему списку
            if kind == "TOTAL_A":
                totals = {c: round(sum(row[c] for row in asset_rows), 2) for c in ALL_COLS}
            else:
                totals = {c: round(sum(row[c] for row in le_rows_idx), 2) for c in ALL_COLS}
            for col, lbl, _ in PERIODS:
                set_cell(ws, f"{col}{r}", totals[col],
                         font=font, fill=KEY_METRIC_FILL, align=C_RIGHT,
                         border=box_thin, number_format=NUMFMT)
            if kind == "TOTAL_A":
                totals_a = totals
            else:
                totals_le = totals

        set_cell(ws, f"T{r}", comment, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 18
        r += 1

    # ===== CONTROL / RECONCILIATION =====
    r += 1
    ws.merge_cells(f"B{r}:T{r}")
    set_cell(ws, f"B{r}", "C.  BALANCE CHECK  (Assets − Liab/Equity; должно быть 0)",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 20
    r += 1

    set_cell(ws, f"B{r}", "C1", font=F_BOLD, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Balance check = Assets − (L+E)", font=F_BOLD, align=C_LEFT, border=box_thin)
    max_abs = 0.0
    for col, lbl, _ in PERIODS:
        diff = round(totals_a[col] - totals_le[col], 2)
        if abs(diff) > max_abs:
            max_abs = abs(diff)
        fill = CHECK_GREEN if abs(diff) < 1.0 else "FCE4D6"
        set_cell(ws, f"{col}{r}", diff,
                 font=F_BOLD, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
    status = "✓ BALANCED" if max_abs < 1.0 else f"⚠ Δ_max = {max_abs:.2f}"
    set_cell(ws, f"T{r}", status, font=F_BOLD, fill=CHECK_GREEN if max_abs<1 else "FCE4D6",
             align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 20
    r += 2

    # Итоговая заметка
    ws.merge_cells(f"B{r}:T{r}")
    content_net_o = max(0, cum_content_capex['O']-cum_prod_amort['O'])
    ppe_net_o = max(0, cum_ppe_capex['O']-cum_da['O'])
    t1_net_o = max(0, cum_t1['O']+dist_inv['O'])
    bs_summary = (
        f"★ Snapshot Q4'28: Assets {totals_a['O']:,.0f} = Cash {cum_cash['O']:,.0f} + "
        f"Content {content_net_o:,.0f} + PP&E {ppe_net_o:,.0f}  |  "
        f"L+E {totals_le['O']:,.0f} = T₁ {t1_net_o:,.0f} + Prod.Eq {cum_prod_eq['O']+dist_prod['O']:,.0f} + "
        f"Retained {cum_ni['O']:,.0f} + FloorAdj {content_floor_adj['O']:,.0f}"
    )
    set_cell(ws, f"B{r}", bs_summary,
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 24

    ws.freeze_panes = "D6"

    return {"totals_a": totals_a, "totals_le": totals_le, "max_diff": max_abs}


def main():
    wb = load_workbook(FILE)
    print(f"Loaded: {FILE}")
    print(f"Sheets before: {len(wb.sheetnames)}")

    print("\n[1/2] Building 10_Cash_Flow …")
    cf_data = build_cash_flow(wb)

    print("[2/2] Building 11_Balance_Sheet …")
    bs_data = build_balance_sheet(wb, cf_data)

    wb.save(FILE)
    print(f"\nSheets after: {wb.sheetnames}")
    print(f"BS max diff: {bs_data['max_diff']:.4f}")
    print(f"Saved: {FILE}")


if __name__ == "__main__":
    main()
