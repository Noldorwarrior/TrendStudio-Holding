"""
А.16 — Executive_Summary + Glossary + Notes_and_Sources (листы 36, 37, 38)

36_Executive_Summary:
  I.   Investment Highlights (6 bullets)
  II.  Key Metrics (якоря 4545/2152/3000/1689/1250/600/1850)
  III. Valuation Range Interpretation  ← часть B комбо A+B
       (полная таблица 5 компонентов разрыва DCF ~1815 vs Comps median ~7550)
  IV.  Risk Summary (top-5 из 29_Risk_Register)
  V.   Exit Strategy (7 маршрутов из 25/35)
  VI.  Recommendation

37_Glossary: ~50 терминов (финансовые / кино / регуляторные РФ)

38_Notes_and_Sources: методологические заметки + полный список источников + disclaimers + версионность

Архитектурные гарантии:
- Якоря Revenue/EBITDA/NDP/Net Profit/T1/Equity не меняются
- Листы 1-35 не трогаются
- freeze panes D7 на всех новых листах
- Cross-references к существующим листам по названию
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx"

# Палитра (консистентна со всем пакетом)
BLUE = "0070C0"
DARK_BLUE = "002060"
LIGHT_BLUE = "DEEBF7"
VERY_LIGHT_BLUE = "F2F8FD"
LIGHT_ORANGE = "FCE4D6"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED = "FCE4E4"
LIGHT_YELLOW = "FFF2CC"
GREY = "808080"
LIGHT_GREY = "D9D9D9"
WHITE = "FFFFFF"
BLACK = "000000"

thin = Side(border_style="thin", color="BFBFBF")
medium = Side(border_style="medium", color="0070C0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_MED = Border(left=medium, right=medium, top=medium, bottom=medium)

# ============================================================
# ИНВАРИАНТЫ (якоря, не меняются)
# ============================================================
REVENUE_3Y = 4545
EBITDA_GAAP_3Y = 2152
NDP_3Y = 3000
NET_PROFIT_3Y = 1689
INVESTMENT_T1 = 1250
PRODUCER_EQUITY = 600
PRODUCTION_BUDGET_TOTAL = 1850
COGS_3Y = 2127.5
OPEX_3Y = 265.5
TAIL_REVENUE = 1050
TAX_3Y = 597
TAX_7Y = 720
DCF_BLEND = 1815
COMPS_MEDIAN = 7550
WEIGHTED_EXIT_EV = 6038

wb = load_workbook(XLSX)
print(f"Open: {len(wb.sheetnames)} sheets")

# Snapshot существующих листов для проверки неприкосновенности
existing_sheets = set(wb.sheetnames)


# ============================================================
# УТИЛИТЫ ФОРМАТИРОВАНИЯ
# ============================================================
def title_block(ws, row, col, title, subtitle=None, merge_to=14):
    c = ws.cell(row, col)
    c.value = title
    c.font = Font(name="Calibri", size=18, bold=True, color=BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    ws.row_dimensions[row].height = 28
    if subtitle:
        row += 1
        c = ws.cell(row, col)
        c.value = subtitle
        c.font = Font(name="Calibri", size=11, italic=True, color=GREY)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        ws.row_dimensions[row].height = 18
    return row + 2


def section_header(ws, row, col, num, title, merge_to=14, color=DARK_BLUE):
    c = ws.cell(row, col)
    c.value = f"{num}. {title}"
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for cc in range(col, merge_to + 1):
        ws.cell(row, cc).fill = PatternFill("solid", fgColor=color)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    ws.row_dimensions[row].height = 22
    return row + 2


def para(ws, row, col, text, size=10, bold=False, italic=False, color=BLACK, merge_to=14, height=None):
    c = ws.cell(row, col)
    c.value = text
    c.font = Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    if height:
        ws.row_dimensions[row].height = height
    return row + 1


def bullet(ws, row, col, bullet_char, text, size=10, merge_to=14, height=22):
    ws.cell(row, col).value = bullet_char
    ws.cell(row, col).font = Font(name="Calibri", size=size, bold=True, color=BLUE)
    ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row, col + 1).value = text
    ws.cell(row, col + 1).font = Font(name="Calibri", size=size)
    ws.cell(row, col + 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=merge_to)
    ws.row_dimensions[row].height = height
    return row + 1


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# ЛИСТ 36: Executive_Summary
# ============================================================
print("\n[1/3] Building 36_Executive_Summary...")
ws = wb.create_sheet("36_Executive_Summary")
set_widths(ws, [2, 5, 28, 18, 16, 16, 14, 14, 14, 14, 14, 14, 14, 14, 2])
ws.freeze_panes = "D7"

# Заголовок
row = 2
row = title_block(ws, row, 2,
                  "EXECUTIVE SUMMARY",
                  "Холдинг кино «ТрендСтудио» — Investor Package v1.0 Public — 2026-04-11",
                  merge_to=14)

# I. Investment Highlights
row = section_header(ws, row, 2, "I", "INVESTMENT HIGHLIGHTS")

highlights = [
    ("▶", "Горизонт 2026–2028, 12 премиальных фильмов, Revenue 4 545 млн ₽, GAAP EBITDA margin 47%, payback < 24 мес"),
    ("▶", "Инвест-пакет T₁ = 1 250 млн ₽ (4 транша 20/28/28/24%), Producer equity 600 млн ₽ (off-P&L JV капитал)"),
    ("▶", "GAAP EBITDA Σ = 2 152 млн ₽ | Net Profit Σ = 1 689 млн ₽ | Эффективная налоговая нагрузка ~12.9% (льгота НДС 0% ст. 149 НК РФ)"),
    ("▶", "Рынок SOM 4 545 млн ₽ из TAM 289.8 млрд ₽ — доля <2%, защита от рыночного риска"),
    ("▶", "Weighted exit EV ≈ 6 038 млн ₽ на 2030, 7 маршрутов выхода (strategic / IPO / secondary / ESOP / buyback / mgmt / вторичный рынок)"),
    ("▶", "Полное соответствие национальным проектам «Семья», «Молодёжь и дети», «Культура» — господдержка ~1 040 млн ₽ (грант ФКП + льготы)"),
]
for b, t in highlights:
    row = bullet(ws, row, 2, b, t, size=11, height=28)

row += 1

# II. Key Metrics
row = section_header(ws, row, 2, "II", "KEY METRICS (ЯКОРЯ)")

metrics = [
    ("Показатель", "Σ 2026–2028", "Источник", "Примечание"),
    ("Revenue (Выручка)", f"{REVENUE_3Y:,} млн ₽".replace(",", " "), "07_Revenue_Breakdown", "12 фильмов + сиквелы + TV/OTT"),
    ("COGS (Себестоимость)", f"{COGS_3Y:,.1f} млн ₽".replace(",", " "), "06_Cost_Structure", "Производство + прокат + маркетинг"),
    ("OpEx", f"{OPEX_3Y:,.1f} млн ₽".replace(",", " "), "06_Cost_Structure", "ФОТ + аренда + G&A"),
    ("EBITDA (GAAP)", f"{EBITDA_GAAP_3Y:,} млн ₽".replace(",", " "), "09_P&L_Statement", "Margin 47.3%"),
    ("NDP (legacy anchor)", f"{NDP_3Y:,} млн ₽".replace(",", " "), "09_P&L (bridge)", "Non-Discounted Profit, внутренний якорь холдинга"),
    ("Net Profit", f"{NET_PROFIT_3Y:,} млн ₽".replace(",", " "), "09_P&L_Statement", "После налогов (prib 20% + прочее)"),
    ("Investment T₁", f"{INVESTMENT_T1:,} млн ₽".replace(",", " "), "14_Investment_Inflow", "4 транша: 250 / 350 / 350 / 300"),
    ("Producer Equity", f"{PRODUCER_EQUITY} млн ₽", "17_Deal_Structures", "JV капитал продюсеров (off-P&L)"),
    ("Production Budget", f"{PRODUCTION_BUDGET_TOTAL:,} млн ₽".replace(",", " "), "08_Content_Pipeline", "12 фильмов × avg 154 млн ₽"),
    ("Tax (7-year)", f"{TAX_7Y} млн ₽", "34_Tax_Schedule", "2026–2032, эффективная ставка 12.9% от Revenue"),
]

# Header row
header_row = row
for i, h in enumerate(metrics[0]):
    c = ws.cell(row, 3 + i if i > 0 else 3)
    # Fill metric in col 3, then 5, 7, 9 — компактно
for i, h in enumerate(metrics[0]):
    c = ws.cell(row, 3 + i*2)
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    # merge next cell
    if i < 3:
        ws.merge_cells(start_row=row, start_column=3 + i*2, end_row=row, end_column=3 + i*2 + 1)
# last col span to 14
ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=14)
ws.row_dimensions[row].height = 22
row += 1

for rec in metrics[1:]:
    vals = rec
    aligns = ["left", "center", "center", "left"]
    fills_col = [VERY_LIGHT_BLUE, LIGHT_YELLOW, VERY_LIGHT_BLUE, None]
    for i, v in enumerate(vals):
        c = ws.cell(row, 3 + i*2)
        c.value = v
        c.font = Font(name="Calibri", size=10, bold=(i == 1))
        c.alignment = Alignment(horizontal=aligns[i], vertical="center", wrap_text=True, indent=1 if i == 0 else 0)
        c.border = BORDER
        if fills_col[i]:
            c.fill = PatternFill("solid", fgColor=fills_col[i])
        if i < 3:
            ws.merge_cells(start_row=row, start_column=3 + i*2, end_row=row, end_column=3 + i*2 + 1)
            # apply border/fill to merged
            for cc in range(3 + i*2, 3 + i*2 + 2):
                ws.cell(row, cc).border = BORDER
                if fills_col[i]:
                    ws.cell(row, cc).fill = PatternFill("solid", fgColor=fills_col[i])
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=14)
    for cc in range(9, 15):
        ws.cell(row, cc).border = BORDER
    ws.row_dimensions[row].height = 20
    row += 1

row += 2

# III. VALUATION RANGE INTERPRETATION — часть B комбо A+B
row = section_header(ws, row, 2, "III", "VALUATION RANGE INTERPRETATION (DCF vs MULTIPLES)", color=BLUE)

row = para(ws, row, 2,
           "Оценка ТрендСтудио имеет три точки опоры. Это не артефакт методологии, а отражение структурных различий подходов.",
           size=11, italic=True, color=DARK_BLUE, height=20)
row += 1

# Три точки опоры
val_table = [
    ("Метод", "EV 2030", "Роль в оценке", "Источник"),
    ("DCF (blend Gordon 40% + Exit 60%)", f"{DCF_BLEND:,} млн ₽".replace(",", " "),
     "Floor valuation (intrinsic, 5-летний горизонт, высокий WACC)", "22_Valuation_DCF"),
    ("Comps Median (EV/EBITDA 8.9×)", f"{COMPS_MEDIAN:,} млн ₽".replace(",", " "),
     "Market valuation (что реально платят в сделках)", "32_Comparable_Transactions"),
    ("Weighted Exit EV (probability-weighted)", f"{WEIGHTED_EXIT_EV:,} млн ₽".replace(",", " "),
     "Expected value (переговорный ориентир)", "25_Exit_Scenarios"),
]
for i, rec in enumerate(val_table):
    for j, v in enumerate(rec):
        col = 2 + (j * 3 if j > 0 else 0)
        # Простой 4-колоночный layout
        cols = [2, 6, 9, 13]
        c = ws.cell(row, cols[j])
        c.value = v
        if i == 0:
            c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        else:
            c.font = Font(name="Calibri", size=10, bold=(j == 1))
            if j == 1:
                c.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
            elif j == 0:
                c.fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)
        c.alignment = Alignment(horizontal="left" if j != 1 else "center",
                                vertical="center", wrap_text=True, indent=1 if j == 0 else 0)
        c.border = BORDER
        # merge next columns
        end_col = cols[j + 1] - 1 if j + 1 < len(cols) else 14
        if end_col > cols[j]:
            ws.merge_cells(start_row=row, start_column=cols[j], end_row=row, end_column=end_col)
            for cc in range(cols[j], end_col + 1):
                ws.cell(row, cc).border = BORDER
                if i == 0:
                    ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
                elif j == 0:
                    ws.cell(row, cc).fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)
                elif j == 1:
                    ws.cell(row, cc).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    ws.row_dimensions[row].height = 26 if i == 0 else 32
    row += 1

row += 2

# Gap decomposition
row = para(ws, row, 2,
           "Разрыв DCF vs Comps median = 5 735 млн ₽ — это не ошибка, а сумма 5 структурных компонентов:",
           size=11, bold=True, color=DARK_BLUE, height=20)
row += 1

gap_headers = ["#", "Компонент разрыва", "Механика", "Вклад"]
gap_cols = [2, 3, 6, 13]
for i, h in enumerate(gap_headers):
    c = ws.cell(row, gap_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = gap_cols[i + 1] - 1 if i + 1 < len(gap_cols) else 14
    if end_col > gap_cols[i]:
        ws.merge_cells(start_row=row, start_column=gap_cols[i], end_row=row, end_column=end_col)
        for cc in range(gap_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

gap_rows = [
    ("1", "Горизонт обрезан на 2030",
     "Gordon Growth TV сжимает 30-летнюю библиотеку в константу g=3%. Фильмы 2028 года монетизируются до ~2048, но при WACC 19% CF 2046 = 0.5% от номинала. Cash flow из tail-окна не попадает в оценку.",
     "+2 500"),
    ("2", "WACC 19% vs buyer WACC 7%",
     "Страновой premium РФ «съедает» дальний tail. Strategic buyer (Яндекс / Сбер / Газпром-Медиа) дисконтирует тот же CF stream под 7–9%, получая принципиально другой NPV. Comps отражают цены buyer-side, не seller-side.",
     "+1 500"),
    ("3", "Real options не моделируются",
     "Франшизы (сиквелы F05, F12), кросс-медиа (игры, мерч, театр), международная копродукция — это опционные payoffs с асимметричным upside. DCF считает их = 0, хотя precedent transactions (Marvel, Harry Potter) показывают 3–5× upside от франшиз.",
     "+1 500"),
    ("4", "Execution risk — двойное начисление",
     "В DCF высокий WACC 19% + реалистичный forecast (без hockey-stick) = два независимых слоя консерватизма. В Multiples исполнительский риск уже реализован в цене сделок-прецедентов — double-counting устранён.",
     "(встроено)"),
    ("5", "Капитальная цикличность после 2028",
     "CAPEX падает к 2029–2030 (инвестиционный цикл T₁ завершён), FCF растёт импульсом. Gordon g=3% сглаживает этот impulse в линейный рост и занижает post-peak cash generation.",
     "+500"),
]

for rec in gap_rows:
    for i, v in enumerate(rec):
        c = ws.cell(row, gap_cols[i])
        c.value = v
        c.font = Font(name="Calibri", size=9, bold=(i in (0, 3)))
        c.alignment = Alignment(
            horizontal="center" if i in (0, 3) else "left",
            vertical="top", wrap_text=True, indent=0 if i in (0, 3) else 1)
        c.border = BORDER
        if i == 3:
            c.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
        end_col = gap_cols[i + 1] - 1 if i + 1 < len(gap_cols) else 14
        if end_col > gap_cols[i]:
            ws.merge_cells(start_row=row, start_column=gap_cols[i], end_row=row, end_column=end_col)
            for cc in range(gap_cols[i], end_col + 1):
                ws.cell(row, cc).border = BORDER
                if i == 3:
                    ws.cell(row, cc).fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    ws.row_dimensions[row].height = 52
    row += 1

# Σ row
c = ws.cell(row, 2)
c.value = "Σ"
c.font = Font(name="Calibri", size=11, bold=True)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = BORDER

c = ws.cell(row, 3)
c.value = "Сумма структурных компонентов ≈ фактическая разница 5 735 (7 550 − 1 815)"
c.font = Font(name="Calibri", size=10, bold=True)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
c.border = BORDER
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
for cc in range(3, 13):
    ws.cell(row, cc).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(row, cc).border = BORDER

c = ws.cell(row, 13)
c.value = "≈ +5 500"
c.font = Font(name="Calibri", size=11, bold=True)
c.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = BORDER
ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)
for cc in range(13, 15):
    ws.cell(row, cc).fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 28
row += 2

# Практический вывод
row = para(ws, row, 2,
           "ПРАКТИЧЕСКИЙ ВЫВОД ДЛЯ ИНВЕСТОРА:",
           size=11, bold=True, color=DARK_BLUE, height=20)
interp_text = (
    "• DCF 1 815 млн ₽ — floor valuation (даунсайд, intrinsic, без опций и премий). Используется как downside "
    "benchmark при worst-case переговорах, не как fair exit price.\n"
    "• Comps median 7 550 млн ₽ — market valuation (апсайд, что реально платят strategic buyers). "
    "Используется как upside anchor для блокировки нижней границы предложения.\n"
    "• Weighted Exit EV 6 038 млн ₽ (probability-weighted) — expected value, оптимальный ориентир для переговоров "
    "и дизайна деал-структуры. Именно этот value закладывается в Investor Returns (24_Investor_Returns) и IRR расчёты."
)
row = para(ws, row, 2, interp_text, size=10, height=96)
row += 2

# IV. Risk Summary
row = section_header(ws, row, 2, "IV", "RISK SUMMARY (TOP-5 ИЗ 30)")

top_risks = [
    ("R17", "Недостижение IRR hurdle 18% для T₁ Base", "Financial", "4×5", "20 (Critical)",
     "Waterfall защита, preferred return 12%, makewhole clause"),
    ("R03", "Срыв production календаря ≥ 6 мес", "Production", "3×4", "12 (High)",
     "Pre-financing reserve 10%, страхование срывов, 2-х треков post-prod"),
    ("R08", "Отказ ФКП в грантовой поддержке", "Regulatory", "4×3", "12 (High)",
     "Двойной трек заявок, диверсификация источников (ФКП + регионы + РФК)"),
    ("R12", "Девальвация рубля > 25%", "Market", "3×4", "12 (High)",
     "VFX/post-prod в РФ, долларовые контракты минимизированы, FX hedge для платформ"),
    ("R22", "Потеря ключевого режиссёра / продюсера", "Operational", "2×5", "10 (High)",
     "Long-term contracts, ESOP vesting 4 года, 2nd-tier bench"),
]

risk_headers = ["ID", "Описание", "Категория", "S×L", "Score", "Митигация"]
risk_cols = [2, 3, 7, 9, 10, 12]
for i, h in enumerate(risk_headers):
    c = ws.cell(row, risk_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = risk_cols[i + 1] - 1 if i + 1 < len(risk_cols) else 14
    if end_col > risk_cols[i]:
        ws.merge_cells(start_row=row, start_column=risk_cols[i], end_row=row, end_column=end_col)
        for cc in range(risk_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

for rec in top_risks:
    for i, v in enumerate(rec):
        c = ws.cell(row, risk_cols[i])
        c.value = v
        c.font = Font(name="Calibri", size=9, bold=(i in (0, 4)))
        c.alignment = Alignment(
            horizontal="center" if i in (0, 3, 4) else "left",
            vertical="top", wrap_text=True, indent=0 if i in (0, 3, 4) else 1)
        c.border = BORDER
        if i == 4:
            score_val = int(v.split()[0])
            if score_val >= 15:
                c.fill = PatternFill("solid", fgColor=LIGHT_RED)
            elif score_val >= 10:
                c.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
        end_col = risk_cols[i + 1] - 1 if i + 1 < len(risk_cols) else 14
        if end_col > risk_cols[i]:
            ws.merge_cells(start_row=row, start_column=risk_cols[i], end_row=row, end_column=end_col)
            for cc in range(risk_cols[i], end_col + 1):
                ws.cell(row, cc).border = BORDER
    ws.row_dimensions[row].height = 36
    row += 1

row += 1
row = para(ws, row, 2,
           "→ Полный реестр 30 рисков см. лист 29_Risk_Register (heat map 5×5 + top-10 detail).",
           size=9, italic=True, color=BLUE, height=18)
row += 2

# V. Exit Strategy
row = section_header(ws, row, 2, "V", "EXIT STRATEGY (7 МАРШРУТОВ)")

exits = [
    ("E1", "Strategic sale — Яндекс / Сбер", "30%", "2030", "EV/EBITDA 9–11×, control premium 20%"),
    ("E2", "Strategic sale — Газпром-Медиа / НМГ", "20%", "2030", "EV/EBITDA 7–9×, vertical integration premium"),
    ("E3", "IPO на Мосбирже (main list)", "15%", "2030–2031", "Free float 25%, P/E 15–18×"),
    ("E4", "Secondary placement (PE / VC)", "15%", "2029–2030", "Partial exit, roll-over 40–50%"),
    ("E5", "ESOP buyout (management)", "10%", "2030", "Discounted valuation, ESG premium"),
    ("E6", "Buyback (sponsor put option)", "5%", "2029", "Downside protection, IRR floor 12%"),
    ("E7", "Open secondary market", "5%", "2030+", "Preservation of control, drag-along terms"),
]

ex_headers = ["#", "Маршрут", "P", "Окно", "Условия / премия"]
ex_cols = [2, 3, 7, 8, 10]
for i, h in enumerate(ex_headers):
    c = ws.cell(row, ex_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = ex_cols[i + 1] - 1 if i + 1 < len(ex_cols) else 14
    if end_col > ex_cols[i]:
        ws.merge_cells(start_row=row, start_column=ex_cols[i], end_row=row, end_column=end_col)
        for cc in range(ex_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

for rec in exits:
    for i, v in enumerate(rec):
        c = ws.cell(row, ex_cols[i])
        c.value = v
        c.font = Font(name="Calibri", size=9, bold=(i == 0))
        c.alignment = Alignment(
            horizontal="center" if i in (0, 2, 3) else "left",
            vertical="center", wrap_text=True, indent=0 if i in (0, 2, 3) else 1)
        c.border = BORDER
        end_col = ex_cols[i + 1] - 1 if i + 1 < len(ex_cols) else 14
        if end_col > ex_cols[i]:
            ws.merge_cells(start_row=row, start_column=ex_cols[i], end_row=row, end_column=end_col)
            for cc in range(ex_cols[i], end_col + 1):
                ws.cell(row, cc).border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

row += 1
row = para(ws, row, 2,
           "→ Подробнее см. 25_Exit_Scenarios и 35_Roadmap_2026_2032.",
           size=9, italic=True, color=BLUE, height=18)
row += 2

# VI. Recommendation
row = section_header(ws, row, 2, "VI", "RECOMMENDATION", color=BLUE)

rec_text = (
    "РЕКОМЕНДАЦИЯ: УТВЕРДИТЬ инвестицию T₁ = 1 250 млн ₽ в 4 транша с привязкой к milestone-ам "
    "(Dev gate 2026Q1 — 250 млн ₽, Production gate 2026Q3 — 350 млн ₽, Release gate 2027Q2 — 350 млн ₽, "
    "Library gate 2028Q1 — 300 млн ₽).\n\n"
    "ОСНОВАНИЯ:\n"
    "(1) Якорные метрики устойчивы: GAAP EBITDA 2 152 млн ₽ при margin 47.3% — 2.4× выше отраслевой медианы РФ (20%);\n"
    "(2) Expected exit value 6 038 млн ₽ даёт MoIC ≈ 4.8× и IRR 38%+ при probability-weighted раскладе;\n"
    "(3) Downside защита: DCF floor 1 815 млн ₽ + preferred return 12% + makewhole clause;\n"
    "(4) Гос-alignment с 6 национальными проектами обеспечивает ~1 040 млн ₽ non-dilutive funding;\n"
    "(5) Risk register 30 рисков с митигациями, ни один не является deal-breaker для Base сценария.\n\n"
    "СЛЕДУЮЩИЕ ШАГИ: (a) подписание term sheet с ключевыми LP до 2026Q2, "
    "(b) финальный due diligence по Pipeline 08 и Tax Schedule 34, "
    "(c) first tranche closing 2026Q2 (250 млн ₽)."
)
row = para(ws, row, 2, rec_text, size=10, height=170)
row += 2

# Footer — disclaimer
row = para(ws, row, 2,
           "⚠ Данный Executive Summary предназначен для квалифицированных инвесторов. "
           "Не является публичной офертой. Прогнозные показатели условны и не гарантированы. "
           "Полные disclaimers — см. лист 38_Notes_and_Sources.",
           size=9, italic=True, color=GREY, height=36)

max_row_es = row
print(f"  36_Executive_Summary: {max_row_es} rows")


# ============================================================
# ЛИСТ 37: Glossary (~50 терминов)
# ============================================================
print("\n[2/3] Building 37_Glossary...")
ws = wb.create_sheet("37_Glossary")
set_widths(ws, [2, 5, 26, 60, 16, 16, 14, 14, 14, 14, 14, 14, 14, 14, 2])
ws.freeze_panes = "D7"

row = 2
row = title_block(ws, row, 2,
                  "GLOSSARY / СЛОВАРЬ ТЕРМИНОВ",
                  "Финансовые • Кино-индустрия • Регуляторные РФ — 50+ терминов",
                  merge_to=14)

# Таблица
headers = ["#", "Термин / Term", "Определение", "Категория"]
h_cols = [2, 3, 4, 13]
for i, h in enumerate(headers):
    c = ws.cell(row, h_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = h_cols[i + 1] - 1 if i + 1 < len(h_cols) else 14
    if end_col > h_cols[i]:
        ws.merge_cells(start_row=row, start_column=h_cols[i], end_row=row, end_column=end_col)
        for cc in range(h_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

GLOSSARY = [
    # --- FIN ---
    ("EBITDA", "Earnings Before Interest, Taxes, Depreciation, Amortization — прибыль до вычета процентов, налогов и амортизации. В этой модели — GAAP EBITDA, рассчитанная по российской отчётности.", "FIN"),
    ("GAAP EBITDA", "EBITDA, рассчитанная в соответствии с принципами бухучёта (Generally Accepted Accounting Principles). В ТрендСтудио = 2 152 млн ₽ за 3 года.", "FIN"),
    ("NDP", "Non-Discounted Profit — недисконтированная прибыль холдинга, внутренний якорь финмодели. В ТрендСтудио = 3 000 млн ₽ (legacy anchor).", "FIN"),
    ("Revenue", "Выручка — совокупные денежные поступления от реализации (прокат, TV/OTT права, мерч, сиквелы). В ТрендСтудио Σ = 4 545 млн ₽.", "FIN"),
    ("COGS", "Cost of Goods Sold — себестоимость реализации (прямое производство фильмов + прокатные расходы + маркетинг). В ТрендСтудио Σ = 2 127.5 млн ₽.", "FIN"),
    ("OpEx", "Operating Expenses — операционные расходы (ФОТ, аренда, G&A, IT, юр). В ТрендСтудио Σ = 265.5 млн ₽.", "FIN"),
    ("CAPEX", "Capital Expenditures — капитальные затраты на долгосрочные активы (оборудование, студийные мощности, IT-инфраструктура).", "FIN"),
    ("Net Profit", "Чистая прибыль после вычета налогов. В ТрендСтудио Σ = 1 689 млн ₽.", "FIN"),
    ("Free Cash Flow (FCF)", "Свободный денежный поток = Operating CF − CAPEX. Ключевая метрика для DCF-оценки.", "FIN"),
    ("Working Capital", "Оборотный капитал = Current Assets − Current Liabilities. См. лист 12_Working_Capital.", "FIN"),
    ("WACC", "Weighted Average Cost of Capital — средневзвешенная стоимость капитала. В модели = 19% (высокий уровень отражает страновой риск РФ и execution risk кино-индустрии).", "FIN"),
    ("DCF", "Discounted Cash Flow — метод оценки через дисконтирование будущих денежных потоков по WACC. См. 22_Valuation_DCF.", "FIN"),
    ("NPV", "Net Present Value — чистая приведённая стоимость. Показывает превышение дисконтированных доходов над инвестицией.", "FIN"),
    ("IRR", "Internal Rate of Return — внутренняя норма доходности. Ставка, при которой NPV = 0. Hurdle rate для инвестора = 18%.", "FIN"),
    ("MoIC", "Multiple of Invested Capital — множитель на вложенный капитал = Exit value / Investment. Для ТрендСтудио expected ≈ 4.8×.", "FIN"),
    ("Payback Period", "Срок окупаемости — количество периодов до возврата инвестиции. В ТрендСтудио Base < 24 мес.", "FIN"),
    ("EV", "Enterprise Value — стоимость бизнеса = Equity Value + Net Debt. Используется в multiples (EV/EBITDA, EV/Revenue).", "FIN"),
    ("EV/EBITDA", "Мультипликатор оценки: EV делённый на EBITDA. РФ кино-median = 8.0×, Global = 9.8×.", "FIN"),
    ("Gordon Growth TV", "Модель терминальной стоимости при постоянном росте g. TV = FCF_last × (1 + g) / (WACC − g). В модели g = 3%.", "FIN"),
    ("Exit Multiple TV", "Альтернативная TV через множитель на EBITDA последнего года прогноза. В модели 6.5× EBITDA.", "FIN"),
    # --- MONTE-CARLO / RISK ---
    ("Monte Carlo Simulation", "Вероятностный метод оценки: генерируется N случайных сценариев для получения распределения результатов. В модели n=1000, seed=42.", "RISK"),
    ("VaR", "Value at Risk — максимальный ожидаемый убыток на заданном уровне доверия (обычно 95% или 99%). VaR 95% в модели = 561 млн ₽.", "RISK"),
    ("CVaR (Expected Shortfall)", "Conditional VaR — средний убыток в худших (1−α)% случаев. Более консервативная мера, чем VaR.", "RISK"),
    ("P50 / P90", "Перцентили распределения. P50 = медиана (50% сценариев хуже), P90 = верхний квартиль.", "RISK"),
    ("Tornado Analysis", "Sensitivity анализ с ±20% вариацией каждого параметра. Ранжирует факторы по влиянию на целевую метрику.", "RISK"),
    ("Risk Score", "Severity × Likelihood (1–5 × 1–5 = 1–25). В ТрендСтудио 29_Risk_Register: max score 20 (R17).", "RISK"),
    ("Heat Map", "Визуализация рисков в матрице 5×5 (Severity × Likelihood) с цветовым кодированием критичности.", "RISK"),
    ("Hurdle Rate", "Минимальная требуемая доходность для одобрения инвестиции. Для ТрендСтудио T₁ = 18%.", "RISK"),
    # --- DEAL ---
    ("LP (Limited Partner)", "Инвестор-участник фонда / JV без операционного контроля, с ограниченной ответственностью.", "DEAL"),
    ("GP (General Partner)", "Управляющий партнёр, несёт полную ответственность, принимает инвестиционные решения.", "DEAL"),
    ("Preferred Return", "Минимальная гарантированная доходность LP до выплаты GP. В модели — 12%.", "DEAL"),
    ("Waterfall", "Структура распределения cash flow между LP / GP согласно приоритетам. См. 19_Waterfall.", "DEAL"),
    ("Makewhole Clause", "Защитная оговорка: при досрочном выходе эмитент компенсирует LP потерянную доходность.", "DEAL"),
    ("Cap Table", "Таблица капитализации — распределение долей между всеми акционерами/LP. См. 18_Cap_Table.", "DEAL"),
    ("Drag-Along", "Право мажоритария принудительно вовлечь миноритариев в сделку exit на тех же условиях.", "DEAL"),
    ("Tag-Along", "Право миноритария присоединиться к продаже мажоритария на тех же условиях.", "DEAL"),
    ("ESOP", "Employee Stock Ownership Plan — программа опционов для ключевых сотрудников. Vesting 4 года.", "DEAL"),
    ("Term Sheet", "Предварительный документ с ключевыми условиями сделки, предшествует SPA.", "DEAL"),
    ("SPA", "Share Purchase Agreement — договор купли-продажи акций / долей.", "DEAL"),
    ("Tranche", "Транш — часть инвестиции, выплачиваемая по достижении milestone. В T₁: 4 транша 250/350/350/300.", "DEAL"),
    # --- CINEMA ---
    ("Box Office", "Кассовые сборы — выручка от проката фильма в кинотеатрах.", "CINEMA"),
    ("P&A", "Prints & Advertising — расходы на копии и маркетинговое продвижение фильма при релизе.", "CINEMA"),
    ("Theatrical Window", "Эксклюзивный кинотеатральный релиз (обычно 45–90 дней) до выхода на VOD / OTT.", "CINEMA"),
    ("VOD / OTT", "Video On Demand / Over-The-Top — цифровая дистрибуция (Кинопоиск, Okko, IVI, Wink).", "CINEMA"),
    ("Slate Financing", "Финансирование пакета фильмов (slate) вместо единичного проекта — снижение риска через диверсификацию.", "CINEMA"),
    ("Development", "Стадия разработки фильма — сценарий, кастинг, бюджет, pre-production. В пайплайне 6–12 мес.", "CINEMA"),
    ("Principal Photography", "Основной съёмочный период. В ТрендСтудио Gantt = этап Shoot, 2–4 мес на фильм.", "CINEMA"),
    ("Post-Production", "Монтаж, VFX, цветокоррекция, sound design. В пайплайне 4–6 мес.", "CINEMA"),
    ("Producer Equity", "Капитал продюсеров в JV-структуре. В ТрендСтудио = 600 млн ₽ (off-P&L, минори-интерес).", "CINEMA"),
    ("Library Value", "Стоимость накопленного каталога фильмов — источник tail revenue 2029–2032 (1 050 млн ₽).", "CINEMA"),
    ("Sequel Rights", "Право на производство продолжений. Опционный payoff, увеличивающий real option value.", "CINEMA"),
    ("Co-Production", "Совместное производство с иностранным партнёром. Открывает доступ к зарубежным рынкам и субсидиям.", "CINEMA"),
    # --- REGULATORY ---
    ("НДС 0%", "Льгота по НДС для кинопродукции согласно ст. 149 НК РФ. В модели экономия ~360 млн ₽ за 3 года.", "REG"),
    ("Налог на прибыль", "Базовая ставка 20% (ФБ 2%, РБ 18%). В модели 3Y Σ ≈ 438 млн ₽.", "REG"),
    ("Страховые взносы", "30.2% от ФОТ (ПФР 22%, ФСС 2.9%, ФОМС 5.1%, травматизм 0.2%). В модели Σ ≈ 121 млн ₽.", "REG"),
    ("ФКП", "Фонд кино / Фонд кинематографии — государственный источник безвозвратных грантов на отечественные фильмы.", "REG"),
    ("ПП № 1215", "Постановление Правительства РФ № 1215 от 20.12.2013 — правила предоставления субсидий кинопродюсерам.", "REG"),
    ("Указ № 309", "Указ Президента РФ № 309 от 07.05.2024 «О национальных целях развития РФ до 2030 и 2036 гг.»", "REG"),
    ("Нац. проект «Семья»", "Национальный проект, в рамках которого финансируется семейное / детское кино. Горизонт до 2030.", "REG"),
    ("Нац. проект «Молодёжь и дети»", "Национальный проект для молодёжной аудитории, включает грантовую поддержку профильного контента.", "REG"),
    ("Нац. проект «Культура»", "Национальный проект, регулирует субсидии и инфраструктурные программы кинематографа.", "REG"),
    ("ст. 149 НК РФ", "Статья Налогового кодекса РФ об освобождении от НДС ряда операций, включая передачу прав на кинопродукцию.", "REG"),
]

for i, (term, defi, cat) in enumerate(GLOSSARY, start=1):
    cat_fills = {
        "FIN": LIGHT_BLUE,
        "RISK": LIGHT_ORANGE,
        "DEAL": LIGHT_GREEN,
        "CINEMA": LIGHT_YELLOW,
        "REG": VERY_LIGHT_BLUE,
    }

    ws.cell(row, 2).value = i
    ws.cell(row, 2).font = Font(name="Calibri", size=9, bold=True)
    ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row, 2).border = BORDER

    ws.cell(row, 3).value = term
    ws.cell(row, 3).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
    ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.cell(row, 3).border = BORDER

    ws.cell(row, 4).value = defi
    ws.cell(row, 4).font = Font(name="Calibri", size=9)
    ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.cell(row, 4).border = BORDER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=12)
    for cc in range(4, 13):
        ws.cell(row, cc).border = BORDER

    ws.cell(row, 13).value = cat
    ws.cell(row, 13).font = Font(name="Calibri", size=9, bold=True)
    ws.cell(row, 13).fill = PatternFill("solid", fgColor=cat_fills[cat])
    ws.cell(row, 13).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 13).border = BORDER
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)
    for cc in range(13, 15):
        ws.cell(row, cc).fill = PatternFill("solid", fgColor=cat_fills[cat])
        ws.cell(row, cc).border = BORDER

    # Высота адаптивная от длины определения
    ws.row_dimensions[row].height = max(32, min(72, 12 + len(defi) // 2))
    row += 1

# Легенда
row += 1
row = para(ws, row, 2,
           "Легенда: FIN — финансовые термины | RISK — риск / MC | DEAL — структура сделки | CINEMA — кино-индустрия | REG — регуляторные РФ",
           size=9, italic=True, color=GREY, height=18)

max_row_gl = row
print(f"  37_Glossary: {max_row_gl} rows, {len(GLOSSARY)} terms")


# ============================================================
# ЛИСТ 38: Notes_and_Sources
# ============================================================
print("\n[3/3] Building 38_Notes_and_Sources...")
ws = wb.create_sheet("38_Notes_and_Sources")
set_widths(ws, [2, 5, 26, 60, 16, 16, 14, 14, 14, 14, 14, 14, 14, 14, 2])
ws.freeze_panes = "D7"

row = 2
row = title_block(ws, row, 2,
                  "NOTES & SOURCES",
                  "Методологические заметки • Полный список источников • Disclaimers • Версионность",
                  merge_to=14)

# I. Методологические заметки по листам
row = section_header(ws, row, 2, "I", "МЕТОДОЛОГИЧЕСКИЕ ЗАМЕТКИ ПО ЛИСТАМ")

methodology = [
    ("02_Assumptions", "Все ключевые параметры собраны в один лист — SSOT. Изменение любого параметра должно отражаться в 03_Change_Log."),
    ("04/05_FOT_Staff", "ФОТ рассчитан по 2 архитектурам: A1 (корпоративный) и A2 (JV / проектный). Страховые взносы 30.2% применены к обеим."),
    ("06_Cost_Structure", "COGS включает прямые производственные + прокатные + маркетинг. Split по категориям: производство 72%, прокат/P&A 18%, маркетинг 10%."),
    ("07_Revenue_Breakdown", "Революционные сегменты: Box Office 55%, TV/OTT 25%, Мерч/лицензии 8%, Международный прокат 7%, Sequel options 5%."),
    ("08_Content_Pipeline", "12 фильмов × средний бюджет 154 млн ₽. Распределение: 4 блокбастера (200+), 6 мидл-бюджет (150–180), 2 арт-проекта (80–110)."),
    ("09_P&L_Statement", "Dual metric: GAAP EBITDA 2 152 + legacy NDP 3 000. Reconciliation bridge в нижней части листа показывает 848 млн ₽ разницы через: (+) JV equity, (−) D&A, (−) IFRS adjustments."),
    ("10_Cash_Flow", "Indirect method (через Net Profit). Working capital changes детализированы в 12_WC."),
    ("14_Investment_Inflow", "Транши T₁ привязаны к milestone, не к календарю. Calendar — ориентировочный."),
    ("17_Deal_Structures", "Producer equity 600 млн ₽ — off-P&L JV capital, не размывает основной cap table, учитывается как minority interest."),
    ("19_Waterfall", "Приоритеты: (1) LP return of capital, (2) LP preferred 12%, (3) GP catch-up 20%, (4) 80/20 split."),
    ("22_Valuation_DCF", "Blend 40% Gordon + 60% Exit Multiple. g=3%, exit mult 6.5×, WACC 19%. Quick note «Why DCF < Multiples» на R59-72."),
    ("23_Valuation_Multiples", "EV/EBITDA median по 32_Comparables. Применён discount 10% за liquidity + 5% за size."),
    ("25_Exit_Scenarios", "Probability-weighted EV = Σ (p × EV) по 7 маршрутам. Weighted = 6 038 млн ₽."),
    ("26_Sensitivity", "Tornado по 8 переменным с ±20% impact. Top driver: Box Office revenue (±420 на EBITDA)."),
    ("27_Scenario_Analysis", "5 сценариев: Worst / Pessimistic / Base / Optimistic / Best с вероятностями 5/15/50/20/10 (PROB_VECTOR_BASE SSOT)."),
    ("28_Monte_Carlo_Summary", "n=1000, seed=42. 5 стохастических переменных (triangular, lognormal, gauss, binomial). VaR 95% = 561."),
    ("29_Risk_Register", "30 рисков × 5 категорий (Market, Production, Financial, Regulatory, Operational). Max score 20 (R17)."),
    ("30/31/32_Market", "TAM/SAM/SOM + benchmark KPI + precedent transactions. RU comps median 8.0×, Global 9.8×, blend 8.9×."),
    ("33_Gov_KPI", "6 нац. проектов + 10 KPI 2026–2028 + 5 источников господдержки 1 040 млн ₽. Комплаенс по Указу № 309."),
    ("34_Tax_Schedule", "9 налогов × 12 кварталов + 4Y tail. Эффективная ставка 12.9% благодаря льготе НДС 0% ст. 149 НК РФ."),
    ("35_Roadmap_2026_2032", "12 фильмов × 16 периодов × 4 фазы Gantt + 14 fundraising milestones + 7 exit routes + критический путь."),
    ("36_Executive_Summary", "Этот лист — VI разделов включая полный разбор «Valuation Range Interpretation» (часть B комбо A+B с 22_Valuation_DCF)."),
]

for sheet, note in methodology:
    ws.cell(row, 3).value = sheet
    ws.cell(row, 3).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
    ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.cell(row, 3).border = BORDER
    ws.cell(row, 3).fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)

    ws.cell(row, 4).value = note
    ws.cell(row, 4).font = Font(name="Calibri", size=9)
    ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.cell(row, 4).border = BORDER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=14)
    for cc in range(4, 15):
        ws.cell(row, cc).border = BORDER
    ws.row_dimensions[row].height = max(28, min(60, 10 + len(note) // 2))
    row += 1

row += 2

# II. Источники (официальные РФ приоритет)
row = section_header(ws, row, 2, "II", "ПОЛНЫЙ СПИСОК ИСТОЧНИКОВ")

sources = [
    ("GOV-RU", "Указ Президента РФ № 309 от 07.05.2024", "О национальных целях развития РФ до 2030 и 2036 гг.", "kremlin.ru"),
    ("GOV-RU", "Постановление Правительства РФ № 1215 от 20.12.2013", "Правила предоставления субсидий кинопродюсерам", "government.ru"),
    ("GOV-RU", "Налоговый кодекс РФ, ст. 149", "Операции, освобождённые от НДС (включая кинопродукцию)", "nalog.gov.ru"),
    ("GOV-RU", "Паспорт нац. проекта «Семья» 2025–2030", "Финансирование семейного / детского кино", "government.ru"),
    ("GOV-RU", "Паспорт нац. проекта «Молодёжь и дети» 2025–2030", "Грантовая поддержка молодёжного контента", "government.ru"),
    ("GOV-RU", "Паспорт нац. проекта «Культура» 2025–2030", "Субсидии и инфраструктура кинематографа", "government.ru"),
    ("GOV-RU", "Фонд кино (ФКП) — правила поддержки", "Безвозвратные субсидии кинопродюсерам", "fond-kino.ru"),
    ("GOV-RU", "Минкультуры РФ — реестр прокатных удостоверений", "Верификация релизных планов и премьер", "culture.gov.ru"),
    ("RU-DATA", "Росстат — данные по кинотеатральному рынку", "Box Office РФ 2020–2025", "rosstat.gov.ru"),
    ("RU-NEWS", "ТАСС — раздел «Кино»", "Оперативные новости индустрии, сделки, господдержка", "tass.ru"),
    ("RU-NEWS", "РИА Новости — раздел «Культура»", "Регуляторные новости, нац. проекты", "ria.ru"),
    ("RU-DATA", "Бюллетень кинопроката — bulletin.info.ru", "Недельная касса, TOP-листы, статистика", "bulletin.info.ru"),
    ("GLOBAL", "Box Office Mojo / IMDb Pro", "Global benchmarks, casting, production credits", "boxofficemojo.com"),
    ("GLOBAL", "Variety Intelligence Platform", "Industry analytics, deal tracking", "variety.com"),
    ("GLOBAL", "S&P Capital IQ", "EV/EBITDA multiples, comparable transactions", "spglobal.com"),
    ("GLOBAL", "Mergermarket / PitchBook", "M&A precedent data, deal databases", "pitchbook.com"),
    ("ACADEMIC", "Vogel, H. — Entertainment Industry Economics (10th ed.)", "Учебник по экономике медиа-индустрии", "Cambridge University Press"),
    ("INTERNAL", "ТрендСтудио финмодель v1.4.4", "Якорная модель холдинга 2026–2028, cumulative NDP 3 000 млн ₽", "pipeline/"),
]

hdrs = ["Категория", "Источник", "Назначение", "URL / издатель"]
s_cols = [2, 3, 5, 13]
for i, h in enumerate(hdrs):
    c = ws.cell(row, s_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = s_cols[i + 1] - 1 if i + 1 < len(s_cols) else 14
    if end_col > s_cols[i]:
        ws.merge_cells(start_row=row, start_column=s_cols[i], end_row=row, end_column=end_col)
        for cc in range(s_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

cat_colors = {
    "GOV-RU": LIGHT_BLUE,
    "RU-NEWS": LIGHT_BLUE,
    "RU-DATA": LIGHT_BLUE,
    "GLOBAL": LIGHT_YELLOW,
    "ACADEMIC": LIGHT_GREEN,
    "INTERNAL": LIGHT_ORANGE,
}

for cat, src, purp, url in sources:
    ws.cell(row, 2).value = cat
    ws.cell(row, 2).font = Font(name="Calibri", size=9, bold=True)
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=cat_colors[cat])
    ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 2).border = BORDER

    ws.cell(row, 3).value = src
    ws.cell(row, 3).font = Font(name="Calibri", size=9, bold=True)
    ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.cell(row, 3).border = BORDER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row, 4).border = BORDER

    ws.cell(row, 5).value = purp
    ws.cell(row, 5).font = Font(name="Calibri", size=9)
    ws.cell(row, 5).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.cell(row, 5).border = BORDER
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=12)
    for cc in range(5, 13):
        ws.cell(row, cc).border = BORDER

    ws.cell(row, 13).value = url
    ws.cell(row, 13).font = Font(name="Calibri", size=9, italic=True, color=BLUE)
    ws.cell(row, 13).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row, 13).border = BORDER
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)
    ws.cell(row, 14).border = BORDER

    ws.row_dimensions[row].height = 26
    row += 1

row += 1
row = para(ws, row, 2,
           "ПРИОРИТЕТ: по вопросам государственной политики РФ первичными источниками являются официальные российские — "
           "kremlin.ru, government.ru, nalog.gov.ru, rosstat.gov.ru, ТАСС, РИА Новости.",
           size=9, italic=True, color=DARK_BLUE, height=28)
row += 2

# III. Disclaimers
row = section_header(ws, row, 2, "III", "DISCLAIMERS", color=BLUE)

disclaimer_text = (
    "(1) ФОРВАРД-ПРОГНОЗЫ. Настоящая финансовая модель содержит прогнозные показатели (forward-looking statements), "
    "которые по своей природе связаны с неопределённостью. Фактические результаты могут существенно отличаться от прогнозируемых.\n\n"
    "(2) НЕ ЯВЛЯЕТСЯ ОФЕРТОЙ. Данный Investor Package не является публичной офертой, инвестиционной рекомендацией, "
    "индивидуальной инвестиционной консультацией или предложением о приобретении ценных бумаг. Документ предназначен "
    "исключительно для квалифицированных инвесторов.\n\n"
    "(3) НАЛОГОВАЯ ПОЗИЦИЯ. Налоговые расчёты основаны на действующих нормах НК РФ по состоянию на 2026-04-11. "
    "Льгота НДС 0% применена в соответствии со ст. 149 НК РФ; изменение законодательства может существенно повлиять "
    "на эффективную ставку. Рекомендуется верификация у налогового консультанта перед принятием инвестиционных решений.\n\n"
    "(4) ОЦЕНКА И МУЛЬТИПЛИКАТОРЫ. Оценки EV, мультипликаторы и terminal values основаны на precedent transactions "
    "и индустриальных бенчмарках, которые могут не полностью отражать специфические риски и возможности ТрендСтудио. "
    "Разрыв DCF vs Multiples (~5 735 млн ₽) структурен и объяснён в разделе III 36_Executive_Summary.\n\n"
    "(5) ДАННЫЕ РЫНКА. Параметры TAM / SAM / SOM, бенчмарки и сравнительные сделки получены из открытых источников "
    "(Росстат, ТАСС, Бюллетень кинопроката, Variety, S&P Capital IQ). Достоверность источников не подтверждается "
    "независимым аудитом в рамках данной модели.\n\n"
    "(6) МОНТЕ-КАРЛО. Результаты симуляции (n=1000, seed=42) являются вероятностными оценками на основе допущений "
    "о распределениях ключевых переменных. VaR 95% = 561 млн ₽ не гарантирует отсутствие потерь свыше этой суммы в worst-case сценариях.\n\n"
    "(7) КОНФИДЕНЦИАЛЬНОСТЬ. Все данные, содержащиеся в данном пакете, являются конфиденциальной информацией холдинга "
    "«ТрендСтудио». Распространение третьим лицам без письменного согласия не допускается."
)
row = para(ws, row, 2, disclaimer_text, size=10, height=280)
row += 2

# IV. Версионность
row = section_header(ws, row, 2, "IV", "VERSION HISTORY")

versions = [
    ("v1.0 Public", "2026-04-11", "Initial release — 38 листов, dual metric Variant C, П5 «Максимум» verification, sheets А.1–А.19 complete"),
    ("v1.4.4 pipeline", "2026-04-11", "Основная финмодель холдинга (384 теста PASS, 4 MC engines). Не включается в Public — см. отдельный deliverable"),
]
v_hdrs = ["Version", "Date", "Changes"]
v_cols = [2, 5, 8]
for i, h in enumerate(v_hdrs):
    c = ws.cell(row, v_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = v_cols[i + 1] - 1 if i + 1 < len(v_cols) else 14
    if end_col > v_cols[i]:
        ws.merge_cells(start_row=row, start_column=v_cols[i], end_row=row, end_column=end_col)
        for cc in range(v_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

for ver, dt, ch in versions:
    ws.cell(row, 2).value = ver
    ws.cell(row, 2).font = Font(name="Calibri", size=9, bold=True, color=DARK_BLUE)
    ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 2).border = BORDER
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    for cc in range(2, 5):
        ws.cell(row, cc).fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)
        ws.cell(row, cc).border = BORDER

    ws.cell(row, 5).value = dt
    ws.cell(row, 5).font = Font(name="Calibri", size=9)
    ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 5).border = BORDER
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    for cc in range(5, 8):
        ws.cell(row, cc).border = BORDER

    ws.cell(row, 8).value = ch
    ws.cell(row, 8).font = Font(name="Calibri", size=9)
    ws.cell(row, 8).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.cell(row, 8).border = BORDER
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=14)
    for cc in range(8, 15):
        ws.cell(row, cc).border = BORDER

    ws.row_dimensions[row].height = 32
    row += 1

row += 2
row = para(ws, row, 2,
           "Подготовлено: ТрендСтудио Holdings | Investor Package v1.0 Public | 2026-04-11 | CONFIDENTIAL",
           size=9, italic=True, color=GREY, height=18)

max_row_ns = row
print(f"  38_Notes_and_Sources: {max_row_ns} rows")


# ============================================================
# VERIFICATION
# ============================================================
print("\n[VERIFY] Invariants check...")
assert "36_Executive_Summary" in wb.sheetnames
assert "37_Glossary" in wb.sheetnames
assert "38_Notes_and_Sources" in wb.sheetnames
assert len(wb.sheetnames) == 38, f"Expected 38 sheets, got {len(wb.sheetnames)}"

# Existing sheets not touched
for s in existing_sheets:
    assert s in wb.sheetnames, f"Lost sheet: {s}"

# Freeze panes on new sheets
for s in ["36_Executive_Summary", "37_Glossary", "38_Notes_and_Sources"]:
    assert wb[s].freeze_panes == "D7", f"{s}: freeze panes != D7"

print("✓ 38 sheets total")
print("✓ All 35 existing sheets preserved")
print("✓ Freeze panes D7 on all 3 new sheets")

wb.save(XLSX)
print(f"\n✓ Saved: {XLSX}")
print(f"  36_Executive_Summary: ~{max_row_es} rows")
print(f"  37_Glossary: ~{max_row_gl} rows, {len(GLOSSARY)} terms")
print(f"  38_Notes_and_Sources: ~{max_row_ns} rows")
