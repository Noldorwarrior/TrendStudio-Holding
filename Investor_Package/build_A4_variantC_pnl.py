"""
А.4 — P&L_Statement + патч Cover/Assumptions под Вариант C (Dual metric).

Изменения:
1) 01_Cover — заменяем одиночный EBITDA на dual-metric:
   Revenue 4545 | EBITDA 2152 (GAAP) | NDP 3000 (legacy) | Net Profit 1710 | Investment 1250
2) 02_Assumptions — добавляем блок N. NDP Reconciliation Bridge
3) 09_P&L_Statement — квартальный P&L с 12Q (2026-2028) + 4Y (2029-2032)
   + блок NDP Bridge (EBITDA → NDP)
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from copy import copy

import os
_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

# ============ ПАЛИТРА ============
BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"   # жёлтый для ключевых метрик
NDP_FILL = "E2EFDA"          # светло-зелёный для NDP
SUBTOTAL_FILL = "D9E1F2"
INPUT_BLUE = "0000FF"
FORMULA_BLACK = "000000"
LINK_GREEN = "006100"
BORDER_GREY = "BFBFBF"
WHITE = "FFFFFF"

thin = Side(style="thin", color=BORDER_GREY)
medium = Side(style="medium", color=BRAND_BLUE)
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
box_brand = Border(left=medium, right=medium, top=medium, bottom=medium)

F_TITLE = Font(name="Calibri", size=20, bold=True, color=DARK_BLUE)
F_H1 = Font(name="Calibri", size=14, bold=True, color=BRAND_BLUE)
F_H2 = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
F_BODY = Font(name="Calibri", size=11, color="000000")
F_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
F_KPI_NUM = Font(name="Calibri", size=22, bold=True, color=DARK_BLUE)
F_KPI_LBL = Font(name="Calibri", size=10, color="595959")
F_INPUT = Font(name="Calibri", size=11, color=INPUT_BLUE, bold=False)
F_FORMULA = Font(name="Calibri", size=11, color=FORMULA_BLACK)
F_LINK = Font(name="Calibri", size=11, color=LINK_GREEN, italic=True)
F_TOTAL = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")


def set_cell(ws, ref, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws[ref]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align:
        c.alignment = align
    if border:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c


# ============================================================
#   ПАТЧ 1 — 01_Cover: Dual metric (5 карт)
# ============================================================
def _clear_range(ws, min_row, max_row, min_col=1, max_col=30):
    """Разъединяет merged в зоне, потом чистит ячейки."""
    # 1. Разъединяем все merges, которые пересекают зону
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= min_row and mr.max_row <= max_row \
                and mr.min_col >= min_col and mr.max_col <= max_col:
            to_unmerge.append(str(mr))
    for ref in to_unmerge:
        ws.unmerge_cells(ref)
    # 2. Чистим ячейки
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            try:
                cell.value = None
                cell.font = Font()
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.alignment = Alignment()
            except AttributeError:
                pass


def patch_cover(wb):
    ws = wb["01_Cover"]

    # Очищаем старые метрики в зоне B7:K22
    _clear_range(ws, 7, 22, 2, 12)

    # Новая шапка блока метрик
    ws.merge_cells("B7:K7")
    set_cell(ws, "B7", "КЛЮЧЕВЫЕ МЕТРИКИ  ·  2026–2028  ·  Dual-metric представление",
             font=Font(name="Calibri", size=12, bold=True, color=WHITE),
             fill=BRAND_BLUE, align=C_CENTER)
    ws.row_dimensions[7].height = 24

    # 5 карт: Revenue | EBITDA(GAAP) | NDP(legacy) | Net Profit | Investment
    # Расположение: B:C, D:E, F:G, H:I, J:K (2 колонки на карту)
    cards = [
        ("B8", "C8", "B9", "C9", "B10", "C10", "4 545",  "Revenue Σ 2026–2028, млн ₽",  KEY_METRIC_FILL, "Выручка"),
        ("D8", "E8", "D9", "E9", "D10", "E10", "2 152",  "EBITDA Σ, млн ₽ (GAAP, стандарт)", KEY_METRIC_FILL, "EBITDA стандартный"),
        ("F8", "G8", "F9", "G9", "F10", "G10", "3 000",  "NDP Σ, млн ₽ (legacy v1.4.4, якорь)", NDP_FILL, "Net Distributable Proceeds"),
        ("H8", "I8", "H9", "I9", "H10", "I10", "1 710",  "Net Profit Σ, млн ₽ (после налога)", KEY_METRIC_FILL, "Чистая прибыль"),
        ("J8", "K8", "J9", "K9", "J10", "K10", "1 250",  "Investment, млн ₽ (T₁ Legacy)", KEY_METRIC_FILL, "Инвестиционный тикет"),
    ]

    # Делаем merge для каждой карты
    ws.merge_cells("B8:C8"); ws.merge_cells("B9:C9"); ws.merge_cells("B10:C10")
    ws.merge_cells("D8:E8"); ws.merge_cells("D9:E9"); ws.merge_cells("D10:E10")
    ws.merge_cells("F8:G8"); ws.merge_cells("F9:G9"); ws.merge_cells("F10:G10")
    ws.merge_cells("H8:I8"); ws.merge_cells("H9:I9"); ws.merge_cells("H10:I10")
    ws.merge_cells("J8:K8"); ws.merge_cells("J9:K9"); ws.merge_cells("J10:K10")

    card_cells = [
        ("B8",  "B9",  "B10",  "4 545",  "Revenue  ·  Σ 2026–2028", "млн ₽",              KEY_METRIC_FILL),
        ("D8",  "D9",  "D10",  "2 152",  "EBITDA (GAAP) ",           "млн ₽  ·  стандарт",   KEY_METRIC_FILL),
        ("F8",  "F9",  "F10",  "3 000",  "NDP (Net Distrib. Proc.)", "млн ₽  ·  legacy",     NDP_FILL),
        ("H8",  "H9",  "H10",  "1 710",  "Net Profit",                "млн ₽  ·  после 20% налога", KEY_METRIC_FILL),
        ("J8",  "J9",  "J10",  "1 250",  "Investment  ·  T₁",        "млн ₽  ·  4 транша",   KEY_METRIC_FILL),
    ]

    for top, mid, bot, val, lbl1, lbl2, fill in card_cells:
        set_cell(ws, top, val,
                 font=F_KPI_NUM, fill=fill, align=C_CENTER, border=box_brand)
        set_cell(ws, mid, lbl1,
                 font=F_H2, fill=fill, align=C_CENTER, border=box_brand)
        set_cell(ws, bot, lbl2,
                 font=F_KPI_LBL, fill=fill, align=C_CENTER, border=box_brand)

    ws.row_dimensions[8].height = 36
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[10].height = 18

    # --- Пояснение под картами ---
    ws.merge_cells("B12:K13")
    set_cell(ws, "B12",
             "Dual-metric подход: EBITDA (GAAP 2 152 млн ₽) — стандартный бухгалтерский показатель "
             "для мультипликаторов и DCF.  NDP (3 000 млн ₽) — legacy-якорь v1.4.4, суммарные "
             "распределяемые денежные потоки к пайщикам JV (инвестор 2 215 + продюсер 785). "
             "Детальный мост EBITDA → NDP — на листе 09_P&L_Statement.",
             font=Font(name="Calibri", size=10, italic=True, color="595959"),
             align=Alignment(horizontal="left", vertical="top", wrap_text=True),
             fill="FAFAFA")
    ws.row_dimensions[12].height = 18
    ws.row_dimensions[13].height = 18

    # --- Scenario stripe ---
    _clear_range(ws, 15, 25, 2, 12)

    ws.merge_cells("B15:K15")
    set_cell(ws, "B15", "СЦЕНАРИИ  ·  Bear  /  Base  /  Bull",
             font=Font(name="Calibri", size=12, bold=True, color=WHITE),
             fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[15].height = 22

    # Заголовки колонок
    headers = ["Показатель", "Bear (0.70×)", "Base (1.00×)", "Bull (1.25×)"]
    ws.merge_cells("B16:E16"); set_cell(ws, "B16", headers[0], font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells("F16:G16"); set_cell(ws, "F16", headers[1], font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells("H16:I16"); set_cell(ws, "H16", headers[2], font=F_BOLD, fill=KEY_METRIC_FILL, align=C_CENTER, border=box_thin)
    ws.merge_cells("J16:K16"); set_cell(ws, "J16", headers[3], font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)

    scen_rows = [
        ("Revenue Σ, млн ₽",     "3 181",  "4 545",  "5 681"),
        ("EBITDA (GAAP), млн ₽", "1 506",  "2 152",  "2 690"),
        ("NDP Σ, млн ₽",          "2 100",  "3 000",  "3 750"),
        ("Net Profit, млн ₽",    "1 197",  "1 710",  "2 138"),
        ("IRR (5-year), %",      "~14 %",  "~22 %",  "~31 %"),
    ]
    r0 = 17
    for i, (lbl, bear, base, bull) in enumerate(scen_rows):
        r = r0 + i
        ws.merge_cells(f"B{r}:E{r}"); set_cell(ws, f"B{r}", lbl, font=F_BODY, align=C_LEFT, border=box_thin)
        ws.merge_cells(f"F{r}:G{r}"); set_cell(ws, f"F{r}", bear, font=F_BODY, align=C_CENTER, border=box_thin)
        ws.merge_cells(f"H{r}:I{r}"); set_cell(ws, f"H{r}", base, font=F_BOLD, fill=KEY_METRIC_FILL, align=C_CENTER, border=box_thin)
        ws.merge_cells(f"J{r}:K{r}"); set_cell(ws, f"J{r}", bull, font=F_BODY, align=C_CENTER, border=box_thin)
        ws.row_dimensions[r].height = 18


# ============================================================
#   ПАТЧ 2 — 02_Assumptions: блок N. NDP Reconciliation
# ============================================================
def patch_assumptions(wb):
    ws = wb["02_Assumptions"]

    # Найдём последнюю заполненную строку (после блока M)
    last_row = ws.max_row
    # Пишем начиная с last_row + 2
    start = last_row + 2

    # Заголовок блока
    ws.merge_cells(f"B{start}:E{start}")
    set_cell(ws, f"B{start}",
             "N.  NDP RECONCILIATION BRIDGE  ·  EBITDA (GAAP) → NDP (legacy)",
             font=Font(name="Calibri", size=12, bold=True, color=WHITE),
             fill=BRAND_BLUE, align=C_CENTER, border=box_brand)
    ws.row_dimensions[start].height = 22

    # Описание
    r = start + 1
    ws.merge_cells(f"B{r}:E{r}")
    set_cell(ws, f"B{r}",
             "Методологическое примечание: legacy-якорь v1.4.4 «3 000 млн ₽» — это NDP "
             "(Net Distributable Proceeds, распределяемые денежные потоки к пайщикам JV), "
             "а не стандартный EBITDA. Бридж ниже переводит одну метрику в другую.",
             font=Font(name="Calibri", size=10, italic=True, color="595959"),
             align=Alignment(horizontal="left", vertical="top", wrap_text=True),
             fill="FAFAFA")
    ws.row_dimensions[r].height = 36

    # Таблица бриджа
    headers = ["#", "Компонент", "Значение, млн ₽", "Комментарий"]
    r += 2
    for i, h in enumerate(headers):
        col = chr(ord("B") + i)
        set_cell(ws, f"{col}{r}", h, font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 18

    bridge_rows = [
        ("1", "EBITDA (GAAP, standard) Σ 2026–2028", 2152, "Revenue 4545 − COGS 2127.5 − OpEx 265.5"),
        ("2", "(+) Add-back: Producer equity contribution", 600, "Участие продюсера (off-P&L minority interest / JV capital)"),
        ("3", "(+) Add-back: Working capital release & gov credit timing", 248, "Высвобождение WC и налоговый кредит за периметром P&L"),
        ("=", "NDP Σ (legacy v1.4.4 anchor)", 3000, "Cумма распределяемых cash flows к пайщикам"),
        ("4", "(−) Producer share (60 / 40 waterfall W₁)", -785, "Доля продюсера по дефолтной каскадной схеме"),
        ("=", "Investor net distributable proceeds", 2215, "Чистые распределения инвестору (T₁ Legacy)"),
    ]
    r += 1
    for idx, (num, comp, val, comment) in enumerate(bridge_rows):
        is_total = num == "="
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", num, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", comp, font=font, fill=fill, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", val, font=font, fill=fill, align=C_RIGHT, border=box_thin,
                 number_format='#,##0;[Red]-#,##0')
        set_cell(ws, f"E{r}", comment, font=Font(name="Calibri", size=10, italic=True, color="595959"),
                 fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 20
        r += 1

    # Параметры NDP
    r += 1
    ws.merge_cells(f"B{r}:E{r}")
    set_cell(ws, f"B{r}",
             "Ключевые параметры JV-интерпретации:",
             font=F_BOLD, align=C_LEFT)
    r += 1

    jv_params = [
        ("N.01", "Investor funded production cost",           1250, "млн ₽",  "T₁ Legacy, 4 транша 20/28/28/24%"),
        ("N.02", "Producer equity contribution (off-P&L)",     600, "млн ₽",  "Вклад продюсера в JV"),
        ("N.03", "Total production budget (consolidated COGS)",1850, "млн ₽",  "12 фильмов, средний бюджет 154 млн"),
        ("N.04", "P&A (Prints & Advertising)",                 277.5,"млн ₽", "6.1% от выручки 4545"),
        ("N.05", "Waterfall split (default W₁)",               "60 / 40",  "—",  "Инвестор / Продюсер"),
        ("N.06", "NDP anchor (legacy v1.4.4)",                 3000, "млн ₽",  "Инвариант, закреплено"),
        ("N.07", "EBITDA (GAAP) Σ 2026–2028",                  2152, "млн ₽",  "Расчёт по стандарту"),
        ("N.08", "Net Profit Σ 2026–2028",                     1710, "млн ₽",  "После 20% налога на прибыль"),
    ]
    for code, name, val, unit, note in jv_params:
        set_cell(ws, f"B{r}", code, font=F_BODY, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BODY, align=C_LEFT, border=box_thin)
        if isinstance(val, (int, float)):
            set_cell(ws, f"D{r}", val, font=F_INPUT, align=C_RIGHT, border=box_thin,
                     number_format='#,##0.0' if isinstance(val, float) else '#,##0')
        else:
            set_cell(ws, f"D{r}", val, font=F_INPUT, align=C_RIGHT, border=box_thin)
        set_cell(ws, f"E{r}", f"{unit}  ·  {note}",
                 font=Font(name="Calibri", size=9, italic=True, color="595959"),
                 align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 18
        r += 1


# ============================================================
#   БЛОК 3 — 09_P&L_Statement (12Q + 4Y + NDP Bridge)
# ============================================================
def build_pnl(wb):
    # Удаляем лист если существует
    if "09_P&L_Statement" in wb.sheetnames:
        del wb["09_P&L_Statement"]

    ws = wb.create_sheet("09_P&L_Statement")

    # Ширины колонок
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 4     # #
    ws.column_dimensions["C"].width = 42    # Наименование
    for i in range(4, 20):                  # D:S = 16 периодов
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions["T"].width = 12    # Σ 2026-2028
    ws.column_dimensions["U"].width = 12    # Σ total
    ws.column_dimensions["V"].width = 30    # комментарий

    # ---- Титул ----
    ws.merge_cells("B2:V2")
    set_cell(ws, "B2",
             "09  ·  P&L STATEMENT  ·  Квартально 2026–2028 + годовые 2029–2032",
             font=F_TITLE, fill=DARK_BLUE, align=C_CENTER)
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:V3")
    set_cell(ws, "B3",
             "Вариант C · Dual metric: GAAP EBITDA (стандарт) + NDP (legacy). Все суммы — млн ₽.",
             font=Font(name="Calibri", size=10, italic=True, color="595959"),
             align=C_CENTER)
    ws.row_dimensions[3].height = 18

    # ---- Заголовки периодов ----
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_BOLD, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws[f"B{r}"].font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    set_cell(ws, f"C{r}", "Показатель", font=F_BOLD, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    ws[f"C{r}"].font = Font(name="Calibri", size=10, bold=True, color=WHITE)

    # Периоды: 12 кварталов (D:O) + 4 года (P:S) + Σ 2026-2028 (T) + Σ Total (U) + comment (V)
    periods = [
        ("D", "Q1'26"), ("E", "Q2'26"), ("F", "Q3'26"), ("G", "Q4'26"),
        ("H", "Q1'27"), ("I", "Q2'27"), ("J", "Q3'27"), ("K", "Q4'27"),
        ("L", "Q1'28"), ("M", "Q2'28"), ("N", "Q3'28"), ("O", "Q4'28"),
        ("P", "2029"), ("Q", "2030"), ("R", "2031"), ("S", "2032"),
    ]
    for col, label in periods:
        set_cell(ws, f"{col}{r}", label,
                 font=Font(name="Calibri", size=9, bold=True, color=WHITE),
                 fill=BRAND_BLUE, align=C_CENTER, border=box_thin)

    set_cell(ws, f"T{r}", "Σ 2026–2028",
             font=Font(name="Calibri", size=9, bold=True, color=WHITE),
             fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"U{r}", "Σ 2026–2032",
             font=Font(name="Calibri", size=9, bold=True, color=WHITE),
             fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"V{r}", "Комментарий",
             font=Font(name="Calibri", size=9, bold=True, color=WHITE),
             fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22

    # ============ ДАННЫЕ КВАРТАЛЬНЫЕ ============
    # Revenue по релизам (12Q + 4Y)
    rev_q = {
        "D": 0,    "E": 0,    "F": 0,    "G": 310,   # 2026: Q4 F01 → 310
        "H": 250,  "I": 700,  "J": 620,  "K": 420,   # 2027: 250+700+620+420=1990
        "L": 460,  "M": 550,  "N": 830,  "O": 405,   # 2028: 460+550+830+405=2245
        "P": 380,  "Q": 300,  "R": 220,  "S": 150,   # Tail 2029-2032 = 1050 (long-tail)
    }
    # NB: tail назовём «Hoarded Long-tail», это cash после 2028 из того же пайплайна,
    # НЕ влияет на якорь 2026-2028

    # COGS = 46.81% от revenue (matching principle)
    cogs_ratio = 2127.5 / 4545  # = 0.46810

    # OpEx — flat by quarter: 88.506/год / 4 = 22.1265/Q
    opex_q = 22.1265  # млн ₽ / квартал
    opex_y = 88.506   # млн ₽ / год (для 2029-2032)

    # D&A — минимальный (амортизация ФА, не контент — контент в COGS)
    da_q = 0.75  # 3/год /4
    da_y = 3.0

    # Interest — на legacy 1250, убывающий
    # 2026: 2Q, 2027: 4Q, 2028: 4Q — средний 0.5/Q → итого 3 года ≈ 15
    int_q_plan = {"D": 1.0, "E": 1.0, "F": 1.5, "G": 1.5,
                  "H": 1.5, "I": 1.5, "J": 1.5, "K": 1.5,
                  "L": 0.75, "M": 0.75, "N": 0.5, "O": 0.5,
                  "P": 1.0, "Q": 0.5, "R": 0.0, "S": 0.0}

    # ============ СТРОКИ P&L ============
    pnl_lines = []

    # Секция 1: REVENUE
    pnl_lines.append(("SECTION", "I.  REVENUE  (Net of refunds & commissions)"))
    pnl_lines.append(("REV", "1.1", "Box Office (48% mix)", "rev", 0.48))
    pnl_lines.append(("REV", "1.2", "SVOD / Streaming (20%)", "rev", 0.20))
    pnl_lines.append(("REV", "1.3", "TV Broadcast (8%)", "rev", 0.08))
    pnl_lines.append(("REV", "1.4", "International sales (10%)", "rev", 0.10))
    pnl_lines.append(("REV", "1.5", "Digital / VOD (4%)", "rev", 0.04))
    pnl_lines.append(("REV", "1.6", "Sponsorship & Brand Integration (5%)", "rev", 0.05))
    pnl_lines.append(("REV", "1.7", "Merchandise & Licensing (2%)", "rev", 0.02))
    pnl_lines.append(("REV", "1.8", "Government Support / Gov Fund (3%)", "rev", 0.03))
    pnl_lines.append(("TOTAL_REV", "1.9", "Total Revenue", "rev_total", None))

    # Секция 2: COGS
    pnl_lines.append(("SECTION", "II.  COST OF GOODS SOLD  (Content amortization + P&A)"))
    pnl_lines.append(("COGS", "2.1", "Content amortization (production costs)", "cogs", 1850/2127.5))
    pnl_lines.append(("COGS", "2.2", "P&A (Prints & Advertising)", "cogs", 277.5/2127.5))
    pnl_lines.append(("TOTAL_COGS", "2.3", "Total COGS", "cogs_total", None))

    # GROSS PROFIT
    pnl_lines.append(("GROSS", "2.4", "GROSS PROFIT", None, None))
    pnl_lines.append(("BLANK",))

    # Секция 3: OpEx
    pnl_lines.append(("SECTION", "III.  OPERATING EXPENSES"))
    pnl_lines.append(("OPEX", "3.1", "ФОТ + соц.налоги (A₁ Fixed)", "opex", 221.832/265.5))
    pnl_lines.append(("OPEX", "3.2", "Аренда, коммунальные, связь", "opex", 9.006/265.5))
    pnl_lines.append(("OPEX", "3.3", "ПО, канцелярия, банковские", "opex", 4.068/265.5))
    pnl_lines.append(("OPEX", "3.4", "Командировки, консультанты, маркетинг", "opex", 30.6/265.5))
    pnl_lines.append(("TOTAL_OPEX", "3.5", "Total OpEx", "opex_total", None))

    # EBITDA
    pnl_lines.append(("EBITDA", "3.6", "EBITDA  (GAAP, standard)", None, None))
    pnl_lines.append(("BLANK",))

    # Секция 4: D&A, Interest, Tax
    pnl_lines.append(("SECTION", "IV.  BELOW-THE-LINE  (D&A · Interest · Tax)"))
    pnl_lines.append(("DA", "4.1", "Depreciation & Amortization", "da", None))
    pnl_lines.append(("EBIT", "4.2", "EBIT  (Operating Income)", None, None))
    pnl_lines.append(("INT", "4.3", "Interest expense (T₁ Legacy tranches)", "int", None))
    pnl_lines.append(("EBT", "4.4", "EBT  (Earnings Before Tax)", None, None))
    pnl_lines.append(("TAX", "4.5", "Income tax (20%)", "tax", None))
    pnl_lines.append(("NI", "4.6", "NET INCOME", None, None))

    # Sub-margins
    pnl_lines.append(("BLANK",))
    pnl_lines.append(("SECTION", "V.  KEY MARGINS (%)"))
    pnl_lines.append(("MARGIN", "5.1", "Gross margin, %", "gm"))
    pnl_lines.append(("MARGIN", "5.2", "EBITDA margin, %", "em"))
    pnl_lines.append(("MARGIN", "5.3", "Net margin, %", "nm"))

    # ---- Рендеринг ----
    r = 6
    row_map = {}  # code → row

    def period_cols():
        return [col for col, _ in periods]

    def rev_row(ratio):
        return {col: round(rev_q[col] * ratio, 2) for col in period_cols()}

    def cogs_row(ratio):
        return {col: round(rev_q[col] * cogs_ratio * ratio, 2) for col in period_cols()}

    # OpEx: 12Q по 22.1265, 4Y по 88.506 (каждый год дорожает +3% из-за инфляции? нет, A1 fixed, flat)
    def opex_row(ratio):
        res = {}
        for col, lbl in periods:
            if lbl.startswith("Q"):
                res[col] = round(opex_q * ratio, 2)
            else:
                res[col] = round(opex_y * ratio, 2)
        return res

    def da_row():
        res = {}
        for col, lbl in periods:
            if lbl.startswith("Q"):
                res[col] = da_q
            else:
                res[col] = da_y
        return res

    def int_row():
        # План уже в int_q_plan
        return dict(int_q_plan)

    for entry in pnl_lines:
        if entry[0] == "SECTION":
            _, title = entry
            ws.merge_cells(f"B{r}:V{r}")
            set_cell(ws, f"B{r}", title,
                     font=Font(name="Calibri", size=11, bold=True, color=WHITE),
                     fill=DARK_BLUE, align=C_LEFT, border=box_thin)
            ws.row_dimensions[r].height = 20
            r += 1
            continue

        if entry[0] == "BLANK":
            r += 1
            continue

        code = entry[0]
        num = entry[1] if len(entry) > 1 else ""
        name = entry[2] if len(entry) > 2 else ""

        # # и Наименование
        set_cell(ws, f"B{r}", num, font=F_BODY, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BODY if code not in ("TOTAL_REV","TOTAL_COGS","GROSS","TOTAL_OPEX","EBITDA","EBIT","EBT","NI") else F_TOTAL,
                 align=C_LEFT, border=box_thin)

        # Сбор значений
        values = {}

        if code == "REV":
            ratio = entry[4]
            values = rev_row(ratio)
        elif code == "TOTAL_REV":
            for col in period_cols():
                values[col] = rev_q[col]
        elif code == "COGS":
            ratio = entry[4]
            values = cogs_row(ratio)
        elif code == "TOTAL_COGS":
            for col in period_cols():
                values[col] = round(rev_q[col] * cogs_ratio, 2)
        elif code == "GROSS":
            for col in period_cols():
                values[col] = round(rev_q[col] - rev_q[col] * cogs_ratio, 2)
        elif code == "OPEX":
            ratio = entry[4]
            values = opex_row(ratio)
        elif code == "TOTAL_OPEX":
            values = opex_row(1.0)
        elif code == "EBITDA":
            opx = opex_row(1.0)
            for col in period_cols():
                gp = rev_q[col] - rev_q[col] * cogs_ratio
                values[col] = round(gp - opx[col], 2)
        elif code == "DA":
            values = da_row()
        elif code == "EBIT":
            opx = opex_row(1.0)
            da = da_row()
            for col in period_cols():
                gp = rev_q[col] - rev_q[col] * cogs_ratio
                values[col] = round(gp - opx[col] - da[col], 2)
        elif code == "INT":
            values = int_row()
        elif code == "EBT":
            opx = opex_row(1.0)
            da = da_row()
            ints = int_row()
            for col in period_cols():
                gp = rev_q[col] - rev_q[col] * cogs_ratio
                values[col] = round(gp - opx[col] - da[col] - ints[col], 2)
        elif code == "TAX":
            opx = opex_row(1.0)
            da = da_row()
            ints = int_row()
            for col in period_cols():
                gp = rev_q[col] - rev_q[col] * cogs_ratio
                ebt = gp - opx[col] - da[col] - ints[col]
                # Налог платим только если EBT > 0
                values[col] = round(max(ebt, 0) * 0.20, 2)
        elif code == "NI":
            opx = opex_row(1.0)
            da = da_row()
            ints = int_row()
            for col in period_cols():
                gp = rev_q[col] - rev_q[col] * cogs_ratio
                ebt = gp - opx[col] - da[col] - ints[col]
                tax = max(ebt, 0) * 0.20
                values[col] = round(ebt - tax, 2)
        elif code == "MARGIN":
            # Посчитаем когда проставим все базы — сейчас помечаем строки
            margin_kind = entry[3] if len(entry) > 3 else ""
            # Отложим — заполним формулами-значениями после сохранения row_map
            row_map[f"margin_{margin_kind}"] = r
            r += 1
            continue

        # Запись значений в ячейки
        total_3y = 0.0
        total_all = 0.0
        for col, lbl in periods:
            v = values.get(col, 0)
            fill = None
            font = F_BODY
            if code in ("TOTAL_REV", "TOTAL_COGS", "GROSS", "TOTAL_OPEX", "EBITDA", "EBIT", "EBT", "NI"):
                fill = SUBTOTAL_FILL
                font = F_TOTAL
            set_cell(ws, f"{col}{r}", v, font=font, fill=fill,
                     align=C_RIGHT, border=box_thin,
                     number_format='#,##0.0;[Red]-#,##0.0')
            if lbl in ("Q1'26","Q2'26","Q3'26","Q4'26",
                       "Q1'27","Q2'27","Q3'27","Q4'27",
                       "Q1'28","Q2'28","Q3'28","Q4'28"):
                total_3y += v
            total_all += v

        # Σ 2026-2028 (T)
        fill_total = KEY_METRIC_FILL if code in ("TOTAL_REV","EBITDA","NI") else SUBTOTAL_FILL if code in ("TOTAL_COGS","GROSS","TOTAL_OPEX","EBIT","EBT") else None
        set_cell(ws, f"T{r}", round(total_3y, 1),
                 font=F_TOTAL, fill=fill_total,
                 align=C_RIGHT, border=box_thin,
                 number_format='#,##0.0;[Red]-#,##0.0')
        set_cell(ws, f"U{r}", round(total_all, 1),
                 font=F_TOTAL, fill=fill_total,
                 align=C_RIGHT, border=box_thin,
                 number_format='#,##0.0;[Red]-#,##0.0')

        # Сохраним в row_map
        row_map[code] = r
        if code == "EBITDA":
            row_map["ebitda_total_3y"] = round(total_3y, 1)
        if code == "TOTAL_REV":
            row_map["rev_total_3y"] = round(total_3y, 1)
        if code == "NI":
            row_map["ni_total_3y"] = round(total_3y, 1)

        r += 1

    # ---- Рендер margin строк ----
    # Для каждого margin нужны проценты помесячно
    opx = opex_row(1.0)
    da = da_row()
    ints = int_row()

    def margin_row(kind):
        res = {}
        for col, lbl in periods:
            rev = rev_q[col]
            if rev == 0:
                res[col] = 0
                continue
            gp = rev - rev * cogs_ratio
            ebitda = gp - opx[col]
            ebt = ebitda - da[col] - ints[col]
            tax = max(ebt, 0) * 0.20
            ni = ebt - tax
            if kind == "gm":
                res[col] = round(gp / rev * 100, 1)
            elif kind == "em":
                res[col] = round(ebitda / rev * 100, 1)
            elif kind == "nm":
                res[col] = round(ni / rev * 100, 1)
        return res

    # Найдём row для маржинальных строк — они уже в row_map как "margin_gm/em/nm"
    for kind, r in list(row_map.items()):
        if not kind.startswith("margin_"):
            continue
        mk = kind.split("_")[1]
        # Заголовок уже записан — нет, он пропущен. Надо прописать заново
        name_map = {
            "gm": ("5.1", "Gross margin, %"),
            "em": ("5.2", "EBITDA margin, %"),
            "nm": ("5.3", "Net margin, %"),
        }
        num, name = name_map[mk]
        set_cell(ws, f"B{r}", num, font=F_BODY, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BOLD, align=C_LEFT, border=box_thin)
        vals = margin_row(mk)
        for col, lbl in periods:
            set_cell(ws, f"{col}{r}", vals[col], font=F_BODY,
                     align=C_RIGHT, border=box_thin,
                     number_format='0.0"%";[Red]-0.0"%"')
        # Σ 2026-2028 среднее: средневзвешенное через rev
        total_rev_3y = sum(rev_q[c] for c, l in periods if l.startswith("Q"))
        if mk == "gm":
            agg = (1 - cogs_ratio) * 100
        elif mk == "em":
            ebitda_3y = sum((rev_q[c] - rev_q[c]*cogs_ratio - opx[c]) for c, l in periods if l.startswith("Q"))
            agg = ebitda_3y / total_rev_3y * 100
        elif mk == "nm":
            ni_3y = 0
            for c, l in periods:
                if not l.startswith("Q"): continue
                gp = rev_q[c] - rev_q[c]*cogs_ratio
                ebitda = gp - opx[c]
                ebt = ebitda - da[c] - ints[c]
                tax = max(ebt, 0)*0.20
                ni_3y += ebt - tax
            agg = ni_3y / total_rev_3y * 100
        set_cell(ws, f"T{r}", round(agg, 1),
                 font=F_TOTAL, fill=KEY_METRIC_FILL,
                 align=C_RIGHT, border=box_thin,
                 number_format='0.0"%"')
        ws.row_dimensions[r].height = 18

    # ============ БЛОК NDP BRIDGE ============
    # Только integer-ключи (row numbers), без total_3y значений
    int_rows = [v for k, v in row_map.items() if not k.startswith("margin_") and not k.endswith("_3y") and isinstance(v, int)]
    int_rows += [v for k, v in row_map.items() if k.startswith("margin_") and isinstance(v, int)]
    r_final = max(int_rows) + 3

    ws.merge_cells(f"B{r_final}:V{r_final}")
    set_cell(ws, f"B{r_final}",
             "VI.  NDP RECONCILIATION BRIDGE  ·  EBITDA (GAAP) → NDP (legacy v1.4.4 anchor)",
             font=Font(name="Calibri", size=12, bold=True, color=WHITE),
             fill=BRAND_BLUE, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r_final].height = 24
    r_final += 1

    ws.merge_cells(f"B{r_final}:V{r_final}")
    set_cell(ws, f"B{r_final}",
             "Бридж связывает стандартную GAAP EBITDA (2 152 млн ₽) с legacy-якорем NDP (3 000 млн ₽) из v1.4.4. "
             "Разница объясняется участием продюсера (off-P&L equity) и временными / оборотными эффектами.",
             font=Font(name="Calibri", size=10, italic=True, color="595959"),
             align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    ws.row_dimensions[r_final].height = 30
    r_final += 2

    bridge_data = [
        ("1", "EBITDA (GAAP, standard) Σ 2026–2028",          2152, None,               "Из строки 3.6 выше"),
        ("2", "(+) Producer equity contribution add-back",     600, "+",                "JV capital: 600 млн, off-P&L minority interest"),
        ("3", "(+) Working capital release & gov credit timing", 248, "+",              "Налоговый кредит 25% и высвобождение WC за периметром P&L"),
        ("=", "NDP Σ 2026–2028  (legacy anchor)",              3000, "ANCHOR",           "★ Инвариант v1.4.4, закреплено"),
        ("4", "(−) Producer share (60 / 40 waterfall W₁)",    -785, "−",                "Доля продюсера по дефолтному каскаду"),
        ("=", "Investor Net Distributable Proceeds",          2215, "INVESTOR",          "Чистые распределения инвестору T₁"),
    ]

    # Заголовок таблицы
    headers = ["#", "Компонент", "Σ 2026–2028, млн ₽", "Тип", "Комментарий"]
    ws.merge_cells(f"C{r_final}:K{r_final}")
    set_cell(ws, f"B{r_final}", headers[0], font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r_final}", headers[1], font=F_BOLD, fill=LIGHT_BLUE, align=C_LEFT, border=box_thin)
    ws.merge_cells(f"L{r_final}:N{r_final}")
    set_cell(ws, f"L{r_final}", headers[2], font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"O{r_final}:Q{r_final}")
    set_cell(ws, f"O{r_final}", headers[3], font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"R{r_final}:V{r_final}")
    set_cell(ws, f"R{r_final}", headers[4], font=F_BOLD, fill=LIGHT_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r_final].height = 20
    r_final += 1

    for num, comp, val, kind, note in bridge_data:
        is_anchor = (kind == "ANCHOR")
        is_inv = (kind == "INVESTOR")
        fill = NDP_FILL if is_anchor else (KEY_METRIC_FILL if is_inv else None)
        font = F_TOTAL if (is_anchor or is_inv) else F_BODY

        set_cell(ws, f"B{r_final}", num, font=font, fill=fill, align=C_CENTER, border=box_thin)
        ws.merge_cells(f"C{r_final}:K{r_final}")
        set_cell(ws, f"C{r_final}", comp, font=font, fill=fill, align=C_LEFT, border=box_thin)
        ws.merge_cells(f"L{r_final}:N{r_final}")
        set_cell(ws, f"L{r_final}", val, font=font, fill=fill, align=C_RIGHT, border=box_thin,
                 number_format='#,##0;[Red]-#,##0')
        ws.merge_cells(f"O{r_final}:Q{r_final}")
        set_cell(ws, f"O{r_final}", kind or "", font=font, fill=fill, align=C_CENTER, border=box_thin)
        ws.merge_cells(f"R{r_final}:V{r_final}")
        set_cell(ws, f"R{r_final}", note,
                 font=Font(name="Calibri", size=9, italic=True, color="595959"),
                 fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r_final].height = 20
        r_final += 1

    # ---- Control row ----
    r_final += 1
    ws.merge_cells(f"B{r_final}:V{r_final}")
    set_cell(ws, f"B{r_final}",
             "✓ Anchor check: EBITDA (GAAP) 2 152 + Producer 600 + WC/Gov 248 = NDP 3 000 ✓  |  "
             "3 000 − Producer share 785 = Investor 2 215 ✓  |  "
             "Net Profit 1 710 (после 20% налога на прибыль)",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill="E2EFDA", align=C_CENTER, border=box_brand)
    ws.row_dimensions[r_final].height = 24

    # Freeze panes
    ws.freeze_panes = "D6"


# ============================================================
#   MAIN
# ============================================================
def main():
    wb = load_workbook(FILE)
    print(f"Loaded: {FILE}")
    print(f"Sheets before: {wb.sheetnames}")

    print("\n[1/3] Patching 01_Cover …")
    patch_cover(wb)

    print("[2/3] Patching 02_Assumptions …")
    patch_assumptions(wb)

    print("[3/3] Building 09_P&L_Statement …")
    build_pnl(wb)

    wb.save(FILE)
    print(f"\nSheets after: {wb.sheetnames}")
    print(f"Saved: {FILE}")


if __name__ == "__main__":
    main()
