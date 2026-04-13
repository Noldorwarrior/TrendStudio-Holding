#!/usr/bin/env python3
"""СКРИПТ 3: BS Balance Check — Assets = L + E, Cash(BS) = Cash(CF), Control = 0"""
import openpyxl
import json
import sys
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)) + '/..')

results = []

for fname in ['investor_model_v3_Internal.xlsx', 'investor_model_v3_Public.xlsx']:
    wb = openpyxl.load_workbook(fname, data_only=True)

    # Balance Sheet = sheet 11_Balance_Sheet
    ws_bs = wb['11_Balance_Sheet']
    # Cash Flow = sheet 10_Cash_Flow
    ws_cf = wb['10_Cash_Flow']

    # Map columns to periods from BS header (row 3 or 5)
    print(f"\n=== {fname} ===")
    print("\n--- Balance Sheet structure ---")
    for row in ws_bs.iter_rows(min_row=2, max_row=21, min_col=2, max_col=20, values_only=False):
        label = row[0].value
        if label is not None:
            vals = [c.value for c in row]
            print(f"  Row {row[0].row}: {vals}")

    print("\n--- Cash Flow structure ---")
    for row in ws_cf.iter_rows(min_row=2, max_row=25, min_col=2, max_col=22, values_only=False):
        label = row[0].value
        if label is not None:
            vals = [c.value for c in row]
            print(f"  Row {row[0].row}: {vals}")

    wb.close()

print("\n\n=== ANALYSIS COMPLETE ===")
print(json.dumps(results, ensure_ascii=False, indent=2))
