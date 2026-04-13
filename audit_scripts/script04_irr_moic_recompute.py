#!/usr/bin/env python3
"""СКРИПТ 4: IRR & MOIC independent recomputation from Waterfall data"""
import openpyxl
import json
import numpy as np
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)) + '/..')

results = []

for fname in ['investor_model_v3_Internal.xlsx', 'investor_model_v3_Public.xlsx']:
    wb = openpyxl.load_workbook(fname, data_only=True)

    # Waterfall = 19_Waterfall
    ws_wf = wb['19_Waterfall']
    print(f"\n=== {fname} — 19_Waterfall ===")
    for row in ws_wf.iter_rows(min_row=2, max_row=47, min_col=2, max_col=10, values_only=False):
        label = row[0].value
        if label is not None:
            vals = [c.value for c in row]
            print(f"  Row {row[0].row}: {vals}")

    # Investor Returns = 24_Investor_Returns
    ws_ir = wb['24_Investor_Returns']
    print(f"\n=== {fname} — 24_Investor_Returns ===")
    for row in ws_ir.iter_rows(min_row=2, max_row=51, min_col=2, max_col=15, values_only=False):
        label = row[0].value
        if label is not None:
            vals = [c.value for c in row]
            print(f"  Row {row[0].row}: {vals}")

    wb.close()

print("\n=== ANALYSIS COMPLETE ===")
