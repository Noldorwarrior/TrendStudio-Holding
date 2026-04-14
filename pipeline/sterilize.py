#!/usr/bin/env python3
"""
pipeline/sterilize.py — Zip-level + openpyxl surgery for Public xlsx sterilization.

Removes:
  - Internal leakage from cell content (F-001, F-002, F-009, F-010, F-011)
  - Filesystem paths / absPath from workbook.xml (F-003)
  - Sensitive metadata from docProps/core.xml (F-020, F-021, F-022)
  - Dangling '=' sign in 17_Deal_Structures (F-036)
  - computer:// reference in changelog (bonus)

Preserves:
  - "Internal Rate of Return" (IRR financial term) in 37_Glossary
  - All formulas, formatting, and sheet structure

Usage:
    python pipeline/sterilize.py <input.xlsx> <output.xlsx>
    python pipeline/sterilize.py <input.xlsx>          # overwrites in-place

R-items addressed: R-001, R-002, R-003, R-004
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl


# ─── Cell Content Sterilization ───────────────────────────────────────────


def scrub_internal_cells(wb: openpyxl.Workbook) -> list[str]:
    """Scrub Internal references from specific cells.

    Returns list of human-readable change descriptions.
    """
    changes: list[str] = []

    # F-009: 03_Change_Log C30
    # "Public export + Internal (Внутренний) с 4 сервисными"
    # → "Public export с 4 сервисными"
    if "03_Change_Log" in wb.sheetnames:
        ws = wb["03_Change_Log"]
        cell = ws["C30"]
        if cell.value and "Internal" in str(cell.value):
            old = str(cell.value)
            cell.value = old.replace(
                "Public export + Internal (Внутренний) с 4 сервисными",
                "Public export с 4 сервисными",
            )
            changes.append(f"F-009: 03_Change_Log!C30: scrubbed Internal reference")

        # Bonus: C32 "Финализация + computer:// ссылки"
        cell32 = ws["C32"]
        if cell32.value and "computer://" in str(cell32.value):
            cell32.value = str(cell32.value).replace(
                "computer:// ссылки", "навигационные ссылки"
            )
            changes.append("Bonus: 03_Change_Log!C32: removed computer:// reference")

    # F-010: 23_Valuation_Multiples B46
    # "Scenario (Сценарий) Base (Базовый) (internal)"
    # → "Scenario (Сценарий) Base (Базовый)"
    if "23_Valuation_Multiples" in wb.sheetnames:
        ws = wb["23_Valuation_Multiples"]
        cell = ws["B46"]
        if cell.value and "internal" in str(cell.value).lower():
            old = str(cell.value)
            cell.value = old.replace(" (internal)", "")
            changes.append(f"F-010: 23_Valuation_Multiples!B46: removed (internal)")

    # F-001: 24_Investor_Returns B49
    # Remove comparative clause referencing Internal W₅ V-D
    if "24_Investor_Returns" in wb.sheetnames:
        ws = wb["24_Investor_Returns"]
        cell = ws["B49"]
        if cell.value and "Internal" in str(cell.value):
            old = str(cell.value)
            # Replace: "Margin (Маржа) к hurdle ... upside capture)."
            # with a neutral statement
            cell.value = re.sub(
                r"Margin \(Маржа\) к hurdle \(порог\) под W₃.*?upside capture\)\.",
                "Margin (Маржа) к hurdle (порог) под W₃ составляет +2.09pp.",
                old,
            )
            changes.append(
                "F-001: 24_Investor_Returns!B49: scrubbed Internal W₅ V-D reference"
            )

    # F-011: 32_Comparable_Transactions E12, E13, J12
    if "32_Comparable_Transactions" in wb.sheetnames:
        ws = wb["32_Comparable_Transactions"]
        for coord, replacement in [
            ("E12", None),  # "Газпром-медиа (internal)" → "Газпром-медиа"
            ("E13", None),  # "Яндекс (internal)" → "Яндекс"
            (
                "J12",
                "Consolidation (Консолидация)",
            ),  # full replacement
        ]:
            cell = ws[coord]
            if cell.value and "internal" in str(cell.value).lower():
                old = str(cell.value)
                if replacement:
                    cell.value = replacement
                else:
                    cell.value = re.sub(
                        r"\s*\(internal\)", "", old, flags=re.IGNORECASE
                    ).strip()
                changes.append(
                    f"F-011: 32_Comparable_Transactions!{coord}: "
                    f'"{old}" → "{cell.value}"'
                )

    # F-036: 17_Deal_Structures B31 — dangling "="
    if "17_Deal_Structures" in wb.sheetnames:
        ws = wb["17_Deal_Structures"]
        cell = ws["B31"]
        if cell.value and str(cell.value).strip() == "=":
            cell.value = "—"
            changes.append('F-036: 17_Deal_Structures!B31: dangling "=" → "—"')

    return changes


# ─── Metadata Sterilization (zip-level) ──────────────────────────────────


def clean_core_xml(xml_content: str) -> str:
    """Clean docProps/core.xml: set Author, clear description/keywords/lastModifiedBy.

    Addresses: F-020 (lastModifiedBy='a'), F-021 (description leaks anchor),
    F-022 (keywords contain 'L3').
    """
    # dc:creator → TrendStudio
    xml_content = re.sub(
        r"(<dc:creator[^>]*>).*?(</dc:creator>)",
        r"\g<1>TrendStudio\g<2>",
        xml_content,
    )
    # dc:description → empty (was leaking anchor value)
    xml_content = re.sub(
        r"(<dc:description[^>]*>).*?(</dc:description>)",
        r"\g<1>\g<2>",
        xml_content,
    )
    # cp:keywords → remove L3
    xml_content = re.sub(
        r"(<cp:keywords>).*?(</cp:keywords>)",
        r"\g<1>investor, financial model, pre-IPO, cinema\g<2>",
        xml_content,
    )
    # cp:lastModifiedBy → TrendStudio
    xml_content = re.sub(
        r"(<cp:lastModifiedBy>).*?(</cp:lastModifiedBy>)",
        r"\g<1>TrendStudio\g<2>",
        xml_content,
    )
    return xml_content


def strip_abspath(xml_content: str) -> str:
    """Remove x15ac:absPath element from xl/workbook.xml.

    Addresses: F-003.
    """
    xml_content = re.sub(r"<x15ac:absPath[^/]*/>\s*", "", xml_content)
    xml_content = re.sub(r'\s*xmlns:x15ac="[^"]*"', "", xml_content)
    return xml_content


# ─── Internal xlsx title fix ─────────────────────────────────────────────


def fix_internal_title(wb: openpyxl.Workbook) -> list[str]:
    """Fix title of Internal xlsx if it incorrectly says 'Public' (R-005)."""
    changes: list[str] = []
    if hasattr(wb.properties, "title") and wb.properties.title:
        title = wb.properties.title
        if "Public" in title and "Internal" not in title:
            wb.properties.title = title.replace("Public", "Internal")
            changes.append(f'R-005: title "{title}" → "{wb.properties.title}"')
    return changes


# ─── Main Pipeline ───────────────────────────────────────────────────────


def sterilize_xlsx(
    input_path: Path,
    output_path: Path,
    *,
    is_internal: bool = False,
) -> list[str]:
    """Full sterilization pipeline for an xlsx file.

    Args:
        input_path: Source xlsx file.
        output_path: Destination xlsx file (can be same as input).
        is_internal: If True, only fix title (R-005), skip Public-specific scrubs.

    Returns:
        List of human-readable change descriptions.
    """
    changes: list[str] = []

    # Step 1: Cell content changes via openpyxl
    wb = openpyxl.load_workbook(str(input_path))

    if is_internal:
        changes.extend(fix_internal_title(wb))
    else:
        changes.extend(scrub_internal_cells(wb))

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()
    wb.save(str(tmp_path))
    wb.close()

    # Step 2: Metadata surgery on the openpyxl-saved file
    final_tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    final_tmp_path = Path(final_tmp.name)
    final_tmp.close()

    with zipfile.ZipFile(str(tmp_path), "r") as zin:
        with zipfile.ZipFile(str(final_tmp_path), "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)

                if item == "docProps/core.xml":
                    xml = data.decode("utf-8")
                    xml = clean_core_xml(xml)
                    data = xml.encode("utf-8")
                    changes.append("F-020/021/022: cleaned core.xml metadata")

                elif item == "xl/workbook.xml":
                    xml = data.decode("utf-8")
                    if "absPath" in xml:
                        xml = strip_abspath(xml)
                        data = xml.encode("utf-8")
                        changes.append("F-003: stripped absPath from workbook.xml")

                zout.writestr(item, data)

    # Move final to output
    tmp_path.unlink()
    shutil.move(str(final_tmp_path), str(output_path))

    return changes


def verify_sterilization(xlsx_path: Path) -> list[str]:
    """Post-sterilization verification. Returns list of issues found."""
    issues: list[str] = []

    with zipfile.ZipFile(str(xlsx_path), "r") as z:
        for name in z.namelist():
            data = z.read(name).decode("utf-8", errors="replace")

            # Check for "Internal" (case-insensitive)
            for m in re.finditer(r"[Ii]nternal", data):
                # Allow "Internal Rate of Return"
                start = max(0, m.start() - 5)
                end = min(len(data), m.end() + 30)
                context = data[start:end]
                if "Rate of Return" in context or "rate of return" in context.lower():
                    continue
                # Allow "Internal (Внутренний) Rate of Return" with entity-encoded text
                end_far = min(len(data), m.end() + 200)
                far_context = data[m.start() : end_far]
                if "Rate of Return" in far_context:
                    continue
                issues.append(f"LEAK: '{name}' contains 'Internal' at pos {m.start()}")

            # Check for absPath
            if "absPath" in data and name == "xl/workbook.xml":
                issues.append(f"LEAK: absPath found in {name}")

            # Check core.xml for sensitive data
            if name == "docProps/core.xml":
                if "L3" in data:
                    issues.append("META: keywords still contain 'L3'")
                if "3 000" in data or "3000" in data:
                    issues.append("META: description still leaks anchor value")
                if "<cp:lastModifiedBy>a</cp:lastModifiedBy>" in data:
                    issues.append("META: lastModifiedBy still 'a'")

    return issues


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sterilize Public xlsx for investor distribution"
    )
    parser.add_argument("input", type=Path, help="Input xlsx file")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=None,
        help="Output xlsx file (default: overwrite input)",
    )
    parser.add_argument(
        "--internal",
        action="store_true",
        help="Treat as Internal xlsx (only fix title, R-005)",
    )
    parser.add_argument(
        "--verify-only",
        action="store_true",
        help="Only verify, do not modify",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: {args.input} not found", file=sys.stderr)
        return 1

    if args.verify_only:
        issues = verify_sterilization(args.input)
        if issues:
            print(f"FAIL: {len(issues)} issue(s) found:")
            for iss in issues:
                print(f"  - {iss}")
            return 1
        print("PASS: no leakage detected")
        return 0

    output = args.output or args.input
    print(f"Sterilizing: {args.input} → {output}")
    changes = sterilize_xlsx(args.input, output, is_internal=args.internal)

    print(f"\n{len(changes)} change(s) applied:")
    for ch in changes:
        print(f"  ✓ {ch}")

    # Verify
    issues = verify_sterilization(output)
    if issues:
        print(f"\nWARNING: {len(issues)} issue(s) remain after sterilization:")
        for iss in issues:
            print(f"  ✗ {iss}")
        return 1

    print(f"\nPASS: sterilization verified — 0 leakage issues")
    return 0


if __name__ == "__main__":
    sys.exit(main())
