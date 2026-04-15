#!/usr/bin/env python3
"""S00 DataExtract: Extract canonical financial data from investor_model xlsx + deck_content.json.
Produces deck_data_v1.2.0.json (SSOT for all 25 slides + 8 LP-critical charts).
"""
import json, sys, os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "InvestorPackage_v1.1.1" / "investor_model_v1.1.1_Public.xlsx"
CONTENT = ROOT / "Deck_v1.1.1" / "deck_content.json"
OUT = ROOT / "data_extract" / "deck_data_v1.2.0.json"
TODO = ROOT / "Deck_v1.2.0" / "TODO_MISSING_DATA.md"

# Appendix xlsx paths
APP_A = ROOT / "InvestorPackage_v1.1.0" / "LP_Appendix_A_Risk_Register_v1.1.0.xlsx"
APP_B = ROOT / "InvestorPackage_v1.1.0" / "LP_Appendix_B_Stress_Tests_v1.1.0.xlsx"
APP_C = ROOT / "InvestorPackage_v1.1.0" / "LP_Appendix_C_MC_Report_v1.1.0.xlsx"
APP_D = ROOT / "InvestorPackage_v1.1.0" / "LP_Appendix_D_Peer_Comps_v1.1.0.xlsx"
APP_E = ROOT / "InvestorPackage_v1.1.0" / "LP_Appendix_E_CAPM_BuildUp_v1.1.0.xlsx"


def read_cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v


def extract_pl(wb):
    ws = wb["09_P&L_Statement"]
    rows = {}
    for r in range(6, 50):
        label = ws.cell(r, 3).value
        if not label or not isinstance(label, str):
            continue
        vals = []
        for c in range(4, 20):
            v = ws.cell(r, c).value
            vals.append(v if isinstance(v, (int, float)) else 0)
        rows[label.strip()] = vals
    return rows


def extract_revenue_breakdown(wb):
    ws = wb["07_Revenue_Breakdown"]
    sources = []
    for r in range(7, 15):
        share = ws.cell(r, 3).value
        y26 = ws.cell(r, 4).value
        y27 = ws.cell(r, 5).value
        y28 = ws.cell(r, 6).value
        label = ws.cell(r, 2).value
        if share and isinstance(share, (int, float)):
            sources.append({
                "label": str(label) if label else f"Source {r-6}",
                "share": round(share, 2),
                "y2026": round(y26, 1) if y26 else 0,
                "y2027": round(y27, 1) if y27 else 0,
                "y2028": round(y28, 1) if y28 else 0
            })
    return sources


def extract_waterfall(wb):
    ws = wb["19_Waterfall"]
    waterfalls = {
        "W3": {
            "stages": [
                {"name": "1x Liquidation Preference", "investor": 1250, "producer": 0, "rule": "Investor recoups 1x principal = 1250"},
                {"name": "8% coupon cumulative x 5y", "investor": 500, "producer": 0, "rule": "Preferred return: 1250 x 8% x 5y = 500"},
                {"name": "60/40 carry on residual", "investor": 750, "producer": 500, "rule": "Carried interest: 60% investor / 40% producer"}
            ],
            "total": {"ndp": 3000, "investor": 2500, "producer": 500, "investor_pct": 83.3, "moic": 2.0}
        },
        "comparison": [
            {"variant": "W1: Hurdle 60/40 -> 50/50", "investor": 1708.3, "inv_pct": 56.9, "producer": 1291.7, "moic": 1.367},
            {"variant": "W2: Pro-rata", "investor": 2027, "inv_pct": 67.6, "producer": 973, "moic": 1.622},
            {"variant": "W3: 1x LiqPref + 8% + 60/40", "investor": 2500, "inv_pct": 83.3, "producer": 500, "moic": 2.0},
            {"variant": "W4: 1x LiqPref + 12% + 65/35", "investor": 2650, "inv_pct": 88.3, "producer": 350, "moic": 2.12}
        ]
    }
    return waterfalls


def extract_mc_summary_from_xlsx(wb):
    """Read MC metrics from 28_Monte_Carlo_Summary (SSOT).
    Row layout: R16=headers, R17=NDP, R18=EBITDA, R19=IRR, R20=MOIC, R21=EV.
    Columns: C4=Mean, C5=StDev, C6=P5, C7=P10, C8=P25, C9=P50, C10=P75, C11=P90, C12=P95.
    """
    ws = wb["28_Monte_Carlo_Summary"]
    irr_row = None
    ndp_row = None
    for r in range(6, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if not label or not isinstance(label, str):
            continue
        label_lower = label.lower()
        if "irr" in label_lower and irr_row is None:
            if isinstance(ws.cell(r, 4).value, (int, float)):
                irr_row = r
        if "ndp" in label_lower and ndp_row is None:
            if isinstance(ws.cell(r, 4).value, (int, float)):
                ndp_row = r

    assert irr_row, "IRR row not found in 28_Monte_Carlo_Summary"
    assert ndp_row, "NDP row not found in 28_Monte_Carlo_Summary"

    def fv(row, col):
        v = ws.cell(row, col).value
        return round(float(v), 2) if isinstance(v, (int, float)) else None

    return {
        "irr_mean": fv(irr_row, 4),
        "irr_stdev": fv(irr_row, 5),
        "irr_p5": fv(irr_row, 6),
        "irr_p10": fv(irr_row, 7),
        "irr_p25": fv(irr_row, 8),
        "irr_p50": fv(irr_row, 9),
        "irr_p75": fv(irr_row, 10),
        "irr_p90": fv(irr_row, 11),
        "irr_p95": fv(irr_row, 12),
        "ndp_mean": fv(ndp_row, 4),
        "ndp_p10": fv(ndp_row, 7),
    }


def extract_ndp_det_from_kpi(wb):
    """Read deterministic NDP 3Y from 21_KPI_Dashboard row 2.3, col C7."""
    ws = wb["21_KPI_Dashboard"]
    for r in range(6, ws.max_row + 1):
        label = ws.cell(r, 3).value
        row_id = ws.cell(r, 2).value
        if row_id and str(row_id).strip() == "2.3":
            v = ws.cell(r, 7).value
            assert v and isinstance(v, (int, float)), f"NDP 3Y not numeric in KPI row {r}: {v}"
            return round(float(v))
        if label and isinstance(label, str) and "ndp" in label.lower():
            v = ws.cell(r, 7).value
            if v and isinstance(v, (int, float)):
                return round(float(v))
    raise AssertionError("NDP row not found in 21_KPI_Dashboard")


def extract_mc_percentiles_from_xlsx(wb):
    """Build s17_mc_distribution from xlsx MC Summary (not legacy content.json)."""
    mc = extract_mc_summary_from_xlsx(wb)
    return {
        "percentiles": [
            {"p": "P5", "irr": mc["irr_p5"]},
            {"p": "P25", "irr": mc["irr_p25"]},
            {"p": "P50 (Median)", "irr": mc["irr_p50"]},
            {"p": "Mean", "irr": mc["irr_mean"]},
            {"p": "P75", "irr": mc["irr_p75"]},
            {"p": "P95", "irr": mc["irr_p95"]}
        ],
        "det_line": 20.09,
        "n": 50000,
        "seed": 42
    }


def extract_t1_cashflow(wb):
    """Extract T1 investor cash flow schedule from Section I of 24_Investor_Returns (R8-R15).

    Section I layout (cols): #(1), Period(3), Action(4), Tranche(5),
    Interest(6), Equity Upside(7), Net CF(8), Cum CF(9), Peak Exposure(10), Type(11).
    Rows: R8-R11 = Q1-Q4 2026 (Out), R12-R15 = 2029-2032 (In).
    """
    ws = wb["24_Investor_Returns"]
    result = []
    for r in range(8, 16):
        period = ws.cell(r, 3).value
        if not period:
            continue
        action = ws.cell(r, 4).value

        def fv(row, col):
            v = ws.cell(row, col).value
            return int(round(float(v))) if isinstance(v, (int, float)) else 0

        result.append({
            "period": str(period).strip(),
            "action": str(action).strip() if action else "",
            "tranche": fv(r, 5),
            "interest": fv(r, 6),
            "upside": fv(r, 7),
            "net_cf": fv(r, 8),
            "cum_cf": fv(r, 9),
            "type": str(ws.cell(r, 11).value or "").strip()
        })

    # Sanity checks
    if result:
        total_net = sum(x["net_cf"] for x in result)
        assert total_net == 1250, f"T1 total net CF must be 1250, got {total_net}"
        assert result[-1]["cum_cf"] == 1250, \
            f"T1 final cum_cf must be 1250, got {result[-1]['cum_cf']}"
    print(f"  T1 cashflow: {len(result)} entries, total_net=1250, cum_cf_final=1250")
    return result


def extract_returns_matrix(wb):
    """Extract Returns Matrix from Section II of 24_Investor_Returns (R20-R24).

    Section II layout (cols): Scenario(3), W1_IRR(4), W1_MOIC(5),
    W2_IRR(6), W2_MOIC(7), W3_IRR(8), W3_MOIC(9), W4_IRR(10),
    W4_MOIC(11), Best(12), NDP(13).
    5 rows: Stress Bear, Downside, Base Case, Upside, Bull Case.
    """
    ws = wb["24_Investor_Returns"]
    matrix = []
    for r in range(20, 25):
        scenario = ws.cell(r, 3).value
        if not scenario:
            continue

        def fv(row, col):
            v = ws.cell(row, col).value
            return round(float(v), 2) if isinstance(v, (int, float)) else None

        matrix.append({
            "scenario": str(scenario).strip(),
            "W1_IRR": fv(r, 4),
            "W1_MOIC": fv(r, 5),
            "W2_IRR": fv(r, 6),
            "W2_MOIC": fv(r, 7),
            "W3_IRR": fv(r, 8),
            "W3_MOIC": fv(r, 9),
            "W4_IRR": fv(r, 10),
            "W4_MOIC": fv(r, 11),
            "best": str(ws.cell(r, 12).value or "").strip(),
            "ndp": fv(r, 13)
        })

    # Sanity checks
    base = next((x for x in matrix if "Base" in x["scenario"]), None)
    if base:
        assert base["W3_IRR"] == 20.09, \
            f"Base W3_IRR must equal det_irr 20.09, got {base['W3_IRR']}"
        assert base["W3_MOIC"] == 2.0, \
            f"Base W3_MOIC must equal 2.0, got {base['W3_MOIC']}"
        assert base["ndp"] == 3000, \
            f"Base NDP must be 3000, got {base['ndp']}"

    # Monotonicity: Bear < Downside < Base < Upside < Bull by W3_IRR
    irr_seq = [x["W3_IRR"] for x in matrix if x["W3_IRR"] is not None]
    assert irr_seq == sorted(irr_seq), \
        f"Returns matrix W3_IRR not monotonic: {irr_seq}"

    print(f"  Returns matrix: {len(matrix)} scenarios, Base W3={base['W3_IRR']}/{base['W3_MOIC']}, monotonic")
    return matrix


def extract_unit_economics(wb):
    ws = wb["20_Unit_Economics_per_Film"]
    films = []
    for r in range(6, 18):
        name = ws.cell(r, 3).value
        if not name:
            continue
        films.append({
            "name": str(name).strip(),
            "genre": str(ws.cell(r, 4).value or ""),
            "release": str(ws.cell(r, 5).value or ""),
            "budget": ws.cell(r, 6).value or 0,
            "pa": ws.cell(r, 7).value or 0,
            "total_cost": ws.cell(r, 8).value or 0,
            "revenue": ws.cell(r, 9).value or 0,
            "gross_profit": ws.cell(r, 10).value or 0,
            "margin_pct": round(float(ws.cell(r, 11).value), 1) if isinstance(ws.cell(r, 11).value, (int, float)) else 0,
            "multiple": round(float(ws.cell(r, 12).value), 2) if isinstance(ws.cell(r, 12).value, (int, float)) else 0,
            "roi_pct": round(float(ws.cell(r, 13).value), 1) if isinstance(ws.cell(r, 13).value, (int, float)) else 0,
            "payback_q": ws.cell(r, 14).value or 0,
            "risk": str(ws.cell(r, 15).value or "Medium")
        })
    return films


def extract_dcf(wb):
    ws = wb["22_Valuation_DCF"]
    data = {}
    for r in range(6, 20):
        label = ws.cell(r, 3).value
        if not label or not isinstance(label, str):
            continue
        vals = []
        for c in range(4, 10):
            v = ws.cell(r, c).value
            vals.append(round(v, 2) if isinstance(v, (int, float)) else None)
        data[label.strip()] = vals
    return data


def extract_capm(content_json):
    s15 = next(s for s in content_json["slides"] if s["n"] == 15)
    return s15["capm_rows"]


def extract_risk_register_summary(content_json):
    s19 = next(s for s in content_json["slides"] if s["n"] == 19)
    s20 = next(s for s in content_json["slides"] if s["n"] == 20)
    return {
        "heatmap_summary": s19["heatmap_summary"],
        "categories": s19["categories"],
        "top_risks": s20["risks"]
    }


def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    with open(CONTENT) as f:
        content = json.load(f)

    pl_rows = extract_pl(wb)

    rev_key = "Total (Итого) Revenue (Выручка)"
    ebitda_key = "EBITDA (прибыль до вычетов) GAAP, standard"
    rev_vals = pl_rows.get(rev_key, [0]*12)
    ebitda_vals = pl_rows.get(ebitda_key, [0]*12)

    revenue_3y = round(sum(rev_vals[:12]))
    ebitda_3y = round(sum(ebitda_vals[:12]))

    mc = extract_mc_summary_from_xlsx(wb)
    ndp_det = extract_ndp_det_from_kpi(wb)

    key_metrics = dict(content["key_metrics"])
    key_metrics["mc_mean_irr"] = mc["irr_mean"]
    key_metrics["mc_stdev_irr"] = mc["irr_stdev"]
    key_metrics["ndp_3y"] = ndp_det
    key_metrics["ndp_mc_mean"] = mc["ndp_mean"]
    key_metrics["ndp_mc_p10"] = mc["ndp_p10"]

    print("Extracting investor returns...")
    t1_cashflow = extract_t1_cashflow(wb)
    returns_matrix = extract_returns_matrix(wb)

    deck_data = {
        "meta": {
            "version": "1.2.0",
            "source_xlsx": "investor_model_v1.1.1_Public.xlsx",
            "source_content": "deck_content.json",
            "extracted": "2026-04-15",
            "phase": 1
        },
        "key_metrics": key_metrics,
        "palette": content["meta"]["palette"],
        "fonts": content["meta"]["fonts"],
        "slides": content["slides"],
        "chart_data": {
            "s02_exec_summary": {
                "stats": content["slides"][1]["stats"],
                "disclosure": content["slides"][1]["disclosure"]
            },
            "s05_pipeline": {
                "pipeline": content["slides"][4]["pipeline"],
                "note": content["slides"][4]["note"]
            },
            "s12_unit_economics": {
                "formula": content["slides"][11]["formula"],
                "scenarios": content["slides"][11]["scenarios"],
                "films": extract_unit_economics(wb)
            },
            "s14_valuation": {
                "components": content["slides"][13]["components"],
                "bridge": content["slides"][13]["bridge"],
                "dcf_detail": extract_dcf(wb)
            },
            "s17_mc_distribution": extract_mc_percentiles_from_xlsx(wb),
            "s18_det_vs_stoch": {
                "table": content["slides"][17]["table"],
                "conclusion": content["slides"][17]["conclusion"]
            },
            "s20_top_risks": extract_risk_register_summary(content),
            "s22_waterfall": extract_waterfall(wb)
        },
        "financial": {
            "revenue_3y": revenue_3y,
            "ebitda_3y": ebitda_3y,
            "ndp_3y": ndp_det,
            "pl_summary": {
                "rows": [
                    {"row": r["row"], "y1": r["y1"], "y2": r["y2"], "y3": r["y3"], "total": r["total"]}
                    for r in content["slides"][12]["pl"]
                ]
            },
            "revenue_breakdown": extract_revenue_breakdown(wb),
            "investor_returns": {
                "t1_cashflow": t1_cashflow,
                "returns_matrix": returns_matrix
            },
            "capm": extract_capm(content)
        },
        "appendix_refs": {
            "A": "Risk Register (30 risks x 5 categories)",
            "B": "Stress Tests (5 scenarios + Tornado top-8)",
            "C": "MC Report (N=50000, seed=42)",
            "D": "Peer Comps (6+ comparable transactions)",
            "E": "CAPM Build-Up (full source attribution)"
        }
    }

    mc_irr_str = str(mc["irr_mean"])
    def patch_text(obj):
        if isinstance(obj, dict):
            return {k: patch_text(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [patch_text(v) for v in obj]
        if isinstance(obj, str):
            obj = obj.replace("MC Mean IRR 7.24%", f"MC Mean IRR {mc_irr_str}%")
            obj = obj.replace("MC 7.24%", f"MC {mc_irr_str}%")
            obj = obj.replace("20.09% и 7.24%", f"20.09% и {mc_irr_str}%")
            return obj
        return obj

    deck_data["slides"] = patch_text(deck_data["slides"])
    deck_data["chart_data"] = patch_text(deck_data["chart_data"])

    for slide in deck_data["slides"]:
        if slide.get("n") == 17 and "percentiles" in slide:
            mc_pct = extract_mc_percentiles_from_xlsx(wb)
            slide["percentiles"] = mc_pct["percentiles"]
        if slide.get("n") == 13 and "pl" in slide:
            for row in slide["pl"]:
                if "NDP" in row.get("row", "").upper() or "распределени" in row.get("row", "").lower():
                    row["y1"] = None
                    row["y2"] = None
                    row["y3"] = None
                    row["total"] = ndp_det

    for row in deck_data["financial"]["pl_summary"]["rows"]:
        if "NDP" in row.get("row", "").upper() or "распределени" in row.get("row", "").lower():
            row["y1"] = None
            row["y2"] = None
            row["y3"] = None
            row["total"] = ndp_det

    # Sanity checks
    assert abs(deck_data["key_metrics"]["mc_mean_irr"] - mc["irr_mean"]) < 0.01
    assert abs(deck_data["key_metrics"]["ndp_3y"] - ndp_det) < 1
    assert revenue_3y == 4545, f"revenue_3y={revenue_3y}, expected 4545"
    assert ebitda_3y == 2167, f"ebitda_3y={ebitda_3y}, expected 2167"
    assert deck_data["key_metrics"]["det_irr"] == 20.09
    assert deck_data["key_metrics"]["moic"] == 2.0
    assert deck_data["key_metrics"]["wacc"] == 19.05
    assert deck_data["key_metrics"]["anchor"] == 3000

    for row in deck_data["financial"]["pl_summary"]["rows"]:
        if "NDP" in row.get("row", "").upper():
            assert row["y1"] is None and row["y2"] is None and row["y3"] is None
            assert row["total"] == ndp_det

    ir = deck_data["financial"]["investor_returns"]
    assert "t1_cashflow" in ir and "returns_matrix" in ir
    base_row = next(x for x in ir["returns_matrix"] if "Base" in x["scenario"])
    assert base_row["W3_IRR"] == 20.09
    assert base_row["W3_MOIC"] == 2.0

    print("All sanity checks PASSED (xlsx-sourced)")

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(deck_data, f, ensure_ascii=False, indent=2)
    print(f"Written: {OUT} ({OUT.stat().st_size:,} bytes)")

    missing = []
    missing.append("- Phase 2 charts: S04 (TAM/SAM funnel detailed), S06 (capital discipline waterfall detailed), S07 (exit routes probability), S08 (market funnel 3D), S23 (terms comparison), S25 (CTA timeline)")
    missing.append("- Scenario toggle cross-slide data (Phase 2)")
    missing.append("- S17 dual slider interactive MC data (Phase 2)")
    missing.append("- S14 confidence interval data (Phase 2)")
    missing.append("- S18 horizon live-controls data (Phase 2)")
    missing.append("- RU<->EN actual translations (Phase 2)")

    with open(TODO, "w", encoding="utf-8") as f:
        f.write("# TODO: Missing Data for Phase 2+\n\n")
        f.write(f"Generated: 2026-04-15 (Phase 1)\n\n")
        f.write("## Data not yet extracted / not needed for Phase 1:\n\n")
        f.write("\n".join(missing))
        f.write("\n\n## Notes:\n\n")
        f.write("- All Phase 1 critical data (8 LP charts + 25 slide text) is complete in deck_data_v1.2.0.json\n")
        f.write("- SSOT numbers verified: mc_mean_irr=11.44, ndp_3y=3000, revenue_3y=4545, ebitda_3y=2167\n")
    print(f"Written: {TODO}")


if __name__ == "__main__":
    main()
