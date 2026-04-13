#!/usr/bin/env python3
"""
Audit extraction script: extracts ALL non-empty cell values from critical
financial sheets in the Public investor model for independent verification.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys

FILE_PATH = "/home/user/TrendStudio-Holding/Investor_Package/investor_model_v1.0_Public.xlsx"

# Sheets to extract (name, optional max_row override — 0 means all rows)
SHEETS = [
    ("09_P&L_Statement", 0),
    ("10_Cash_Flow", 0),
    ("11_Balance_Sheet", 0),
    ("19_Waterfall", 0),
    ("24_Investor_Returns", 0),
    ("28_Monte_Carlo_Summary", 0),
    ("22_Valuation_DCF", 0),
    ("02_Assumptions", 0),
    ("06_Cost_Structure", 0),
    ("07_Revenue_Breakdown", 0),
]


def fmt_value(v):
    """Format a cell value for readable output."""
    if v is None:
        return None
    if isinstance(v, float):
        # Show full precision for financial numbers
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return f"{v}"
    return str(v)


def extract_sheet(ws, sheet_name, max_row_override=0):
    """Extract and print every non-empty cell from a worksheet."""
    max_row = max_row_override if max_row_override > 0 else ws.max_row
    max_col = ws.max_column

    print(f"\n{'='*80}")
    print(f"=== SHEET: {sheet_name} === (rows={max_row}, cols={max_col})")
    print(f"{'='*80}")

    # Print column header index for reference
    col_headers = []
    for c in range(1, max_col + 1):
        col_headers.append(f"{get_column_letter(c)}")
    print(f"Column index: {', '.join(col_headers)}")
    print()

    for r in range(1, max_row + 1):
        cells = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None and str(val).strip() != "":
                ref = f"{get_column_letter(c)}{r}"
                formatted = fmt_value(val)
                cells.append(f"{ref}={formatted}")
        if cells:
            print(f"Row {r:>3}: {', '.join(cells)}")

    print()


def main():
    print(f"Loading workbook: {FILE_PATH}")
    print(f"Mode: data_only=True (cached formula results)")
    print()

    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)

    available = wb.sheetnames
    print(f"Available sheets ({len(available)}): {', '.join(available)}")
    print()

    for sheet_name, max_row_override in SHEETS:
        if sheet_name in available:
            ws = wb[sheet_name]
            extract_sheet(ws, sheet_name, max_row_override)
        else:
            print(f"\n*** WARNING: Sheet '{sheet_name}' NOT FOUND in workbook ***\n")

    print("\n" + "="*80)
    print("EXTRACTION COMPLETE")
    print("="*80)


if __name__ == "__main__":
    main()
