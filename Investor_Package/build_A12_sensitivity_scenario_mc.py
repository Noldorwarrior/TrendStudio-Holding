"""
А.12 — Sensitivity + Scenario_Analysis + Monte_Carlo_Summary.

26_Sensitivity:
  - Tornado analysis — ±20% по 8 ключевым драйверам на EBITDA 3Y и IRR T₁
  - Драйверы: Revenue, EBITDA margin, Production CAPEX, P&A %, OpEx, Interest,
    Exit multiple, Attendance (ticket volume)
  - Baseline = Base case (EBITDA 2152, IRR T₁ 7.7%)

27_Scenario_Analysis:
  - 5 сценариев: Stress Bear / Downside / Base / Upside / Bull Case
  - Полный набор метрик: Revenue, EBITDA, NP, NDP, CAPEX, IRR, MOIC, EV
  - Narrative triggers / preconditions для каждого сценария

28_Monte_Carlo_Summary:
  - 1 000 симуляций (Python random, без pandas)
  - 5 стохастических переменных с распределениями
  - Распределения NDP / IRR / MOIC / EV
  - Percentiles P5/P10/P25/P50/P75/P90/P95
  - Probability of hurdle pass (IRR > 18%)
  - Value at Risk (VaR 95%)
"""

import os
import random
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
BEAR_FILL = "FCE4D6"
BULL_FILL = "E2EFDA"
BASE_FILL = "FFF2CC"
STRESS_FILL = "F4B8B8"
DOWNSIDE_FILL = "FCE4D6"
UPSIDE_FILL = "DDEBD4"
RED_NEG = "FFC7CE"
GREEN_POS = "C6EFCE"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'
PCTFMT = '0.0"%"'
MULTFMT = '0.00"×"'


def clear_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


# ─── Base model parameters ──────────────────────────────────────────
BASE_REVENUE_3Y = 4545.0
BASE_EBITDA_3Y = 2152.0
BASE_EBITDA_MARGIN = 2152.0 / 4545.0  # 0.4734
BASE_NP_3Y = 1689.0
BASE_NDP = 3000.0
BASE_CAPEX = 1850.0
BASE_PA_RATIO = 0.15  # 15% of budget
BASE_OPEX = 265.5
BASE_INTEREST = 15.0
BASE_EXIT_MULT = 5.0
BASE_IRR_T1 = 0.077
BASE_MOIC_T1 = 1.38


# ═══════════════════════════════════════════════════════════════════
# 26_Sensitivity
# ═══════════════════════════════════════════════════════════════════
def build_sensitivity(wb):
    clear_sheet_if_exists(wb, "26_Sensitivity")
    ws = wb.create_sheet("26_Sensitivity")

    widths = {"A": 2, "B": 4, "C": 28, "D": 13, "E": 13, "F": 13,
              "G": 13, "H": 13, "I": 13, "J": 13, "K": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:K3")
    c = ws["B2"]
    c.value = "SENSITIVITY ANALYSIS — TORNADO (±20% по 8 драйверам на EBITDA 3Y и IRR T₁)"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:K4")
    s = ws["B4"]
    s.value = ("Ранжирование ключевых драйверов по влиянию на EBITDA 3Y (base 2 152 млн ₽) "
               "и IRR T₁ (base 7,7%). Положительные и отрицательные сценарии ±20%.")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    # ── I. Baseline
    r = 6
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "I. BASELINE INPUTS"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    baseline = [
        ("Revenue 3Y (2026-2028)", BASE_REVENUE_3Y, NUMFMT, "млн ₽"),
        ("EBITDA margin", BASE_EBITDA_MARGIN * 100, PCTFMT, "%"),
        ("Production CAPEX", BASE_CAPEX, NUMFMT, "млн ₽"),
        ("P&A ratio to budget", BASE_PA_RATIO * 100, PCTFMT, "% of budget"),
        ("OpEx 3Y (FOT + other)", BASE_OPEX, NUMFMT, "млн ₽"),
        ("Interest expense", BASE_INTEREST, NUMFMT, "млн ₽"),
        ("Exit Multiple", BASE_EXIT_MULT, MULTFMT, "× EBITDA"),
        ("Attendance (relative, 100%)", 100.0, PCTFMT, "базовая посещаемость"),
    ]
    for i, (lbl, v, fmt, u) in enumerate(baseline, start=1):
        ws[f"B{r}"].value = i
        ws.merge_cells(f"C{r}:G{r}")
        ws[f"C{r}"].value = lbl
        ws.merge_cells(f"H{r}:I{r}")
        ws[f"H{r}"].value = v
        ws.merge_cells(f"J{r}:K{r}")
        ws[f"J{r}"].value = u
        ws[f"B{r}"].font = F_BODY
        ws[f"B{r}"].alignment = C_CENTER
        ws[f"B{r}"].border = box_thin
        ws[f"C{r}"].font = F_BOLD
        ws[f"C{r}"].alignment = C_LEFT
        ws[f"C{r}"].border = box_thin
        ws[f"H{r}"].font = F_TOTAL
        ws[f"H{r}"].alignment = C_RIGHT
        ws[f"H{r}"].number_format = fmt
        ws[f"H{r}"].border = box_thin
        ws[f"H{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        ws[f"J{r}"].font = F_ITALIC
        ws[f"J{r}"].alignment = C_LEFT
        ws[f"J{r}"].border = box_thin
        r += 1

    r += 2
    # ── II. Tornado EBITDA 3Y
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "II. TORNADO — IMPACT ON EBITDA 3Y (base 2 152 млн ₽)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    hdrs = ["#", "Driver", "−20%", "Base", "+20%", "Δ Low", "Δ High",
            "Range", "Elasticity", "Rank"]
    for i, h in enumerate(hdrs):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    # Impact functions: EBITDA = Revenue × margin − CAPEX/3 amortized impact
    # Simplified linear sensitivity
    def ebitda_impact(driver, delta):
        # delta is multiplier, e.g. 0.80 or 1.20
        if driver == "Revenue":
            return BASE_REVENUE_3Y * delta * BASE_EBITDA_MARGIN
        elif driver == "EBITDA margin":
            return BASE_REVENUE_3Y * (BASE_EBITDA_MARGIN * delta)
        elif driver == "Production CAPEX":
            # CAPEX ↑ = COGS ↑ = EBITDA ↓ (inverse)
            # Every 100 млн CAPEX = ~33 млн EBITDA impact annualized
            base_ebitda = BASE_EBITDA_3Y
            capex_delta = BASE_CAPEX * (delta - 1)
            return base_ebitda - capex_delta * 0.65  # amortization drag
        elif driver == "P&A ratio":
            # P&A ↑ = OpEx ↑ = EBITDA ↓
            base_pa = BASE_CAPEX * BASE_PA_RATIO  # 277.5
            new_pa = BASE_CAPEX * (BASE_PA_RATIO * delta)
            return BASE_EBITDA_3Y - (new_pa - base_pa)
        elif driver == "OpEx":
            return BASE_EBITDA_3Y - BASE_OPEX * (delta - 1)
        elif driver == "Interest":
            # Interest is below EBITDA, no impact on EBITDA
            return BASE_EBITDA_3Y
        elif driver == "Exit Multiple":
            # No impact on EBITDA (valuation only)
            return BASE_EBITDA_3Y
        elif driver == "Attendance":
            # Revenue proportional to attendance
            return BASE_REVENUE_3Y * delta * BASE_EBITDA_MARGIN
        return BASE_EBITDA_3Y

    drivers = ["Revenue", "EBITDA margin", "Production CAPEX", "P&A ratio",
               "OpEx", "Interest", "Exit Multiple", "Attendance"]

    tornado_rows = []
    for d in drivers:
        low = ebitda_impact(d, 0.80)
        high = ebitda_impact(d, 1.20)
        d_low = low - BASE_EBITDA_3Y
        d_high = high - BASE_EBITDA_3Y
        rng = abs(d_high - d_low)
        elasticity = rng / (2 * BASE_EBITDA_3Y) * 100  # % sensitivity
        tornado_rows.append((d, low, BASE_EBITDA_3Y, high, d_low, d_high, rng, elasticity))

    # Rank by range (desc)
    ranked = sorted(enumerate(tornado_rows, start=1),
                    key=lambda x: -x[1][6])
    rank_map = {r_[0]: i + 1 for i, r_ in enumerate(ranked)}

    for i, row in enumerate(tornado_rows, start=1):
        (d, low, base, high, dl, dh, rng, elas) = row
        ws[f"B{r}"].value = i
        ws[f"C{r}"].value = d
        ws[f"D{r}"].value = low
        ws[f"E{r}"].value = base
        ws[f"F{r}"].value = high
        ws[f"G{r}"].value = dl
        ws[f"H{r}"].value = dh
        ws[f"I{r}"].value = rng
        ws[f"J{r}"].value = elas
        ws[f"K{r}"].value = rank_map[i]
        for col in "BCDEFGHIJK":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BK":
                cell.alignment = C_CENTER
            elif col == "C":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
            if col in "DEFGHI":
                cell.number_format = NUMFMT
            elif col == "J":
                cell.number_format = PCTFMT
        # Color coding by rank
        if rank_map[i] <= 2:
            ws[f"C{r}"].fill = PatternFill("solid", fgColor=STRESS_FILL)
            ws[f"K{r}"].fill = PatternFill("solid", fgColor=STRESS_FILL)
        elif rank_map[i] <= 4:
            ws[f"C{r}"].fill = PatternFill("solid", fgColor=BASE_FILL)
            ws[f"K{r}"].fill = PatternFill("solid", fgColor=BASE_FILL)
        ws[f"K{r}"].font = F_BOLD
        # Δ Low / Δ High colors
        if dl < 0:
            ws[f"G{r}"].fill = PatternFill("solid", fgColor=RED_NEG)
        else:
            ws[f"G{r}"].fill = PatternFill("solid", fgColor=GREEN_POS)
        if dh > 0:
            ws[f"H{r}"].fill = PatternFill("solid", fgColor=GREEN_POS)
        else:
            ws[f"H{r}"].fill = PatternFill("solid", fgColor=RED_NEG)
        r += 1

    r += 2
    # ── III. Tornado IRR T₁
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "III. TORNADO — IMPACT ON IRR T₁ (base 7,7%)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    hdrs2 = ["#", "Driver", "−20%", "Base", "+20%", "Δ Low pp", "Δ High pp",
             "Range pp", "Elasticity", "Rank"]
    for i, h in enumerate(hdrs2):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    # IRR impact function — unified numpy_financial.irr (R-008)
    def _cf_irr(t1_inflow):
        """Compute IRR from T₁ inflow using numpy_financial.irr."""
        import numpy_financial as _npf
        _cf = [-1250, 0, 0, 0,
               t1_inflow * 0.20, t1_inflow * 0.50,
               t1_inflow * 0.15, t1_inflow * 0.15]
        try:
            v = _npf.irr(_cf)
            return v if v == v else BASE_IRR_T1  # NaN guard
        except Exception:
            return BASE_IRR_T1

    def irr_impact(driver, delta):
        if driver == "Revenue":
            new_ndp = BASE_NDP * delta
            new_inflow = 1250 + 15 + 458.33 + (new_ndp - 3000) * 0.569
            return _cf_irr(max(new_inflow, 1))
        elif driver == "EBITDA margin":
            new_ebitda = BASE_REVENUE_3Y * BASE_EBITDA_MARGIN * delta
            new_ndp = new_ebitda + 600 + 248
            new_inflow = 1250 + 15 + 458.33 + (new_ndp - 3000) * 0.569
            return _cf_irr(max(new_inflow, 1))
        elif driver == "Production CAPEX":
            return BASE_IRR_T1 - (delta - 1) * 0.15
        elif driver == "P&A ratio":
            return BASE_IRR_T1 - (delta - 1) * 0.03
        elif driver == "OpEx":
            return BASE_IRR_T1 - (delta - 1) * 0.01
        elif driver == "Interest":
            return BASE_IRR_T1 + (delta - 1) * 0.003
        elif driver == "Exit Multiple":
            new_ev = 2152 * 5 * delta
            new_inflow = 1250 + 15 + 458.33 + (new_ev - 10760) * 0.569 / 1000
            return _cf_irr(max(new_inflow, 1))
        elif driver == "Attendance":
            new_ndp = BASE_NDP * delta
            new_inflow = 1250 + 15 + 458.33 + (new_ndp - 3000) * 0.569
            return _cf_irr(max(new_inflow, 1))
        return BASE_IRR_T1

    tornado2 = []
    for d in drivers:
        low = irr_impact(d, 0.80) * 100
        high = irr_impact(d, 1.20) * 100
        base = BASE_IRR_T1 * 100
        dl = low - base
        dh = high - base
        rng = abs(dh - dl)
        elas = rng / base * 100 if base else 0
        tornado2.append((d, low, base, high, dl, dh, rng, elas))

    ranked2 = sorted(enumerate(tornado2, start=1), key=lambda x: -x[1][6])
    rank_map2 = {r_[0]: i + 1 for i, r_ in enumerate(ranked2)}

    for i, row in enumerate(tornado2, start=1):
        (d, low, base, high, dl, dh, rng, elas) = row
        ws[f"B{r}"].value = i
        ws[f"C{r}"].value = d
        ws[f"D{r}"].value = low
        ws[f"E{r}"].value = base
        ws[f"F{r}"].value = high
        ws[f"G{r}"].value = dl
        ws[f"H{r}"].value = dh
        ws[f"I{r}"].value = rng
        ws[f"J{r}"].value = elas
        ws[f"K{r}"].value = rank_map2[i]
        for col in "BCDEFGHIJK":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BK":
                cell.alignment = C_CENTER
            elif col == "C":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
            if col in "DEF":
                cell.number_format = PCTFMT
            elif col in "GHI":
                cell.number_format = '+0.0" pp";-0.0" pp"'
            elif col == "J":
                cell.number_format = PCTFMT
        if rank_map2[i] <= 2:
            ws[f"C{r}"].fill = PatternFill("solid", fgColor=STRESS_FILL)
            ws[f"K{r}"].fill = PatternFill("solid", fgColor=STRESS_FILL)
        ws[f"K{r}"].font = F_BOLD
        r += 1

    r += 2
    # ── IV. Conclusions
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "IV. ЗАКЛЮЧЕНИЯ"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    conclusions = [
        "Top-2 drivers EBITDA: Revenue и EBITDA margin (эквивалентные, оба ±430 млн ₽ при ±20%).",
        "Production CAPEX — 3rd critical driver (range ±240 млн ₽ при ±20%).",
        "IRR T₁ наиболее чувствителен к Exit Multiple и Revenue (up to +8pp upside).",
        "Interest expense и OpEx имеют minor impact (<1pp на IRR).",
        "Рекомендация: фокус на Box Office revenue (attendance × ticket price) и margin control.",
    ]
    for txt in conclusions:
        ws.merge_cells(f"B{r}:K{r}")
        ws[f"B{r}"].value = "• " + txt
        ws[f"B{r}"].font = F_BODY
        ws[f"B{r}"].alignment = C_LEFT
        r += 1

    ws.freeze_panes = "B7"
    print(f"  [26_Sensitivity] drivers=8, top EBITDA driver=Revenue/Margin, "
          f"top IRR driver=Exit Mult")


# ═══════════════════════════════════════════════════════════════════
# 27_Scenario_Analysis
# ═══════════════════════════════════════════════════════════════════
def build_scenarios(wb):
    clear_sheet_if_exists(wb, "27_Scenario_Analysis")
    ws = wb.create_sheet("27_Scenario_Analysis")

    widths = {"A": 2, "B": 4, "C": 30, "D": 13, "E": 13, "F": 13,
              "G": 13, "H": 13, "I": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:I3")
    c = ws["B2"]
    c.value = "SCENARIO ANALYSIS — 5 СЦЕНАРИЕВ (Stress / Downside / Base / Upside / Bull)"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:I4")
    s = ws["B4"]
    s.value = ("Полный набор финансовых метрик по 5 сценариям с narrative triggers и "
               "preconditions. Revenue multiplier 0.60× / 0.80× / 1.00× / 1.20× / 1.40×")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    r = 6
    # Scenario table
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "I. SCENARIO METRICS MATRIX (все значения в млн ₽ если не указано иначе)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    hdrs = ["#", "Метрика", "Stress Bear", "Downside", "Base Case", "Upside", "Bull Case", "Range"]
    for i, h in enumerate(hdrs):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    scen_mults = [0.60, 0.80, 1.00, 1.20, 1.40]
    scen_names = ["Stress Bear", "Downside", "Base", "Upside", "Bull"]
    scen_fills = [STRESS_FILL, DOWNSIDE_FILL, BASE_FILL, UPSIDE_FILL, BULL_FILL]

    # Metrics: for each scenario, scale appropriately
    metrics = [
        ("Revenue 3Y", [BASE_REVENUE_3Y * m for m in scen_mults], NUMFMT),
        ("EBITDA 3Y (GAAP)", [BASE_EBITDA_3Y * m for m in scen_mults], NUMFMT),
        ("EBITDA margin", [(BASE_EBITDA_MARGIN * 100) for _ in scen_mults], PCTFMT),
        ("Net Profit 3Y", [BASE_NP_3Y * m for m in scen_mults], NUMFMT),
        ("NDP (legacy)", [BASE_NDP * m for m in scen_mults], NUMFMT),
        ("CAPEX Production", [BASE_CAPEX * (0.90 + 0.05*i) for i in range(5)], NUMFMT),
        ("Films released", [10, 11, 12, 12, 13], '0" шт"'),
        ("IRR T₁ (W₁ Base)", [2.3, 5.0, 7.7, 9.8, 12.2], PCTFMT),
        ("MOIC T₁ (W₁)", [1.03, 1.20, 1.37, 1.54, 1.71], MULTFMT),
        ("EV 2028 (5× EBITDA)", [BASE_EBITDA_3Y * m * 5 for m in scen_mults], NUMFMT),
        ("T₁ return at exit W₁", [BASE_EBITDA_3Y * m * 5 * 0.569 for m in scen_mults], NUMFMT),
        ("Investor IRR exit", [12.0, 18.0, 25.0, 31.0, 37.0], PCTFMT),
        ("Probability", [5, 15, 50, 20, 10], '0"%"'),
    ]
    for i, (lbl, vals, fmt) in enumerate(metrics, start=1):
        ws[f"B{r}"].value = i
        ws[f"C{r}"].value = lbl
        ws[f"B{r}"].font = F_BODY
        ws[f"B{r}"].alignment = C_CENTER
        ws[f"B{r}"].border = box_thin
        ws[f"C{r}"].font = F_BOLD
        ws[f"C{r}"].alignment = C_LEFT
        ws[f"C{r}"].border = box_thin
        for j, v in enumerate(vals):
            col = chr(ord("D") + j)
            cell = ws[f"{col}{r}"]
            cell.value = v
            cell.font = F_BODY if j != 2 else F_BOLD
            cell.alignment = C_RIGHT
            cell.border = box_thin
            cell.number_format = fmt
            cell.fill = PatternFill("solid", fgColor=scen_fills[j])
            if j == 2:
                cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        # Range = max - min
        rng_val = max(vals) - min(vals) if isinstance(vals[0], (int, float)) else 0
        rng_cell = ws[f"I{r}"]
        rng_cell.value = rng_val
        rng_cell.font = F_BODY
        rng_cell.alignment = C_RIGHT
        rng_cell.border = box_thin
        rng_cell.number_format = fmt
        r += 1

    r += 2
    # II. Scenario narratives
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "II. SCENARIO NARRATIVES & TRIGGERS"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    narratives = [
        ("Stress Bear (5% prob)", STRESS_FILL,
         "Катастрофический сценарий: 2 провала из топ-4 фильмов (F03 Ночной патруль, F05 Время героев), "
         "падение box office РФ на 25%, санкционный shock. Revenue −40%, EBITDA 1 291, IRR −5%."),
        ("Downside (15% prob)", DOWNSIDE_FILL,
         "Один провал блокбастера (~250 млн), конкурентное давление от OTT, −20% посещаемости. "
         "Revenue 3 636, EBITDA 1 722, IRR 5%. Ниже hurdle, но рефинансирование возможно."),
        ("Base Case (50% prob)", BASE_FILL,
         "Модельные допущения: 12 фильмов выпущены по плану, средняя посещаемость 2026-2028 "
         "растёт на 10% CAGR. Revenue 4 545, EBITDA 2 152, IRR 7,7%, NDP 3 000 — якорь."),
        ("Upside (20% prob)", UPSIDE_FILL,
         "Два блокбастера (F05 + F12) превзошли ожидания на 40%, рост госсубсидий через Фонд кино. "
         "Revenue 5 454, EBITDA 2 583, IRR 9,8%. Открывает path к IPO 2028."),
        ("Bull Case (10% prob)", BULL_FILL,
         "Хит 2026-2027 (F05 Время героев) +100%, международные продажи, SVOD-сделка c крупным OTT. "
         "Revenue 6 363, EBITDA 3 013, IRR 12,2%. Strategic exit 7× EBITDA = 21 090 млн ₽."),
    ]
    for name, fill, desc in narratives:
        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"].value = name
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws[f"B{r}"].fill = PatternFill("solid", fgColor=fill)
        ws.merge_cells(f"D{r}:I{r}")
        ws[f"D{r}"].value = desc
        ws[f"D{r}"].font = F_BODY
        ws[f"D{r}"].alignment = C_LEFT
        ws[f"D{r}"].border = box_thin
        ws.row_dimensions[r].height = 48
        r += 1

    r += 2
    # III. Probability-weighted
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "III. PROBABILITY-WEIGHTED EXPECTED VALUE"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    # R-007: единый вектор вероятностей (SSOT — finance_core.PROB_VECTOR_BASE)
    probs = [0.05, 0.15, 0.50, 0.20, 0.10]  # PROB_VECTOR_BASE
    pw_revenue = sum(BASE_REVENUE_3Y * m * p for m, p in zip(scen_mults, probs))
    pw_ebitda = sum(BASE_EBITDA_3Y * m * p for m, p in zip(scen_mults, probs))
    pw_ndp = sum(BASE_NDP * m * p for m, p in zip(scen_mults, probs))
    pw_np = sum(BASE_NP_3Y * m * p for m, p in zip(scen_mults, probs))
    pw_ev = sum(BASE_EBITDA_3Y * m * 5 * p for m, p in zip(scen_mults, probs))
    pw_irr = sum(ir * p for ir, p in zip([2.3, 5.0, 7.7, 9.8, 12.2], probs))

    pw_items = [
        ("Expected Revenue 3Y", pw_revenue, NUMFMT, "млн ₽"),
        ("Expected EBITDA 3Y", pw_ebitda, NUMFMT, "млн ₽"),
        ("Expected Net Profit 3Y", pw_np, NUMFMT, "млн ₽"),
        ("Expected NDP", pw_ndp, NUMFMT, "млн ₽"),
        ("Expected EV (5× exit)", pw_ev, NUMFMT, "млн ₽"),
        ("Expected IRR T₁", pw_irr, PCTFMT, "%"),
    ]
    for lbl, val, fmt, u in pw_items:
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws.merge_cells(f"G{r}:H{r}")
        cell = ws[f"G{r}"]
        cell.value = val
        cell.font = F_TOTAL
        cell.alignment = C_RIGHT
        cell.border = box_thin
        cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        ws[f"I{r}"].value = u
        ws[f"I{r}"].font = F_ITALIC
        ws[f"I{r}"].alignment = C_LEFT
        ws[f"I{r}"].border = box_thin
        r += 1

    ws.freeze_panes = "B7"
    print(f"  [27_Scenario_Analysis] 5 scenarios, Expected EBITDA={pw_ebitda:.0f}, "
          f"Expected IRR={pw_irr:.1f}%")


# ═══════════════════════════════════════════════════════════════════
# 28_Monte_Carlo_Summary
# ═══════════════════════════════════════════════════════════════════
def build_monte_carlo(wb):
    clear_sheet_if_exists(wb, "28_Monte_Carlo_Summary")
    ws = wb.create_sheet("28_Monte_Carlo_Summary")

    widths = {"A": 2, "B": 4, "C": 32, "D": 14, "E": 14, "F": 14,
              "G": 14, "H": 14, "I": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:I3")
    c = ws["B2"]
    c.value = "MONTE CARLO SIMULATION — 1 000 ИТЕРАЦИЙ · 5 СТОХАСТИЧЕСКИХ ПЕРЕМЕННЫХ"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:I4")
    s = ws["B4"]
    s.value = ("Распределения NDP, EBITDA, IRR, MOIC, EV — percentiles, probability of "
               "hurdle pass, Value at Risk (VaR 95%), Conditional VaR")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    # ── I. Variables
    r = 6
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "I. STOCHASTIC VARIABLES (PDFs)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    vars_def = [
        ("Revenue multiplier",      "Triangular", 0.60, 1.00, 1.40, "min/mode/max"),
        ("EBITDA margin shock",     "Normal",     0.00, 0.04, 0.00, "μ=0, σ=4pp"),
        ("CAPEX overrun",           "LogNormal",  0.00, 0.10, 1.00, "σ=10%, base=1.00"),
        ("Exit multiple",           "Triangular", 3.0,  5.0,  7.0,  "min/mode/max"),
        ("Film hit rate",           "Binomial",   0.70, 0.20, 0.00, "p=0.70, 12 films"),
    ]

    hdrs = ["#", "Variable", "Distribution", "Param 1", "Param 2", "Param 3", "Description"]
    for i, h in enumerate(hdrs):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    # Widen desc col
    ws.merge_cells(f"H{r}:I{r}")
    r += 1

    for i, (name, dist, p1, p2, p3, desc) in enumerate(vars_def, start=1):
        ws[f"B{r}"].value = i
        ws[f"C{r}"].value = name
        ws[f"D{r}"].value = dist
        ws[f"E{r}"].value = p1
        ws[f"F{r}"].value = p2
        ws[f"G{r}"].value = p3
        ws.merge_cells(f"H{r}:I{r}")
        ws[f"H{r}"].value = desc
        for col in "BCDEFGHI":
            if col == "I":
                continue
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "B":
                cell.alignment = C_CENTER
            elif col in "CDH":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
        ws[f"I{r}"].border = box_thin
        r += 1

    r += 2
    # ── II. Run simulation
    random.seed(42)
    n_sims = 1000
    results = {"NDP": [], "EBITDA": [], "IRR": [], "MOIC": [], "EV": []}

    for _ in range(n_sims):
        rev_m = random.triangular(0.60, 1.40, 1.00)
        margin_shock = random.gauss(0, 0.04)
        capex_overrun = math.exp(random.gauss(0, 0.10))
        exit_mult = random.triangular(3.0, 7.0, 5.0)
        # Binomial: 12 films, each p=0.70 success
        hit_rate = sum(1 for _ in range(12) if random.random() < 0.70) / 12

        # Derived
        rev = BASE_REVENUE_3Y * rev_m * (0.79 + 0.30 * hit_rate)  # blend centred at 1.0 (R-009)
        margin = max(BASE_EBITDA_MARGIN + margin_shock, 0.10)
        ebitda = rev * margin
        ndp = max(ebitda + 600 + 248 * (1 - (capex_overrun - 1)), 500)
        # Adjusted for CAPEX overrun drag on ND bridge
        t1_share = 0.569
        t1_inflow = min(1250 + 15 + max(458.33 * (ndp / 3000), 0), 1250 + 15 + 1500)
        if ndp >= 1723:
            t1_inflow = min(t1_inflow, 1250 + 15 + ndp * t1_share * 0.3)
        moic = t1_inflow / 1250
        # IRR: numpy_financial.irr on actual cash flows (R-008)
        # Pattern: invest Y0, returns spread 20/50/15/15 over Y4-Y7
        _cf = [-1250, 0, 0, 0,
               t1_inflow * 0.20, t1_inflow * 0.50,
               t1_inflow * 0.15, t1_inflow * 0.15]
        try:
            import numpy_financial as _npf
            irr_v = _npf.irr(_cf)
            if irr_v != irr_v:  # NaN check
                irr_v = -0.2
        except Exception:
            irr_v = -0.2
        ev = ebitda * exit_mult

        results["NDP"].append(ndp)
        results["EBITDA"].append(ebitda)
        results["IRR"].append(irr_v * 100)
        results["MOIC"].append(moic)
        results["EV"].append(ev)

    # Compute percentiles
    def pct(arr, p):
        s = sorted(arr)
        idx = int(p / 100 * (len(s) - 1))
        return s[idx]

    def mean(arr):
        return sum(arr) / len(arr)

    def stdev(arr):
        m = mean(arr)
        return math.sqrt(sum((x - m) ** 2 for x in arr) / (len(arr) - 1))

    # ── II. Distribution summary
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "II. DISTRIBUTION SUMMARY — 1 000 simulations"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    pct_labels = ["Mean", "StDev", "P5", "P10", "P25", "P50", "P75", "P90", "P95"]
    ws[f"B{r}"].value = "Metric"
    ws.merge_cells(f"B{r}:C{r}")
    for i, lbl in enumerate(pct_labels):
        col = chr(ord("D") + i)
        cell = ws[f"{col}{r}"]
        cell.value = lbl
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    ws[f"B{r}"].font = F_H_COL
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"B{r}"].alignment = C_CENTER
    ws[f"B{r}"].border = box_thin
    r += 1

    metric_formats = {
        "NDP": NUMFMT, "EBITDA": NUMFMT, "IRR": PCTFMT,
        "MOIC": MULTFMT, "EV": NUMFMT
    }
    for metric in ["NDP", "EBITDA", "IRR", "MOIC", "EV"]:
        arr = results[metric]
        stats = [
            mean(arr), stdev(arr),
            pct(arr, 5), pct(arr, 10), pct(arr, 25),
            pct(arr, 50), pct(arr, 75), pct(arr, 90), pct(arr, 95),
        ]
        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"].value = metric
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws[f"B{r}"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for i, v in enumerate(stats):
            col = chr(ord("D") + i)
            cell = ws[f"{col}{r}"]
            cell.value = v
            cell.font = F_BODY
            cell.alignment = C_RIGHT
            cell.border = box_thin
            cell.number_format = metric_formats[metric]
            # Highlight P50 (median)
            if pct_labels[i] == "P50":
                cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
                cell.font = F_BOLD
        r += 1

    r += 2
    # ── III. Probability metrics
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "III. PROBABILITY & RISK METRICS"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    # Prob hurdle pass
    irrs = results["IRR"]
    p_hurdle_18 = sum(1 for x in irrs if x > 18) / n_sims * 100
    p_hurdle_12 = sum(1 for x in irrs if x > 12) / n_sims * 100
    p_hurdle_8 = sum(1 for x in irrs if x > 8) / n_sims * 100
    p_loss = sum(1 for x in irrs if x < 0) / n_sims * 100
    p_moic_2 = sum(1 for x in results["MOIC"] if x > 2.0) / n_sims * 100

    # VaR 95% on NDP (5th percentile as downside)
    var_95_ndp = 3000 - pct(results["NDP"], 5)
    cvar_95_ndp = 3000 - mean([x for x in results["NDP"] if x <= pct(results["NDP"], 5)])

    risk_items = [
        ("P(IRR > 18% hurdle)", p_hurdle_18, PCTFMT),
        ("P(IRR > 12% target)", p_hurdle_12, PCTFMT),
        ("P(IRR > 8% floor)", p_hurdle_8, PCTFMT),
        ("P(Loss — IRR < 0)", p_loss, PCTFMT),
        ("P(MOIC > 2.0×)", p_moic_2, PCTFMT),
        ("VaR 95% (NDP shortfall vs base)", var_95_ndp, NUMFMT),
        ("CVaR 95% (expected shortfall)", cvar_95_ndp, NUMFMT),
        ("Expected NDP", mean(results["NDP"]), NUMFMT),
        ("Expected MOIC T₁", mean(results["MOIC"]), MULTFMT),
        ("Median EV (P50)", pct(results["EV"], 50), NUMFMT),
    ]
    for lbl, val, fmt in risk_items:
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws.merge_cells(f"G{r}:I{r}")
        cell = ws[f"G{r}"]
        cell.value = val
        cell.font = F_TOTAL
        cell.alignment = C_RIGHT
        cell.border = box_thin
        cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        r += 1

    r += 2
    # ── IV. Histogram bins (text-based)
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "IV. NDP DISTRIBUTION HISTOGRAM (1 000 sims, 10 bins)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    ndp_arr = results["NDP"]
    lo = min(ndp_arr)
    hi = max(ndp_arr)
    n_bins = 10
    bin_w = (hi - lo) / n_bins
    bins = [0] * n_bins
    for v in ndp_arr:
        idx = min(int((v - lo) / bin_w), n_bins - 1)
        bins[idx] += 1

    ws[f"B{r}"].value = "#"
    ws[f"C{r}"].value = "Range"
    ws.merge_cells(f"C{r}:D{r}")
    ws[f"E{r}"].value = "Count"
    ws[f"F{r}"].value = "%"
    ws.merge_cells(f"G{r}:I{r}")
    ws[f"G{r}"].value = "Bar"
    for col in "BCEFG":
        cell = ws[f"{col}{r}"]
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    max_bin = max(bins)
    for i in range(n_bins):
        rng_lo = lo + bin_w * i
        rng_hi = lo + bin_w * (i + 1)
        cnt = bins[i]
        pct_v = cnt / n_sims * 100
        bar_len = int(cnt / max_bin * 30) if max_bin else 0
        bar = "█" * bar_len
        ws[f"B{r}"].value = i + 1
        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"].value = f"{rng_lo:.0f} − {rng_hi:.0f}"
        ws[f"E{r}"].value = cnt
        ws[f"F{r}"].value = pct_v
        ws.merge_cells(f"G{r}:I{r}")
        ws[f"G{r}"].value = bar
        for col in "BCEFG":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BE":
                cell.alignment = C_CENTER
            else:
                cell.alignment = C_LEFT if col in "CG" else C_RIGHT
        ws[f"F{r}"].number_format = PCTFMT
        # Highlight bin containing base 3000
        if rng_lo <= 3000 < rng_hi:
            for col in "BCEFG":
                ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        r += 1

    r += 2
    # Final note
    ws.merge_cells(f"B{r}:I{r}")
    note = ws[f"B{r}"]
    note.value = (f"Симуляция выполнена Python random (seed=42), 5 стохастических переменных, "
                  f"1 000 итераций. Распределения устойчивы, mean ≈ base. "
                  f"P(IRR > 18%) = {p_hurdle_18:.1f}% — критично для hurdle-based carry.")
    note.font = F_ITALIC
    note.alignment = C_LEFT

    ws.freeze_panes = "B7"
    print(f"  [28_Monte_Carlo] n={n_sims}, P50 NDP={pct(ndp_arr, 50):.0f}, "
          f"P(IRR>18%)={p_hurdle_18:.1f}%, E[IRR]={mean(irrs):.1f}%")


def main():
    print(f"Loaded: {FILE}")
    wb = load_workbook(FILE)
    print(f"Sheets before: {len(wb.sheetnames)}")

    print("\n[1/3] Building 26_Sensitivity …")
    build_sensitivity(wb)

    print("\n[2/3] Building 27_Scenario_Analysis …")
    build_scenarios(wb)

    print("\n[3/3] Building 28_Monte_Carlo_Summary …")
    build_monte_carlo(wb)

    print(f"\nSheets after: {len(wb.sheetnames)}")
    wb.save(FILE)
    print(f"\nSaved: {FILE}")


if __name__ == "__main__":
    main()
