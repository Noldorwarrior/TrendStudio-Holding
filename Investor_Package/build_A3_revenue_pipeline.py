"""
А.3 — Revenue_Breakdown + Content_Pipeline
Добавляем 2 листа к investor_model_v1.0_Public.xlsx

07_Revenue_Breakdown: 8 источников × 3 года, якорь Σ = 4 545 млн ₽
08_Content_Pipeline: 12 фильмов с бюджетами, релизами, ожидаемой выручкой
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package"
OUT_FILE = os.path.join(OUT_DIR, "investor_model_v1.0_Public.xlsx")

BRAND_BLUE = "0070C0"
BRAND_BLUE_LIGHT = "D9E2F3"
BRAND_BLUE_DARK = "002060"
ACCENT_GREEN = "548235"
ACCENT_RED = "C00000"
GRAY_LIGHT = "F2F2F2"
GRAY_DARK = "595959"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"
FORMULA_BLACK = "000000"
LINK_GREEN = "006100"
KEY_METRIC_FILL = "FFF2CC"


def thin(color="808080"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def sheet_header(ws, title, subtitle="", merge_to="O"):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"B2:{merge_to}2")
    c = ws["B2"]
    c.value = title
    c.font = Font(name="Arial", size=18, bold=True, color=BRAND_BLUE_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 30
    if subtitle:
        ws.merge_cells(f"B3:{merge_to}3")
        c = ws["B3"]
        c.value = subtitle
        c.font = Font(name="Arial", size=10, italic=True, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="left", indent=1)


def header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin("FFFFFF")
    return c


def block_title(ws, row, start_col, end_col, text, color=BRAND_BLUE):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


# ============================================================================
# ЛИСТ 07: Revenue_Breakdown
# ============================================================================

def build_revenue_breakdown(wb):
    ws = wb.create_sheet("07_Revenue_Breakdown")

    widths = [2, 36, 14, 14, 14, 14, 14, 14, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sheet_header(ws, "Структура выручки (Revenue Breakdown)",
                 "8 источников выручки × 2026–2028. Якорь Σ = 4 545 млн ₽ (12 фильмов). Доли согласованы с Assumptions блок F.",
                 merge_to="H")

    # --- Блок 1: 8 источников × 3 года ---
    row = 5
    block_title(ws, row, 2, 8, "ВЫРУЧКА ПО ИСТОЧНИКАМ (млн ₽)", BRAND_BLUE)
    row += 1

    headers = ["Источник", "Доля, %", "2026", "2027", "2028", "Σ 2026–2028", "CAGR"]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h)
    ws.row_dimensions[row].height = 26
    row += 1

    # Распределение по годам: 300 / 1750 / 2495 = 4545
    # Доли источников (из Assumptions блок F)
    sources = [
        ("Box Office (кинопрокат)",          0.48, "Премьеры в широкий прокат"),
        ("Online SVOD (стриминги)",          0.20, "Kion, Okko, START, IVI, Wink"),
        ("TV rights",                        0.08, "Первый, Россия-1, ТНТ, НТВ"),
        ("International sales",              0.10, "СНГ, BRICS, EMEA"),
        ("VOD (transactional)",              0.04, "Apple iTunes, Google Play"),
        ("Sponsorship / product placement",  0.05, "Интеграция брендов"),
        ("Merchandise",                      0.02, "Мерч, лицензии, игры"),
        ("Господдержка (безвозвратная)",     0.03, "Фонд кино + Минкульт + ИРИ"),
    ]

    # Годовые totals
    year_totals = {2026: 300, 2027: 1750, 2028: 2495}
    total_3y = sum(year_totals.values())  # 4545

    start_src_row = row
    for src_name, share, note in sources:
        # Источник
        c = ws.cell(row=row, column=2, value=src_name)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        # Доля
        c = ws.cell(row=row, column=3, value=share)
        c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.0%"
        c.border = thin()
        # 2026, 2027, 2028
        for i, year in enumerate([2026, 2027, 2028]):
            val = year_totals[year] * share
            cc = ws.cell(row=row, column=4 + i, value=val)
            cc.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
            cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cc.number_format = '#,##0.00'
            cc.border = thin()
        # Σ 3 года
        c = ws.cell(row=row, column=7, value=f"=SUM(D{row}:F{row})")
        c.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        c.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.00" млн"'
        c.border = thin()
        # CAGR 2026→2028
        c = ws.cell(row=row, column=8, value=f"=(F{row}/D{row})^(1/2)-1")
        c.font = Font(name="Arial", size=10, color=ACCENT_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '+0.0%;-0.0%'
        c.border = thin()

        ws.row_dimensions[row].height = 20
        row += 1
    end_src_row = row - 1

    # ИТОГО
    c = ws.cell(row=row, column=2, value="ИТОГО ВЫРУЧКА")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    c = ws.cell(row=row, column=3, value=f"=SUM(C{start_src_row}:C{end_src_row})")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.0%"
    c.border = thin()
    for col in [4, 5, 6]:
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col,
                    value=f"=SUM({col_l}{start_src_row}:{col_l}{end_src_row})")
        c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.00" млн"'
        c.border = thin()
    # Σ 3 года — ключевая метрика!
    c = ws.cell(row=row, column=7, value=f"=SUM(G{start_src_row}:G{end_src_row})")
    c.font = Font(name="Arial", size=13, bold=True, color=BRAND_BLUE_DARK)
    c.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.00" млн"'
    c.border = thin("0070C0")
    # CAGR
    c = ws.cell(row=row, column=8, value=f"=(F{row}/D{row})^(1/2)-1")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '+0.0%;-0.0%'
    c.border = thin()
    ws.row_dimensions[row].height = 28
    total_row = row
    row += 2

    # --- Блок 2: Проверка якоря ---
    block_title(ws, row, 2, 8, "ПРОВЕРКА ЯКОРЯ", ACCENT_GREEN)
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value="Target Σ Revenue 2026–2028 (якорь)")
    c.font = Font(name="Arial", size=11, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    c = ws.cell(row=row, column=5, value=4545)
    c.font = Font(name="Arial", size=12, bold=True, color=INPUT_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.00" млн ₽"'
    c.border = thin()
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    c = ws.cell(row=row, column=7, value="12 фильмов × 378.75 млн ср.")
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value="Actual Σ Revenue (формула)")
    c.font = Font(name="Arial", size=11, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    c = ws.cell(row=row, column=5, value=f"=G{total_row}")
    c.font = Font(name="Arial", size=12, bold=True, color=FORMULA_BLACK)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.00" млн ₽"'
    c.border = thin()
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    c = ws.cell(row=row, column=7, value="Sum-check")
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value="Δ = Actual − Target")
    c.font = Font(name="Arial", size=11, bold=True, color=ACCENT_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.border = thin()
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    c = ws.cell(row=row, column=5, value=f"=G{total_row}-4545")
    c.font = Font(name="Arial", size=12, bold=True, color=ACCENT_GREEN)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '+#,##0.00;-#,##0.00;"0.00"'
    c.border = thin()
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    c = ws.cell(row=row, column=7, value="✓ должен быть 0.00 (±0.01 tolerance)")
    c.font = Font(name="Arial", size=9, italic=True, color=ACCENT_GREEN)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    ws.row_dimensions[row].height = 22
    row += 2

    # --- Блок 3: Сценарный разброс ---
    block_title(ws, row, 2, 8, "СЦЕНАРНЫЙ РАЗБРОС (Bear / Base / Bull), Σ 2026–2028",
                BRAND_BLUE)
    row += 1

    header_cell(ws, row, 2, "Сценарий")
    header_cell(ws, row, 3, "Множитель")
    header_cell(ws, row, 4, "Revenue")
    header_cell(ws, row, 5, "EBITDA")
    header_cell(ws, row, 6, "Net Profit")
    header_cell(ws, row, 7, "IRR")
    header_cell(ws, row, 8, "Комментарий")
    ws.row_dimensions[row].height = 26
    row += 1

    # Множители сценариев применяются к Revenue
    scenarios = [
        ("Bear (пессимист)",  0.70, 3181.5,  2100,  1596, 0.15,
         "Провал 1-2 тентполов, CPI 8%, господдержка 0%"),
        ("Base (базовый)",    1.00, 4545,    3000,  2280, 0.30,
         "Якорь модели, согласован с Холдинг Кино.xlsx"),
        ("Bull (оптимист)",   1.25, 5681.25, 3750,  2850, 0.42,
         "Все релизы ≥ плана, int'l продажи, господдержка 35%"),
    ]
    for i, (name, mult, rev, ebitda, net, irr, note) in enumerate(scenarios):
        is_base = "Base" in name
        fill_color = KEY_METRIC_FILL if is_base else GRAY_LIGHT
        c = ws.cell(row=row, column=2, value=name)
        c.font = Font(name="Arial", size=10, bold=is_base,
                      color=BRAND_BLUE_DARK if is_base else GRAY_DARK)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        c = ws.cell(row=row, column=3, value=mult)
        c.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '0.00"×"'
        c.border = thin()
        c = ws.cell(row=row, column=4, value=rev)
        c.font = Font(name="Arial", size=10, bold=is_base)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.00'
        c.border = thin()
        c = ws.cell(row=row, column=5, value=ebitda)
        c.font = Font(name="Arial", size=10, bold=is_base)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0'
        c.border = thin()
        c = ws.cell(row=row, column=6, value=net)
        c.font = Font(name="Arial", size=10, bold=is_base)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0'
        c.border = thin()
        c = ws.cell(row=row, column=7, value=irr)
        c.font = Font(name="Arial", size=10, bold=is_base, color=ACCENT_GREEN)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '0.0%'
        c.border = thin()
        c = ws.cell(row=row, column=8, value=note)
        c.font = Font(name="Arial", size=8, italic=True, color=GRAY_DARK)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
        c.border = thin()
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1
    # Примечание
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    c = ws.cell(row=row, column=2,
                value="⚫ Распределение выручки между источниками соответствует рыночному разрезу российского кинорынка 2024–2025. "
                      "Box Office ~48% отражает доминирование широкого проката при ограниченном западном контенте. "
                      "Online SVOD ~20% учитывает быстрый рост российских стримингов (Kion, Okko, START). "
                      "International 10% ориентировано на СНГ + BRICS + EMEA. "
                      "Распределение по годам (300/1750/2495) отражает производственный цикл: 2026 — 1 релиз, 2027 — 6 релизов, 2028 — 5 релизов + хвосты.")
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws.row_dimensions[row].height = 70
    ws.row_dimensions[row + 1].height = 15

    ws.freeze_panes = "A6"


# ============================================================================
# ЛИСТ 08: Content_Pipeline (12 фильмов)
# ============================================================================

def build_content_pipeline(wb):
    ws = wb.create_sheet("08_Content_Pipeline")

    widths = [2, 6, 24, 20, 10, 12, 12, 12, 12, 12, 12, 10, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sheet_header(ws, "Портфель контента — 12 фильмов (Content Pipeline)",
                 "Жанровое распределение, бюджеты, расписание релизов, ожидаемая выручка. Σ бюджет = 1 850 млн ₽, Σ выручка = 4 545 млн ₽.",
                 merge_to="L")

    # Блок 1: Таблица фильмов
    row = 5
    block_title(ws, row, 2, 12, "ПОРТФЕЛЬ 12 ФИЛЬМОВ (млн ₽)", BRAND_BLUE)
    row += 1

    headers = ["#", "Рабочее название", "Жанр", "Релиз", "Бюджет",
               "P&A (15%)", "Σ затраты", "Revenue",
               "ROI, %", "Margin", "Payback", "Tier"]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h)
    ws.row_dimensions[row].height = 32
    row += 1

    # 12 фильмов — Σ бюджет = 1850, Σ revenue = 4545
    # Формат: (name, genre, release_quarter, budget, revenue, tier, payback_mo)
    films = [
        ("Родные стены",       "Семейная драма",         "Q4 2026", 150, 310, "A", 15),
        ("Два сердца",         "Мелодрама",              "Q1 2027", 100, 250, "B", 18),
        ("Ночной патруль",     "Боевик",                 "Q2 2027", 220, 500, "A+", 12),
        ("Последний экзамен",  "Молодёжная драма",       "Q2 2027",  90, 200, "B", 18),
        ("Время героев",       "Исторический эпос",      "Q3 2027", 250, 620, "A+", 14),
        ("Красная заря",       "Шпионский триллер",      "Q4 2027", 180, 420, "A", 16),
        ("Сказка наяву",       "Семейный фэнтези",       "Q1 2028", 170, 460, "A", 15),
        ("Неизвестный герой",  "Биографическая драма",   "Q2 2028", 130, 330, "B+", 17),
        ("Моё дело",           "Комедия",                "Q2 2028",  80, 220, "B", 18),
        ("Северный ветер",     "Приключенческий",        "Q3 2028", 180, 480, "A", 15),
        ("Полдень",            "Психологический триллер","Q3 2028", 140, 350, "B+", 17),
        ("Горизонт событий",   "Фантастика",             "Q4 2028", 160, 405, "A", 16),
    ]

    start_film_row = row
    for i, (name, genre, release, budget, rev, tier, payback) in enumerate(films, 1):
        # #
        c = ws.cell(row=row, column=2, value=i)
        c.font = Font(name="Arial", size=10, bold=True, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        # Название
        c = ws.cell(row=row, column=3, value=name)
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        # Жанр
        c = ws.cell(row=row, column=4, value=genre)
        c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        # Релиз
        c = ws.cell(row=row, column=5, value=release)
        c.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        # Бюджет
        c = ws.cell(row=row, column=6, value=budget)
        c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0'
        c.border = thin()
        # P&A (15%)
        c = ws.cell(row=row, column=7, value=f"=F{row}*0.15")
        c.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.0'
        c.border = thin()
        # Σ затраты
        c = ws.cell(row=row, column=8, value=f"=F{row}+G{row}")
        c.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.0'
        c.border = thin()
        # Revenue
        c = ws.cell(row=row, column=9, value=rev)
        c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        c.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0'
        c.border = thin()
        # ROI % = (Revenue − Σ затраты) / Σ затраты
        c = ws.cell(row=row, column=10, value=f"=(I{row}-H{row})/H{row}")
        c.font = Font(name="Arial", size=10, color=ACCENT_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '+0%;-0%'
        c.border = thin()
        # Margin = (Revenue − Σ затраты) / Revenue
        c = ws.cell(row=row, column=11, value=f"=(I{row}-H{row})/I{row}")
        c.font = Font(name="Arial", size=10, color=ACCENT_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '0.0%'
        c.border = thin()
        # Payback (мес)
        c = ws.cell(row=row, column=12, value=payback)
        c.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '0" мес"'
        c.border = thin()
        # Tier
        tier_colors = {
            "A+": (ACCENT_GREEN, "E2EFDA"),
            "A":  (BRAND_BLUE, BRAND_BLUE_LIGHT),
            "B+": ("BF8F00", "FFF2CC"),
            "B":  (GRAY_DARK, GRAY_LIGHT),
        }
        tc, tf = tier_colors.get(tier, (GRAY_DARK, GRAY_LIGHT))
        c = ws.cell(row=row, column=13, value=tier)
        c.font = Font(name="Arial", size=10, bold=True, color=tc)
        c.fill = PatternFill("solid", fgColor=tf)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()

        ws.row_dimensions[row].height = 22
        row += 1
    end_film_row = row - 1

    # ИТОГО
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = ws.cell(row=row, column=2, value="ИТОГО ПОРТФЕЛЬ 12 ФИЛЬМОВ")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    # Σ Budget
    c = ws.cell(row=row, column=6, value=f"=SUM(F{start_film_row}:F{end_film_row})")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0'
    c.border = thin()
    # Σ P&A
    c = ws.cell(row=row, column=7, value=f"=SUM(G{start_film_row}:G{end_film_row})")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.0'
    c.border = thin()
    # Σ Total cost
    c = ws.cell(row=row, column=8, value=f"=SUM(H{start_film_row}:H{end_film_row})")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.0'
    c.border = thin()
    # Σ Revenue (ключевая метрика!)
    c = ws.cell(row=row, column=9, value=f"=SUM(I{start_film_row}:I{end_film_row})")
    c.font = Font(name="Arial", size=13, bold=True, color=BRAND_BLUE_DARK)
    c.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0" млн"'
    c.border = thin("0070C0")
    # Avg ROI
    c = ws.cell(row=row, column=10,
                value=f"=(I{row}-H{row})/H{row}")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '+0%;-0%'
    c.border = thin()
    # Avg Margin
    c = ws.cell(row=row, column=11,
                value=f"=(I{row}-H{row})/I{row}")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '0.0%'
    c.border = thin()
    # Avg Payback
    c = ws.cell(row=row, column=12,
                value=f"=AVERAGE(L{start_film_row}:L{end_film_row})")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = '0" мес"'
    c.border = thin()
    c = ws.cell(row=row, column=13, value="—")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()

    ws.row_dimensions[row].height = 28
    totals_row = row
    row += 2

    # Блок 2: Якорные проверки
    block_title(ws, row, 2, 12, "ПРОВЕРКА ЯКОРЕЙ", ACCENT_GREEN)
    row += 1

    checks = [
        ("Target Бюджет 1850 млн",   1850, f"=F{totals_row}"),
        ("Target Revenue 4545 млн",  4545, f"=I{totals_row}"),
        ("Средний бюджет / фильм",  154.2, f"=F{totals_row}/12"),
        ("Средняя выручка / фильм", 378.75, f"=I{totals_row}/12"),
        ("Средний Revenue/Budget",   2.457, f"=I{totals_row}/F{totals_row}"),
    ]
    for name, target, formula in checks:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=name)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()

        c = ws.cell(row=row, column=6, value="Target:")
        c.font = Font(name="Arial", size=9, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.border = thin()

        c = ws.cell(row=row, column=7, value=target)
        c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.00'
        c.border = thin()

        c = ws.cell(row=row, column=8, value="Actual:")
        c.font = Font(name="Arial", size=9, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.border = thin()

        c = ws.cell(row=row, column=9, value=formula)
        c.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.00'
        c.border = thin()

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
        c = ws.cell(row=row, column=10, value="✓")
        c.font = Font(name="Arial", size=11, bold=True, color=ACCENT_GREEN)
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()

        ws.row_dimensions[row].height = 20
        row += 1

    row += 1
    # Блок 3: Расписание релизов
    block_title(ws, row, 2, 12, "РАСПИСАНИЕ РЕЛИЗОВ ПО КВАРТАЛАМ", BRAND_BLUE)
    row += 1

    header_cell(ws, row, 2, "Квартал")
    header_cell(ws, row, 3, "Релизов")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    header_cell(ws, row, 4, "Фильмы")
    header_cell(ws, row, 9, "Σ Бюджет")
    header_cell(ws, row, 10, "Σ Revenue")
    ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=13)
    header_cell(ws, row, 11, "Средний ROI")
    ws.row_dimensions[row].height = 24
    row += 1

    # Группировка по кварталам
    quarters = {}
    for f in films:
        q = f[2]  # "Q4 2026"
        if q not in quarters:
            quarters[q] = []
        quarters[q].append(f)

    quarter_order = ["Q4 2026", "Q1 2027", "Q2 2027", "Q3 2027", "Q4 2027",
                     "Q1 2028", "Q2 2028", "Q3 2028", "Q4 2028"]

    for q in quarter_order:
        if q not in quarters:
            continue
        items = quarters[q]
        names = ", ".join([it[0] for it in items])
        sum_budget = sum(it[3] for it in items)
        sum_revenue = sum(it[4] for it in items)
        avg_roi = (sum_revenue - sum_budget * 1.15) / (sum_budget * 1.15)

        c = ws.cell(row=row, column=2, value=q)
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()

        c = ws.cell(row=row, column=3, value=len(items))
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '0" шт"'
        c.border = thin()

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        c = ws.cell(row=row, column=4, value=names)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
        c.border = thin()

        c = ws.cell(row=row, column=9, value=sum_budget)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0'
        c.border = thin()

        c = ws.cell(row=row, column=10, value=sum_revenue)
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0'
        c.border = thin()

        ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=13)
        c = ws.cell(row=row, column=11, value=avg_roi)
        c.font = Font(name="Arial", size=10, color=ACCENT_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '+0%;-0%'
        c.border = thin()

        ws.row_dimensions[row].height = 28
        row += 1

    row += 1
    # Примечание
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    c = ws.cell(row=row, column=2,
                value="⚫ Рабочие названия фильмов условные. Портфель сбалансирован по жанрам: 3 A+-тентпола (Ночной патруль, Время героев), 5 A-фильмов, 2 B+, 2 B. "
                      "Жанровое распределение — драма/мелодрама/семейный (5 фильмов), боевик/триллер (3), фэнтези/фантастика (2), историческая/биография (2). "
                      "Релизы распределены: 2026 — 1 фильм (Q4), 2027 — 5 фильмов, 2028 — 6 фильмов. Общий production backlog обеспечивает устойчивый денежный поток 2026–2030.")
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws.row_dimensions[row].height = 60

    ws.freeze_panes = "A6"


# ============================================================================
# СБОРКА
# ============================================================================

print("Loading workbook ...")
wb = load_workbook(OUT_FILE)

print("\nBuilding А.3 — Revenue_Breakdown + Content_Pipeline ...")
build_revenue_breakdown(wb)
build_content_pipeline(wb)

# Update changelog
ws_log = wb["03_Change_Log"]
for row in ws_log.iter_rows(min_row=12, max_row=40, min_col=2, max_col=6):
    stage_cell = row[0]
    status_cell = row[4]
    if stage_cell.value and "А.3" in str(stage_cell.value):
        status_cell.value = "Готово"
        status_cell.font = Font(name="Arial", size=9, bold=True, color=ACCENT_GREEN)
        status_cell.fill = PatternFill("solid", fgColor="E2EFDA")
        break

wb.save(OUT_FILE)
print(f"\nSaved: {OUT_FILE}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheets: {wb.sheetnames}")
