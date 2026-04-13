import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/home/user/TrendStudio-Holding/audit_scripts/audit_public_v1_findings.xlsx'
wb = openpyxl.load_workbook(OUT)
thin = Side(style='thin')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="003366")
wrap = Alignment(wrap_text=True, vertical="top")

def add_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = wrap; c.border = border
    ws.freeze_panes = "A2"

def add_row(ws, row_num, data):
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.alignment = wrap; c.border = border

# === SHEET 6: By Design ===
ws6 = wb.create_sheet("By Design")
add_header(ws6, ["#","Item","Rationale","Risk","Mitigation"])
by_design = [
[1,"Anchor-based model (97.5% constants)","Ensures iron convergence on NDP=3000","Analyst sees 'static spreadsheet'","v2.0: convert to formula-based"],
[2,"Dual-metric GAAP/NDP","GAAP EBITDA=2167 vs NDP=3000 serve different audiences","Confusion if not labeled","Prefix all EBITDA with (GAAP) or (NDP)"],
[3,"W3 vs W5 waterfall difference","W3 (Public) conservative, W5 (Internal) full terms","IRR gap 4.66pp","Disclosed in manifest"],
[4,"MC mean < deterministic IRR","Hit_rate, frontload pattern reduce expected returns","P(IRR>hurdle)=13.6%","Prominent disclosure, narrative"],
[5,"Revenue pattern A+C 33/45/9/13","Front-loaded for time-value advantage","Sensitive to front-end shocks","Bear scenario documented"],
[6,"Back-solved COGS/OpEx split","COGS 2008.8 + OpEx 368.8 = 2377.6 (Revenue-EBITDA)","Visible as fitting","Document as target margin approach"],
]
for i, d in enumerate(by_design, 2):
    add_row(ws6, i, d)
for col in range(1,6):
    ws6.column_dimensions[get_column_letter(col)].width = [5,35,40,35,35][col-1]

# === SHEET 7: Stress Tests ===
ws7 = wb.create_sheet("Stress Tests")
add_header(ws7, ["#","Scenario","Shock","NDP Impact","IRR Impact","BS Status","Covenant"])
stress = [
[1,"Recession","Revenue -20%","NDP ~2400","IRR ~15.14%","Balanced","Within limits"],
[2,"Inflation/FOT","FOT +15%","NDP ~2850","IRR ~18.5%","Balanced","Within limits"],
[3,"Rate Shock","WACC +3pp","NPV -25%","IRR unchanged, NPV drops","Balanced","N/A"],
[4,"Regulatory","NDP -50%","NDP 1500","IRR ~2%","Cash tight","Breach risk"],
[5,"Box Office Fail","3/12 films fail","NDP ~2250","IRR ~12%","Balanced","Within limits"],
[6,"FX Shock","Ruble -30%","Rev+10%/Cost+20%","IRR ~16%","Balanced","Monitor"],
]
for i, s in enumerate(stress, 2):
    add_row(ws7, i, s)
for col in range(1,8):
    ws7.column_dimensions[get_column_letter(col)].width = [5,18,18,18,18,14,14][col-1]

# === SHEET 8: Peer Benchmarks ===
ws8 = wb.create_sheet("Peer Benchmarks")
add_header(ws8, ["Company","EV/Revenue","EV/EBITDA","Source","Notes"])
peers = [
["Kinopoisk","0.80","5.0","Model (no citation)","Sber subsidiary"],
["Okko","0.81","5.2","Model (no citation)","Acquired by Sber 2023"],
["ivi","0.80","4.8","Model (no citation)","Independent OTT"],
["START","0.80","5.5","Model (no citation)","MTS subsidiary"],
["Premier","0.81","5.0","Model (no citation)","Gazprom-Media"],
["Mosfilm","1.05","6.0","Model (no citation)","State studio"],
["","","","",""],
["AUDIT NOTE:","","","","5 of 6 peers show suspiciously uniform EV/Revenue (0.80-0.81). No source citations. DD team will request Pitchbook/Mergermarket verification."],
]
for i, p in enumerate(peers, 2):
    add_row(ws8, i, p)
for col in range(1,6):
    ws8.column_dimensions[get_column_letter(col)].width = [18,12,12,25,40][col-1]

# === SHEET 9: Script Results ===
ws9 = wb.create_sheet("Script Results")
add_header(ws9, ["Script","Block","Purpose","Result","Key Finding"])
scripts = [
["step0_map_public.py","A","Map Public xlsx structure","42 sheets, 226 formulas, 8794 values","2.5% formula density"],
["step0_map_internal.py","A","Map Internal xlsx structure","42 sheets, 230 formulas, 9160 values","Title says 'Public' on Internal"],
["priority1_leakage_scan.py","C","Leakage scan","23 findings (9 CRIT, 9 HIGH, 5 MED)","W5 V-D waterfall exposed"],
["priority2_anchor_reconcile.py","C","Anchor reconciliation","10/10 PASS, 1 INFO","All shared anchors aligned"],
["priority3_xml_integrity.py","A","OOXML integrity","Clean - no corruption","CalcChain 211 entries on sheets 4-8"],
["extract_all_financials.py","B","Extract financial data","10 sheets extracted","Revenue 4545, EBITDA 2167.4, NDP 3000"],
["block_g_pipeline_audit.py","G","Pipeline code quality","16621 LOC, 0 TODO, 0 security issues","100% type hints on generators"],
["Pipeline pytest","G","Run test suite","323/328 pass, 5 skip, 0 fail","6.29s runtime, fully deterministic"],
["Pipeline verify.py","G","Verification script","6/7 pass, 1 false-negative","Bug: hardcoded count 14 should be 18"],
["Pipeline verify_full.py","G","Full P5 verification","29/32 pass, 3 environmental","Missing bundle files in clean clone"],
]
for i, s in enumerate(scripts, 2):
    add_row(ws9, i, s)
for col in range(1,6):
    ws9.column_dimensions[get_column_letter(col)].width = [30,8,30,35,40][col-1]

# === SHEET 10: Arguments Ranking ===
ws10 = wb.create_sheet("Arguments Ranking")
add_header(ws10, ["Rank","Argument","Strength","Sources","Uncertainty","Category"])
args = [
[1,"NDP 3000M anchor with tolerance ±1%","STRONG","Pipeline-verified, 323 tests","Low","Structural"],
[2,"Revenue 4545M from 5 segments + 12 films","STRONG","Pipeline bottom-up","Medium","Revenue"],
[3,"Pipeline deterministic build (seed=42)","STRONG","ADR-008, test_03","Low","Technical"],
[4,"Government support 25% of budget","MODERATE","Fond Kino policy","Medium","Revenue"],
[5,"EBITDA margin 47.7% GAAP","MODERATE","Anchor-based, not derived","High","Profitability"],
[6,"IRR 20.09% (W3 deterministic)","MODERATE","Verified by 3 methods","Medium","Returns"],
[7,"Peer comps median EV 7.5B","WEAK","No source citations","Very High","Valuation"],
[8,"MC Mean IRR 11.44%","WEAK","Below hurdle 18%","High","Returns"],
[9,"DCF EV 1.8B vs Comps 7.5B","WEAK","6x unexplained gap","Very High","Valuation"],
[10,"Exit via IPO 2029-2030","WEAK","Russia IPO market uncertain","Very High","Exit"],
]
for i, a in enumerate(args, 2):
    add_row(ws10, i, a)
    c = ws10.cell(row=i, column=3)
    colors = {"STRONG":"00AA00","MODERATE":"FF8C00","WEAK":"FF0000"}
    c.font = Font(bold=True, color=colors.get(a[2],"000000"))
for col in range(1,7):
    ws10.column_dimensions[get_column_letter(col)].width = [6,40,12,25,14,14][col-1]

# === SHEET 11: Red Team Attack ===
ws11 = wb.create_sheet("Red Team Attack")
add_header(ws11, ["#","VC Kill Question","Why Dangerous","Recommended Answer","Severity"])
redteam = [
[1,"Why is 97% of the model hardcoded constants?","Instant credibility loss","'Anchor-based for version control; v2.0 will have live formulas'","CRITICAL"],
[2,"Your MC shows only 13.6% chance of beating hurdle. Why should I invest?","Kills the deal","'Deterministic base beats hurdle by 2.09pp; MC reflects hit-rate variance which is mitigated by 12-film diversification'","CRITICAL"],
[3,"Why is there a 6x gap between DCF and Comps valuation?","Questions all valuations","'DCF is conservative floor (5yr horizon, 19% WACC); Comps reflect market premiums for content IP. We position DCF as floor, not target.'","HIGH"],
[4,"Your internal model shows better returns (24.75% vs 20.09%). What are you hiding?","W5 V-D leaked in Public xlsx","IMMEDIATE: scrub leakage. Then: 'W3 is the offered structure; W5 was a negotiation benchmark.'","CRITICAL"],
[5,"Where did you get these peer multiples? They look fabricated.","5/6 at 0.80-0.81 EV/Rev","'We will provide Pitchbook citations in updated CIM.'","HIGH"],
[6,"What happens if Yandex/VK launch a competing studio?","Market risk","'Our 12-film slate is genre-diversified; our govt relationships provide 25% cost subsidy competitors lack.'","MEDIUM"],
[7,"FOT is 87% of OpEx. What's your bus factor?","Team concentration","'Key person insurance in place; 3 senior producers can independently run slates.'","MEDIUM"],
[8,"Can you actually IPO in Russia in 2029?","Exit uncertainty","'IPO is one of 5 exit routes (15-25% probability each). Strategic M&A and Secondary are alternatives.'","HIGH"],
[9,"Your D&A jumps from 3M to 500M in one year. Explain.","DCF integrity","'2029 marks transition from production to exploitation phase. D&A = content amortization of 1.85B library.'","HIGH"],
[10,"If NDP is 3000 but EBITDA is 2167, where does the extra 833 come from?","Bridge transparency","'NDP bridge: EBITDA 2167 + Producer equity 600 + WC/Gov 233 = 3000. Documented in P&L reconciliation.'","MEDIUM"],
]
for i, r in enumerate(redteam, 2):
    add_row(ws11, i, r)
for col in range(1,6):
    ws11.column_dimensions[get_column_letter(col)].width = [5,45,35,55,12][col-1]

wb.save(OUT)
print(f"Part 2 done: sheets 6-11 added. Total sheets: {len(wb.sheetnames)}")
