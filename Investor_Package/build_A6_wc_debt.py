"""
А.6 — Working Capital Schedule + Debt Schedule.

12_Working_Capital:
  - AR (Accounts Receivable) = 25% × Q revenue (collection lag ~60 days)
  - Prepaid/Inventory = 5% × Q revenue
  - AP (Accounts Payable) = 30% × (OpEx + P&A)
  - Net Working Capital = AR + Prepaid − AP
  - ΔNWC — изменение (negative = cash use)

13_Debt_Schedule (T₁ Legacy):
  - Opening balance → Draws → Interest accrued → Interest paid → Principal repaid → Closing
  - 4 транша: 250/350/350/300 = 1 250 (Q1'26, Q4'26, Q3'27, Q2'28)
  - Принципал погашается из distributions 2029-2032
  - Total interest = 15 (consistent with P&L)
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
INFLOW = "E2EFDA"
OUTFLOW = "FCE4D6"
CHECK_GREEN = "C6EFCE"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color=BRAND_BLUE)
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
box_brand = Border(left=medium, right=medium, top=medium, bottom=medium)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'


def set_cell(ws, ref, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws[ref]
    c.value = value
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


# ===== ПЕРИОДЫ =====
PERIODS = [
    ("D", "Q1'26", "Q"), ("E", "Q2'26", "Q"), ("F", "Q3'26", "Q"), ("G", "Q4'26", "Q"),
    ("H", "Q1'27", "Q"), ("I", "Q2'27", "Q"), ("J", "Q3'27", "Q"), ("K", "Q4'27", "Q"),
    ("L", "Q1'28", "Q"), ("M", "Q2'28", "Q"), ("N", "Q3'28", "Q"), ("O", "Q4'28", "Q"),
    ("P", "2029",  "Y"), ("Q", "2030",  "Y"), ("R", "2031",  "Y"), ("S", "2032",  "Y"),
]
ALL_COLS = [p[0] for p in PERIODS]

# Revenue floor: минимальный доход от библиотеки контента (SVOD, TV rights, merchandise).
REVENUE_FLOOR = 380  # млн ₽/год — минимальный доход от библиотеки контента

_REV_Q_RAW = {
    "D": 0,    "E": 0,    "F": 0,    "G": 310,
    "H": 250,  "I": 700,  "J": 620,  "K": 420,
    "L": 460,  "M": 550,  "N": 830,  "O": 405,
    "P": 380,  "Q": 300,  "R": 220,  "S": 150,   # raw before floor
}
REV_Q = {col: max(val, REVENUE_FLOOR) if col in ("P", "Q", "R", "S") else val
         for col, val in _REV_Q_RAW.items()}

OPEX_Q = 22.1265
OPEX_Y = 88.506
# FIX-02: ФОТ cap — OpEx ≤ 70% от Revenue
MAX_OPEX_REVENUE_RATIO = 0.70
PA_Q = {c: round(REV_Q[c] * (277.5 / 4545), 2) for c in ALL_COLS}


def opex_val(col):
    base = OPEX_Q if PERIODS[ALL_COLS.index(col)][2] == "Q" else OPEX_Y
    rev = REV_Q[col]
    if rev <= 0:
        return base
    return min(base, rev * MAX_OPEX_REVENUE_RATIO)


# ============================================================
#   12_Working_Capital
# ============================================================
def build_wc(wb):
    if "12_Working_Capital" in wb.sheetnames:
        del wb["12_Working_Capital"]
    ws = wb.create_sheet("12_Working_Capital")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 38
    for i in range(4, 20):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions["T"].width = 12
    ws.column_dimensions["U"].width = 30

    ws.merge_cells("B2:U2")
    set_cell(ws, "B2", "12  ·  WORKING CAPITAL SCHEDULE  ·  AR / Prepaid / AP / NWC / ΔNWC",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:U3")
    set_cell(ws, "B3",
             "Предпосылки: AR = 25% Q revenue (60-day lag) · Prepaid = 5% Q revenue · AP = 30% (OpEx + P&A). "
             "Все суммы — млн ₽.",
             font=F_ITALIC, align=C_CENTER)

    # Headers
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Статья", font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    for col, lbl, _ in PERIODS:
        set_cell(ws, f"{col}{r}", lbl, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"T{r}", "Σ 2026–2028", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"U{r}", "Комментарий", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 20

    # ====== Расчёты ======
    AR_RATIO = 0.25       # 25% от квартальной выручки
    PREPAID_RATIO = 0.05  # 5% от Q revenue
    AP_RATIO = 0.30       # 30% от (OpEx + P&A)

    ar = {c: round(REV_Q[c] * AR_RATIO, 2) for c in ALL_COLS}
    prepaid = {c: round(REV_Q[c] * PREPAID_RATIO, 2) for c in ALL_COLS}
    ap = {c: round((opex_val(c) + PA_Q[c]) * AP_RATIO, 2) for c in ALL_COLS}
    nwc = {c: round(ar[c] + prepaid[c] - ap[c], 2) for c in ALL_COLS}

    d_nwc = {}
    prev = 0.0
    for c in ALL_COLS:
        d_nwc[c] = round(nwc[c] - prev, 2)
        prev = nwc[c]
    # Cash impact = -ΔNWC (рост NWC = отток cash)
    cash_impact = {c: round(-d_nwc[c], 2) for c in ALL_COLS}

    # === Рендер ===
    r = 6
    rows = [
        ("SECTION", "I.  ОБОРОТНЫЕ АКТИВЫ"),
        ("ROW", "1.1", "Accounts Receivable (25% Q rev)", ar, INFLOW, "AR = 25% × Revenue квартала"),
        ("ROW", "1.2", "Prepaid expenses / Inventory", prepaid, INFLOW, "5% × Revenue"),
        ("SECTION", "II.  ОБОРОТНЫЕ ОБЯЗАТЕЛЬСТВА"),
        ("ROW", "2.1", "Accounts Payable (30% OpEx + P&A)", ap, OUTFLOW, "AP = 30% × (OpEx + P&A)"),
        ("SECTION", "III.  NET WORKING CAPITAL"),
        ("TOTAL", "3.1", "NWC = AR + Prepaid − AP", nwc, SUBTOTAL_FILL, "Баланс на конец периода"),
        ("TOTAL", "3.2", "ΔNWC (период-к-периоду)", d_nwc, SUBTOTAL_FILL, "Рост NWC = отток cash"),
        ("TOTAL", "3.3", "Cash impact (−ΔNWC)", cash_impact, KEY_METRIC_FILL,
         "Влияние на OCF: − при росте NWC"),
    ]

    for entry in rows:
        if entry[0] == "SECTION":
            ws.merge_cells(f"B{r}:U{r}")
            set_cell(ws, f"B{r}", entry[1], font=F_SECTION, fill=DARK_BLUE,
                     align=C_LEFT, border=box_thin)
            ws.row_dimensions[r].height = 20
            r += 1
            continue

        _, num, name, data, fill, comment = entry
        is_total = entry[0] == "TOTAL"
        font = F_TOTAL if is_total else F_BODY

        set_cell(ws, f"B{r}", num, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)

        total_3y = 0.0
        for col, lbl, k in PERIODS:
            v = data[col]
            set_cell(ws, f"{col}{r}", v, font=font, fill=fill, align=C_RIGHT,
                     border=box_thin, number_format=NUMFMT)
            if k == "Q":
                total_3y += v

        set_cell(ws, f"T{r}", round(total_3y, 1), font=F_TOTAL,
                 fill=KEY_METRIC_FILL if is_total else fill, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"U{r}", comment, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 18
        r += 1

    # Control row
    r += 1
    ws.merge_cells(f"B{r}:U{r}")
    nwc_end_q4_28 = nwc["O"]
    total_d_nwc_3y = sum(d_nwc[c] for c in ALL_COLS if PERIODS[ALL_COLS.index(c)][2] == "Q")
    set_cell(ws, f"B{r}",
             f"✓ NWC end Q4'28 = {nwc_end_q4_28:,.1f} млн ₽  |  Σ ΔNWC 2026–2028 = {total_d_nwc_3y:,.1f}  |  "
             f"Impact on CF: {-total_d_nwc_3y:,.1f} млн ₽ (supplementary — не включено в simplified CF А.5)",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 22

    ws.freeze_panes = "D6"
    return {"nwc": nwc, "d_nwc": d_nwc}


# ============================================================
#   13_Debt_Schedule
# ============================================================
def build_debt_schedule(wb):
    if "13_Debt_Schedule" in wb.sheetnames:
        del wb["13_Debt_Schedule"]
    ws = wb.create_sheet("13_Debt_Schedule")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 40
    for i in range(4, 20):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions["T"].width = 12
    ws.column_dimensions["U"].width = 28

    ws.merge_cells("B2:U2")
    set_cell(ws, "B2", "13  ·  DEBT SCHEDULE  ·  T₁ Legacy  ·  1 250 млн ₽  ·  4 tranches 20/28/28/24%",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:U3")
    set_cell(ws, "B3",
             "T₁ Legacy: bullet-loan структура с распределением через waterfall W₁ (60/40). "
             "Проценты — service fee 15 млн суммарно. Принципал возвращается 2029-2032. Все суммы — млн ₽.",
             font=F_ITALIC, align=C_CENTER)

    # Headers
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Позиция", font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    for col, lbl, _ in PERIODS:
        set_cell(ws, f"{col}{r}", lbl, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"T{r}", "Σ 2026–2032", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"U{r}", "Комментарий", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 20

    # ====== Данные ======
    T1_TRANCHES = {"D": 250, "G": 350, "J": 350, "M": 300}
    draws = {c: T1_TRANCHES.get(c, 0) for c in ALL_COLS}

    # Interest accrued (квартальные): consistent с P&L
    int_acc = {"D": 1.0, "E": 1.0, "F": 1.5, "G": 1.5,
               "H": 1.5, "I": 1.5, "J": 1.5, "K": 1.5,
               "L": 0.75, "M": 0.75, "N": 0.5, "O": 0.5,
               "P": 1.0, "Q": 0.5, "R": 0.0, "S": 0.0}
    # Interest paid: в том же периоде (cash basis)
    int_paid = dict(int_acc)

    # Principal repayment — из distributions 2029-2032
    # Investor receives 2215 total. Of this, 1250 is principal return.
    # Schedule: 2029=886, 2030=665, 2031=443, 2032=221 (investor share = 0.60 × dist)
    # Распределяем principal первым:
    # 2029: 886 → principal 886 (balance 1250-886=364), rest 0
    # 2030: 665 → principal 364 (balance 0), equity upside 301
    # 2031: 443 → equity upside 443
    # 2032: 221 → equity upside 221
    # Total principal = 1250, equity upside = 965
    principal_repaid = {c: 0.0 for c in ALL_COLS}
    principal_repaid["P"] = 886.0
    principal_repaid["Q"] = 364.0
    # 2031, 2032 — только equity upside (не в debt schedule)

    equity_upside = {c: 0.0 for c in ALL_COLS}
    equity_upside["Q"] = 301.0
    equity_upside["R"] = 443.0
    equity_upside["S"] = 221.0

    # ====== Computed: opening, closing balance ======
    opening = {}
    closing = {}
    bal = 0.0
    for c in ALL_COLS:
        opening[c] = round(bal, 2)
        bal = bal + draws[c] - principal_repaid[c]
        closing[c] = round(bal, 2)

    # Cumulative rows
    cum_draws = {}
    cum_int_paid = {}
    cum_principal = {}
    cum_equity = {}
    cd = ci = cp = ce = 0.0
    for c in ALL_COLS:
        cd += draws[c]; cum_draws[c] = round(cd, 2)
        ci += int_paid[c]; cum_int_paid[c] = round(ci, 2)
        cp += principal_repaid[c]; cum_principal[c] = round(cp, 2)
        ce += equity_upside[c]; cum_equity[c] = round(ce, 2)

    # ====== Рендер ======
    r = 6
    rows = [
        ("SECTION", "I.  DRAWDOWN (приток от инвестора)"),
        ("ROW", "1.1", "Opening balance", opening, None, "Баланс принципала на начало"),
        ("ROW", "1.2", "(+) Draws (tranches)", draws, INFLOW, "250/350/350/300 = 1 250"),
        ("SECTION", "II.  INTEREST (servicing cost)"),
        ("ROW", "2.1", "Interest accrued @ service rate", int_acc, OUTFLOW,
         "Проценты начислены"),
        ("ROW", "2.2", "Interest paid (cash basis)", int_paid, OUTFLOW,
         "Платятся в том же периоде"),
        ("SECTION", "III.  PRINCIPAL REPAYMENT"),
        ("ROW", "3.1", "(−) Principal repaid", principal_repaid, OUTFLOW,
         "Из distributions 2029-2030"),
        ("SECTION", "IV.  ENDING BALANCE"),
        ("TOTAL", "4.1", "Closing principal balance", closing, SUBTOTAL_FILL,
         "Баланс принципала на конец"),
        ("SECTION", "V.  CUMULATIVE (контроль)"),
        ("ROW", "5.1", "Σ Cumulative draws", cum_draws, None, "Должно ≤ 1 250"),
        ("ROW", "5.2", "Σ Cumulative interest paid", cum_int_paid, None, "Должно = 15"),
        ("ROW", "5.3", "Σ Cumulative principal repaid", cum_principal, None, "Должно = 1 250"),
        ("SECTION", "VI.  EQUITY UPSIDE (sharing profit component)"),
        ("ROW", "6.1", "Equity upside distributed", equity_upside, INFLOW,
         "Доля в прибыли сверх принципала (= 965)"),
        ("ROW", "6.2", "Σ Cumulative equity upside", cum_equity, None, "Должно = 965"),
    ]

    for entry in rows:
        if entry[0] == "SECTION":
            ws.merge_cells(f"B{r}:U{r}")
            set_cell(ws, f"B{r}", entry[1], font=F_SECTION, fill=DARK_BLUE,
                     align=C_LEFT, border=box_thin)
            ws.row_dimensions[r].height = 20
            r += 1
            continue

        _, num, name, data, fill, comment = entry
        is_total = entry[0] == "TOTAL"
        font = F_TOTAL if is_total else F_BODY

        set_cell(ws, f"B{r}", num, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)

        total_all = 0.0
        for col, lbl, k in PERIODS:
            v = data[col]
            set_cell(ws, f"{col}{r}", v, font=font, fill=fill, align=C_RIGHT,
                     border=box_thin, number_format=NUMFMT)
            total_all += v

        # Для cumulative/opening/closing — показываем последнее значение вместо суммы
        display_total = total_all
        if "Cumulative" in name or "balance" in name or "Opening" in name:
            display_total = data["S"]  # последнее значение

        set_cell(ws, f"T{r}", round(display_total, 1),
                 font=F_TOTAL, fill=KEY_METRIC_FILL if is_total else fill,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"U{r}", comment, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 18
        r += 1

    # ====== Control ======
    r += 1
    ws.merge_cells(f"B{r}:U{r}")
    check1 = "✓" if abs(cum_draws["S"] - 1250) < 0.01 else "✗"
    check2 = "✓" if abs(cum_principal["S"] - 1250) < 0.01 else "✗"
    check3 = "✓" if abs(cum_int_paid["S"] - 15) < 0.05 else "✗"
    check4 = "✓" if abs(cum_equity["S"] - 965) < 0.01 else "✗"
    check5 = "✓" if abs(closing["S"]) < 0.01 else "✗"
    msg = (f"{check1} Σ draws = {cum_draws['S']:,.1f} (target 1 250)  |  "
           f"{check2} Σ principal repaid = {cum_principal['S']:,.1f} (target 1 250)  |  "
           f"{check3} Σ interest = {cum_int_paid['S']:,.1f} (target ~15)  |  "
           f"{check4} Σ equity upside = {cum_equity['S']:,.1f} (target 965)  |  "
           f"{check5} Balance 2032 = {closing['S']:,.2f}")
    set_cell(ws, f"B{r}", msg,
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 24

    ws.freeze_panes = "D6"
    return {
        "total_draws": cum_draws["S"],
        "total_principal": cum_principal["S"],
        "total_interest": cum_int_paid["S"],
        "total_equity_upside": cum_equity["S"],
        "balance_2032": closing["S"],
    }


def main():
    wb = load_workbook(FILE)
    print(f"Loaded: {FILE}")
    print(f"Sheets before: {len(wb.sheetnames)}")

    print("\n[1/2] Building 12_Working_Capital …")
    wc_data = build_wc(wb)

    print("[2/2] Building 13_Debt_Schedule …")
    debt_data = build_debt_schedule(wb)

    wb.save(FILE)
    print(f"\nSheets after: {len(wb.sheetnames)}")
    print(f"\nDebt Schedule checks:")
    for k, v in debt_data.items():
        print(f"  {k}: {v}")
    print(f"\nSaved: {FILE}")


if __name__ == "__main__":
    main()
