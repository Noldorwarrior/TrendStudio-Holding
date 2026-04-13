"""
А.7 — Investment_Inflow + Use_of_Funds + CAPEX_Schedule.

14_Investment_Inflow:
  - 4 транша T₁ Legacy: 250/350/350/300 = 1 250
  - Milestones, cumulative inflow, utilization %

15_Use_of_Funds:
  - Распределение 1 250 млн по 5 категориям
  - Pre-production 5%, Principal Photography 65%, Post-production 20%,
    Contingency 10%. Σ = 100% = 1 250

16_CAPEX_Schedule:
  - 12 фильмов × timeline по кварталам
  - Для каждого: pre-prod (T-3Q), prod (T-2Q, T-1Q), post (T-0Q)
  - Total production CAPEX = 1 850 (consistent с 10_Cash_Flow)
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
INFLOW = "E2EFDA"
OUTFLOW = "FCE4D6"
CHECK_GREEN = "C6EFCE"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color=BRAND_BLUE)
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
box_brand = Border(left=medium, right=medium, top=medium, bottom=medium)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'
PCTFMT = '0.0"%"'


def set_cell(ws, ref, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws[ref]
    c.value = value
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


PERIODS = [
    ("D", "Q1'26", "Q"), ("E", "Q2'26", "Q"), ("F", "Q3'26", "Q"), ("G", "Q4'26", "Q"),
    ("H", "Q1'27", "Q"), ("I", "Q2'27", "Q"), ("J", "Q3'27", "Q"), ("K", "Q4'27", "Q"),
    ("L", "Q1'28", "Q"), ("M", "Q2'28", "Q"), ("N", "Q3'28", "Q"), ("O", "Q4'28", "Q"),
]  # А.7 — только 12 кварталов 2026-2028
ALL_COLS = [p[0] for p in PERIODS]


# ============================================================
#   14_Investment_Inflow
# ============================================================
def build_inflow(wb):
    if "14_Investment_Inflow" in wb.sheetnames:
        del wb["14_Investment_Inflow"]
    ws = wb.create_sheet("14_Investment_Inflow")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 40
    for i in range(4, 16):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions["P"].width = 14
    ws.column_dimensions["Q"].width = 30

    # Title
    ws.merge_cells("B2:Q2")
    set_cell(ws, "B2",
             "14  ·  INVESTMENT INFLOW SCHEDULE  ·  T₁ Legacy Tranche Structure",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:Q3")
    set_cell(ws, "B3",
             "4 tranches × (20 / 28 / 28 / 24) %  =  1 250 млн ₽. Каждый транш привязан "
             "к производственному milestone. Все суммы — млн ₽.",
             font=F_ITALIC, align=C_CENTER)

    # Headers
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Статья", font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    for col, lbl, _ in PERIODS:
        set_cell(ws, f"{col}{r}", lbl, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"P{r}", "Σ 2026–2028", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"Q{r}", "Milestone / Комментарий", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22

    # === Tranches ===
    tranches_data = [
        ("T₁.1", "Транш №1 — 20% (Pre-production launch)",  "D", 250, "Q1'26: запуск pre-production F01-F02 (сценарии, кастинг, площадки)"),
        ("T₁.2", "Транш №2 — 28% (Production Wave I)",       "G", 350, "Q4'26: старт съёмок F01-F03, постпродакшн F01, первый релиз F01"),
        ("T₁.3", "Транш №3 — 28% (Production Wave II)",      "J", 350, "Q3'27: съёмки F05-F08, кассовые сборы F03-F05"),
        ("T₁.4", "Транш №4 — 24% (Production Wave III)",     "M", 300, "Q2'28: завершение производства F09-F12, финальные релизы"),
    ]

    r = 6
    # Individual tranches
    ws.merge_cells(f"B{r}:Q{r}")
    set_cell(ws, f"B{r}", "I.  TRANCHE SCHEDULE", font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 20
    r += 1

    for code, name, col_trigger, amount, milestone in tranches_data:
        set_cell(ws, f"B{r}", code, font=F_BODY, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BODY, align=C_LEFT, border=box_thin)
        for col, lbl, _ in PERIODS:
            val = amount if col == col_trigger else 0
            fill = INFLOW if val > 0 else None
            set_cell(ws, f"{col}{r}", val, font=F_BODY if val == 0 else F_BOLD,
                     fill=fill, align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"P{r}", amount, font=F_TOTAL, fill=INFLOW,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"Q{r}", milestone, font=F_ITALIC, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    # Σ Tranches row
    set_cell(ws, f"B{r}", "Σ", font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Σ TRANCHES  (Т₁ Legacy total)", font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_LEFT, border=box_thin)
    t1_sched = {"D": 250, "G": 350, "J": 350, "M": 300}
    total = 0
    for col, lbl, _ in PERIODS:
        v = t1_sched.get(col, 0)
        total += v
        set_cell(ws, f"{col}{r}", v, font=F_TOTAL, fill=KEY_METRIC_FILL,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"P{r}", total, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"Q{r}", "★ Invariant: 1 250 млн ₽",
             font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 24
    r += 2

    # === Cumulative ===
    ws.merge_cells(f"B{r}:Q{r}")
    set_cell(ws, f"B{r}", "II.  CUMULATIVE INFLOW  &  COVERAGE %",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 20
    r += 1

    # Cumulative inflow
    set_cell(ws, f"B{r}", "2.1", font=F_BODY, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Σ Cumulative inflow", font=F_BOLD, align=C_LEFT, border=box_thin)
    running = 0
    cum_by_col = {}
    for col, lbl, _ in PERIODS:
        running += t1_sched.get(col, 0)
        cum_by_col[col] = running
        set_cell(ws, f"{col}{r}", running, font=F_BOLD, fill=LIGHT_BLUE,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"P{r}", 1250, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"Q{r}", "Накопленный приток, млн ₽",
             font=F_ITALIC, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 18
    r += 1

    # Coverage %
    set_cell(ws, f"B{r}", "2.2", font=F_BODY, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Coverage %  (drawn / total)", font=F_BOLD, align=C_LEFT, border=box_thin)
    for col, lbl, _ in PERIODS:
        pct = cum_by_col[col] / 1250 * 100
        set_cell(ws, f"{col}{r}", round(pct, 1), font=F_BOLD, fill=LIGHT_BLUE,
                 align=C_RIGHT, border=box_thin, number_format=PCTFMT)
    set_cell(ws, f"P{r}", 100.0, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=PCTFMT)
    set_cell(ws, f"Q{r}", "Доля использованного тикета",
             font=F_ITALIC, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 18
    r += 2

    # === Conditions & Governance ===
    ws.merge_cells(f"B{r}:Q{r}")
    set_cell(ws, f"B{r}", "III.  GOVERNANCE  &  CONDITIONS",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 20
    r += 1

    governance = [
        ("3.1", "Release condition для транша №1",     "Подписание SPA + Investment Agreement, KYC/AML clearance"),
        ("3.2", "Release condition для транша №2",     "Завершение pre-production F01-F03, утверждение production budgets"),
        ("3.3", "Release condition для транша №3",     "Первый релиз (F01, Q4'26) — проверка фактической выручки ≥ 80% плана"),
        ("3.4", "Release condition для транша №4",     "Релиз ≥ 5 фильмов с совокупным RoI ≥ 1.2× (Q1-Q2 2028)"),
        ("3.5", "Reporting cadence",                    "Ежеквартальный financial report + KPI dashboard"),
        ("3.6", "Information rights",                   "Read-only доступ к производственным треккерам и бухгалтерии"),
        ("3.7", "Termination clause",                   "При срыве milestone 2-х кварталов подряд — right to exit по fair value"),
    ]
    for code, name, desc in governance:
        set_cell(ws, f"B{r}", code, font=F_BODY, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BOLD, align=C_LEFT, border=box_thin)
        ws.merge_cells(f"D{r}:Q{r}")
        set_cell(ws, f"D{r}", desc, font=F_ITALIC, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    # Control row
    r += 1
    ws.merge_cells(f"B{r}:Q{r}")
    set_cell(ws, f"B{r}",
             "✓ Σ Tranches = 1 250 млн ₽  (20% + 28% + 28% + 24% = 100%)  |  "
             "Matches T₁ Legacy invariant & 10_Cash_Flow 3.1",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 22

    ws.freeze_panes = "D6"


# ============================================================
#   15_Use_of_Funds
# ============================================================
def build_use_of_funds(wb):
    if "15_Use_of_Funds" in wb.sheetnames:
        del wb["15_Use_of_Funds"]
    ws = wb.create_sheet("15_Use_of_Funds")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 30

    ws.merge_cells("B2:I2")
    set_cell(ws, "B2",
             "15  ·  USE OF FUNDS  ·  Распределение 1 250 млн ₽ инвестиций",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:I3")
    set_cell(ws, "B3",
             "Разбивка инвестиционного тикета T₁ Legacy (1 250) по производственным категориям "
             "для 12 фильмов (средний бюджет 154 млн, всего 1 850 млн — инвестор покрывает ~67.6%).",
             font=F_ITALIC, align=C_CENTER)

    r = 5
    # Headers
    headers = ["#", "Категория расхода", "% от T₁", "Сумма, млн ₽", "Min (−10%)", "Max (+10%)", "Allowance", "Описание"]
    for i, h in enumerate(headers):
        col = chr(ord("B") + i)
        set_cell(ws, f"{col}{r}", h, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Categories (% from 1250)
    use_cats = [
        ("U.1", "Pre-production  (script, casting, location)",    0.05,  62.5,
         "Сценарии, раскадровки, подбор актёров, выбор площадок"),
        ("U.2", "Principal photography  (съёмочный период)",      0.65, 812.5,
         "Съёмочный процесс, аренда площадок, оплата съёмочных групп"),
        ("U.3", "Post-production  (VFX, sound, editing)",         0.20, 250.0,
         "Монтаж, цветокоррекция, VFX, озвучка, сведение звука"),
        ("U.4", "Contingency reserve  (10% страховой резерв)",    0.10, 125.0,
         "Резерв на перерасходы, форс-мажор, изменения сценария"),
    ]

    for code, name, pct, amount, desc in use_cats:
        set_cell(ws, f"B{r}", code, font=F_BODY, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BODY, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", pct * 100, font=F_BOLD, fill=LIGHT_BLUE,
                 align=C_RIGHT, border=box_thin, number_format=PCTFMT)
        set_cell(ws, f"E{r}", amount, font=F_BOLD, fill=INFLOW,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"F{r}", round(amount * 0.9, 1), font=F_BODY,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"G{r}", round(amount * 1.1, 1), font=F_BODY,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        allowance = "★ Flex ±10%" if code != "U.4" else "✓ Capped 10%"
        set_cell(ws, f"H{r}", allowance, font=F_ITALIC,
                 align=C_CENTER, border=box_thin)
        set_cell(ws, f"I{r}", desc, font=F_ITALIC, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 24
        r += 1

    # Total row
    set_cell(ws, f"B{r}", "Σ", font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "ИТОГО  ·  T₁ Legacy Investment",
             font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", 100.0, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=PCTFMT)
    set_cell(ws, f"E{r}", 1250.0, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"F{r}", 1125.0, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"G{r}", 1375.0, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"H{r}", "★ ANCHOR", font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_CENTER, border=box_thin)
    set_cell(ws, f"I{r}", "Invariant: 1 250 млн ₽",
             font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 26
    r += 2

    # === Дополнительно: мост к полному производственному бюджету 1850 ===
    ws.merge_cells(f"B{r}:I{r}")
    set_cell(ws, f"B{r}",
             "II.  МОСТ  ·  T₁ Legacy (1 250) → Full Production Budget (1 850)",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 20
    r += 1

    bridge_rows = [
        ("B.1", "T₁ Legacy Investor contribution",                  1250, "Основной тикет инвестора (67.6%)"),
        ("B.2", "(+) Producer equity (JV capital)",                  600, "Собственные средства продюсера (32.4%)"),
        ("=",  "Full production budget  (12 films × avg 154)",      1850, "★ Σ расходов на производство"),
        ("B.3", "(+) P&A expenditure (separate from CAPEX)",        277.5, "Prints & Advertising, финансируется из cashflow"),
        ("=",  "Total content-related costs",                      2127.5, "= COGS в P&L"),
    ]
    for entry in bridge_rows:
        code, name, amount, note = entry
        is_total = code == "="
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        ws.merge_cells(f"C{r}:D{r}")
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        ws.merge_cells(f"E{r}:G{r}")
        set_cell(ws, f"E{r}", amount, font=font, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"H{r}:I{r}")
        set_cell(ws, f"H{r}", note, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    # Control
    r += 1
    ws.merge_cells(f"B{r}:I{r}")
    set_cell(ws, f"B{r}",
             "✓ Σ Use of Funds = 1 250 (5+65+20+10%)  |  Bridge: 1 250 + 600 = 1 850 (budget)  |  "
             "+ P&A 277.5 = 2 127.5 (COGS в P&L)",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 22


# ============================================================
#   16_CAPEX_Schedule
# ============================================================
def build_capex(wb):
    if "16_CAPEX_Schedule" in wb.sheetnames:
        del wb["16_CAPEX_Schedule"]
    ws = wb.create_sheet("16_CAPEX_Schedule")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 8
    for i in range(5, 17):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.column_dimensions["Q"].width = 11
    ws.column_dimensions["R"].width = 10

    ws.merge_cells("B2:R2")
    set_cell(ws, "B2",
             "16  ·  CAPEX SCHEDULE  ·  12 фильмов × timeline (pre-prod / prod / post) по кварталам",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:R3")
    set_cell(ws, "B3",
             "Распределение производственных затрат по фильмам и кварталам. "
             "Σ production CAPEX = 1 850 млн ₽ (consistent с 10_Cash_Flow 2.1). Все суммы — млн ₽.",
             font=F_ITALIC, align=C_CENTER)

    # Headers
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Film", font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", "Бюджет", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)

    for col, lbl, _ in PERIODS:
        set_cell(ws, f"{col}{r}", lbl, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"P{r}", "Σ CAPEX", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"Q{r}", "Релиз", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"R{r}", "Категория", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22

    # Films data: (code, name, budget, release_col, category)
    films = [
        ("F01", "Родные стены",        150, "G", "A",   "Семейная драма"),
        ("F02", "Два сердца",          100, "H", "B",   "Мелодрама"),
        ("F03", "Ночной патруль",      220, "I", "A+",  "Боевик"),
        ("F04", "Последний экзамен",    90, "I", "B",   "Молодёжная драма"),
        ("F05", "Время героев",        250, "J", "A+",  "Исторический эпос"),
        ("F06", "Красная заря",        180, "K", "A",   "Шпионский триллер"),
        ("F07", "Сказка наяву",        170, "L", "A",   "Семейный фэнтези"),
        ("F08", "Неизвестный герой",   130, "M", "B+",  "Биография"),
        ("F09", "Моё дело",             80, "M", "B",   "Комедия"),
        ("F10", "Северный ветер",      180, "N", "A",   "Приключения"),
        ("F11", "Полдень",             140, "N", "B+",  "Триллер"),
        ("F12", "Горизонт событий",    160, "O", "A",   "Фантастика"),
    ]

    # Каждому фильму: распределение budget по 4 стадиям перед релизом
    # pre-prod: T-3Q (5%), prod: T-2Q (45%), prod: T-1Q (35%), post: T-0Q (15%)
    # Если T-3Q < Q1'26, стадия сжимается в доступные кварталы
    def distribute_capex(budget, release_col):
        """Распределить budget по 4 кварталам перед и в релиз."""
        release_idx = ALL_COLS.index(release_col)
        stages = [(-3, 0.05), (-2, 0.45), (-1, 0.35), (0, 0.15)]
        result = {c: 0.0 for c in ALL_COLS}
        for offset, frac in stages:
            target = release_idx + offset
            if target < 0:
                # Сдвигаем в Q1'26 — первый доступный квартал
                target = 0
            result[ALL_COLS[target]] += budget * frac
        return result

    r = 6
    sum_by_period = {c: 0.0 for c in ALL_COLS}
    total_budget = 0

    for code, name, budget, release_col, cat, genre in films:
        set_cell(ws, f"B{r}", code, font=F_BOLD, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BODY, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", budget, font=F_BOLD, fill=LIGHT_BLUE,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)

        capex = distribute_capex(budget, release_col)
        row_sum = 0.0
        for col, lbl, _ in PERIODS:
            v = round(capex[col], 1)
            sum_by_period[col] += v
            row_sum += v
            fill = OUTFLOW if v > 0 else None
            set_cell(ws, f"{col}{r}", v if v > 0 else "",
                     font=F_BODY, fill=fill, align=C_RIGHT,
                     border=box_thin, number_format=NUMFMT)

        set_cell(ws, f"P{r}", round(row_sum, 1), font=F_TOTAL, fill=SUBTOTAL_FILL,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)

        release_lbl = dict((c, l) for c, l, _ in PERIODS).get(release_col, release_col)
        set_cell(ws, f"Q{r}", release_lbl, font=F_BODY,
                 align=C_CENTER, border=box_thin)
        set_cell(ws, f"R{r}", cat, font=F_BOLD,
                 fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
        ws.row_dimensions[r].height = 20
        total_budget += budget
        r += 1

    # Σ row
    set_cell(ws, f"B{r}", "Σ", font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "ИТОГО Production CAPEX", font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", total_budget, font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=NUMFMT)

    total_all = 0
    for col, lbl, _ in PERIODS:
        v = round(sum_by_period[col], 1)
        total_all += v
        set_cell(ws, f"{col}{r}", v, font=F_TOTAL, fill=KEY_METRIC_FILL,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"P{r}", round(total_all, 1), font=F_TOTAL, fill=KEY_METRIC_FILL,
             align=C_RIGHT, border=box_thin, number_format=NUMFMT)
    set_cell(ws, f"Q{r}", "—", font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_CENTER, border=box_thin)
    set_cell(ws, f"R{r}", "★", font=F_TOTAL, fill=KEY_METRIC_FILL, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 24
    r += 2

    # Year totals
    year_totals = {}
    for col, lbl, _ in PERIODS:
        year = "20" + lbl[-2:]  # 26, 27, 28 → 2026, 2027, 2028
        year_totals[year] = year_totals.get(year, 0) + sum_by_period[col]

    ws.merge_cells(f"B{r}:R{r}")
    set_cell(ws, f"B{r}",
             f"✓ Σ Production CAPEX = {total_all:,.1f} млн ₽ (target: 1 850)  |  "
             f"2026: {year_totals.get('2026', 0):,.0f}  ·  "
             f"2027: {year_totals.get('2027', 0):,.0f}  ·  "
             f"2028: {year_totals.get('2028', 0):,.0f}",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 24

    ws.freeze_panes = "E6"
    return total_all, year_totals


def main():
    wb = load_workbook(FILE)
    print(f"Loaded: {FILE}")
    print(f"Sheets before: {len(wb.sheetnames)}")

    print("\n[1/3] Building 14_Investment_Inflow …")
    build_inflow(wb)

    print("[2/3] Building 15_Use_of_Funds …")
    build_use_of_funds(wb)

    print("[3/3] Building 16_CAPEX_Schedule …")
    total_capex, year_totals = build_capex(wb)

    wb.save(FILE)
    print(f"\nSheets after: {len(wb.sheetnames)}")
    print(f"\nCAPEX totals:")
    print(f"  Total: {total_capex:.1f} (target: 1850)")
    for y, v in sorted(year_totals.items()):
        print(f"  {y}: {v:.1f}")
    print(f"\nSaved: {FILE}")


if __name__ == "__main__":
    main()
