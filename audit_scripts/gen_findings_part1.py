import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
thin = Side(style='thin')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="003366")
wrap = Alignment(wrap_text=True, vertical="top")
sev_colors = {"CRITICAL": "FF0000", "HIGH": "FF8C00", "MEDIUM": "FFD700", "LOW": "90EE90"}

def add_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = wrap; c.border = border
    ws.freeze_panes = "A2"

def add_row(ws, row_num, data):
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.alignment = wrap; c.border = border
        if i == 3 and val in sev_colors:
            c.fill = PatternFill("solid", fgColor=sev_colors[val])
            c.font = Font(bold=True, color="FFFFFF" if val in ["CRITICAL","HIGH"] else "000000")

# === SHEET 1: Findings ===
ws1 = wb.active; ws1.title = "Findings"
add_header(ws1, ["ID","Block","Severity","Category","Description","Location","Evidence","Remediation","Status"])
findings = [
["F-001","C","CRITICAL","Leakage","Internal W5 V-D waterfall exposed in Public","24_Investor_Returns!B49","Text references Internal W5 V-D comparison with margin values","Scrub cell content","OPEN"],
["F-002","C","CRITICAL","Leakage","8 'Internal' keyword occurrences in sharedStrings","sharedStrings.xml","Multiple references to Internal model found","Remove all Internal references","OPEN"],
["F-003","A","CRITICAL","Leakage","Developer filesystem path in workbook.xml","xl/workbook.xml absPath","/Users/noldorwarrior/Documents/Claude/Projects/","Re-save from clean environment","OPEN"],
["F-004","B","CRITICAL","Architecture","Model is 97.5% static constants","All 42 sheets","226 formulas / 9000+ values = 2.5%","Convert top-5 sheets to live formulas","OPEN"],
["F-005","B","CRITICAL","Architecture","Anchor-based back-solving visible","P&L, Cost Structure","4545 = 2008.8 + 368.8 + 2167.4 exact","Document as design choice","OPEN"],
["F-006","B","HIGH","Math","Three inconsistent IRR methods","build_A11, A12, patches","Newton ~7.7%, MOIC^(1/6.5), numpy 20.09%","Standardize on numpy_financial.irr","OPEN"],
["F-007","D","HIGH","Valuation","Valuation spread 6x between methods","22_DCF, 23_Mult, 28_MC","DCF 1.8B vs Comps 7.5B vs MC 11.2B","Reconcile to common assumptions","OPEN"],
["F-008","D","HIGH","Risk","MC P(IRR>hurdle) = 13.6% only","28_Monte_Carlo_Summary","Mean IRR 11.44% vs deterministic 20.09%","Prominent disclosure","OPEN"],
["F-009","C","HIGH","Leakage","Change Log mentions 'Internal'","03_Change_Log!C30","'Public export + Internal with 4 service'","Remove reference","OPEN"],
["F-010","C","HIGH","Leakage","Valuation labels '(internal)'","23_Valuation_Multiples!B46","Scenario Base (internal)","Remove label","OPEN"],
["F-011","C","HIGH","Leakage","Comparable Transactions '(internal)'","32_Comparable_Transactions","Gazprom-media/Yandex (internal)","Remove annotations","OPEN"],
["F-012","B","HIGH","Math","D&A jumps 167x in DCF","22_Valuation_DCF","3M (2028) to 500M (2029)","Justify or smooth","OPEN"],
["F-013","B","HIGH","Math","MoIC contradiction 4.8x vs 1.39x","36_ES vs 28_MC","Aggregate vs T1 cash-on-cash","Label both metrics","OPEN"],
["F-014","B","HIGH","Consistency","Scenario probability mismatch","build_A12 vs manifest","5/15/50/20/10 vs 10/20/40/20/10","Reconcile","OPEN"],
["F-015","D","HIGH","Valuation","Peer comps fabrication risk","23_Valuation_Multiples","5/6 peers EV/Revenue ~0.80-0.81","Add sources","OPEN"],
["F-016","B","HIGH","Math","MC revenue blend +6% upward bias","build_A12","0.85+0.30*hit_rate at mode=1.06","Correct to center at 1.0","OPEN"],
["F-017","G","HIGH","Pipeline","3 hardcoded macOS paths","build_memo/onepager/presentation","/Users/noldorwarrior/Downloads","Use env var","OPEN"],
["F-018","G","HIGH","Pipeline","requirements.txt not pinned","requirements.txt","All >= not ==","Pin with ==","OPEN"],
["F-019","G","HIGH","Pipeline","Missing deps in requirements.txt","requirements.txt","plotly, python-pptx, rakhman_docs","Add to requirements","OPEN"],
["F-020","A","MEDIUM","Metadata","Last Modified By = 'a'","docProps/core.xml","Truncated username","Clean metadata","OPEN"],
["F-021","A","MEDIUM","Metadata","Description leaks anchor value","docProps/core.xml","cumulative EBITDA = 3000","Remove","OPEN"],
["F-022","A","MEDIUM","Metadata","Keywords contain 'L3'","docProps/core.xml","Internal classification","Remove L3","OPEN"],
["F-023","C","MEDIUM","Metadata","Internal file title says 'Public'","Internal xlsx title","Copy artifact","Fix title","OPEN"],
["F-024","A","MEDIUM","Metadata","Email address in cells","sharedStrings","team@trendstudio.ru","Evaluate","OPEN"],
["F-025","C","MEDIUM","Structure","Cash Flow formula diff","Public vs Internal","0 vs 5 formulas","Investigate","OPEN"],
["F-026","B","MEDIUM","Math","NDP bridge uses plug number","09_P&L bridge","WC/Gov 233 manual","Document derivation","OPEN"],
["F-027","B","MEDIUM","Tax","VAT vs Profit Tax confusion","P&L/Tax","20% without distinction","Separate","OPEN"],
["F-028","D","MEDIUM","Risk","Risk scoring uncalibrated","29_Risk_Register","No rubric 1-5","Add rubric","OPEN"],
["F-029","B","MEDIUM","Structure","Producer equity 600M ambiguous","17_Deal_Structures","Unclear hierarchy","Create Capital Stack","OPEN"],
["F-030","B","MEDIUM","Valuation","WACC single point no CAPM","22_Valuation_DCF","19.05% without decomp","Add CAPM","OPEN"],
["F-031","A","MEDIUM","Integrity","Cached values are None","Multiple sheets","Not recalculated","Open/recalculate/save","OPEN"],
["F-032","F","MEDIUM","Docs","Template placeholders","42_Cover_Letter","[Name of LP], [Date]","Fill","OPEN"],
["F-033","G","MEDIUM","Pipeline","verify.py bug: hardcoded 14","scripts/verify.py:161","Should be 18","Fix to len()","OPEN"],
["F-034","B","MEDIUM","Anchor","EBITDA anchor mismatch","09_P&L","Model 2167.4 vs spec 2076.1","Clarify","OPEN"],
["F-035","A","LOW","Cosmetic","Counter 62 vs 45","40_Investor_Checklist","Heading mismatch","Fix","OPEN"],
["F-036","A","LOW","Integrity","Dangling = in cell","17_Deal!R31C2","Bare equals sign","Remove","OPEN"],
["F-037","A","LOW","Cosmetic","Print_Area not set","All sheets","No print area","Set","OPEN"],
["F-038","G","LOW","Docs","README says 78 tests, actual 287+","pipeline/README.md","Version mismatch","Update","OPEN"],
["F-039","A","LOW","Hygiene","20+ backup files","Investor_Package/","*.bak files","Remove from release","OPEN"],
]
for i, f in enumerate(findings, 2):
    add_row(ws1, i, f)
for col in range(1, 10):
    ws1.column_dimensions[get_column_letter(col)].width = [8,6,10,12,45,30,40,30,8][col-1]

# === SHEET 2: Summary ===
ws2 = wb.create_sheet("Summary")
add_header(ws2, ["Category","Count","Details"])
summary = [
["CRITICAL",5,"Leakage (3), Architecture (2)"],
["HIGH",14,"Math (5), Leakage (3), Valuation (2), Pipeline (3), Consistency (1)"],
["MEDIUM",15,"Metadata (5), Math (3), Structure (2), Tax (1), Risk (1), Docs (1), Pipeline (1), Anchor (1)"],
["LOW",5,"Cosmetic (2), Integrity (1), Hygiene (1), Docs (1)"],
["TOTAL",39,""],
["","",""],
["Block","Count",""],
["A - xlsx Technical",11,""],
["B - Financial Math",12,""],
["C - Public/Internal",6,""],
["D - Investor Readiness",4,""],
["E - Going Concern",0,"GREEN - no findings"],
["F - Cover Docs",1,""],
["G - Pipeline",5,""],
["","",""],
["VERDICT","CONDITIONAL FAIL","Requires remediation before investor distribution"],
]
for i, s in enumerate(summary, 2):
    add_row(ws2, i, s)
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 50

# === SHEET 3: Roadmap ===
ws3 = wb.create_sheet("Roadmap")
add_header(ws3, ["Phase","Timeline","Items","Effort","Description"])
roadmap = [
["Quick Wins","0-2 days","F-001-003,009-011,020-024,032,035-036,039","8-16h","Leakage scrub, metadata cleanup, cosmetics"],
["Structural","1-3 weeks","F-004-008,012-016,027,029-030","80-120h","Live formulas, IRR standardization, valuation reconciliation"],
["Architectural","4-6 weeks","Full formula rebuild, independent verification","160-240h","Formula-based architecture, second-engineer verification, MC upgrade"],
]
for i, r in enumerate(roadmap, 2):
    add_row(ws3, i, r)
for col in range(1,6):
    ws3.column_dimensions[get_column_letter(col)].width = [15,12,35,12,50][col-1]

# === SHEET 4: Improvements ===
ws4 = wb.create_sheet("Improvements")
add_header(ws4, ["#","Area","Recommendation","Priority","Effort"])
improvements = [
[1,"Architecture","Convert from anchor-based to formula-based model","HIGH","3 weeks"],
[2,"IRR","Standardize on numpy_financial.irr across all scripts","HIGH","2 days"],
[3,"MC","Upgrade to vectorized numpy MC with proper cashflow-based IRR","HIGH","1 week"],
[4,"Comps","Add Pitchbook/Mergermarket source citations","HIGH","3 days"],
[5,"WACC","Full CAPM build-up with sourced components","MEDIUM","2 days"],
[6,"Tax","Separate VAT pass-through from profit tax","MEDIUM","1 day"],
[7,"D&A","Add asset depreciation schedule to justify DCF D&A","MEDIUM","2 days"],
[8,"Pipeline","Pin all dependencies with ==","MEDIUM","1 hour"],
[9,"Pipeline","Add plotly, python-pptx to requirements.txt","LOW","30 min"],
[10,"Docs","Complete Cover Letter, remove all placeholders","LOW","2 hours"],
]
for i, imp in enumerate(improvements, 2):
    add_row(ws4, i, imp)
for col in range(1,6):
    ws4.column_dimensions[get_column_letter(col)].width = [5,15,55,10,12][col-1]

# === SHEET 5: DD Readiness ===
ws5 = wb.create_sheet("DD Readiness")
add_header(ws5, ["#","Category","Status","Key Issue","Action Required"])
dd = [
[1,"Model Integrity","FAIL","97.5% constants, not live","Convert to formula-based"],
[2,"Leakage Control","FAIL","23 leakage findings","Scrub all Internal refs"],
[3,"Financial Math","CONDITIONAL","Anchors verified but back-solved","Document methodology"],
[4,"Valuation","FAIL","6x spread, fabricated comps","Reconcile, add sources"],
[5,"MC Simulation","CONDITIONAL","Results verified, underpowered","Increase N, fix bias"],
[6,"Documentation","PASS","34-item self-audit, 8 ADRs","Minor updates"],
[7,"Pipeline","PASS","16K LOC, 323/328 tests","Pin deps"],
[8,"Cover Docs","CONDITIONAL","Honest but incomplete","Fill placeholders"],
[9,"Stress Testing","CONDITIONAL","Base passes, tails weak","Add scenario detail"],
[10,"Legal/Disclaimers","CONDITIONAL","Present but templates","Complete"],
[11,"Metadata","FAIL","Dev path, L3 classification","Clean all"],
[12,"Version Control","PASS","Manifest, SHA-256, backups","None"],
]
for i, d in enumerate(dd, 2):
    add_row(ws5, i, d)
    c = ws5.cell(row=i, column=3)
    colors = {"PASS":"00AA00","CONDITIONAL":"FF8C00","FAIL":"FF0000"}
    c.font = Font(bold=True, color=colors.get(d[2],"000000"))
for col in range(1,6):
    ws5.column_dimensions[get_column_letter(col)].width = [5,20,14,40,30][col-1]

OUT = '/home/user/TrendStudio-Holding/audit_scripts/audit_public_v1_findings.xlsx'
wb.save(OUT)
print(f"Part 1 done: 5 sheets saved to {OUT}")
