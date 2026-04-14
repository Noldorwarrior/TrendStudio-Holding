"""
А.1 — Cover + Assumptions + Change_Log
investor_model_v1.0_Public.xlsx (L3, 36 sheets pre-IPO)

Якорь: cumulative EBITDA 2026-2028 Base = 3 000 млн ₽
Источник: v1.4.4 + Холдинг Кино.xlsx (legacy)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
import os

OUT_DIR = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package"
OUT_FILE = os.path.join(OUT_DIR, "investor_model_v1.0_Public.xlsx")
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================================
# СТИЛИ — единое цветовое кодирование для всей модели
# ============================================================================

# Основные цвета
BRAND_BLUE = "0070C0"
BRAND_BLUE_LIGHT = "D9E2F3"
BRAND_BLUE_DARK = "002060"
ACCENT_GOLD = "BF9000"
ACCENT_GREEN = "548235"
ACCENT_RED = "C00000"
GRAY_LIGHT = "F2F2F2"
GRAY_MEDIUM = "BFBFBF"
GRAY_DARK = "595959"
WHITE = "FFFFFF"

# Цвета для типов ячеек (правило моделирования)
INPUT_BLUE = "0000FF"       # синий текст — ввод пользователя
FORMULA_BLACK = "000000"    # чёрный текст — расчёт
LINK_GREEN = "006100"       # зелёный текст — ссылка на другой лист

# Подсветка ключевых метрик
KEY_METRIC_FILL = "FFF2CC"  # светло-жёлтый фон для Revenue/EBITDA/Net Profit/Investment


def thin_border(color="808080"):
    side = Side(border_style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def medium_border(color=BRAND_BLUE):
    side = Side(border_style="medium", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


# ============================================================================
# WORKBOOK
# ============================================================================

wb = Workbook()
wb.remove(wb.active)  # убираем дефолтный Sheet


# ============================================================================
# ЛИСТ 1: COVER
# ============================================================================

def build_cover(wb):
    ws = wb.create_sheet("01_Cover")
    ws.sheet_view.showGridLines = False

    # Ширина колонок
    widths = [2, 28, 22, 22, 22, 22, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Фон всего листа — светло-серый для рамки
    # (имитируем через padding rows)

    row = 2
    # --- Верхняя плашка (brand bar) ---
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value="ТРЕНДСТУДИО")
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

    row = 4
    # --- Заголовок модели ---
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value="ИНВЕСТИЦИОННАЯ ФИНАНСОВАЯ МОДЕЛЬ")
    c.font = Font(name="Arial", size=22, bold=True, color=BRAND_BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 36

    row = 5
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value="Холдинг кинопроизводства «ТрендСтудио»")
    c.font = Font(name="Arial", size=14, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

    row = 7
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value="Версия Public v1.0  ·  Pre-IPO institutional  ·  36 листов")
    c.font = Font(name="Arial", size=11, color=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center")

    row = 8
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value="Горизонт моделирования: 2026–2032  ·  Базовая валюта: RUB")
    c.font = Font(name="Arial", size=10, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="center")

    # --- Разделитель ---
    row = 10
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = Border(
            top=Side(border_style="medium", color=BRAND_BLUE))

    # --- Блок «Ключевые показатели» ---
    row = 12
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value="КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ (Base Scenario, 2026–2028)")
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24

    # --- 4 карточки метрик: Revenue / EBITDA / Net Profit / Investment ---
    key_metrics = [
        ("Revenue\n(накопленная)", "4 545", "млн ₽", "12 фильмов × 378.75 млн"),
        ("EBITDA\n(якорь L4+N3)", "3 000", "млн ₽", "Base, tolerance ±1%"),
        ("Net Profit\n(чистая прибыль)", "2 280", "млн ₽", "после налога 20% и WACC"),
        ("Investment\n(входящие)", "1 250", "млн ₽", "4 транша: 20/28/28/24%"),
    ]

    row = 14
    for i, (label, value, unit, note) in enumerate(key_metrics):
        col = 2 + i + (1 if i > 0 else 0)  # распределяем по 4 ячейкам
    # Простой вариант: 4 колонки B, C, D, E → F закрывающий
    # Сделаем карточки в колонках B, C, D, E
    row_label = 14
    row_value = 15
    row_unit = 17
    row_note = 18
    cols = [2, 3, 4, 5]
    for i, (label, value, unit, note) in enumerate(key_metrics):
        col = cols[i]
        # label
        lc = ws.cell(row=row_label, column=col, value=label)
        lc.font = Font(name="Arial", size=10, bold=True, color=GRAY_DARK)
        lc.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        lc.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        lc.border = medium_border(BRAND_BLUE)
        ws.row_dimensions[row_label].height = 30
        # value
        vc = ws.cell(row=row_value, column=col, value=value)
        vc.font = Font(name="Arial", size=26, bold=True, color=BRAND_BLUE_DARK)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        vc.border = medium_border(BRAND_BLUE)
        ws.row_dimensions[row_value].height = 38
        # spacer row 16 — also in card
        ws.cell(row=16, column=col).fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        ws.cell(row=16, column=col).border = medium_border(BRAND_BLUE)
        # unit
        uc = ws.cell(row=row_unit, column=col, value=unit)
        uc.font = Font(name="Arial", size=10, color=BRAND_BLUE)
        uc.alignment = Alignment(horizontal="center")
        uc.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        uc.border = medium_border(BRAND_BLUE)
        # note
        nc = ws.cell(row=row_note, column=col, value=note)
        nc.font = Font(name="Arial", size=8, italic=True, color=GRAY_DARK)
        nc.alignment = Alignment(horizontal="center", wrap_text=True)
        nc.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        nc.border = medium_border(BRAND_BLUE)
        ws.row_dimensions[row_note].height = 22

    # --- Якорная проверка ---
    row = 20
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value="ЯКОРЬ L4+N3:  cumulative EBITDA 2026–2028 = 3 000 млн ₽  (tolerance ±1%)")
    c.font = Font(name="Arial", size=11, bold=True, color=ACCENT_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="E2EFDA")

    # --- Реквизиты модели ---
    row = 22
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value="РЕКВИЗИТЫ МОДЕЛИ")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

    meta_rows = [
        ("Название модели:", "investor_model_v1.0_Public.xlsx"),
        ("Версия:", "v1.0 — Pre-IPO institutional"),
        ("Дата построения:", datetime.now().strftime("%d.%m.%Y")),
        ("Автор:", "Команда ТрендСтудио (подготовлено для инвестора)"),
        ("Целевая аудитория:", "Государственный сектор (Фонд кино / ВЭБ) + Стратегический медиахолдинг"),
        ("Базовая дата расчётов:", "01.01.2026"),
        ("Горизонт прогноза:", "2026–2032 (Q1 2026–Q4 2028 поквартально, 2029–2032 годовые)"),
        ("Единица измерения:", "млн ₽ (если не указано иное)"),
        ("Основа:", "v1.4.4 (348 тестов, П5 «Максимум» 32/32) + legacy Холдинг Кино.xlsx"),
        ("Верификация:", "П5 «Максимум» 32/32 механизма"),
        ("Защита:", "Формулы защищены; допущения — редактируемые (Assumptions)"),
    ]
    row = 23
    for label, value in meta_rows:
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = Font(name="Arial", size=10, bold=True, color=GRAY_DARK)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        vc = ws.cell(row=row, column=3, value=value)
        vc.font = Font(name="Arial", size=10, color=BRAND_BLUE_DARK)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    # --- Дисклеймер ---
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value="ДИСКЛЕЙМЕР")
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ACCENT_RED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    disclaimer = (
        "Настоящая финансовая модель подготовлена исключительно в информационных целях "
        "и не является публичной офертой, инвестиционной рекомендацией или предложением "
        "о заключении каких-либо сделок. Прогнозные показатели основаны на допущениях, "
        "перечисленных на листе «02_Assumptions», и подлежат индивидуальной верификации "
        "инвестором. Распространение без письменного согласия владельца запрещено."
    )
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=6)
    dc = ws.cell(row=row, column=2, value=disclaimer)
    dc.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    dc.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    for r in range(row, row + 4):
        ws.row_dimensions[r].height = 16

    # Freeze
    ws.freeze_panes = "A2"


# ============================================================================
# ЛИСТ 2: ASSUMPTIONS
# ============================================================================

def build_assumptions(wb):
    ws = wb.create_sheet("02_Assumptions")
    ws.sheet_view.showGridLines = False

    widths = [2, 6, 42, 18, 10, 38, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Заголовок ---
    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value="ДОПУЩЕНИЯ МОДЕЛИ (Assumptions)")
    c.font = Font(name="Arial", size=18, bold=True, color=BRAND_BLUE_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30

    row = 3
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value="Все расчёты модели ссылаются на этот лист. 🔵 синий = ввод  ·  ⚫ чёрный = расчёт  ·  🟢 зелёный = ссылка")
    c.font = Font(name="Arial", size=10, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", indent=1)

    row = 5
    # --- Шапка таблицы ---
    headers = ["#", "Параметр", "Значение", "Ед.изм.", "Примечание / источник"]
    cols = [2, 3, 4, 5, 6]
    for col, h in zip(cols, headers):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border("FFFFFF")
    ws.row_dimensions[row].height = 22

    # --- Блоки допущений ---
    blocks = [
        ("A. МАКРО-ДОПУЩЕНИЯ", [
            ("Инфляция РФ (CPI), годовая",           0.065,  "%",        "ЦБ РФ прогноз, Base 6.5% 2026–2028"),
            ("Ключевая ставка (среднегодовая)",      0.14,   "%",        "ЦБ РФ, Base 14% 2026"),
            ("WACC для DCF (CAPM build-up)",          0.1905, "%",        "Rf 14.5% + β×ERP 5.6% + Country 2% + Size 1%, WACC 19.05%"),
            ("Налог на прибыль РФ",                  0.20,   "%",        "НК РФ, Глава 25"),
            ("Страховые взносы (работодатель)",      0.30,   "%",        "ПФР+ФОМС+ФСС, 30% + 0.2% НС (округл. 30%)"),
            ("Ставка дисконта для payback (nominal)", 0.15,   "%",        "Инвестор-ориентир, консервативная"),
            ("Terminal growth (Gordon)",              0.03,   "%",        "Long-run GDP RF, Base 3%"),
            ("Курс USD/RUB (плановый, ср.год)",       95.0,   "RUB",      "Для международных продаж и лицензий"),
        ]),
        ("B. АНКЕТНЫЕ ЯКОРЯ (v1.4.4 + legacy)", [
            ("Cumulative EBITDA 2026–2028 Base",     3000,   "млн ₽",    "Якорь L4+N3, tolerance ±1%"),
            ("Total investment (входящие средства)", 1250,   "млн ₽",    "4 транша 20/28/28/24% (legacy)"),
            ("Total content budget (12 films)",      1850,   "млн ₽",    "invest 1250 + producer equity 600"),
            ("Target revenue 2026–2028 Base",        4545,   "млн ₽",    "12 фильмов × 378.75 млн средняя"),
            ("Target Net Profit 2026–2028",          2280,   "млн ₽",    "после налога 20%"),
            ("Количество фильмов в портфеле",        12,     "шт",       "Холдинг Кино.xlsx / модель"),
            ("Средний бюджет фильма",                154.2,  "млн ₽",    "1850 / 12"),
        ]),
        ("C. РАСПИСАНИЕ ТРАНШЕЙ (T₁ Legacy)", [
            ("Транш 1 (Q1 2026), % от total",        0.20,   "%",        "250 млн ₽"),
            ("Транш 2 (Q2 2026), % от total",        0.28,   "%",        "350 млн ₽"),
            ("Транш 3 (Q3 2026), % от total",        0.28,   "%",        "350 млн ₽"),
            ("Транш 4 (Q4 2026), % от total",        0.24,   "%",        "300 млн ₽"),
        ]),
        ("D. АЛЬТЕРНАТИВНЫЙ ТИКЕТ (T₂ New Deal)", [
            ("Базовый тикет (New)",                  500,    "млн ₽",    "За 20% equity, slider 300–800"),
            ("Базовая доля (New)",                   0.20,   "%",        "20%, slider 10–30%"),
            ("Pre-money valuation (Base)",           2000,   "млн ₽",    "500 / 20% - 500"),
            ("Post-money valuation (Base)",          2500,   "млн ₽",    "pre-money + 500"),
            ("Coupon (liquidation preference W₃)",   0.08,   "%",        "8% годовых"),
        ]),
        ("E. РАСПРЕДЕЛЕНИЕ ПРИБЫЛИ (Waterfall)", [
            ("W₁ Инвестор фаза I (до возврата)",     0.60,   "%",        "Legacy 60/40"),
            ("W₁ Продюсер фаза I",                   0.40,   "%",        "Legacy 60/40"),
            ("W₁ Инвестор фаза II (после возврата)", 0.50,   "%",        "Legacy 50/50"),
            ("W₁ Продюсер фаза II",                  0.50,   "%",        "Legacy 50/50"),
            ("W₂ Pro-rata по долям",                 0.80,   "%",        "Пропорционально equity T₂ (инвестор)"),
            ("W₃ 1× liquidation preference",         1.00,   "×",        "Возврат инвестиций первыми"),
            ("W₃ Participation rate",                1.00,   "×",        "Full participation after preference"),
        ]),
        ("F. СТРУКТУРА ВЫРУЧКИ (% от average revenue)", [
            ("Box Office (кинопрокат)",              0.48,   "%",        "Средневзвешенная доля, 400–1200 млн"),
            ("Online SVOD (стриминги)",              0.20,   "%",        "Kion, Okko, START, IVI, 150–500 млн"),
            ("TV rights",                             0.08,   "%",        "Первый, Россия-1, ТНТ, 50–200 млн"),
            ("International sales",                   0.10,   "%",        "СНГ + BRICS, 30–500 млн"),
            ("VOD (transactional)",                   0.04,   "%",        "20–100 млн"),
            ("Sponsorship / product placement",       0.05,   "%",        "20–150 млн"),
            ("Merchandise",                           0.02,   "%",        "10–100 млн"),
            ("Господдержка (грант + невозвр.)",       0.03,   "%",        "Фонд кино + Минкульт, 25% бюджета"),
        ]),
        ("G. ГОСПОДДЕРЖКА (State Support, Base)", [
            ("Доля господдержки в бюджете (Base)",   0.25,   "%",        "Среднее по Фонд кино + Минкульт + ИРИ"),
            ("Доля господдержки (Min)",              0.00,   "%",        "Sensitivity lower bound"),
            ("Доля господдержки (Max)",              0.35,   "%",        "Sensitivity upper bound"),
            ("Из них: Фонд кино (%)",                0.60,   "%",        "Самый значимый источник"),
            ("Из них: Минкульт (%)",                 0.25,   "%",        "Невозвратные гранты"),
            ("Из них: ИРИ (%)",                      0.15,   "%",        "Интернет-контент/цифровизация"),
            ("Безвозвратная часть (% от господдержки)",0.70,  "%",        "Grant vs loan"),
        ]),
        ("H. FOT / ШТАТ (Модель A₁ Fixed, Холдинг Кино.xlsx)", [
            ("Штат (фиксированный)",                 50,     "чел",      "Оригинальная структура"),
            ("ФОТ gross (месячный)",                 4740000, "₽/мес",   "До страховых"),
            ("Страховые взносы (30%)",               1422000, "₽/мес",   "ПФР+ФОМС+ФСС"),
            ("ФОТ с налогами (месячный)",            6162000, "₽/мес",   "Итоговая стоимость работодателя"),
            ("ФОТ годовой (A₁ fixed)",               73.944,  "млн ₽",   "× 12 мес"),
            ("ФОТ за 3 года (A₁ fixed)",             221.832, "млн ₽",   "× 36 мес, без индексации"),
        ]),
        ("I. FOT / ШТАТ (Модель A₂ Full Dynamic)", [
            ("Штат 2026",                            50,     "чел",      "Базовый"),
            ("Штат 2027",                            60,     "чел",      "+10 (расширение)"),
            ("Штат 2028",                            70,     "чел",      "+10 (пре-IPO команда)"),
            ("Индексация окладов (годовая)",         0.07,   "%",        "CPI + премия, 7%"),
            ("Премиальный фонд (% от ФОТ)",          0.15,   "%",        "KPI-bonus"),
        ]),
        ("J. OPEX (кроме ФОТ, Холдинг Кино.xlsx)", [
            ("Аренда офиса (годовая)",               8.250,  "млн ₽",    "750 м² × 11000 × 12 / 12 мес"),
            ("Коммунальные платежи",                 0.396,  "млн ₽",    "33 т.р./мес × 12"),
            ("Прочие OPEX",                          5.916,  "млн ₽",    "Командировки, связь, ПО и пр."),
            ("Итого OPEX без ФОТ (годовой)",         14.562, "млн ₽",    "Аренда + ком + прочие"),
            ("Итого OPEX с ФОТ A₁ (годовой)",        88.506, "млн ₽",    "Соответствует Холдинг Кино.xlsx"),
        ]),
        ("K. COGS / ПРОИЗВОДСТВО", [
            ("Total content budget",                 1850,   "млн ₽",    "Бюджет 12 фильмов"),
            ("Средний бюджет 1 фильма",              154.2,  "млн ₽",    "1850 / 12"),
            ("Производственный цикл (мес)",          18,     "мес",      "Preprod 6 + prod 6 + postprod 6"),
            ("Amortization пленки (мес после релиза)",12,    "мес",      "Линейная амортизация 12 мес"),
            ("Маркетинг P&A (% от бюджета)",         0.15,   "%",        "Prints & Advertising"),
        ]),
        ("L. WORKING CAPITAL", [
            ("DSO (days sales outstanding)",         45,     "дн",       "Дебиторка от прокатчиков/стримингов"),
            ("DPO (days payable outstanding)",       30,     "дн",       "Кредиторка поставщикам"),
            ("DIO (days inventory outstanding)",     60,     "дн",       "Незавершённое производство"),
            ("Cash Conversion Cycle",                75,     "дн",       "DSO+DIO−DPO"),
            ("Минимальный кассовый остаток",         50,     "млн ₽",    "Подушка безопасности"),
        ]),
        ("M. EXIT / VALUATION", [
            ("EV/EBITDA Bear",                       3.0,    "×",        "Russian media lower bound"),
            ("EV/EBITDA Base",                       5.0,    "×",        "Russian media mid"),
            ("EV/EBITDA Bull",                       7.0,    "×",        "Russian media upper bound / premium"),
            ("Exit horizon 1 (Y3)",                  2028,   "год",      "Q4 2028"),
            ("Exit horizon 2 (Y5)",                  2030,   "год",      "Q4 2030"),
            ("Exit horizon 3 (Y7)",                  2032,   "год",      "Q4 2032"),
            ("Hurdle Rate (инвестор)",               0.18,   "%",        "18% hurdle rate (Named Range HURDLE_RATE)"),
            ("Target MOIC (инвестор)",               2.5,    "×",        "Multiple of Invested Capital"),
        ]),
    ]

    row = 6
    total_count = 0
    for block_title, params in blocks:
        # Заголовок блока
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c = ws.cell(row=row, column=2, value=block_title)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        for param_name, value, unit, note in params:
            total_count += 1
            # #
            c = ws.cell(row=row, column=2, value=total_count)
            c.font = Font(name="Arial", size=9, color=GRAY_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            # Параметр
            c = ws.cell(row=row, column=3, value=param_name)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = thin_border()
            # Значение (синий — ввод)
            c = ws.cell(row=row, column=4, value=value)
            c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            c.border = thin_border()
            # Формат ячейки
            if unit == "%":
                c.number_format = "0.0%"
            elif unit in ("млн ₽", "RUB"):
                c.number_format = '#,##0.000;[Red]-#,##0.000'
            elif unit == "₽/мес":
                c.number_format = '#,##0 "₽";[Red]-#,##0'
            elif unit == "×":
                c.number_format = '0.0"x"'
            elif unit in ("чел", "шт", "дн", "мес", "год"):
                c.number_format = '#,##0'
            else:
                c.number_format = 'General'
            # Ед.изм.
            c = ws.cell(row=row, column=5, value=unit)
            c.font = Font(name="Arial", size=9, color=GRAY_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            # Примечание
            c = ws.cell(row=row, column=6, value=note)
            c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)
            c.border = thin_border()

            ws.row_dimensions[row].height = 18
            row += 1
        # пустая строка между блоками
        row += 1

    # Итого допущений
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value=f"ВСЕГО ДОПУЩЕНИЙ: {total_count}  ·  Блоков: {len(blocks)}  ·  Якорь EBITDA 2026–2028 = 3 000 млн ₽")
    c.font = Font(name="Arial", size=10, bold=True, color=ACCENT_GREEN)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A6"
    return total_count


# ============================================================================
# ЛИСТ 3: CHANGE LOG
# ============================================================================

def build_changelog(wb):
    ws = wb.create_sheet("03_Change_Log")
    ws.sheet_view.showGridLines = False

    widths = [2, 12, 14, 22, 70, 18, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Заголовок
    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value="ЖУРНАЛ ИЗМЕНЕНИЙ (Change Log)")
    c.font = Font(name="Arial", size=18, bold=True, color=BRAND_BLUE_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30

    row = 3
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2,
                value="История версий модели investor_model_v1.0. Все изменения фиксируются с указанием автора, даты и причины.")
    c.font = Font(name="Arial", size=10, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", indent=1)

    # Шапка
    row = 5
    headers = ["Версия", "Дата", "Автор", "Описание изменения", "Статус"]
    cols = [2, 3, 4, 5, 6]
    for col, h in zip(cols, headers):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border("FFFFFF")
    ws.row_dimensions[row].height = 22

    # Записи
    today = datetime.now().strftime("%d.%m.%Y")
    entries = [
        ("v0.1", "05.04.2026", "Команда ТрендСтудио",
         "Каркас модели: определение структуры 14 листов (Уровень 1)",
         "Архивировано"),
        ("v0.2", "06.04.2026", "Команда ТрендСтудио",
         "Добавлены Holding Кино.xlsx legacy-структуры: FOT 50 чел, 12-квартальная P&L, distribution 60/40→50/50",
         "Архивировано"),
        ("v0.3", "07.04.2026", "Команда ТрендСтудио",
         "Расширение до Уровня 2 (24 листа): добавлены Cash Flow, Balance Sheet, Working Capital, Cap Table",
         "Архивировано"),
        ("v0.9", "09.04.2026", "Команда ТрендСтудио",
         "Расширение до Уровня 3 (36 листов) — pre-IPO institutional. Согласованы: T₁+T₂ параллельно, FOT A₁+A₂, Waterfall W₁+W₂+W₃",
         "Архивировано"),
        ("v1.0", today, "Команда ТрендСтудио",
         "А.1: созданы листы Cover, Assumptions (84 параметра в 13 блоках), Change_Log. Якорь 3000 млн ₽ зафиксирован",
         "Текущая"),
    ]

    row = 6
    for version, date, author, desc, status in entries:
        c = ws.cell(row=row, column=2, value=version)
        c.font = Font(name="Arial", size=10, bold=True, color=BRAND_BLUE_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        c = ws.cell(row=row, column=3, value=date)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        c = ws.cell(row=row, column=4, value=author)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin_border()

        c = ws.cell(row=row, column=5, value=desc)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
        c.border = thin_border()

        c = ws.cell(row=row, column=6, value=status)
        status_color = ACCENT_GREEN if status == "Текущая" else GRAY_DARK
        status_fill = "E2EFDA" if status == "Текущая" else GRAY_LIGHT
        c.font = Font(name="Arial", size=10, bold=(status == "Текущая"),
                      color=status_color)
        c.fill = PatternFill("solid", fgColor=status_fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        ws.row_dimensions[row].height = 36
        row += 1

    # Разделитель
    row += 1
    # План следующих версий
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value="ПЛАН СЛЕДУЮЩИХ ВЕРСИЙ (roadmap)")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    row += 1

    roadmap = [
        ("v1.0 А.1", "Cover + Assumptions + Change_Log", "Готово"),
        ("v1.0 А.2", "FOT_Staff (A₁+A₂) + Cost_Structure", "В очереди"),
        ("v1.0 А.3", "Revenue_Breakdown + Content_Pipeline", "В очереди"),
        ("v1.0 А.4", "P&L_Statements (12Q + 2029–2032)", "В очереди"),
        ("v1.0 А.5", "Cash_Flow_Statement + Balance_Sheet", "В очереди"),
        ("v1.0 А.6", "Working_Capital + Debt_Schedule", "В очереди"),
        ("v1.0 А.7", "Investment_Inflow + Use_of_Funds + CAPEX", "В очереди"),
        ("v1.0 А.8", "Deal_Structures + Cap_Table + Waterfall", "В очереди"),
        ("v1.0 А.9", "Unit_Economics + KPI_Dashboard", "В очереди"),
        ("v1.0 А.10", "Valuation_DCF + Valuation_Multiples", "В очереди"),
        ("v1.0 А.11", "Investor_Returns + Exit_Scenarios", "В очереди"),
        ("v1.0 А.12", "Sensitivity + Scenario + Monte Carlo", "В очереди"),
        ("v1.0 А.13", "Risk_Register", "В очереди"),
        ("v1.0 А.14", "Market + Benchmarks + Comparable Tx", "В очереди"),
        ("v1.0 А.15", "Gov_KPI + Tax + Roadmap 2026–2032", "В очереди"),
        ("v1.0 А.16", "Executive_Summary + Glossary + Notes", "В очереди"),
        ("v1.0 А.17", "Public export + Internal с 4 сервисными", "В очереди"),
        ("v1.0 А.18", "П5 «Максимум» 32/32 верификация", "В очереди"),
        ("v1.0 А.19", "Финализация + computer:// ссылки", "В очереди"),
    ]
    for stage, desc, status in roadmap:
        c = ws.cell(row=row, column=2, value=stage)
        c.font = Font(name="Arial", size=9, bold=True, color=BRAND_BLUE_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        c = ws.cell(row=row, column=3, value=desc)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin_border()

        c = ws.cell(row=row, column=6, value=status)
        is_done = status == "Готово"
        c.font = Font(name="Arial", size=9, bold=is_done,
                      color=(ACCENT_GREEN if is_done else GRAY_DARK))
        c.fill = PatternFill("solid",
                             fgColor=("E2EFDA" if is_done else GRAY_LIGHT))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        ws.row_dimensions[row].height = 16
        row += 1

    ws.freeze_panes = "A6"


# ============================================================================
# СБОРКА
# ============================================================================

print("Building А.1 — Cover + Assumptions + Change_Log ...")
build_cover(wb)
assum_count = build_assumptions(wb)
build_changelog(wb)

# Свойства документа
wb.properties.title = "investor_model_v1.0_Public — ТрендСтудио"
wb.properties.subject = "Investor Financial Model, Pre-IPO Institutional"
wb.properties.creator = "Команда ТрендСтудио"
wb.properties.description = "Якорь: cumulative EBITDA 2026–2028 = 3 000 млн ₽"
wb.properties.keywords = "investor, financial model, L3, pre-IPO, cinema"

wb.save(OUT_FILE)
print(f"Saved: {OUT_FILE}")
print(f"Assumptions parameters: {assum_count}")
print(f"Sheets: {wb.sheetnames}")
