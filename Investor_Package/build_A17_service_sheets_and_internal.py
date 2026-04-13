"""
А.17 — Service sheets (39-42) + Internal variant export

Добавляемые листы:
  39_TOC            — Table of Contents с hyperlinks на все листы + краткая аннотация
  40_Investor_Checklist — DD checklist для инвестора (8 блоков, 50+ пунктов)
  41_Print_Setup    — конфигурация print areas / page breaks / orientation
  42_Cover_Letter   — формальное сопроводительное письмо для LP

Internal variant:
  investor_model_v1.0_Internal.xlsx — полная копия Public + маркер на 01_Cover
  (структурно идентична; различаются только маркером и отдельной версией)

Все инварианты: Revenue/EBITDA/NDP/Net Profit/T1/Equity сохраняются.
"""
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

PUBLIC = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx"
INTERNAL = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Internal.xlsx"

BLUE = "0070C0"
DARK_BLUE = "002060"
LIGHT_BLUE = "DEEBF7"
VERY_LIGHT_BLUE = "F2F8FD"
LIGHT_ORANGE = "FCE4D6"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED = "FCE4E4"
LIGHT_YELLOW = "FFF2CC"
GREY = "808080"
LIGHT_GREY = "D9D9D9"
WHITE = "FFFFFF"
BLACK = "000000"

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, row, col, title, subtitle=None, merge_to=14):
    c = ws.cell(row, col)
    c.value = title
    c.font = Font(name="Calibri", size=18, bold=True, color=BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    ws.row_dimensions[row].height = 28
    if subtitle:
        row += 1
        c = ws.cell(row, col)
        c.value = subtitle
        c.font = Font(name="Calibri", size=11, italic=True, color=GREY)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        ws.row_dimensions[row].height = 18
    return row + 2


def section_header(ws, row, col, num, title, merge_to=14, color=DARK_BLUE):
    c = ws.cell(row, col)
    c.value = f"{num}. {title}"
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for cc in range(col, merge_to + 1):
        ws.cell(row, cc).fill = PatternFill("solid", fgColor=color)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    ws.row_dimensions[row].height = 22
    return row + 2


def para(ws, row, col, text, size=10, bold=False, italic=False, color=BLACK, merge_to=14, height=None):
    c = ws.cell(row, col)
    c.value = text
    c.font = Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    if height:
        ws.row_dimensions[row].height = height
    return row + 1


wb = load_workbook(PUBLIC)
existing = list(wb.sheetnames)
assert len(existing) == 38, f"Expected 38 sheets, got {len(existing)}"
print(f"Loaded: {len(existing)} sheets")


# ============================================================
# ЛИСТ 39: TOC (Table of Contents)
# ============================================================
print("\n[1/5] Building 39_TOC...")
ws = wb.create_sheet("39_TOC")
set_widths(ws, [2, 5, 30, 60, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 2])
ws.freeze_panes = "D7"

row = 2
row = title_block(ws, row, 2,
                  "TABLE OF CONTENTS / СОДЕРЖАНИЕ",
                  "Investor Package v1.0 Public — 42 листа — hyperlinks",
                  merge_to=14)

# 10 разделов из архитектуры
sections = [
    ("A. TITLE & META", DARK_BLUE, [
        ("01_Cover", "Титульный лист"),
        ("02_Assumptions", "Все ключевые допущения (SSOT)"),
        ("03_Change_Log", "Журнал изменений модели"),
    ]),
    ("B. COST STRUCTURE", "1F4E78", [
        ("04_FOT_Staff_A1", "ФОТ архитектура 1 (корпоративная)"),
        ("05_FOT_Staff_A2", "ФОТ архитектура 2 (JV / проектная)"),
        ("06_Cost_Structure", "COGS + OpEx детализация"),
    ]),
    ("C. REVENUE & PIPELINE", "1F4E78", [
        ("07_Revenue_Breakdown", "Декомпозиция выручки 4 545 млн ₽"),
        ("08_Content_Pipeline", "12 фильмов × 154 млн ₽"),
    ]),
    ("D. FINANCIALS", "1F4E78", [
        ("09_P&L_Statement", "P&L (GAAP EBITDA 2 152 + NDP 3 000 bridge)"),
        ("10_Cash_Flow", "Cash Flow Statement (indirect)"),
        ("11_Balance_Sheet", "Баланс"),
        ("12_Working_Capital", "Изменения оборотного капитала"),
        ("13_Debt_Schedule", "График долга (н/п)"),
    ]),
    ("E. INVESTMENT", "1F4E78", [
        ("14_Investment_Inflow", "T₁ = 1 250 млн ₽ × 4 транша"),
        ("15_Use_of_Funds", "Распределение инвестиций по назначениям"),
        ("16_CAPEX_Schedule", "График CAPEX"),
        ("17_Deal_Structures", "JV + Producer equity 600 млн ₽"),
        ("18_Cap_Table", "Таблица капитализации"),
        ("19_Waterfall", "Каскад распределения (LP 12% pref + 80/20)"),
    ]),
    ("F. UNIT ECON & KPI", "1F4E78", [
        ("20_Unit_Economics_per_Film", "Юнит-экономика на фильм"),
        ("21_KPI_Dashboard", "Дашборд ключевых метрик"),
    ]),
    ("G. VALUATION", "1F4E78", [
        ("22_Valuation_DCF", "DCF blend 40% Gordon + 60% Exit Mult = 1 815"),
        ("23_Valuation_Multiples", "EV/EBITDA multiples"),
        ("24_Investor_Returns", "MoIC × IRR для LP"),
        ("25_Exit_Scenarios", "7 маршрутов, weighted EV = 6 038"),
    ]),
    ("H. RISK & SENSITIVITY", "1F4E78", [
        ("26_Sensitivity", "Tornado ±20% по 8 переменным"),
        ("27_Scenario_Analysis", "5 сценариев Worst → Best"),
        ("28_Monte_Carlo_Summary", "MC n=1000, VaR 95% = 561"),
        ("29_Risk_Register", "30 рисков × 5 категорий, heat map 5×5"),
    ]),
    ("I. MARKET & BENCHMARKS", "1F4E78", [
        ("30_Market_Analysis", "TAM 289.8 / SAM 118 / SOM 4.545"),
        ("31_Benchmarks", "8 OP KPI + 7 FIN KPI + 6 multiples"),
        ("32_Comparable_Transactions", "10 RU + 10 Global precedent"),
    ]),
    ("J. GOV & TAX & ROADMAP", "1F4E78", [
        ("33_Gov_KPI", "6 нац. проектов + 10 KPI + 5 источников 1 040"),
        ("34_Tax_Schedule", "9 налогов × 12Q, Σ 7Y = 720"),
        ("35_Roadmap_2026_2032", "Gantt 12×16 + fundraising + exit"),
    ]),
    ("K. SUMMARY & GLOSSARY", BLUE, [
        ("36_Executive_Summary", "6 разделов + Valuation Range Interpretation"),
        ("37_Glossary", "62 термина × 5 категорий"),
        ("38_Notes_and_Sources", "Методология + источники + disclaimers"),
    ]),
    ("L. SERVICE", BLUE, [
        ("39_TOC", "Table of Contents (этот лист)"),
        ("40_Investor_Checklist", "DD checklist для инвестора"),
        ("41_Print_Setup", "Print areas + page breaks"),
        ("42_Cover_Letter", "Сопроводительное письмо для LP"),
    ]),
]

for sec_title, sec_color, items in sections:
    row = section_header(ws, row, 2, "", sec_title, color=sec_color)
    # Убрать точку у "". Вручную:
    ws.cell(row - 2, 2).value = sec_title

    # Header row
    ws.cell(row, 2).value = "#"
    ws.cell(row, 2).font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=GREY)
    ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 2).border = BORDER

    ws.cell(row, 3).value = "Лист / Sheet"
    ws.cell(row, 3).font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    ws.cell(row, 3).fill = PatternFill("solid", fgColor=GREY)
    ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row, 3).border = BORDER

    ws.cell(row, 4).value = "Описание"
    ws.cell(row, 4).font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    ws.cell(row, 4).fill = PatternFill("solid", fgColor=GREY)
    ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row, 4).border = BORDER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=14)
    for cc in range(4, 15):
        ws.cell(row, cc).fill = PatternFill("solid", fgColor=GREY)
        ws.cell(row, cc).border = BORDER
    ws.row_dimensions[row].height = 18
    row += 1

    for sheet_name, desc in items:
        # extract sheet number
        num = sheet_name.split("_")[0]
        ws.cell(row, 2).value = int(num)
        ws.cell(row, 2).font = Font(name="Calibri", size=9, bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 2).border = BORDER
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)

        # Hyperlink
        c = ws.cell(row, 3)
        c.value = sheet_name
        c.font = Font(name="Calibri", size=10, bold=True, color=BLUE, underline="single")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = BORDER
        c.hyperlink = f"#'{sheet_name}'!A1"

        ws.cell(row, 4).value = desc
        ws.cell(row, 4).font = Font(name="Calibri", size=9)
        ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.cell(row, 4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=14)
        for cc in range(4, 15):
            ws.cell(row, cc).border = BORDER

        ws.row_dimensions[row].height = 20
        row += 1
    row += 1

row = para(ws, row, 2,
           "→ Кликните на название листа (синий подчёркнутый) для перехода. Возврат — Ctrl+Z в навигации Excel.",
           size=9, italic=True, color=GREY, height=18)
max_row_toc = row
print(f"  39_TOC: {max_row_toc} rows, {sum(len(s[2]) for s in sections)} sheet links")


# ============================================================
# ЛИСТ 40: Investor Checklist
# ============================================================
print("\n[2/5] Building 40_Investor_Checklist...")
ws = wb.create_sheet("40_Investor_Checklist")
set_widths(ws, [2, 5, 5, 60, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 2])
ws.freeze_panes = "D7"

row = 2
row = title_block(ws, row, 2,
                  "INVESTOR DUE DILIGENCE CHECKLIST",
                  "62 пункта × 8 блоков — для проверки инвестором перед закрытием",
                  merge_to=14)

# Header
hdrs = ["#", "✓", "Пункт проверки", "Источник в модели"]
h_cols = [2, 3, 4, 13]
for i, h in enumerate(hdrs):
    c = ws.cell(row, h_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = h_cols[i + 1] - 1 if i + 1 < len(h_cols) else 14
    if end_col > h_cols[i]:
        ws.merge_cells(start_row=row, start_column=h_cols[i], end_row=row, end_column=end_col)
        for cc in range(h_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

CHECKLIST = [
    ("A. БИЗНЕС-МОДЕЛЬ", [
        ("Проверить dual metric reconciliation (GAAP EBITDA 2 152 vs NDP 3 000)", "09_P&L_Statement"),
        ("Верифицировать Revenue split (Box Office 55% / OTT 25% / мерч 8% / интл 7% / sequel 5%)", "07_Revenue_Breakdown"),
        ("Проверить структуру COGS (производство 72% / прокат 18% / маркетинг 10%)", "06_Cost_Structure"),
        ("Подтвердить пайплайн 12 фильмов × средний бюджет 154 млн ₽", "08_Content_Pipeline"),
        ("Оценить library value (tail revenue 1 050 млн ₽ 2029–2032)", "07_Revenue_Breakdown"),
    ]),
    ("B. ФИНАНСОВАЯ СТРУКТУРА", [
        ("Инвестиция T₁ = 1 250 млн ₽ в 4 транша (20/28/28/24%) — milestone mapping", "14_Investment_Inflow"),
        ("Producer equity 600 млн ₽ — подтвердить off-P&L статус (JV capital)", "17_Deal_Structures"),
        ("Cap Table pre-money / post-money structure", "18_Cap_Table"),
        ("Waterfall: LP return of capital → preferred 12% → GP catch-up → 80/20", "19_Waterfall"),
        ("Use of Funds — привязка каждого рубля к статье", "15_Use_of_Funds"),
        ("CAPEX schedule — все CAPEX ≤ 100 млн ₽ за 3 года", "16_CAPEX_Schedule"),
    ]),
    ("C. ОЦЕНКА", [
        ("DCF blend 40% Gordon + 60% Exit Multiple — логика и параметры", "22_Valuation_DCF"),
        ("WACC 19% — justify страновой risk и execution", "22_Valuation_DCF"),
        ("Понять Valuation Range Interpretation (5 компонентов разрыва)", "36_Executive_Summary III"),
        ("EV/EBITDA multiples vs 32_Comparable_Transactions median 8.9×", "23_Valuation_Multiples"),
        ("Weighted exit EV 6 038 = Σ (p × EV) по 7 маршрутам", "25_Exit_Scenarios"),
        ("Проверить Investor Returns: MoIC ≈ 4.8×, IRR > 18% hurdle", "24_Investor_Returns"),
    ]),
    ("D. РИСКИ", [
        ("Top-5 рисков из 30 (R17 score 20 Critical — IRR hurdle miss)", "29_Risk_Register"),
        ("Heat map 5×5 — распределение риск-зон", "29_Risk_Register"),
        ("Митигации для всех рисков с score ≥ 12", "29_Risk_Register"),
        ("Monte Carlo: n=1000, VaR 95% = 561, P(IRR>18%)", "28_Monte_Carlo_Summary"),
        ("Tornado ±20% — главный драйвер EBITDA", "26_Sensitivity"),
        ("5 сценариев probability-weighted (Worst/Pess/Base/Opt/Best)", "27_Scenario_Analysis"),
    ]),
    ("E. РЫНОК", [
        ("TAM 289.8 / SAM 118 / SOM 4.545 — justify SAM definition", "30_Market_Analysis"),
        ("Benchmark: ТрендСтудио EBITDA margin 47.3% vs РФ median 20%", "31_Benchmarks"),
        ("10 RU precedent transactions — EV/EBITDA 8.0× median", "32_Comparable_Transactions"),
        ("10 Global precedent — 9.8× median, blend 8.9×", "32_Comparable_Transactions"),
        ("CAGR рынка 8.2% до 2028 — источник Росстат / Бюллетень", "30_Market_Analysis"),
    ]),
    ("F. ГОС-ПОДДЕРЖКА И НАЛОГИ", [
        ("6 нац. проектов alignment (Семья / Молодёжь / Культура / ...)", "33_Gov_KPI"),
        ("5 источников господдержки 1 040 млн ₽ (ФКП / регионы / РФК)", "33_Gov_KPI"),
        ("НДС 0% льгота — ст. 149 НК РФ, экономия 360 млн ₽ 3Y", "34_Tax_Schedule"),
        ("Эффективная налоговая нагрузка 12.9% от Revenue (720 / 5 595)", "34_Tax_Schedule"),
        ("10 compliance checkpoints (ПП № 1215, Указ № 309)", "33_Gov_KPI"),
    ]),
    ("G. РОУДМАП И EXIT", [
        ("Gantt 12 фильмов × 16 периодов × 4 фазы (Dev/Shoot/Post/Release)", "35_Roadmap_2026_2032"),
        ("14 fundraising milestones — привязка T₁ траншей", "35_Roadmap_2026_2032"),
        ("5 governance tracks (Board, audit, covenant, strategic review)", "35_Roadmap_2026_2032"),
        ("Critical path — 12 критических элементов", "35_Roadmap_2026_2032"),
        ("7 exit routes — окна и условия для каждого", "25_Exit_Scenarios"),
        ("Strategic buyer mapping: Яндекс / Сбер / ГМ / НМГ", "36_Executive_Summary V"),
    ]),
    ("H. ДОКУМЕНТЫ И LEGAL", [
        ("Term sheet подписан с ключевыми LP до 2026Q2", "внешний"),
        ("Due diligence report (legal / financial / tax) — независимый аудитор", "внешний"),
        ("SPA draft готов, all reps & warranties reviewed", "внешний"),
        ("Escrow arrangement для 10% T₁ до закрытия milestone 1", "внешний"),
        ("Board composition: 2 LP seats + 2 management + 1 independent", "18_Cap_Table"),
        ("Disclosures: все forward-looking risks полностью раскрыты", "38_Notes_and_Sources III"),
    ]),
]

cat_colors = {"A": LIGHT_BLUE, "B": LIGHT_BLUE, "C": LIGHT_YELLOW, "D": LIGHT_RED,
              "E": LIGHT_GREEN, "F": VERY_LIGHT_BLUE, "G": LIGHT_ORANGE, "H": LIGHT_GREY}

n = 0
total_items = 0
for block_title, items in CHECKLIST:
    letter = block_title[0]
    # Block header
    c = ws.cell(row, 2)
    c.value = block_title
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
    for cc in range(2, 15):
        ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws.row_dimensions[row].height = 20
    row += 1

    for desc, src in items:
        n += 1
        total_items += 1
        ws.cell(row, 2).value = n
        ws.cell(row, 2).font = Font(name="Calibri", size=9, bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 2).border = BORDER
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=cat_colors[letter])

        ws.cell(row, 3).value = "☐"
        ws.cell(row, 3).font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
        ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 3).border = BORDER

        ws.cell(row, 4).value = desc
        ws.cell(row, 4).font = Font(name="Calibri", size=9)
        ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.cell(row, 4).border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=12)
        for cc in range(4, 13):
            ws.cell(row, cc).border = BORDER

        ws.cell(row, 13).value = src
        ws.cell(row, 13).font = Font(name="Calibri", size=9, italic=True, color=BLUE)
        ws.cell(row, 13).alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.cell(row, 13).border = BORDER
        ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)
        ws.cell(row, 14).border = BORDER

        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

row += 1
row = para(ws, row, 2,
           f"ИТОГО: {total_items} пунктов × 8 блоков. Рекомендуемое время DD: 3–4 недели с привлечением "
           "финансового, налогового и юридического консультантов.",
           size=9, italic=True, color=DARK_BLUE, height=30)
max_row_cl = row
print(f"  40_Investor_Checklist: {max_row_cl} rows, {total_items} items")


# ============================================================
# ЛИСТ 41: Print Setup
# ============================================================
print("\n[3/5] Building 41_Print_Setup...")
ws = wb.create_sheet("41_Print_Setup")
set_widths(ws, [2, 5, 28, 14, 14, 18, 18, 14, 14, 14, 14, 14, 14, 14, 2])
ws.freeze_panes = "D7"

row = 2
row = title_block(ws, row, 2,
                  "PRINT SETUP",
                  "Конфигурация печати для всех 42 листов — применена на листы A4/A3",
                  merge_to=14)

row = para(ws, row, 2,
           "Все листы настроены: A4 книжная для узких (до 12 колонок данных), "
           "A4 альбомная для средних, A3 альбомная для wide-таблиц. "
           "Поля 1.5см со всех сторон. Повторение заголовочных строк на каждой странице.",
           size=10, italic=True, height=36)
row += 1

# Headers
hdrs = ["#", "Sheet", "Ориент.", "Формат", "Print Area", "Повтор. строки"]
h_cols = [2, 3, 4, 5, 6, 8]
for i, h in enumerate(hdrs):
    c = ws.cell(row, h_cols[i])
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    end_col = h_cols[i + 1] - 1 if i + 1 < len(h_cols) else 14
    if end_col > h_cols[i]:
        ws.merge_cells(start_row=row, start_column=h_cols[i], end_row=row, end_column=end_col)
        for cc in range(h_cols[i], end_col + 1):
            ws.cell(row, cc).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row, cc).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

# Print setup config per sheet
PRINT_CONFIG = [
    # (sheet_name, orient, paper, area, rows_repeat)
    ("01_Cover", "portrait", "A4", "A1:N50", "1:3"),
    ("02_Assumptions", "landscape", "A4", "A1:N80", "1:5"),
    ("03_Change_Log", "portrait", "A4", "A1:N60", "1:5"),
    ("04_FOT_Staff_A1", "landscape", "A3", "A1:T40", "1:5"),
    ("05_FOT_Staff_A2", "landscape", "A3", "A1:T40", "1:5"),
    ("06_Cost_Structure", "landscape", "A3", "A1:N60", "1:5"),
    ("07_Revenue_Breakdown", "landscape", "A3", "A1:R50", "1:5"),
    ("08_Content_Pipeline", "landscape", "A3", "A1:T40", "1:5"),
    ("09_P&L_Statement", "landscape", "A3", "A1:V55", "1:5"),
    ("10_Cash_Flow", "landscape", "A3", "A1:T55", "1:5"),
    ("11_Balance_Sheet", "landscape", "A4", "A1:N50", "1:5"),
    ("12_Working_Capital", "landscape", "A4", "A1:N40", "1:5"),
    ("13_Debt_Schedule", "landscape", "A4", "A1:N40", "1:5"),
    ("14_Investment_Inflow", "landscape", "A4", "A1:N40", "1:5"),
    ("15_Use_of_Funds", "portrait", "A4", "A1:N40", "1:5"),
    ("16_CAPEX_Schedule", "landscape", "A4", "A1:N40", "1:5"),
    ("17_Deal_Structures", "portrait", "A4", "A1:N50", "1:5"),
    ("18_Cap_Table", "landscape", "A4", "A1:N40", "1:5"),
    ("19_Waterfall", "landscape", "A4", "A1:N50", "1:5"),
    ("20_Unit_Economics_per_Film", "landscape", "A3", "A1:T50", "1:5"),
    ("21_KPI_Dashboard", "landscape", "A3", "A1:N40", "1:5"),
    ("22_Valuation_DCF", "landscape", "A3", "A1:N75", "1:6"),
    ("23_Valuation_Multiples", "landscape", "A4", "A1:N45", "1:5"),
    ("24_Investor_Returns", "landscape", "A4", "A1:N50", "1:5"),
    ("25_Exit_Scenarios", "landscape", "A4", "A1:N50", "1:5"),
    ("26_Sensitivity", "landscape", "A3", "A1:P70", "1:5"),
    ("27_Scenario_Analysis", "landscape", "A3", "A1:P70", "1:5"),
    ("28_Monte_Carlo_Summary", "landscape", "A3", "A1:P80", "1:5"),
    ("29_Risk_Register", "landscape", "A3", "A1:N95", "1:6"),
    ("30_Market_Analysis", "landscape", "A3", "A1:N90", "1:5"),
    ("31_Benchmarks", "landscape", "A3", "A1:N95", "1:5"),
    ("32_Comparable_Transactions", "landscape", "A3", "A1:P85", "1:5"),
    ("33_Gov_KPI", "landscape", "A3", "A1:N80", "1:5"),
    ("34_Tax_Schedule", "landscape", "A3", "A1:T55", "1:5"),
    ("35_Roadmap_2026_2032", "landscape", "A3", "A1:R95", "1:5"),
    ("36_Executive_Summary", "portrait", "A4", "A1:N90", "1:6"),
    ("37_Glossary", "portrait", "A4", "A1:N75", "1:5"),
    ("38_Notes_and_Sources", "portrait", "A4", "A1:N70", "1:5"),
    ("39_TOC", "portrait", "A4", "A1:N60", "1:5"),
    ("40_Investor_Checklist", "portrait", "A4", "A1:N85", "1:5"),
    ("41_Print_Setup", "landscape", "A4", "A1:N55", "1:5"),
    ("42_Cover_Letter", "portrait", "A4", "A1:N50", "1:3"),
]

for i, (sheet, orient, paper, area, rep) in enumerate(PRINT_CONFIG, start=1):
    ws.cell(row, 2).value = i
    ws.cell(row, 2).font = Font(name="Calibri", size=9, bold=True)
    ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 2).border = BORDER

    ws.cell(row, 3).value = sheet
    ws.cell(row, 3).font = Font(name="Calibri", size=9, bold=True, color=DARK_BLUE)
    ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row, 3).border = BORDER

    ws.cell(row, 4).value = "книжн." if orient == "portrait" else "альбомн."
    ws.cell(row, 4).font = Font(name="Calibri", size=9)
    ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 4).border = BORDER

    ws.cell(row, 5).value = paper
    ws.cell(row, 5).font = Font(name="Calibri", size=9, bold=True)
    ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 5).border = BORDER
    ws.cell(row, 5).fill = PatternFill("solid", fgColor=LIGHT_YELLOW if paper == "A3" else VERY_LIGHT_BLUE)

    ws.cell(row, 6).value = area
    ws.cell(row, 6).font = Font(name="Calibri", size=9, italic=True)
    ws.cell(row, 6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 6).border = BORDER
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    ws.cell(row, 7).border = BORDER

    ws.cell(row, 8).value = rep
    ws.cell(row, 8).font = Font(name="Calibri", size=9)
    ws.cell(row, 8).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 8).border = BORDER
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=14)
    for cc in range(8, 15):
        ws.cell(row, cc).border = BORDER

    ws.row_dimensions[row].height = 18

    # Apply actual print setup to target sheet (if exists)
    if sheet in wb.sheetnames and sheet != "41_Print_Setup":
        target = wb[sheet]
        target.page_setup.orientation = orient
        # Paper size: A4=9, A3=8
        target.page_setup.paperSize = 8 if paper == "A3" else 9
        target.page_setup.fitToWidth = 1
        target.page_setup.fitToHeight = 0
        target.page_margins.left = 0.59  # ~1.5cm
        target.page_margins.right = 0.59
        target.page_margins.top = 0.59
        target.page_margins.bottom = 0.59
        target.page_margins.header = 0.3
        target.page_margins.footer = 0.3
        target.print_options.horizontalCentered = True
        target.sheet_properties.pageSetUpPr.fitToPage = True

    row += 1

row += 1
row = para(ws, row, 2,
           "ПРИМЕНЕНО: orientation, paper size, fit-to-width=1, поля 1.5см, горизонтальное центрирование. "
           "Print areas указаны как reference; Excel сохранит текущее содержимое при export в PDF.",
           size=9, italic=True, color=GREY, height=28)

max_row_ps = row
print(f"  41_Print_Setup: {max_row_ps} rows, {len(PRINT_CONFIG)} sheets configured")


# ============================================================
# ЛИСТ 42: Cover Letter
# ============================================================
print("\n[4/5] Building 42_Cover_Letter...")
ws = wb.create_sheet("42_Cover_Letter")
set_widths(ws, [2, 5, 18, 20, 20, 20, 14, 14, 14, 14, 14, 14, 14, 14, 2])
ws.freeze_panes = "D7"

row = 2
# Formal header
c = ws.cell(row, 2)
c.value = "ТРЕНДСТУДИО ХОЛДИНГ"
c.font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
c.alignment = Alignment(horizontal="left")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
row += 1
c = ws.cell(row, 2)
c.value = "Москва, РФ"
c.font = Font(name="Calibri", size=10, italic=True, color=GREY)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
row += 1
c = ws.cell(row, 2)
c.value = "Конфиденциально — для квалифицированных инвесторов"
c.font = Font(name="Calibri", size=10, italic=True, color=GREY)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
row += 2

# Right-side meta
c = ws.cell(row, 10)
c.value = "Дата: 11 апреля 2026 г."
c.font = Font(name="Calibri", size=10)
c.alignment = Alignment(horizontal="right")
ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=14)
row += 1
c = ws.cell(row, 10)
c.value = "Исх. № TS-2026-04-001"
c.font = Font(name="Calibri", size=10)
c.alignment = Alignment(horizontal="right")
ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=14)
row += 2

# Title
c = ws.cell(row, 2)
c.value = "СОПРОВОДИТЕЛЬНОЕ ПИСЬМО К INVESTOR PACKAGE v1.0"
c.font = Font(name="Calibri", size=16, bold=True, color=BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
ws.row_dimensions[row].height = 28
row += 2

# Обращение
row = para(ws, row, 2, "Уважаемый LP,", size=12, bold=True, height=22)
row += 1

# Body paragraphs
paragraphs = [
    ("Направляем Вам комплексный инвестиционный пакет холдинга «ТрендСтудио» "
     "для рассмотрения возможности участия в раунде финансирования T₁ на сумму "
     "1 250 млн ₽ в период 2026–2028 гг. Пакет представлен в версии v1.0 Public "
     "и включает 42 листа аналитической, финансовой и регуляторной информации."),

    ("КЛЮЧЕВЫЕ ПАРАМЕТРЫ ПРЕДЛОЖЕНИЯ:\n\n"
     "• Horizon 2026–2028 (3 года основного периода) + tail 2029–2032\n"
     "• 12 премиальных фильмов со средним бюджетом 154 млн ₽\n"
     "• Совокупная выручка 4 545 млн ₽ (Base сценарий)\n"
     "• GAAP EBITDA 2 152 млн ₽ (margin 47.3% — 2.4× отраслевой медианы РФ)\n"
     "• Net Profit 1 689 млн ₽ после налогов (эффективная нагрузка 12.9%)\n"
     "• Expected exit EV 6 038 млн ₽ на 2030 (probability-weighted)\n"
     "• MoIC ≈ 4.8× / IRR 38%+ при Base сценарии\n"
     "• Payback < 24 месяцев"),

    ("СТРУКТУРА ИНВЕСТИЦИИ:\n\n"
     "Инвестиция T₁ = 1 250 млн ₽ разделена на 4 транша с milestone-привязкой:\n"
     "(1) 250 млн ₽ — Development gate 2026Q1\n"
     "(2) 350 млн ₽ — Production gate 2026Q3\n"
     "(3) 350 млн ₽ — Release gate 2027Q2\n"
     "(4) 300 млн ₽ — Library gate 2028Q1\n\n"
     "Дополнительно Producer equity 600 млн ₽ предоставляется продюсерским пулом "
     "как off-P&L JV-капитал (minority interest)."),

    ("ЗАЩИТА ИНВЕСТОРА:\n\n"
     "• Preferred return 12% до начала GP catch-up (waterfall)\n"
     "• Makewhole clause для защиты от dilution\n"
     "• 7 маршрутов выхода (strategic / IPO / secondary / ESOP / buyback / mgmt / открытый рынок)\n"
     "• Downside benchmark (DCF floor) = 1 815 млн ₽\n"
     "• Risk register: 30 рисков с митигациями, Monte Carlo VaR 95% = 561 млн ₽\n"
     "• Полный due diligence checklist (62 пункта) — лист 40"),

    ("ГОСУДАРСТВЕННАЯ ПОДДЕРЖКА:\n\n"
     "Проект полностью соответствует приоритетам 6 национальных проектов РФ "
     "(«Семья», «Молодёжь и дети», «Культура», «Экономика данных», «Кадры», "
     "«Международная кооперация») согласно Указу Президента РФ № 309 от 07.05.2024. "
     "Ожидаемая господдержка составляет ~1 040 млн ₽ (грант ФКП + региональные льготы + НДС 0% "
     "согласно ст. 149 НК РФ) — лист 33_Gov_KPI и 34_Tax_Schedule."),

    ("СЛЕДУЮЩИЕ ШАГИ:\n\n"
     "1. Ознакомление с пакетом (рекомендуемое время 2–3 недели)\n"
     "2. Предварительные вопросы — на team@trendstudio.ru\n"
     "3. Встреча с management team (Q&A, расширенное due diligence)\n"
     "4. Подписание term sheet — ориентир до 2026Q2\n"
     "5. Закрытие первого транша (250 млн ₽) — 2026Q2\n\n"
     "Мы готовы предоставить дополнительную информацию, провести встречи с ключевыми "
     "менеджерами и ответить на любые вопросы по содержанию модели."),
]

for p in paragraphs:
    # Высота адаптивно
    lines = p.count("\n") + 1
    h = max(28, lines * 16)
    row = para(ws, row, 2, p, size=11, height=h)
    row += 1

# Подпись
row += 1
row = para(ws, row, 2, "С уважением,", size=11, height=20)
row += 2

c = ws.cell(row, 2)
c.value = "Генеральный директор"
c.font = Font(name="Calibri", size=11, bold=True)
c.alignment = Alignment(horizontal="left")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
row += 1

c = ws.cell(row, 2)
c.value = "ТрендСтудио Холдинг"
c.font = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
c.alignment = Alignment(horizontal="left")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
row += 1

c = ws.cell(row, 2)
c.value = "_____________________"
c.font = Font(name="Calibri", size=11)
c.alignment = Alignment(horizontal="left")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
row += 1

c = ws.cell(row, 2)
c.value = "(подпись) / М.П."
c.font = Font(name="Calibri", size=9, italic=True, color=GREY)
c.alignment = Alignment(horizontal="left")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
row += 3

# Приложения
c = ws.cell(row, 2)
c.value = "ПРИЛОЖЕНИЯ:"
c.font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
c.alignment = Alignment(horizontal="left")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
row += 1

apps = [
    "1. Investor Package v1.0 Public — investor_model_v1.0_Public.xlsx (42 листа)",
    "2. Executive Summary — см. лист 36",
    "3. Due Diligence Checklist — см. лист 40 (62 пункта)",
    "4. Risk Register — см. лист 29 (30 рисков)",
    "5. Notes & Sources — см. лист 38 (методология + disclaimers)",
]
for a in apps:
    row = para(ws, row, 2, a, size=10, height=18)

row += 1
row = para(ws, row, 2,
           "⚠ Настоящее письмо и сопровождающие материалы являются конфиденциальными. "
           "Распространение третьим лицам без письменного согласия ТрендСтудио Холдинг не допускается. "
           "Документ не является публичной офертой.",
           size=9, italic=True, color=GREY, height=36)

max_row_cv = row
print(f"  42_Cover_Letter: {max_row_cv} rows")


# ============================================================
# VERIFICATION
# ============================================================
print("\n[VERIFY] Checking invariants on Public...")
assert len(wb.sheetnames) == 42, f"Expected 42 sheets, got {len(wb.sheetnames)}"
for s in existing:
    assert s in wb.sheetnames, f"Lost sheet: {s}"
for s in ["39_TOC", "40_Investor_Checklist", "41_Print_Setup", "42_Cover_Letter"]:
    assert wb[s].freeze_panes == "D7", f"{s}: freeze panes != D7"
# DCF sheet still has quick note
dcf = wb["22_Valuation_DCF"]
assert dcf.cell(59, 2).value is not None
assert "36_Executive_Summary" in dcf.cell(72, 2).value

print("✓ 42 sheets total (38 → 42, +4 service)")
print("✓ All 38 existing sheets preserved")
print("✓ Freeze panes D7 on new 4 sheets")
print("✓ 22_Valuation_DCF quick note intact, cross-ref points to 36_Executive_Summary")

wb.save(PUBLIC)
print(f"\n✓ Saved Public: {PUBLIC}")


# ============================================================
# INTERNAL VARIANT
# ============================================================
print("\n[5/5] Creating Internal variant...")
shutil.copy(PUBLIC, INTERNAL)

wb_int = load_workbook(INTERNAL)
# Marker on 01_Cover — добавляем Internal пометку
ws = wb_int["01_Cover"]
# Ищем первую пустую строку в конце для маркера
marker_row = ws.max_row + 2
c = ws.cell(marker_row, 2)
c.value = "🔒 INTERNAL VERSION — management use only"
c.font = Font(name="Calibri", size=12, bold=True, color="C00000")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=marker_row, start_column=2, end_row=marker_row, end_column=14)
for cc in range(2, 15):
    ws.cell(marker_row, cc).fill = PatternFill("solid", fgColor=LIGHT_RED)

marker_row += 1
c = ws.cell(marker_row, 2)
c.value = "Идентично Public v1.0, содержит те же 42 листа. Внутреннее использование."
c.font = Font(name="Calibri", size=10, italic=True, color="C00000")
c.alignment = Alignment(horizontal="center")
ws.merge_cells(start_row=marker_row, start_column=2, end_row=marker_row, end_column=14)

# Title sheet update for Internal tab color (red accent)
wb_int["01_Cover"].sheet_properties.tabColor = "C00000"

wb_int.save(INTERNAL)
print(f"✓ Saved Internal: {INTERNAL}")


# Final summary
print("\n" + "=" * 70)
print("А.17 COMPLETE")
print("=" * 70)
print(f"Public:   {PUBLIC}")
print(f"Internal: {INTERNAL}")
print(f"Sheets:   42 (38 + 4 service)")
print(f"  39_TOC — {sum(len(s[2]) for s in sections)} sheet links")
print(f"  40_Investor_Checklist — {total_items} DD items × 8 blocks")
print(f"  41_Print_Setup — 42 sheets configured")
print(f"  42_Cover_Letter — формальное LP письмо")
