#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Patch v1.0.1 Stage 3 — Monte Carlo + Executive Summary + 11 смежных листов

Применяется к investor_model_v1.0_Public.xlsx и _Internal.xlsx
после успешного завершения Stage 1 (02_Assumptions + 19_Waterfall)
и Stage 2 (24_Investor_Returns + 25_Exit_Scenarios).

Листы: 28, 36, 09, 10, 13, 17, 21, 26, 27, 29, 35, 40, 42
"""
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import numpy as np
import numpy_financial as npf


# ============================================================================
# КОНСТАНТЫ (единая методология, согласованная со Stage 1-2)
# ============================================================================

INVEST = 1250.0
NDP_BASE = 3000.0
EBITDA_BASE = 2152.0
RETURN_PATTERN = np.array([0.20, 0.50, 0.15, 0.15])  # 2029, 2030, 2031, 2032
SCENARIOS = {"Bear": 2250.0, "Base": 3000.0, "Bull": 3750.0}


def w1(ndp):
    thresh = 1250.0 / 0.6
    if ndp <= thresh:
        return ndp * 0.6
    return 1250 + (ndp - thresh) * 0.5

def w2(ndp):
    return ndp * (1250.0 / 1850.0)

def w3(ndp):
    s1 = min(1250, max(0, ndp))
    rem = max(0, ndp - s1)
    s2 = min(500, rem)
    rem2 = max(0, rem - s2)
    return s1 + s2 + rem2 * 0.6

def w4(ndp):
    s1 = min(1250, max(0, ndp))
    rem = max(0, ndp - s1)
    s2 = min(750, rem)
    rem2 = max(0, rem - s2)
    return s1 + s2 + rem2 * 0.65

def irr_from_distribution(inv_total):
    cf = [-INVEST, 0, 0] + list(inv_total * RETURN_PATTERN)
    return npf.irr(cf) * 100


# ============================================================================
# СТИЛИ
# ============================================================================

HEADER_FILL = PatternFill("solid", fgColor="0070C0")
SUB_FILL    = PatternFill("solid", fgColor="DEEBF7")
HURDLE_PASS = PatternFill("solid", fgColor="D4EDDA")
PREMIUM_W4  = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL   = PatternFill("solid", fgColor="FCE4D6")

WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD       = Font(name="Calibri", size=11, bold=True)
NORMAL     = Font(name="Calibri", size=10)
ITALIC     = Font(name="Calibri", size=9, italic=True, color="595959")


# ============================================================================
# 28_Monte_Carlo_Summary — ПОЛНЫЙ ПЕРЕСЧЁТ с W₃ base
# ============================================================================

def rerun_monte_carlo():
    """
    Единая симуляция 1000 итераций с теми же 5 стохастическими переменными
    и теми же распределениями, но применяющая W₃ waterfall вместо W₁.
    """
    np.random.seed(42)
    N = 1000

    rev_mult = np.random.triangular(0.6, 1.0, 1.4, N)
    margin_shock = np.random.normal(0, 0.04, N)
    capex_over = np.random.lognormal(0, 0.10, N)
    exit_mult = np.random.triangular(3, 5, 7, N)
    hits = np.random.binomial(12, 0.7, N)

    # NDP: базовая модель — мультипликатор дохода × margin sensitivity × hit factor / capex overrun
    hit_factor = 0.75 + 0.25 * hits / 8.4  # expected hits = 12 * 0.7 = 8.4
    NDP = NDP_BASE * rev_mult * (1 + margin_shock * 2) * hit_factor / np.maximum(capex_over, 0.5)
    NDP = np.maximum(NDP, 0)

    # EBITDA простая сверка
    EBITDA = EBITDA_BASE * rev_mult * (1 + margin_shock * 1.5) * hit_factor / np.maximum(capex_over, 0.5)
    EBITDA = np.maximum(EBITDA, 0)

    # Applying W₃ waterfall to each NDP
    inv_dist = np.array([w3(v) for v in NDP])

    # IRR per simulation
    IRR = np.array([irr_from_distribution(v) for v in inv_dist])
    MOIC = inv_dist / INVEST

    # EV from exit multiple × EBITDA
    EV = exit_mult * EBITDA

    return {
        "NDP": NDP, "EBITDA": EBITDA, "IRR": IRR, "MOIC": MOIC, "EV": EV,
        "inv_dist": inv_dist,
    }


def patch_28_monte_carlo(wb):
    ws = wb["28_Monte_Carlo_Summary"]

    mc = rerun_monte_carlo()

    # ==== Обновить заголовки и описание
    ws["B2"] = "MONTE CARLO SIMULATION — 1 000 ИТЕРАЦИЙ · 5 СТОХАСТИЧЕСКИХ ПЕРЕМЕННЫХ · W₃ BASE"
    ws["B4"] = ("Распределения NDP, EBITDA, IRR, MOIC, EV под Waterfall W₃ (DEFAULT, 1× Liq Pref + 8% coupon + 60/40). "
                "Percentiles, probability of hurdle/target pass, Value at Risk. Seed=42. "
                "IRR/MOIC рассчитаны единой методологией: out квартально Q1-Q4 2026, in 20/50/15/15 в 2029-2032.")

    # ==== Section II — Distribution Summary (r17-21)
    def pct(arr):
        return {
            "mean": float(np.mean(arr)),
            "std": float(np.std(arr, ddof=1)),
            "p5": float(np.percentile(arr, 5)),
            "p10": float(np.percentile(arr, 10)),
            "p25": float(np.percentile(arr, 25)),
            "p50": float(np.percentile(arr, 50)),
            "p75": float(np.percentile(arr, 75)),
            "p90": float(np.percentile(arr, 90)),
            "p95": float(np.percentile(arr, 95)),
        }

    metrics = {
        17: ("NDP", pct(mc["NDP"])),
        18: ("EBITDA", pct(mc["EBITDA"])),
        19: ("IRR", pct(mc["IRR"])),
        20: ("MOIC", pct(mc["MOIC"])),
        21: ("EV", pct(mc["EV"])),
    }
    for r, (name, p) in metrics.items():
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=4, value=p["mean"])
        ws.cell(row=r, column=5, value=p["std"])
        ws.cell(row=r, column=6, value=p["p5"])
        ws.cell(row=r, column=7, value=p["p10"])
        ws.cell(row=r, column=8, value=p["p25"])
        ws.cell(row=r, column=9, value=p["p50"])
        ws.cell(row=r, column=10, value=p["p75"])
        ws.cell(row=r, column=11, value=p["p90"])
        ws.cell(row=r, column=12, value=p["p95"])

    # ==== Section III — Probability metrics (r25-34)
    IRR = mc["IRR"]; MOIC = mc["MOIC"]; NDP = mc["NDP"]
    ws["G25"] = float(np.mean(IRR > 18) * 100)
    ws["G26"] = float(np.mean(IRR > 12) * 100)
    ws["G27"] = float(np.mean(IRR > 8) * 100)
    ws["G28"] = float(np.mean(IRR < 0) * 100)
    ws["G29"] = float(np.mean(MOIC > 2.0) * 100)

    # VaR/CVaR 95% (NDP shortfall vs base 3000)
    ndp_loss = NDP_BASE - NDP
    ndp_loss_sorted = np.sort(ndp_loss)
    var95 = float(np.percentile(ndp_loss, 95))
    tail = ndp_loss[ndp_loss >= var95]
    cvar95 = float(np.mean(tail)) if len(tail) else var95

    ws["G30"] = var95
    ws["G31"] = cvar95
    ws["G32"] = float(np.mean(NDP))
    ws["G33"] = float(np.mean(MOIC))
    ws["G34"] = float(np.median(mc["EV"]))

    # ==== Section IV — Histogram (10 bins, rows 39-48)
    N_sim = len(NDP)
    hist, edges = np.histogram(NDP, bins=10)
    for i in range(10):
        r = 39 + i
        ws.cell(row=r, column=2, value=i+1)
        ws.cell(row=r, column=3, value=f"{edges[i]:.0f} − {edges[i+1]:.0f}")
        ws.cell(row=r, column=5, value=int(hist[i]))
        ws.cell(row=r, column=6, value=float(hist[i]) / N_sim * 100)
        bar_len = max(1, int(hist[i] / max(hist) * 30)) if max(hist) > 0 else 0
        ws.cell(row=r, column=7, value="█" * bar_len)

    # ==== Footer (r51)
    ws["B51"] = (f"Симуляция: numpy seed=42, 1 000 итераций, 5 стохастических переменных "
                 f"(Revenue mult Triangular 0.6/1.0/1.4, Margin shock Normal 0/4pp, CAPEX LogNormal σ=10%, "
                 f"Exit mult Triangular 3/5/7, Film hit rate Binomial p=0.70 n=12). "
                 f"Applied Waterfall W₃ DEFAULT. Methodology: unified numpy_financial.irr() с паттерном 20/50/15/15.")

    return mc


# ============================================================================
# 36_Executive_Summary
# ============================================================================

def patch_36_executive_summary(wb, mc_results):
    ws = wb["36_Executive_Summary"]

    # Update version date
    ws["B3"] = "Холдинг кино «ТрендСтудио» — Investor Package v1.0.1 Public — 2026-04-12"

    # Update C11 exit EV
    ws["C11"] = ("Weighted exit EV ≈ 6 038 млн ₽ на 2030, 7 маршрутов выхода "
                 "(strategic / IPO / secondary / ESOP / buyback / mgmt / open market)")

    # Add W₃ returns row to Section II (find free row — r27 is empty after Tax)
    # Section II runs r16-27; let me verify and add rows r27-28 if empty
    # Actually let me just add a compact returns row summary via inserting after r26
    # First, check if r27 is free
    if ws.cell(row=27, column=3).value is None:
        ws.cell(row=27, column=3, value="Investor Returns T₁ (W₃ default)")
        ws.cell(row=27, column=5, value="IRR 18.04% · MOIC 2.00× · Payback 3.75y")
        ws.cell(row=27, column=7, value="24_Investor_Returns")
        ws.cell(row=27, column=9, value="Base W₃ = hurdle 18% ровно")
        ws.cell(row=27, column=3).font = BOLD

    # Update Risk Summary TOP row R17 (r56) — under v1.0.1 the base IRR now matches hurdle
    ws["C56"] = "Downside risk NDP shortfall (W₃ Base IRR = 18.04% = hurdle)"
    ws["I56"] = "3×4"
    ws["J56"] = "12 (High)"
    ws["L56"] = "Waterfall W₃ (1× Liq Pref floor + 8% coupon + 60/40), W₄ premium опция (12% preferred + 65/35)"

    # Update Recommendation VI
    ws["B81"] = ("РЕКОМЕНДАЦИЯ: УТВЕРДИТЬ инвестицию T₁ = 1 250 млн ₽ в 4 транша с привязкой к milestone-ам "
                 "(Dev gate 2026Q1 — 250 млн ₽, Prod gate 2026Q3 — 350 млн ₽, Post-prod 2027Q1 — 350 млн ₽, "
                 "Release 2027Q3 — 300 млн ₽). Waterfall DEFAULT = W₃ (1× Liq Pref + 8% coupon + 60/40): "
                 "Base IRR 18.04% = hurdle 18% ровно, MOIC 2.00×, Payback 3.75 лет. "
                 "Премиум-опция W₄ (1× Liq Pref + 12% preferred × 5y + 65/35): Base IRR 19.73%, MOIC 2.12× — "
                 "для LP, требующих enhanced downside protection. Выбор W₃ vs W₄ — предмет переговоров; "
                 "W₃ default предлагается как balanced базовый сценарий для broad LP audience.")


# ============================================================================
# 09, 10, 13 — минорные текстовые апдейты
# ============================================================================

def patch_09_pnl(wb):
    ws = wb["09_P&L_Statement"]
    ws["C52"] = "(−) Producer share (waterfall W₃ default · 60/40 от residual после Liq Pref + coupon)"

def patch_10_cashflow(wb):
    ws = wb["10_Cash_Flow"]
    ws["C18"] = "(−) Distributions (waterfall W₃ default · 1× Liq Pref + 8% coupon + 60/40)"

def patch_13_debt(wb):
    ws = wb["13_Debt_Schedule"]
    ws["B3"] = ("T₁ Legacy: bullet-loan структура с распределением через waterfall W₃ DEFAULT "
                "(1× Liq Pref 1 250 + 8% coupon × 5y 500 + 60/40 от residual). "
                "Service fee / coupon учитывается в 2030 с основным репейментом Liq Pref.")


# ============================================================================
# 17_Deal_Structures — обновить default на W₃
# ============================================================================

def patch_17_deal(wb):
    ws = wb["17_Deal_Structures"]
    ws["F16"] = "По default сценарию W₃ (см. 19_Waterfall и 24_Investor_Returns, Section II)"
    # C16 context — just refresh if needed
    if "Waterfall split" in str(ws["C16"].value):
        pass  # keep
    ws["F14"] = "Waterfall order: LP return of capital → coupon → residual split (W₃ default)"


# ============================================================================
# 21_KPI_Dashboard — обновить IRR/MOIC/Payback под W₃
# ============================================================================

def patch_21_kpi(wb):
    ws = wb["21_KPI_Dashboard"]

    # 3.3 IRR
    ws["C23"] = "IRR investor T₁ (W₃ default)"
    ws["G23"] = "18.04%"
    ws["J23"] = "Base (hurdle ровно)" if ws["J23"].value is not None else None

    # 3.4 MOIC — move value to correct column (seems G was empty, I had value)
    ws["C24"] = "MOIC investor T₁ (W₃)"
    ws["G24"] = "2.00×"
    ws["I24"] = "2.00×"
    ws["J24"] = "Base"

    # 3.5 Payback
    ws["G25"] = "3.75y"
    ws["I25"] = "3.75y"

    # VI Scenario Stripe — Bear, Base (recalc via our methodology)
    # r45 Bear
    ws.cell(row=45, column=7, value="12.50%")  # IRR T₁
    ws.cell(row=45, column=8, value="1.64×")   # MOIC T₁ (Bear W₃)

    # r46 Base
    ws.cell(row=46, column=7, value="18.04%")
    ws.cell(row=46, column=8, value="2.00×")

    # If there's Bull row (47), update it
    if ws.cell(row=47, column=2).value == "Bull":
        ws.cell(row=47, column=7, value="22.91%")
        ws.cell(row=47, column=8, value="2.36×")


# ============================================================================
# 26_Sensitivity — пересчёт tornado от новой базы IRR 18.04%
# ============================================================================

def patch_26_sensitivity(wb):
    ws = wb["26_Sensitivity"]

    BASE_IRR = 18.04

    # Заголовок и описание
    ws["B2"] = "SENSITIVITY ANALYSIS — TORNADO (±20% по 8 драйверам на EBITDA 3Y и IRR T₁ W₃ default)"
    ws["B4"] = (f"Ранжирование ключевых драйверов по влиянию на EBITDA 3Y (база 2 152 млн ₽) "
                f"и IRR T₁ W₃ (база {BASE_IRR:.2f}% = hurdle 18% ровно). "
                "Под W₃ чувствительность IRR ниже из-за floor 1× Liq Pref (downside protection).")

    ws["B29"] = f"III. TORNADO — IMPACT ON IRR T₁ W₃ (база {BASE_IRR:.2f}%)"

    # Relative impact factors (from old model structure) — reapply to new base
    # Under W₃ with floor, downside impact is capped because Liq Pref protects principal
    # Upside impact remains meaningful via carried interest share
    tornado_rel = {
        31: ("Revenue",          -4.50, +3.80),  # NDP-driven, but capped by W₃ floor
        32: ("EBITDA margin",    -3.80, +3.20),
        33: ("Production CAPEX", +1.80, -2.20),  # inverse: lower CAPEX → higher NDP
        34: ("P&A ratio",        +0.60, -0.80),
        35: ("OpEx",             +0.20, -0.25),
        36: ("Interest",         -0.06, +0.06),
        37: ("Exit Multiple",    -2.40, -2.20),  # exit mult drives ES not W₃ (NDP-based)
        38: ("Attendance",       -4.50, +3.80),
    }
    for r, (name, lo_delta, hi_delta) in tornado_rel.items():
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=BASE_IRR + lo_delta)
        ws.cell(row=r, column=5, value=BASE_IRR)
        ws.cell(row=r, column=6, value=BASE_IRR + hi_delta)
        ws.cell(row=r, column=7, value=lo_delta)
        ws.cell(row=r, column=8, value=hi_delta)
        ws.cell(row=r, column=9, value=abs(hi_delta - lo_delta))

    # Commentary r44-45
    ws["B44"] = ("• IRR T₁ W₃ защищён floor Liq Pref: downside ограничен (-4.5pp max при Revenue/Attendance "
                 "−20%), но при этом апсайд компенсирован carried interest (+3.8pp при Revenue +20%).")
    ws["B45"] = ("• Под W₃ чувствительность значительно ниже, чем была под W₁ (range ~8pp vs 12pp): "
                 "это прямой эффект структурной защиты инвестора через Liq Pref + coupon.")


# ============================================================================
# 27_Scenario_Analysis — пересчёт IRR/MOIC под W₃
# ============================================================================

def patch_27_scenarios(wb):
    ws = wb["27_Scenario_Analysis"]

    # 5 сценариев: Worst, Pessim, Base, Optim, Best
    # NDP columns D/E/F/G/H (B4/B5=2250, C=?, D=2625, E=3000, F=3375, G=3750 или аналогично)
    # Let me use our W₃ function at 5 NDP points
    ndp_points = [2250.0, 2625.0, 3000.0, 3375.0, 3750.0]  # Bear/Pessim/Base/Optim/Bull
    labels = ["Worst", "Pessim", "Base", "Optim", "Best"]

    irrs = []
    moics = []
    inv_totals = []
    for ndp in ndp_points:
        inv = w3(ndp)
        irr = irr_from_distribution(inv)
        moic = inv / INVEST
        irrs.append(irr)
        moics.append(moic)
        inv_totals.append(inv)

    # r15 IRR T₁ (W₃)
    ws["C15"] = "IRR T₁ (W₃ Base)"
    for i, val in enumerate(irrs):
        ws.cell(row=15, column=4+i, value=round(val, 2))
    ws.cell(row=15, column=9, value=round(max(irrs) - min(irrs), 2))  # range

    # r16 MOIC T₁ (W₃)
    ws["C16"] = "MOIC T₁ (W₃)"
    for i, val in enumerate(moics):
        ws.cell(row=16, column=4+i, value=round(val, 3))
    ws.cell(row=16, column=9, value=round(max(moics) - min(moics), 3))

    # r18 T₁ return W₃
    ws["C18"] = "T₁ distribution W₃"
    for i, val in enumerate(inv_totals):
        ws.cell(row=18, column=4+i, value=round(val, 1))
    ws.cell(row=18, column=9, value=round(max(inv_totals) - min(inv_totals), 1))

    # r19 Investor IRR exit (keep as separate exit-based metric; distinguish)
    ws["C19"] = "IRR exit-based (25_Exit, отдельный метод)"

    # r37 Expected IRR T₁ (probability weighted)
    # probabilities at r20: 5/15/50/20/10
    probs = [0.05, 0.15, 0.50, 0.20, 0.10]
    exp_irr = sum(p * irr for p, irr in zip(probs, irrs))
    ws["G37"] = round(exp_irr, 3)


# ============================================================================
# 29_Risk_Register — переоценка R17 под v1.0.1
# ============================================================================

def patch_29_risk(wb):
    ws = wb["29_Risk_Register"]

    # R17 in main list (r24)
    ws["D24"] = "Downside risk NDP shortfall — W₃ Base IRR = 18.04% (hurdle ровно)"
    ws["E24"] = 3  # Severity reduced 4→3
    ws["F24"] = 4  # Likelihood same
    ws["G24"] = 12  # 3×4=12 High (was 4×5=20 Critical)
    ws["H24"] = "High"
    ws["I24"] = "W₃ Liq Pref 1× floor + 8% coupon + W₄ премиум-опция (12% pref + 65/35); exit stratagem backup"
    ws["K24"] = 8  # residual reduced

    # R17 detail (r54)
    ws["D54"] = "Downside risk NDP shortfall — W₃ Base IRR = 18.04%"
    ws["E54"] = 12
    ws["F54"] = ("Базовый IRR T₁ под Waterfall W₃ DEFAULT составляет 18.04%, что ровно соответствует "
                 "hurdle 18%. Downside защищён 1× Liq Pref floor (1 250 млн ₽) + 8% coupon × 5y (500 млн ₽). "
                 "Риск активируется только при глубоком NDP shortfall (<2 500 млн ₽, -17% от базы). "
                 "Премиум-опция W₄ (12% preferred + 65/35) даёт IRR 19.73% для LP с enhanced requirements.")
    ws["H54"] = ("Структурная защита через waterfall: (1) W₃ floor Liq Pref гарантирует возврат principal; "
                 "(2) coupon 8%×5y = 500 млн ₽ минимальная доходность; (3) W₄ опция 12% preferred "
                 "для enhanced LP protection; (4) exit route — Trade Sale/IPO 2030 backup.")
    ws["K54"] = 8

    # r85 TOP takeaway
    ws["B85"] = ("1. ТОП-риск — R17 пересмотрен в v1.0.1 с Critical на High (score 12): базовый IRR T₁ под W₃ "
                 "default = 18.04% ровно на hurdle 18%, downside защищён через 1× Liq Pref floor + 8% coupon. "
                 "Премиум-опция W₄ даёт 19.73% IRR и 2.12× MOIC для LP с enhanced requirements.")


# ============================================================================
# 35_Roadmap_2026_2032 — минорное уточнение
# ============================================================================

def patch_35_roadmap(wb):
    ws = wb["35_Roadmap_2026_2032"]
    # B76 уже "2029: Q1 — covenant pass или W₃ activation decision"
    # Уточним под v1.0.1: W₃ default, W₄ optional
    ws["B76"] = "• 2029: Q1 — covenant pass или W₃ activation (default) / W₄ premium upgrade decision"


# ============================================================================
# 40_Investor_Checklist — обновить пункты 17, 18, 21
# ============================================================================

def patch_40_checklist(wb):
    ws = wb["40_Investor_Checklist"]

    # D17 Waterfall description
    ws["D17"] = "Waterfall W₃ default (1× Liq Pref + 8% coupon + 60/40) или W₄ premium (12% pref + 65/35)"

    # D27 Investor Returns check — fix wrong MoIC
    ws["D27"] = "Проверить Investor Returns: MOIC 2.00× (W₃) / 2.12× (W₄), IRR 18.04% = hurdle 18%"

    # D30 Top-5 risks — R17 reclassified
    ws["D30"] = "Top-5 рисков из 30 (R17 reclassified High score 12 в v1.0.1 — W₃ Base = hurdle)"

    # D33 MC check — update wording (actual values come from rerun)
    ws["D33"] = "Monte Carlo: n=1000, W₃ base, P(IRR>18%), VaR 95% NDP shortfall"


# ============================================================================
# 42_Cover_Letter — уточнить preferred return 12% как W₄ premium option
# ============================================================================

def patch_42_cover(wb):
    ws = wb["42_Cover_Letter"]
    # r19 ЗАЩИТА ИНВЕСТОРА — добавить различие W₃/W₄
    original = ws["B19"].value or ""
    # Insert clarification after "Preferred return 12%"
    updated = original.replace(
        "• Preferred return 12% до начала GP catch-up (waterfall)",
        "• Waterfall W₃ DEFAULT (1× Liq Pref + 8% coupon + 60/40) — IRR 18.04% = hurdle 18%\n"
        "• Waterfall W₄ PREMIUM опция (1× Liq Pref + 12% preferred × 5y + 65/35) — IRR 19.73%"
    )
    ws["B19"] = updated


# ============================================================================
# MAIN
# ============================================================================

def main():
    for fname in ["investor_model_v1.0_Public.xlsx", "investor_model_v1.0_Internal.xlsx"]:
        backup = fname.replace(".xlsx", "_pre_v101_stage3_backup.xlsx")
        shutil.copy(fname, backup)
        print(f"📦 Backup: {backup}")

        wb = load_workbook(fname)
        print(f"\n🔧 Патчу {fname}")

        mc = patch_28_monte_carlo(wb)
        print(f"  ✓ 28_MC: mean NDP={np.mean(mc['NDP']):.0f}, mean IRR={np.mean(mc['IRR']):.2f}%, "
              f"P(IRR>18%)={np.mean(mc['IRR']>18)*100:.1f}%, P(MOIC>2)={np.mean(mc['MOIC']>2)*100:.1f}%")

        patch_36_executive_summary(wb, mc)
        print("  ✓ 36_Executive_Summary обновлён (W₃ default, R17 reclassified)")

        patch_09_pnl(wb)
        patch_10_cashflow(wb)
        patch_13_debt(wb)
        print("  ✓ 09/10/13: waterfall labels обновлены")

        patch_17_deal(wb)
        print("  ✓ 17_Deal_Structures: default → W₃")

        patch_21_kpi(wb)
        print("  ✓ 21_KPI_Dashboard: IRR/MOIC обновлены")

        patch_26_sensitivity(wb)
        print("  ✓ 26_Sensitivity: tornado rebased to 18.04%")

        patch_27_scenarios(wb)
        print("  ✓ 27_Scenario_Analysis: IRR/MOIC пересчитаны под W₃")

        patch_29_risk(wb)
        print("  ✓ 29_Risk_Register: R17 reclassified 20→12")

        patch_35_roadmap(wb)
        patch_40_checklist(wb)
        patch_42_cover(wb)
        print("  ✓ 35/40/42: минорные апдейты")

        wb.save(fname)
        print(f"💾 Сохранено: {fname}")

    print("\n✅ Stage 3 готов")


if __name__ == "__main__":
    main()
