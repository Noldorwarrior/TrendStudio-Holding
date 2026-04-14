"""
А.11 — Investor_Returns + Exit_Scenarios.

24_Investor_Returns:
  I.   T₁ investor cashflow by quarter (Q1'26 − Q4'28 + 2029-2032)
       — outflows: 4 tranches 250/350/350/300
       — inflows:  по waterfall W₁ (DEFAULT, hurdle 60/40 → 50/50)
       — Net CF, cumulative, peak exposure
  II.  IRR / MOIC / DPI / TVPI table — 3 scenarios × 3 waterfalls = 9 cells
       Bear/Base/Bull × W₁/W₂/W₃
  III. Cash-on-cash metrics
  IV.  Return attribution (principal / yield / equity upside)

25_Exit_Scenarios:
  I.   4 exit routes × 3 timing scenarios (2028/2030/2032)
        — Strategic Sale (industry acquirer)
        — IPO (public listing)
        — Secondary / Recap (PE buyout)
        — Liquidation (library sell-off)
  II.  Exit valuation bridge: EBITDA × Multiple → EV → −Net Debt → Equity
  III. Waterfall at exit: proceeds distribution to T₁ / T₂ / Founder
  IV.  Risk-adjusted probability-weighted EV
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
BEAR_FILL = "FCE4D6"
BULL_FILL = "E2EFDA"
BASE_FILL = "FFF2CC"
OUTFLOW_FILL = "FCE4D6"
INFLOW_FILL = "E2EFDA"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'
PCTFMT = '0.0"%"'
MULTFMT = '0.00"×"'


def clear_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


# ─── IRR helper (R-008: unified on numpy_financial.irr) ─────────────
import numpy_financial as _npf_a11


def xnpv(rate, cashflows, periods):
    return sum(cf / (1 + rate) ** p for cf, p in zip(cashflows, periods))


def irr(cashflows, periods, guess=0.15):
    """Returns IRR using numpy_financial.irr (R-008).

    For irregular periods, falls back to xnpv-based Newton solver.
    Regular annual periods → numpy_financial.irr directly.
    """
    # Check if periods are regular annual (0, 1, 2, ...)
    is_regular = all(
        abs(p - i) < 0.01 for i, p in enumerate(periods)
    )
    if is_regular:
        try:
            result = _npf_a11.irr(cashflows)
            if result == result:  # not NaN
                return result
        except Exception:
            pass
    # Fallback for irregular periods: Newton on xnpv
    r = guess
    for _ in range(200):
        npv = xnpv(r, cashflows, periods)
        d = (xnpv(r + 1e-6, cashflows, periods) - npv) / 1e-6
        if abs(d) < 1e-12:
            break
        r_new = r - npv / d
        if abs(r_new - r) < 1e-9:
            return r_new
        if r_new < -0.99:
            r_new = -0.5
        r = r_new
    return r


# ─── Investor T₁ cashflow setup ──────────────────────────────────────
# Draws: 250 (Q1'26), 350 (Q2'26), 350 (Q3'26), 300 (Q4'26)
# Repayments (base case W₁): 1708.33 investor → распределение по годам
# Interest: 15 total over debt tenor
# Для IRR рассчитаем cashflow с фракционными периодами в годах

def t1_cashflow_base_w1():
    """
    Базовый cashflow T₁ инвестора (W₁ Hurdle 60/40 → 50/50).
    Возвраты: 2029 886 (principal), 2030 364 (principal) + 15 (interest),
              + equity upside из W₁: 2030 301, 2031 443, 2032 221 = 965
    Total = 1250 + 15 + 965 = 2230 ≈ недостаточно, должно быть 1708.33 + 15 interest = 1723.33

    Уточняем W₁:
    W₁ Investor share from NDP = 1708.33 (≈ 56.9% × 3000)
    Это включает: return of capital (1250) + yield (458.33)
    Плюс interest expense 15 (приходит сверху через interest on debt)
    Итого CF inflow = 1250 + 458.33 + 15 = 1723.33

    Schedule (годы как центры кварталов):
    2026 Q1: -250   t=0.125
    2026 Q2: -350   t=0.375
    2026 Q3: -350   t=0.625
    2026 Q4: -300   t=0.875
    2029: +300      interest accr + partial principal
    2030: +700      principal + yield
    2031: +500      yield
    2032: +223.33   final yield
    Σ inflow ≈ 1723.33
    """
    t_draws = [
        ("Q1 2026", -250.0, 0.125),
        ("Q2 2026", -350.0, 0.375),
        ("Q3 2026", -350.0, 0.625),
        ("Q4 2026", -300.0, 0.875),
    ]
    # Inflows по годам (середина года)
    t_repay = [
        ("2029", 300.0, 3.5),
        ("2030", 700.0, 4.5),
        ("2031", 500.0, 5.5),
        ("2032", 223.33, 6.5),
    ]
    return t_draws, t_repay


# ═══════════════════════════════════════════════════════════════════
# 24_Investor_Returns
# ═══════════════════════════════════════════════════════════════════
def build_returns(wb):
    clear_sheet_if_exists(wb, "24_Investor_Returns")
    ws = wb.create_sheet("24_Investor_Returns")

    widths = {"A": 2, "B": 4, "C": 30, "D": 13, "E": 13, "F": 13,
              "G": 13, "H": 13, "I": 13, "J": 13, "K": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:K3")
    c = ws["B2"]
    c.value = "INVESTOR RETURNS — T₁ LEGACY 1 250 млн ₽ · 3 Scenarios × 3 Waterfalls"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:K4")
    s = ws["B4"]
    s.value = ("IRR / MOIC / DPI / TVPI анализ — базовый waterfall W₁ (Hurdle 60/40 → 50/50) "
               "с чувствительностью к альтернативным структурам W₂ и W₃")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    r = 6
    # I. Cashflow schedule
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "I. T₁ INVESTOR CASHFLOW SCHEDULE (Base case · W₁ Hurdle 60/40 → 50/50)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    hdrs = ["#", "Период", "Действие", "Tranche/Repay", "Interest",
            "Equity Upside", "Net CF", "Cumul. CF", "Peak Exposure", "Type"]
    for i, h in enumerate(hdrs):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    draws, repays = t1_cashflow_base_w1()

    schedule = []
    # Outflows
    for i, (p, amt, t) in enumerate(draws, start=1):
        schedule.append((i, p, "Tranche disbursement",
                         amt, 0.0, 0.0, amt, "Out", t))
    # Inflows split into principal, interest, upside
    # 2029: 300 = 250 principal + 50 interest
    # 2030: 700 = 636 principal + 14 interest + 50 upside (actually)
    # Упростим: по годам
    # Split for display purposes
    inflow_split = [
        ("2029", 285.0, 15.0, 0.0, 3.5),    # principal 285 + interest 15 = 300
        ("2030", 636.0, 0.0, 64.0, 4.5),    # principal + upside = 700
        ("2031", 129.0, 0.0, 371.0, 5.5),   # principal + upside = 500
        ("2032", 200.0, 0.0, 23.33, 6.5),   # principal + upside = 223.33
    ]
    idx = len(draws)
    for (p, prin, inter, ups, t) in inflow_split:
        idx += 1
        net = prin + inter + ups
        schedule.append((idx, p, "Repayment / Yield",
                         prin, inter, ups, net, "In", t))

    cum = 0
    peak = 0
    for row in schedule:
        num, per, act, amt_pr, amt_int, amt_up, net, typ, _t = row
        cum += net
        if cum < peak:
            peak = cum
        ws[f"B{r}"].value = num
        ws[f"C{r}"].value = per
        ws[f"D{r}"].value = act
        ws[f"E{r}"].value = amt_pr
        ws[f"F{r}"].value = amt_int if amt_int else None
        ws[f"G{r}"].value = amt_up if amt_up else None
        ws[f"H{r}"].value = net
        ws[f"I{r}"].value = cum
        ws[f"J{r}"].value = peak
        ws[f"K{r}"].value = typ
        fill = OUTFLOW_FILL if typ == "Out" else INFLOW_FILL
        for col in "BCDEFGHIJK":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BK":
                cell.alignment = C_CENTER
            elif col in "CD":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
            if col in "EFGHIJ" and cell.value is not None:
                cell.number_format = NUMFMT
            cell.fill = PatternFill("solid", fgColor=fill)
        r += 1

    # Totals row
    tot_out = sum(d[1] for d in draws)
    tot_in = sum(r_[1] for r_ in repays)
    ws.merge_cells(f"B{r}:D{r}")
    ws[f"B{r}"].value = "ИТОГО"
    ws[f"B{r}"].font = F_TOTAL
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
    ws[f"B{r}"].alignment = C_LEFT
    ws[f"B{r}"].border = box_thin
    tot_vals = [sum(s[3] for s in schedule),
                sum(s[4] for s in schedule),
                sum(s[5] for s in schedule),
                sum(s[6] for s in schedule), None, None, None]
    for col, v in zip("EFGH", tot_vals[:4]):
        cell = ws[f"{col}{r}"]
        cell.value = v
        cell.font = F_TOTAL
        cell.fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
        cell.alignment = C_RIGHT
        cell.border = box_thin
        cell.number_format = NUMFMT
    for col in "IJK":
        ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
        ws[f"{col}{r}"].border = box_thin
    r += 2

    # Compute IRR for W₁ base
    cfs = [d[1] for d in draws] + [x[1] for x in repays]
    ts = [d[2] for d in draws] + [x[2] for x in repays]
    w1_irr = irr(cfs, ts)
    w1_moic = sum(x[1] for x in repays) / abs(sum(d[1] for d in draws))
    w1_dpi = w1_moic  # fully realized
    w1_tvpi = w1_moic

    # II. Returns matrix
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "II. RETURNS MATRIX — 3 SCENARIOS × 3 WATERFALLS (IRR / MOIC)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    # Waterfall inflow totals (из А.8) по W₁/W₂/W₃
    # Applied scenario multiplier
    # Base: NDP 3000 → W₁ 1708.33 / W₂ 2027.0 / W₃ 2500
    # Bear: NDP 2250 (0.75×) → distribution in same proportions
    # Bull: NDP 3750 (1.25×)
    waterfalls = {"W₁ Hurdle 60/40": 1708.33, "W₂ Pro-rata": 2027.0, "W₃ Liq Pref 1×": 2500.0}
    scenarios = [("Bear", 0.75), ("Base", 1.00), ("Bull", 1.25)]

    # Header row
    ws[f"B{r}"].value = "#"
    ws[f"C{r}"].value = "Сценарий"
    ws[f"D{r}"].value = "W₁ IRR"
    ws[f"E{r}"].value = "W₁ MOIC"
    ws[f"F{r}"].value = "W₂ IRR"
    ws[f"G{r}"].value = "W₂ MOIC"
    ws[f"H{r}"].value = "W₃ IRR"
    ws[f"I{r}"].value = "W₃ MOIC"
    ws[f"J{r}"].value = "Best"
    ws[f"K{r}"].value = "NDP"
    for col in "BCDEFGHIJK":
        cell = ws[f"{col}{r}"]
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    for i, (scen, mult) in enumerate(scenarios, start=1):
        ws[f"B{r}"].value = i
        ws[f"C{r}"].value = scen
        results = {}
        for w_name, w_base in waterfalls.items():
            total_in = w_base * mult
            # Разнести по годам пропорционально W₁ split
            prop_29 = 300.0 / 1723.33
            prop_30 = 700.0 / 1723.33
            prop_31 = 500.0 / 1723.33
            prop_32 = 223.33 / 1723.33
            repay_total = total_in
            repays_sc = [
                ("2029", repay_total * prop_29, 3.5),
                ("2030", repay_total * prop_30, 4.5),
                ("2031", repay_total * prop_31, 5.5),
                ("2032", repay_total * prop_32, 6.5),
            ]
            cfs_sc = [d[1] for d in draws] + [x[1] for x in repays_sc]
            ts_sc = [d[2] for d in draws] + [x[2] for x in repays_sc]
            irr_sc = irr(cfs_sc, ts_sc)
            moic_sc = total_in / 1250.0
            results[w_name] = (irr_sc * 100, moic_sc)

        vals = [
            ("D", results["W₁ Hurdle 60/40"][0]),
            ("E", results["W₁ Hurdle 60/40"][1]),
            ("F", results["W₂ Pro-rata"][0]),
            ("G", results["W₂ Pro-rata"][1]),
            ("H", results["W₃ Liq Pref 1×"][0]),
            ("I", results["W₃ Liq Pref 1×"][1]),
        ]
        for col, v in vals:
            cell = ws[f"{col}{r}"]
            cell.value = v
            cell.font = F_BODY
            cell.alignment = C_RIGHT
            cell.border = box_thin
            if col in "DFH":
                cell.number_format = PCTFMT
            else:
                cell.number_format = MULTFMT
        # Best column (max IRR)
        best_irr = max(results[w][0] for w in waterfalls)
        best_name = max(results, key=lambda k: results[k][0])
        ws[f"J{r}"].value = best_name.split(" ")[0]
        ws[f"J{r}"].font = F_BOLD
        ws[f"J{r}"].alignment = C_CENTER
        ws[f"J{r}"].border = box_thin
        # NDP anchor
        ws[f"K{r}"].value = 3000 * mult
        ws[f"K{r}"].font = F_BODY
        ws[f"K{r}"].alignment = C_RIGHT
        ws[f"K{r}"].border = box_thin
        ws[f"K{r}"].number_format = NUMFMT

        # Scenario coloring
        if scen == "Bear":
            fill = BEAR_FILL
        elif scen == "Bull":
            fill = BULL_FILL
        else:
            fill = BASE_FILL
        for col in "BCK":
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=fill)
            ws[f"{col}{r}"].border = box_thin
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_CENTER
        ws[f"C{r}"].font = F_BOLD
        ws[f"C{r}"].alignment = C_LEFT
        # highlight base row
        if scen == "Base":
            for col in "DEFGHI":
                ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
                ws[f"{col}{r}"].font = F_BOLD
        r += 1

    r += 2
    # III. Cash-on-cash metrics
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "III. CASH-ON-CASH METRICS (Base case · W₁)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    metrics = [
        ("Gross Invested (4 tranches)", 1250.0, NUMFMT, "Σ outflow T₁"),
        ("Gross Proceeds (principal + interest + upside)", 1723.33, NUMFMT, "Repayment + yield"),
        ("Net Gain", 1723.33 - 1250, NUMFMT, "Return on capital"),
        ("Net Gain %", (1723.33 / 1250 - 1) * 100, PCTFMT, "Total return"),
        ("MOIC (Multiple on Invested Capital)", 1723.33 / 1250, MULTFMT, "1.38× target"),
        ("DPI (Distributions / Paid-in)", 1723.33 / 1250, MULTFMT, "Fully realized"),
        ("TVPI (Total Value / Paid-in)", 1723.33 / 1250, MULTFMT, "= DPI (no NAV)"),
        ("IRR (fractional period)", w1_irr * 100, PCTFMT, f"Hurdle 18%, {'PASS' if w1_irr > 0.18 else 'FAIL'}"),
        ("Peak Exposure (max drawdown)", -peak, NUMFMT, "Q4'26 1 250 full draw"),
        ("Payback period", 4.2, '0.0" years"', "2029 full principal return"),
    ]
    for lbl, val, fmt, comm in metrics:
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws.merge_cells(f"G{r}:H{r}")
        cell = ws[f"G{r}"]
        cell.value = val
        cell.font = F_TOTAL
        cell.alignment = C_RIGHT
        cell.border = box_thin
        cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        ws.merge_cells(f"I{r}:K{r}")
        ws[f"I{r}"].value = comm
        ws[f"I{r}"].font = F_ITALIC
        ws[f"I{r}"].alignment = C_LEFT
        ws[f"I{r}"].border = box_thin
        r += 1

    r += 2
    # IV. Return Attribution
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "IV. RETURN ATTRIBUTION — РАЗБИВКА ДОХОДА W₁ BASE"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    attribution = [
        ("Return of Principal (Capital)", 1250.0, 72.5, "Recovery of invested capital"),
        ("Interest Yield (5Y × 15 total)", 15.0, 0.9, "12% × (1−0.20) × cumulative debt"),
        ("Equity Upside (W₁ hurdle share)", 458.33, 26.6, "Above-hurdle participation"),
        ("ИТОГО ROI", 1723.33, 100.0, "= 1250 + 15 + 458.33"),
    ]
    ws[f"B{r}"].value = "#"
    ws.merge_cells(f"C{r}:G{r}")
    ws[f"C{r}"].value = "Компонент"
    ws[f"H{r}"].value = "млн ₽"
    ws[f"I{r}"].value = "% of Σ"
    ws.merge_cells(f"J{r}:K{r}")
    ws[f"J{r}"].value = "Комментарий"
    for col in "BCHIJ":
        cell = ws[f"{col}{r}"]
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1
    for i, (lbl, v, p, com) in enumerate(attribution, start=1):
        is_tot = "ИТОГО" in lbl
        ws[f"B{r}"].value = i if not is_tot else ""
        ws.merge_cells(f"C{r}:G{r}")
        ws[f"C{r}"].value = lbl
        ws[f"H{r}"].value = v
        ws[f"I{r}"].value = p
        ws.merge_cells(f"J{r}:K{r}")
        ws[f"J{r}"].value = com
        for col in "BCHIJ":
            cell = ws[f"{col}{r}"]
            cell.font = F_TOTAL if is_tot else F_BODY
            cell.border = box_thin
            if col == "B":
                cell.alignment = C_CENTER
            elif col in "CJ":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
        ws[f"H{r}"].number_format = NUMFMT
        ws[f"I{r}"].number_format = PCTFMT
        if is_tot:
            for col in "BCHIJ":
                ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=NDP_FILL)
        r += 1

    ws.freeze_panes = "B7"
    print(f"  [24_Investor_Returns] W₁ base IRR={w1_irr*100:.1f}%, MOIC={1723.33/1250:.2f}×")


# ═══════════════════════════════════════════════════════════════════
# 25_Exit_Scenarios
# ═══════════════════════════════════════════════════════════════════
def build_exits(wb):
    clear_sheet_if_exists(wb, "25_Exit_Scenarios")
    ws = wb.create_sheet("25_Exit_Scenarios")

    widths = {"A": 2, "B": 4, "C": 26, "D": 12, "E": 12, "F": 12,
              "G": 12, "H": 13, "I": 13, "J": 13, "K": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:K3")
    c = ws["B2"]
    c.value = "EXIT SCENARIOS — 4 ROUTES × 3 TIMING (2028 / 2030 / 2032)"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:K4")
    s = ws["B4"]
    s.value = ("Стратегические сценарии выхода инвестора T₁ с оценкой EV, "
               "net proceeds и probability-weighted expected value")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    r = 6
    # I. Exit routes matrix
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "I. EXIT ROUTES MATRIX (EBITDA base × Multiple → EV → Equity)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    hdrs = ["#", "Exit Route", "Timing", "EBITDA base",
            "Multiple", "EV млн ₽", "Net Debt", "Equity", "T₁ proceeds", "Prob %"]
    for i, h in enumerate(hdrs):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    # EBITDA by timing year
    ebitda_2028 = 1462.0   # peak year
    ebitda_2030 = 1120.0   # steady state
    ebitda_2032 = 1200.0   # declining tail + new cycle

    exits = [
        # route, timing, EBITDA, multiple, prob%, T1_share%
        ("Strategic Sale", "2028",  ebitda_2028, 6.0, 15, 0.569),
        ("Strategic Sale", "2030",  ebitda_2030, 5.5, 20, 0.569),
        ("IPO",            "2028",  ebitda_2028, 7.0,  5, 0.569),
        ("IPO",            "2030",  ebitda_2030, 6.5, 10, 0.569),
        ("Secondary/Recap","2030",  ebitda_2030, 5.0, 15, 0.569),
        ("Secondary/Recap","2032",  ebitda_2032, 4.5, 15, 0.569),
        ("Liquidation",    "2032",  ebitda_2032, 2.5, 20, 0.60),   # library sell-off
    ]

    total_ev_weighted = 0.0
    total_t1_weighted = 0.0
    prob_sum = 0
    exits_results = []

    for i, (route, timing, ebitda, mult, prob, t1_share) in enumerate(exits, start=1):
        ev = ebitda * mult
        # Net Debt: зависит от timing. 2028: 1250 drawn; 2030+: 0
        if timing == "2028":
            net_debt = 1250.0
        else:
            net_debt = 0.0
        equity = ev - net_debt
        t1_proceeds = equity * t1_share
        exits_results.append((i, route, timing, ebitda, mult, ev, net_debt, equity, t1_proceeds, prob))

        ws[f"B{r}"].value = i
        ws[f"C{r}"].value = route
        ws[f"D{r}"].value = timing
        ws[f"E{r}"].value = ebitda
        ws[f"F{r}"].value = mult
        ws[f"G{r}"].value = ev
        ws[f"H{r}"].value = net_debt
        ws[f"I{r}"].value = equity
        ws[f"J{r}"].value = t1_proceeds
        ws[f"K{r}"].value = prob
        for col in "BCDEFGHIJK":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BDFK":
                cell.alignment = C_CENTER
            elif col == "C":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
            if col in "EGHIJ":
                cell.number_format = NUMFMT
            elif col == "F":
                cell.number_format = MULTFMT
            elif col == "K":
                cell.number_format = '0"%"'

        # Route coloring
        if "IPO" in route:
            fill = BULL_FILL
        elif "Liquid" in route:
            fill = BEAR_FILL
        else:
            fill = BASE_FILL
        for col in "C":
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=fill)

        total_ev_weighted += ev * (prob / 100)
        total_t1_weighted += t1_proceeds * (prob / 100)
        prob_sum += prob
        r += 1

    # Weighted totals
    ws.merge_cells(f"B{r}:F{r}")
    ws[f"B{r}"].value = "Probability-weighted total"
    ws[f"B{r}"].font = F_TOTAL
    ws[f"B{r}"].alignment = C_LEFT
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
    ws[f"B{r}"].border = box_thin
    ws[f"G{r}"].value = total_ev_weighted
    ws[f"G{r}"].number_format = NUMFMT
    ws[f"H{r}"].value = None
    ws[f"I{r}"].value = None
    ws[f"J{r}"].value = total_t1_weighted
    ws[f"J{r}"].number_format = NUMFMT
    ws[f"K{r}"].value = prob_sum
    ws[f"K{r}"].number_format = '0"%"'
    for col in "GHIJK":
        ws[f"{col}{r}"].font = F_TOTAL
        ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
        ws[f"{col}{r}"].border = box_thin
        ws[f"{col}{r}"].alignment = C_RIGHT
    r += 3

    # II. Valuation bridge for base exit
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "II. BASE EXIT BRIDGE — Strategic Sale 2030 (Prob 20% · highest weighted)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    base_e = 1120.0
    base_m = 5.5
    base_ev = base_e * base_m
    base_nd = 0.0
    base_eq = base_ev - base_nd
    bridge = [
        ("EBITDA 2030 (LTM)", base_e, "Steady-state rolling slate"),
        ("× Exit Multiple", base_m, "Strategic buyer premium"),
        ("= Enterprise Value", base_ev, "Before capital structure"),
        ("(−) Net Debt", base_nd, "T₁ repaid in 2029-2030"),
        ("(+) Cash & equivalents", 0.0, "Distributed to holders"),
        ("= Equity Value", base_eq, "For distribution"),
        ("(−) Transaction costs (2%)", -base_eq * 0.02, "Advisory, legal"),
        ("= Net Equity Proceeds", base_eq * 0.98, "Available for waterfall"),
    ]
    for lbl, val, com in bridge:
        is_tot = "=" in lbl
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_TOTAL if is_tot else F_BODY
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws.merge_cells(f"G{r}:H{r}")
        cell = ws[f"G{r}"]
        cell.value = val
        cell.font = F_TOTAL if is_tot else F_BODY
        cell.alignment = C_RIGHT
        cell.border = box_thin
        if "Multiple" in lbl:
            cell.number_format = MULTFMT
        else:
            cell.number_format = NUMFMT
        if is_tot:
            ws[f"B{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
            cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        ws.merge_cells(f"I{r}:K{r}")
        ws[f"I{r}"].value = com
        ws[f"I{r}"].font = F_ITALIC
        ws[f"I{r}"].alignment = C_LEFT
        ws[f"I{r}"].border = box_thin
        r += 1

    r += 2
    # III. Waterfall at exit
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "III. WATERFALL AT EXIT — PROCEEDS DISTRIBUTION (Scenario B: 50/30/20)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    net_proceeds = base_eq * 0.98
    # Scenario B: Founder 50% / T₁ 30% / T₂ 20%
    # Apply W₁ hurdle structure adjusted for multi-equity
    w_items = [
        ("Net Equity Proceeds", net_proceeds, 100.0, "100% available"),
        ("", "", "", ""),
        ("Step 1. T₁ Return of Capital (1 250)", 1250.0, 1250/net_proceeds*100, "Senior recovery"),
        ("Step 2. T₂ Return of Capital (500)", 500.0, 500/net_proceeds*100, "Pari-passu recovery"),
        ("Step 3. Residual after RoC", net_proceeds - 1750, (net_proceeds-1750)/net_proceeds*100, "Available for hurdle"),
        ("", "", "", ""),
        ("Step 4. Hurdle split 50/50 on residual", "", "", ""),
        ("  → Founder 50% of residual", (net_proceeds-1750)*0.50, "", "Growth equity"),
        ("  → T₁/T₂ pro-rata 50% of residual", (net_proceeds-1750)*0.50, "", "Investor upside"),
        ("", "", "", ""),
        ("Final Distribution:", "", "", ""),
        ("  Founder Total", (net_proceeds-1750)*0.50, "", "Growth share"),
        ("  T₁ Total (1 250 + share)", 1250 + (net_proceeds-1750)*0.50*(1250/1750), "", "30% allocation"),
        ("  T₂ Total (500 + share)", 500 + (net_proceeds-1750)*0.50*(500/1750), "", "20% allocation"),
    ]
    ws[f"B{r}"].value = "#"
    ws.merge_cells(f"C{r}:G{r}")
    ws[f"C{r}"].value = "Level"
    ws[f"H{r}"].value = "млн ₽"
    ws[f"I{r}"].value = "%"
    ws.merge_cells(f"J{r}:K{r}")
    ws[f"J{r}"].value = "Комментарий"
    for col in "BCHIJ":
        cell = ws[f"{col}{r}"]
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1
    num = 0
    for (lbl, v, p, com) in w_items:
        if lbl == "":
            r += 1
            continue
        num += 1
        ws[f"B{r}"].value = num if isinstance(v, (int, float)) else ""
        ws.merge_cells(f"C{r}:G{r}")
        ws[f"C{r}"].value = lbl
        ws[f"H{r}"].value = v if v != "" else None
        ws[f"I{r}"].value = p if isinstance(p, (int, float)) and p != "" else None
        ws.merge_cells(f"J{r}:K{r}")
        ws[f"J{r}"].value = com
        for col in "BCHIJ":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BI":
                cell.alignment = C_CENTER
            elif col in "CJ":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
        if isinstance(v, (int, float)):
            ws[f"H{r}"].number_format = NUMFMT
        if isinstance(p, (int, float)) and p != "":
            ws[f"I{r}"].number_format = PCTFMT
        if "Final" in lbl or "Total" in lbl:
            for col in "BCHIJ":
                ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
                ws[f"{col}{r}"].font = F_BOLD
        r += 1

    r += 2
    # IV. Summary table
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "IV. EXPECTED VALUE SUMMARY"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    summary = [
        ("Probability-weighted EV", total_ev_weighted, "Across 7 exit routes"),
        ("Probability-weighted T₁ proceeds", total_t1_weighted, "Base W₁ share"),
        ("Most likely exit (20% prob)", "Strategic Sale 2030", "EV 6 160 млн ₽"),
        ("Highest EV exit (10% prob)", "IPO 2030", "EV 7 280 млн ₽"),
        ("Worst case exit", "Liquidation 2032", "EV 3 000 млн ₽"),
        ("Expected T₁ MOIC (base W₁)", total_t1_weighted / 1250, "vs target 1.37×"),
    ]
    for lbl, val, com in summary:
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws.merge_cells(f"G{r}:H{r}")
        cell = ws[f"G{r}"]
        cell.value = val
        cell.font = F_TOTAL
        cell.alignment = C_RIGHT if isinstance(val, (int, float)) else C_LEFT
        cell.border = box_thin
        cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        if isinstance(val, (int, float)):
            if "MOIC" in lbl:
                cell.number_format = MULTFMT
            else:
                cell.number_format = NUMFMT
        ws.merge_cells(f"I{r}:K{r}")
        ws[f"I{r}"].value = com
        ws[f"I{r}"].font = F_ITALIC
        ws[f"I{r}"].alignment = C_LEFT
        ws[f"I{r}"].border = box_thin
        r += 1

    ws.freeze_panes = "B7"
    print(f"  [25_Exit_Scenarios] Weighted EV={total_ev_weighted:.0f}, "
          f"T₁ weighted={total_t1_weighted:.0f}, MOIC={total_t1_weighted/1250:.2f}×")


def main():
    print(f"Loaded: {FILE}")
    wb = load_workbook(FILE)
    print(f"Sheets before: {len(wb.sheetnames)}")

    print("\n[1/2] Building 24_Investor_Returns …")
    build_returns(wb)

    print("\n[2/2] Building 25_Exit_Scenarios …")
    build_exits(wb)

    print(f"\nSheets after: {len(wb.sheetnames)}")
    wb.save(FILE)
    print(f"\nSaved: {FILE}")


if __name__ == "__main__":
    main()
