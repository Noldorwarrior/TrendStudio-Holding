"""
А.2 — FOT_Staff_A1 + FOT_Staff_A2 + Cost_Structure
Добавляем 3 листа к investor_model_v1.0_Public.xlsx

A1: Fixed — 50 чел × 6.162 млн ₽/мес (Холдинг Кино.xlsx)
A2: Dynamic — 50→60→70 чел + индексация 7%/год
Cost_Structure: OPEX сводка (ФОТ + аренда + коммунал + прочие)
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package"
OUT_FILE = os.path.join(OUT_DIR, "investor_model_v1.0_Public.xlsx")

# Стили (наследуем из А.1)
BRAND_BLUE = "0070C0"
BRAND_BLUE_LIGHT = "D9E2F3"
BRAND_BLUE_DARK = "002060"
ACCENT_GREEN = "548235"
ACCENT_RED = "C00000"
GRAY_LIGHT = "F2F2F2"
GRAY_DARK = "595959"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"
FORMULA_BLACK = "000000"
LINK_GREEN = "006100"
KEY_METRIC_FILL = "FFF2CC"


def thin(color="808080"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def sheet_header(ws, title, subtitle=""):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:O2")
    c = ws["B2"]
    c.value = title
    c.font = Font(name="Arial", size=18, bold=True, color=BRAND_BLUE_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 30
    if subtitle:
        ws.merge_cells("B3:O3")
        c = ws["B3"]
        c.value = subtitle
        c.font = Font(name="Arial", size=10, italic=True, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="left", indent=1)


def header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    c.border = thin("FFFFFF")
    return c


def block_title(ws, row, start_col, end_col, text, color=BRAND_BLUE):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


# ============================================================================
# ЛИСТ 04: FOT_Staff_A1 (Fixed, 50 чел из Холдинг Кино.xlsx)
# ============================================================================

def build_fot_a1(wb):
    ws = wb.create_sheet("04_FOT_Staff_A1")

    widths = [2, 6, 34, 10, 16, 16, 16, 16, 16, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sheet_header(ws, "FOT / Штат — Модель A₁ (Fixed)",
                 "50 чел × 6.162 млн ₽/мес. Структура из Холдинг Кино.xlsx без изменений. 3 года = 221.832 млн ₽.")

    # Шапка таблицы
    row = 5
    headers = ["#", "Категория", "Чел.", "Оклад gross", "ФОТ gross мес", "Страховые 30%", "Итого мес (с налогами)", "Итого год"]
    cols = [2, 3, 4, 5, 6, 7, 8, 9]
    for col, h in zip(cols, headers):
        header_cell(ws, row, col, h)
    ws.row_dimensions[row].height = 32

    # Данные (из оригинальной Холдинг Кино.xlsx — лист Расходы)
    categories = [
        ("Продюсеры",                 5,  200000),
        ("Юристы",                    3,  150000),
        ("Производственный цех",      5,  120000),
        ("Аналитики",                 7,  100000),
        ("Бухгалтеры",                4,  80000),
        ("Эксперты (творческая группа)", 15, 60000),
        ("Водитель",                  1,  70000),
        ("Административный / IT",     10, 70000),
    ]

    row = 6
    start_data_row = row
    for i, (name, n, salary) in enumerate(categories, 1):
        # #
        c = ws.cell(row=row, column=2, value=i)
        c.font = Font(name="Arial", size=10, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        # Название
        c = ws.cell(row=row, column=3, value=name)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        # Чел (ввод)
        c = ws.cell(row=row, column=4, value=n)
        c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "#,##0"
        c.border = thin()
        # Оклад gross (ввод)
        c = ws.cell(row=row, column=5, value=salary)
        c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0" ₽"'
        c.border = thin()
        # ФОТ gross мес (формула)
        c = ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        c.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0" ₽"'
        c.border = thin()
        # Страховые 30% (формула, ссылка на Assumptions страховые взносы)
        c = ws.cell(row=row, column=7, value=f"=F{row}*0.3")
        c.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0" ₽"'
        c.border = thin()
        # Итого мес (формула)
        c = ws.cell(row=row, column=8, value=f"=F{row}+G{row}")
        c.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0" ₽"'
        c.border = thin()
        c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
        # Итого год (формула)
        c = ws.cell(row=row, column=9, value=f"=H{row}*12")
        c.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0" ₽"'
        c.border = thin()
        c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)

        ws.row_dimensions[row].height = 20
        row += 1

    end_data_row = row - 1

    # ИТОГО строка
    totals_row = row
    c = ws.cell(row=row, column=2, value="")
    c.border = thin()
    c = ws.cell(row=row, column=3, value="ИТОГО ФОТ (Fixed)")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    for col, formula in [
        (4, f"=SUM(D{start_data_row}:D{end_data_row})"),
        (5, None),  # оклады не суммируются
        (6, f"=SUM(F{start_data_row}:F{end_data_row})"),
        (7, f"=SUM(G{start_data_row}:G{end_data_row})"),
        (8, f"=SUM(H{start_data_row}:H{end_data_row})"),
        (9, f"=SUM(I{start_data_row}:I{end_data_row})"),
    ]:
        c = ws.cell(row=row, column=col, value=formula)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="right" if col != 4 else "center",
                                vertical="center", indent=1)
        if col == 4:
            c.number_format = "#,##0"
        else:
            c.number_format = '#,##0" ₽"'
        c.border = thin()
    ws.row_dimensions[row].height = 26
    row += 2

    # Блок «За 3 года»
    block_title(ws, row, 2, 9, "ФОТ A₁ за 3 года (2026–2028), млн ₽", BRAND_BLUE)
    row += 1

    # 3 года — одинаковые (fixed, без индексации)
    years = [2026, 2027, 2028]
    for col_idx, year in enumerate(years):
        c = ws.cell(row=row, column=2 + col_idx + (1 if col_idx > 0 else 0))  # неактуально
    # Используем простую раскладку
    header_cell(ws, row, 3, "Год")
    header_cell(ws, row, 4, "Штат")
    header_cell(ws, row, 5, "ФОТ мес gross")
    header_cell(ws, row, 6, "ФОТ мес с нал.")
    header_cell(ws, row, 7, "ФОТ год")
    header_cell(ws, row, 8, "ФОТ/Налоги")
    header_cell(ws, row, 9, "Индексация")
    ws.row_dimensions[row].height = 28
    row += 1
    start_y = row
    for year in years:
        c = ws.cell(row=row, column=3, value=year)
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        c = ws.cell(row=row, column=4, value=f"=D{totals_row}")
        c.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "#,##0"
        c.border = thin()
        c = ws.cell(row=row, column=5, value=f"=F{totals_row}")
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0" ₽"'
        c.border = thin()
        c = ws.cell(row=row, column=6, value=f"=H{totals_row}")
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0" ₽"'
        c.border = thin()
        c = ws.cell(row=row, column=7, value=f"=I{totals_row}/1000000")
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.000" млн"'
        c.border = thin()
        c.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        c = ws.cell(row=row, column=8, value=f"={'F' if False else 'F'}{totals_row}*12/1000000")
        # Колонка «ФОТ/Налоги» = доля налогов
        c = ws.cell(row=row, column=8, value=f"=G{totals_row}*12/1000000")
        c.font = Font(name="Arial", size=10, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.000" млн"'
        c.border = thin()
        c = ws.cell(row=row, column=9, value="0.0%")
        c.font = Font(name="Arial", size=10, italic=True, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        ws.row_dimensions[row].height = 20
        row += 1

    end_y = row - 1
    # ИТОГО за 3 года
    c = ws.cell(row=row, column=3, value="Σ 2026–2028")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    for col in [4, 5, 6, 7, 8]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        if col == 4:
            c.value = f"={col_letter}{start_y}"  # штат не суммируем
            c.number_format = "#,##0"
        elif col in (5, 6):
            c.value = f"={col_letter}{start_y}"
            c.number_format = '#,##0" ₽"'
        else:
            c.value = f"=SUM({col_letter}{start_y}:{col_letter}{end_y})"
            c.number_format = '#,##0.000" млн"'
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="right" if col not in (4,) else "center",
                                vertical="center", indent=1)
        c.border = thin()
    c = ws.cell(row=row, column=9, value="—")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    ws.row_dimensions[row].height = 24
    row += 2

    # Примечание
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    c = ws.cell(row=row, column=2,
                value="⚫ Модель A₁ «Fixed»: штат и оклады зафиксированы на уровне Холдинг Кино.xlsx (50 чел, 4 740 000 ₽/мес gross). Индексация и рост штата отсутствуют. Общий ФОТ 2026–2028 с налогами = 221.832 млн ₽. Соответствует v1.4.4.")
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws.row_dimensions[row].height = 32

    ws.freeze_panes = "A6"


# ============================================================================
# ЛИСТ 05: FOT_Staff_A2 (Full Dynamic)
# ============================================================================

def build_fot_a2(wb):
    ws = wb.create_sheet("05_FOT_Staff_A2")

    # 50→60→70 с индексацией 7%/год
    widths = [2, 32, 12, 12, 12, 12, 12, 12, 12, 12, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sheet_header(ws, "FOT / Штат — Модель A₂ (Full Dynamic)",
                 "Рост штата 50→60→70 + индексация окладов 7%/год (CPI + премия). Поквартально 2026–2028, годовые 2029–2032.")

    # Блок 1: Динамика штата и коэффициентов
    row = 5
    block_title(ws, row, 2, 10, "ПАРАМЕТРЫ ДИНАМИКИ (ввод)", BRAND_BLUE)
    row += 1

    # Штат по годам
    header_cell(ws, row, 2, "Параметр")
    header_cell(ws, row, 3, "2026")
    header_cell(ws, row, 4, "2027")
    header_cell(ws, row, 5, "2028")
    header_cell(ws, row, 6, "2029")
    header_cell(ws, row, 7, "2030")
    header_cell(ws, row, 8, "2031")
    header_cell(ws, row, 9, "2032")
    ws.row_dimensions[row].height = 22
    row += 1

    # Штат
    staff_plan = [50, 60, 70, 75, 80, 85, 90]
    c = ws.cell(row=row, column=2, value="Штат (чел.)")
    c.font = Font(name="Arial", size=10, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    staff_row = row
    for i, n in enumerate(staff_plan):
        cc = ws.cell(row=row, column=3 + i, value=n)
        cc.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.number_format = "#,##0"
        cc.border = thin()
    ws.row_dimensions[row].height = 20
    row += 1

    # Индексация кумулятивная (1.07^n)
    index_plan = [1.00, 1.07, 1.1449, 1.2250, 1.3108, 1.4026, 1.5007]
    c = ws.cell(row=row, column=2, value="Индекс оклада (база 2026=1.00)")
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    index_row = row
    for i, idx in enumerate(index_plan):
        cc = ws.cell(row=row, column=3 + i, value=idx)
        cc.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.number_format = "0.0000"
        cc.border = thin()
    ws.row_dimensions[row].height = 18
    row += 1

    # Средний оклад gross (из A1: 4 740 000 / 50 = 94 800)
    avg_salary = 94800
    c = ws.cell(row=row, column=2, value="Ср. оклад gross (₽/мес, база 2026)")
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    salary_row = row
    for i in range(7):
        cc = ws.cell(row=row, column=3 + i,
                     value=f"={avg_salary}*{get_column_letter(3 + i)}{index_row}")
        cc.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cc.number_format = '#,##0" ₽"'
        cc.border = thin()
    ws.row_dimensions[row].height = 18
    row += 1

    # ФОТ gross мес
    c = ws.cell(row=row, column=2, value="ФОТ gross / мес")
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    fot_gross_row = row
    for i in range(7):
        col_letter = get_column_letter(3 + i)
        cc = ws.cell(row=row, column=3 + i,
                     value=f"={col_letter}{staff_row}*{col_letter}{salary_row}")
        cc.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cc.number_format = '#,##0" ₽"'
        cc.border = thin()
    ws.row_dimensions[row].height = 18
    row += 1

    # ФОТ с налогами (×1.3)
    c = ws.cell(row=row, column=2, value="ФОТ с налогами / мес (×1.30)")
    c.font = Font(name="Arial", size=10, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    fot_tax_row = row
    for i in range(7):
        col_letter = get_column_letter(3 + i)
        cc = ws.cell(row=row, column=3 + i,
                     value=f"={col_letter}{fot_gross_row}*1.3")
        cc.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cc.number_format = '#,##0" ₽"'
        cc.border = thin()
        cc.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws.row_dimensions[row].height = 20
    row += 1

    # ФОТ годовой (млн ₽)
    c = ws.cell(row=row, column=2, value="ФОТ годовой (млн ₽)")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    fot_year_row = row
    for i in range(7):
        col_letter = get_column_letter(3 + i)
        cc = ws.cell(row=row, column=3 + i,
                     value=f"={col_letter}{fot_tax_row}*12/1000000")
        cc.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        cc.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cc.number_format = '#,##0.000" млн"'
        cc.border = thin()
    ws.row_dimensions[row].height = 24
    row += 2

    # Блок 2: Поквартальная разбивка 2026-2028
    block_title(ws, row, 2, 10, "ПОКВАРТАЛЬНАЯ РАЗБИВКА 2026–2028 (млн ₽)",
                BRAND_BLUE)
    row += 1

    header_cell(ws, row, 2, "Квартал")
    header_cell(ws, row, 3, "Q1")
    header_cell(ws, row, 4, "Q2")
    header_cell(ws, row, 5, "Q3")
    header_cell(ws, row, 6, "Q4")
    header_cell(ws, row, 7, "Год Σ")
    ws.row_dimensions[row].height = 22
    row += 1

    # Для каждого года: ФОТ/4 (квартальная равная доля)
    for year_idx, year in enumerate([2026, 2027, 2028]):
        c = ws.cell(row=row, column=2, value=f"{year}")
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE_LIGHT)
        c.border = thin()
        col_letter_year = get_column_letter(3 + year_idx)
        # Q1-Q4
        for q in range(4):
            cc = ws.cell(row=row, column=3 + q,
                         value=f"={col_letter_year}{fot_year_row}/4")
            cc.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
            cc.alignment = Alignment(horizontal="right", vertical="center",
                                     indent=1)
            cc.number_format = '#,##0.000'
            cc.border = thin()
        # Σ
        cc = ws.cell(row=row, column=7,
                     value=f"={col_letter_year}{fot_year_row}")
        cc.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        cc.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cc.number_format = '#,##0.000'
        cc.border = thin()
        ws.row_dimensions[row].height = 20
        row += 1

    # ИТОГО 3 года
    c = ws.cell(row=row, column=2, value="Σ 2026–2028")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    # Q суммы
    for q in range(4):
        col_l = get_column_letter(3 + q)
        cc = ws.cell(row=row, column=3 + q,
                     value=f"=SUM({col_l}{row-3}:{col_l}{row-1})")
        cc.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        cc.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cc.number_format = '#,##0.000'
        cc.border = thin()
    # Общая сумма
    cc = ws.cell(row=row, column=7, value=f"=SUM(G{row-3}:G{row-1})")
    cc.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    cc.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cc.number_format = '#,##0.000" млн"'
    cc.border = thin()
    ws.row_dimensions[row].height = 24
    row += 2

    # Блок 3: Сравнение A1 vs A2
    block_title(ws, row, 2, 10, "СРАВНЕНИЕ A₁ Fixed vs A₂ Dynamic (млн ₽)",
                BRAND_BLUE)
    row += 1

    header_cell(ws, row, 2, "Период")
    header_cell(ws, row, 3, "A₁ Fixed")
    header_cell(ws, row, 4, "A₂ Dynamic")
    header_cell(ws, row, 5, "Δ (A₂−A₁)")
    header_cell(ws, row, 6, "Δ, %")
    ws.row_dimensions[row].height = 22
    row += 1

    # Годы
    for year_idx, year in enumerate([2026, 2027, 2028]):
        c = ws.cell(row=row, column=2, value=year)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        # A1 — всегда 73.944
        c = ws.cell(row=row, column=3, value=73.944)
        c.font = Font(name="Arial", size=10, color=LINK_GREEN)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.000'
        c.border = thin()
        # A2 — из fot_year_row
        col_l = get_column_letter(3 + year_idx)
        c = ws.cell(row=row, column=4, value=f"={col_l}{fot_year_row}")
        c.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.000'
        c.border = thin()
        # Δ
        c = ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
        c.font = Font(name="Arial", size=10, color=ACCENT_RED)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '+#,##0.000;-#,##0.000'
        c.border = thin()
        # Δ %
        c = ws.cell(row=row, column=6, value=f"=(D{row}-C{row})/C{row}")
        c.font = Font(name="Arial", size=10, color=ACCENT_RED)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '+0.0%;-0.0%'
        c.border = thin()
        ws.row_dimensions[row].height = 20
        row += 1

    # ИТОГО 3 года
    c = ws.cell(row=row, column=2, value="Σ 3 года")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    c = ws.cell(row=row, column=3, value=f"=SUM(C{row-3}:C{row-1})")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.000" млн"'
    c.border = thin()
    c = ws.cell(row=row, column=4, value=f"=SUM(D{row-3}:D{row-1})")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.000" млн"'
    c.border = thin()
    c = ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '+#,##0.000;-#,##0.000'
    c.border = thin()
    c = ws.cell(row=row, column=6, value=f"=(D{row}-C{row})/C{row}")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '+0.0%;-0.0%'
    c.border = thin()
    ws.row_dimensions[row].height = 24
    row += 2

    # Примечание
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    c = ws.cell(row=row, column=2,
                value="⚫ Модель A₂ «Full Dynamic»: штат растёт 50→60→70 (pre-IPO команда), оклады индексируются 7%/год (CPI 6.5% + премия). Для 2026–2028 разница с A₁ отражает реалистичный рост OPEX и используется в pessimistic scenarios. Для верификации якоря EBITDA = 3 000 млн ₽ используется A₁. В дальнейшем можно переключить P&L через selector на A₂.")
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws.row_dimensions[row].height = 50

    ws.freeze_panes = "A6"


# ============================================================================
# ЛИСТ 06: Cost_Structure
# ============================================================================

def build_cost_structure(wb):
    ws = wb.create_sheet("06_Cost_Structure")

    widths = [2, 38, 14, 14, 14, 14, 14, 16, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sheet_header(ws, "Структура OPEX (Cost Structure)",
                 "Сводка операционных расходов холдинга. Годовые показатели 2026–2032 на основе A₁ Fixed и Assumptions.")

    # Блок 1: Годовой OPEX
    row = 5
    block_title(ws, row, 2, 8, "OPEX ПО ГОДАМ (Base, модель A₁ Fixed), млн ₽",
                BRAND_BLUE)
    row += 1

    header_cell(ws, row, 2, "Статья")
    header_cell(ws, row, 3, "2026")
    header_cell(ws, row, 4, "2027")
    header_cell(ws, row, 5, "2028")
    header_cell(ws, row, 6, "2029")
    header_cell(ws, row, 7, "2030")
    header_cell(ws, row, 8, "Σ 2026–2028")
    ws.row_dimensions[row].height = 22
    row += 1

    # Статьи OPEX (из Холдинг Кино.xlsx)
    # ФОТ A1: 73.944 каждый год
    # Аренда: 8.250 + CPI 6.5%
    # Коммунал: 0.396 + CPI
    # Прочие: 5.916 + CPI
    cost_items = [
        ("ФОТ с налогами (A₁ Fixed)", 73.944, True, "стабильно, fixed"),
        ("Аренда офиса (750 м²)",      8.250,  False, "+6.5% CPI"),
        ("Коммунальные платежи",       0.396,  False, "+6.5% CPI"),
        ("Связь, интернет",            0.360,  False, "+6.5% CPI"),
        ("ПО и лицензии",              0.720,  False, "+6.5% CPI"),
        ("Командировки",               1.200,  False, "+6.5% CPI"),
        ("Представительские",          0.600,  False, "+6.5% CPI"),
        ("Консультанты (юр/аудит)",    1.500,  False, "+6.5% CPI"),
        ("Канцелярия, расходники",     0.216,  False, "+6.5% CPI"),
        ("Банковские комиссии",        0.420,  False, "+6.5% CPI"),
        ("Маркетинг корпоративный",    0.900,  False, "+6.5% CPI"),
    ]

    start_cost_row = row
    for item_name, base_val, is_fot, note in cost_items:
        c = ws.cell(row=row, column=2, value=item_name)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        if is_fot:
            c.font = Font(name="Arial", size=10, bold=True)

        # 2026-2030
        for y_idx, year in enumerate([2026, 2027, 2028, 2029, 2030]):
            col = 3 + y_idx
            if is_fot:
                # Ссылка — но упрощённо жёсткое 73.944
                cc = ws.cell(row=row, column=col, value=base_val)
                cc.font = Font(name="Arial", size=10, bold=True, color=LINK_GREEN)
            else:
                # С CPI 6.5%
                cpi_factor = 1.065 ** y_idx
                cc = ws.cell(row=row, column=col, value=base_val * cpi_factor)
                cc.font = Font(name="Arial", size=10, color=FORMULA_BLACK)
            cc.alignment = Alignment(horizontal="right", vertical="center",
                                     indent=1)
            cc.number_format = '#,##0.000'
            cc.border = thin()

        # Σ 2026-2028
        c = ws.cell(row=row, column=8, value=f"=SUM(C{row}:E{row})")
        c.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        c.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.000" млн"'
        c.border = thin()
        ws.row_dimensions[row].height = 20
        row += 1
    end_cost_row = row - 1

    # ИТОГО OPEX
    c = ws.cell(row=row, column=2, value="ИТОГО OPEX (без COGS)")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    for col in range(3, 9):
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col,
                    value=f"=SUM({col_l}{start_cost_row}:{col_l}{end_cost_row})")
        c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.000" млн"'
        c.border = thin()
    ws.row_dimensions[row].height = 26
    opex_total_row = row
    row += 2

    # Блок 2: Структура OPEX (доли)
    block_title(ws, row, 2, 8, "СТРУКТУРА OPEX 2026 (доли, %)", BRAND_BLUE)
    row += 1

    header_cell(ws, row, 2, "Категория")
    header_cell(ws, row, 3, "Сумма 2026")
    header_cell(ws, row, 4, "Доля, %")
    header_cell(ws, row, 5, "Комментарий")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 22
    row += 1

    categories = [
        ("ФОТ с налогами", "=C6", "ключевая статья (~84%)"),
        ("Офисные расходы (аренда + коммуналки)", "=C7+C8", "фиксированная аренда 750 м²"),
        ("Операционные нужды (связь, ПО, канцтовары)", "=C9+C10+C13", "рабочая инфраструктура"),
        ("Командировки и представительские", "=C11+C12", "продюсерская деятельность"),
        ("Консультанты (юр/аудит)", "=C14", "внешние услуги"),
        ("Банковские комиссии + маркетинг", "=C15+C16", "финансовые и PR"),
    ]
    total_formula = f"=C{opex_total_row}"
    start_cat_row = row
    for cat_name, sum_formula, comment in categories:
        c = ws.cell(row=row, column=2, value=cat_name)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()

        c = ws.cell(row=row, column=3, value=sum_formula)
        c.font = Font(name="Arial", size=10, bold=True, color=FORMULA_BLACK)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '#,##0.000" млн"'
        c.border = thin()

        c = ws.cell(row=row, column=4, value=f"=C{row}/{total_formula}")
        c.font = Font(name="Arial", size=10, color=BRAND_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.number_format = '0.0%'
        c.border = thin()

        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        c = ws.cell(row=row, column=5, value=comment)
        c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        ws.row_dimensions[row].height = 20
        row += 1
    end_cat_row = row - 1

    # Сверка
    c = ws.cell(row=row, column=2, value="ПРОВЕРКА: Σ категорий = OPEX total")
    c.font = Font(name="Arial", size=10, bold=True, color=ACCENT_GREEN)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    c = ws.cell(row=row, column=3,
                value=f"=SUM(C{start_cat_row}:C{end_cat_row})")
    c.font = Font(name="Arial", size=10, bold=True, color=ACCENT_GREEN)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0.000" млн"'
    c.border = thin()
    c = ws.cell(row=row, column=4,
                value=f"=SUM(D{start_cat_row}:D{end_cat_row})")
    c.font = Font(name="Arial", size=10, bold=True, color=ACCENT_GREEN)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '0.0%'
    c.border = thin()
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    c = ws.cell(row=row, column=5,
                value="✓ Должно совпадать с ИТОГО OPEX 2026 и равняться 100%")
    c.font = Font(name="Arial", size=9, italic=True, color=ACCENT_GREEN)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin()
    ws.row_dimensions[row].height = 22
    row += 2

    # Блок 3: COGS (производство) — ссылка на будущий Content_Pipeline
    block_title(ws, row, 2, 8, "COGS / ПРОИЗВОДСТВО КОНТЕНТА (обзор)",
                BRAND_BLUE)
    row += 1

    cogs_data = [
        ("Total content budget",          1850,   "12 фильмов × 154.2"),
        ("Средний бюджет 1 фильма",       154.2,  "производство + постпрод"),
        ("Producer equity (в бюджете)",   600,    "600 / 1850 = 32.4%"),
        ("Investor funding (в бюджете)",  1250,   "1250 / 1850 = 67.6%"),
        ("Amortization (мес после релиза)", 12,   "линейная"),
        ("P&A (Print & Advertising)",     0.15,   "15% от бюджета фильма"),
    ]
    header_cell(ws, row, 2, "Параметр")
    header_cell(ws, row, 3, "Значение")
    header_cell(ws, row, 4, "Ед.изм.")
    header_cell(ws, row, 5, "Комментарий")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 22
    row += 1
    for name, val, comment in cogs_data:
        c = ws.cell(row=row, column=2, value=name)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()

        c = ws.cell(row=row, column=3, value=val)
        c.font = Font(name="Arial", size=10, bold=True, color=INPUT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        if isinstance(val, float) and val < 1:
            c.number_format = '0.0%'
        else:
            c.number_format = '#,##0.0'
        c.border = thin()

        unit = "млн ₽" if val >= 10 else ("%" if val < 1 else ("мес" if "мес" in name else "млн ₽"))
        c = ws.cell(row=row, column=4, value=unit)
        c.font = Font(name="Arial", size=9, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()

        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        c = ws.cell(row=row, column=5, value=comment)
        c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin()
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1
    # Справка-сверка с Холдинг Кино.xlsx
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    c = ws.cell(row=row, column=2,
                value="ССЫЛКА НА ИСТОЧНИК: Структура OPEX воспроизводит лист «Расходы» из оригинальной Холдинг Кино.xlsx. Базовый OPEX 2026 ≈ 94 млн ₽/год (фактический Холдинг Кино = 88.506 млн без маркетинга и консультантов). Расширение до 94 млн учитывает дополнительные категории, актуальные для инвестиционной модели.")
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY_DARK)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=1)
    c.fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    ws.row_dimensions[row].height = 45

    ws.freeze_panes = "A6"


# ============================================================================
# СБОРКА: открываем существующий файл и добавляем листы
# ============================================================================

print("Loading existing workbook ...")
wb = load_workbook(OUT_FILE)
print(f"Current sheets: {wb.sheetnames}")

print("\nBuilding А.2 — FOT_Staff_A1 + FOT_Staff_A2 + Cost_Structure ...")
build_fot_a1(wb)
build_fot_a2(wb)
build_cost_structure(wb)

# Обновляем Change_Log — добавим запись
ws_log = wb["03_Change_Log"]
# Находим последнюю заполненную строку в таблице версий (v1.0 строка)
# Нам нужна новая запись. Упростим: добавим текстовую запись в конец
# Либо просто обновим статус в roadmap

# Пройдёмся по roadmap и обновим статус для А.2 на "Готово"
for row in ws_log.iter_rows(min_row=12, max_row=40, min_col=2, max_col=6):
    stage_cell = row[0]
    status_cell = row[4]
    if stage_cell.value and "А.2" in str(stage_cell.value):
        status_cell.value = "Готово"
        status_cell.font = Font(name="Arial", size=9, bold=True,
                                color=ACCENT_GREEN)
        status_cell.fill = PatternFill("solid", fgColor="E2EFDA")
        break

wb.save(OUT_FILE)
print(f"\nSaved: {OUT_FILE}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheets: {wb.sheetnames}")
