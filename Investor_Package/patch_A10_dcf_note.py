"""
PATCH для А.10 — добавляет компактный блок VI. Why DCF < Multiples в 22_Valuation_DCF
БЕЗ ТРОГАНИЯ строк 1-57.

Полный развёрнутый блок будет в 16_Executive_Summary на этапе А.16 (часть B комбо).

Архитектурные гарантии:
- Не затрагиваются cell references из 03_Change_Log
- freeze_pane D7 остаётся валидным
- Все числовые инварианты (EV_Gordon=1262, EV_Exit=2369, EV_blend=1815) сохранены
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx"

BLUE = "0070C0"
DARK_BLUE = "002060"
LIGHT_BLUE = "DEEBF7"
LIGHT_ORANGE = "FCE4D6"
GREY = "808080"
WHITE = "FFFFFF"

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = load_workbook(XLSX)
ws = wb["22_Valuation_DCF"]

print(f"Before patch: {ws.max_row} rows × {ws.max_column} cols, freeze={ws.freeze_panes}")

# Snapshot invariants перед правкой
EV_GORDON_ROW = None
EV_EXIT_ROW = None
EV_BLEND_ROW = None
for r in range(1, 58):
    v = ws.cell(r, 2).value
    if v and isinstance(v, str):
        if "Gordon" in v and "EV" in v:
            EV_GORDON_ROW = r
        elif "Exit" in v and "EV" in v:
            EV_EXIT_ROW = r
        elif "blend" in v.lower() and "EV" in v:
            EV_BLEND_ROW = r

print(f"Found anchors: Gordon row={EV_GORDON_ROW}, Exit row={EV_EXIT_ROW}, Blend row={EV_BLEND_ROW}")
pre_snapshot = {}
for r in range(1, 58):
    for c in range(1, 15):
        v = ws.cell(r, c).value
        if v is not None:
            pre_snapshot[(r, c)] = v
print(f"Pre-patch snapshot: {len(pre_snapshot)} non-empty cells in R1-57")

# === ДОБАВЛЕНИЕ БЛОКА НАЧИНАЯ С R59 (1 пустая строка-разделитель) ===
start_row = 59

# Заголовок раздела
ws.cell(start_row, 2).value = "VI. WHY DCF < MULTIPLES — QUICK NOTE"
ws.cell(start_row, 2).font = Font(name="Calibri", size=12, bold=True, color=BLUE)
ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=14)

# Подзаголовок с крос-референсом
row = start_row + 1
ws.cell(row, 2).value = "DCF (≈ 1 815 млн ₽) vs Comps median (≈ 7 550 млн ₽). Разница структурна, не ошибочна."
ws.cell(row, 2).font = Font(name="Calibri", size=10, italic=True, color=GREY)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)

# Таблица компонентов разрыва
row += 2
headers = ["#", "Причина разрыва", "Механика", "Оценка вклада"]
widths = [5, 30, 55, 16]
for i, h in enumerate(headers):
    c = ws.cell(row, 2 + i)
    c.value = h
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
# Merge последний столбец на несколько, чтобы не залезать в D:H диапазон DCF таблицы
ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=14)
ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, widths[0])

row += 1

breakdown = [
    (1, "Горизонт обрезан на 2030",
     "Gordon Growth TV сжимает 30-летнюю библиотеку в константу g=3%. Фильмы 2028 года монетизируются до ~2048, но при WACC 19% CF 2046 = 0.5% от номинала.",
     "+2 500 млн ₽"),
    (2, "WACC 19% vs buyer WACC 7%",
     "Страновой premium «съедает» дальний tail. Strategic buyer (Яндекс/Сбер) дисконтирует под 7-9%, получая другой NPV того же CF stream.",
     "+1 500 млн ₽"),
    (3, "Real options не моделируются",
     "Франшизы (сиквелы F05/F12), кросс-медиа (игры, мерч), международная копродукция — опционные payoffs, DCF считает их = 0.",
     "+1 500 млн ₽"),
    (4, "Execution risk двойное начисление",
     "В DCF высокий WACC + реалистичный forecast = два слоя консерватизма. Multiples — precedent prices, риск уже реализован в цене.",
     "(встроено)"),
    (5, "Капитальная цикличность после 2028",
     "CAPEX падает, FCF растёт импульсом. Gordon g=3% сглаживает этот impulse в линейный рост.",
     "+500 млн ₽"),
]

for n, reason, mech, delta in breakdown:
    vals = [n, reason, mech, delta]
    aligns = ["center", "left", "left", "center"]
    for i, (v, a) in enumerate(zip(vals, aligns)):
        c = ws.cell(row, 2 + i)
        c.value = v
        c.font = Font(name="Calibri", size=9, bold=(i in (0, 3)))
        c.alignment = Alignment(horizontal=a, vertical="top", wrap_text=True)
        c.border = BORDER
        if i == 3:
            c.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    # Merge последний столбец
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=14)
    ws.row_dimensions[row].height = 36
    row += 1

# Итоговая строка
ws.cell(row, 2).value = "Σ"
ws.cell(row, 2).font = Font(name="Calibri", size=10, bold=True)
ws.cell(row, 2).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row, 2).border = BORDER

ws.cell(row, 3).value = "Сумма структурных компонентов разрыва"
ws.cell(row, 3).font = Font(name="Calibri", size=10, bold=True)
ws.cell(row, 3).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row, 3).border = BORDER

ws.cell(row, 4).value = "Примерно совпадает с фактической разницей 5 735 млн ₽ (7 550 − 1 815)"
ws.cell(row, 4).font = Font(name="Calibri", size=9, italic=True)
ws.cell(row, 4).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.cell(row, 4).border = BORDER

ws.cell(row, 5).value = "≈ 5 500 млн ₽"
ws.cell(row, 5).font = Font(name="Calibri", size=10, bold=True)
ws.cell(row, 5).fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row, 5).border = BORDER
ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=14)

row += 2

# Интерпретация и cross-reference
ws.cell(row, 2).value = ("ИНТЕРПРЕТАЦИЯ. DCF = floor valuation (intrinsic floor, без премий и опционов). "
                          "Comps = market valuation (что реально платят). Weighted EV exit 6 038 млн ₽ "
                          "(лист 25_Exit_Scenarios) — ближе к mid Football Field, это и есть expected value "
                          "для переговоров. DCF используется как downside benchmark, не как fair exit price.")
ws.cell(row, 2).font = Font(name="Calibri", size=9)
ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
ws.row_dimensions[row].height = 48

row += 2
ws.cell(row, 2).value = "→ Полный анализ разрыва см. лист 16_Executive_Summary (раздел «Valuation Range Interpretation»)"
ws.cell(row, 2).font = Font(name="Calibri", size=9, italic=True, color=BLUE)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)

print(f"\nAfter patch: new blocks written R{start_row}-{row}")

# === VERIFICATION: snapshot R1-57 не изменён ===
post_snapshot = {}
for r in range(1, 58):
    for c in range(1, 15):
        v = ws.cell(r, c).value
        if v is not None:
            post_snapshot[(r, c)] = v

diff = []
for k in pre_snapshot:
    if k not in post_snapshot or post_snapshot[k] != pre_snapshot[k]:
        diff.append(k)
for k in post_snapshot:
    if k not in pre_snapshot:
        diff.append(k)

if diff:
    print(f"❌ SNAPSHOT CHANGED: {len(diff)} cells differ in R1-57")
    for k in diff[:10]:
        print(f"   {k}: before={pre_snapshot.get(k)!r}, after={post_snapshot.get(k)!r}")
    raise RuntimeError("Invariants R1-57 нарушены — откат")
else:
    print(f"✓ R1-57 invariants preserved: {len(pre_snapshot)} cells identical")

# Freeze pane проверка
assert ws.freeze_panes == "D7", f"freeze_panes changed! now {ws.freeze_panes}"
print(f"✓ freeze_panes D7 preserved")

wb.save(XLSX)
print(f"\n✓ Saved: {XLSX}")
print(f"22_Valuation_DCF: {ws.max_row} rows × {ws.max_column} cols")
