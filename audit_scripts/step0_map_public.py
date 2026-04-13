#!/usr/bin/env python3
"""
DD-Grade Audit: Step 0 — Map the complete structure of the PUBLIC xlsx file.
Reads investor_model_v1.0_Public.xlsx with openpyxl (both formula and cached-value modes).
Outputs a full JSON structure and a human-readable summary.
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

FILE = Path("/home/user/TrendStudio-Holding/Investor_Package/investor_model_v1.0_Public.xlsx")

# ── helpers ──────────────────────────────────────────────────────────────────

ERROR_PATTERNS = re.compile(r"#REF!|#N/A|#VALUE!|#DIV/0!|#NAME\?|#NULL!|#NUM!")
EXTERNAL_LINK_RE = re.compile(r"\[.*?\]")  # e.g. [OtherBook.xlsx]Sheet1!A1


def visibility_label(state):
    if state == "visible" or state is None:
        return "visible"
    return str(state)  # "hidden" or "veryHidden"


def cell_repr(cell):
    """Return a string representation of a cell's value/formula."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str) and v.startswith("="):
        return v  # formula
    return v


def is_formula(cell):
    v = cell.value
    return isinstance(v, str) and v.startswith("=")


# ── load workbooks ───────────────────────────────────────────────────────────

print(f"Loading workbook (formulas): {FILE}")
wb_formula = openpyxl.load_workbook(FILE, data_only=False, read_only=False)

print(f"Loading workbook (cached values): {FILE}")
wb_cached = openpyxl.load_workbook(FILE, data_only=True, read_only=False)

# ── document properties ──────────────────────────────────────────────────────

props = wb_formula.properties
doc_properties = {
    "creator": props.creator,
    "last_modified_by": props.lastModifiedBy,
    "created": str(props.created) if props.created else None,
    "modified": str(props.modified) if props.modified else None,
    "title": props.title,
    "description": props.description,
    "subject": props.subject,
    "keywords": props.keywords,
    "category": props.category,
    "company": None,  # filled below
    "manager": None,
    "revision": props.revision,
    "version": props.version,
    "content_status": props.contentStatus,
    "identifier": props.identifier,
    "language": props.language,
}

# Try to get extended properties (company, manager)
try:
    ext = wb_formula.extended_properties if hasattr(wb_formula, 'extended_properties') else None
    if ext:
        doc_properties["company"] = getattr(ext, "company", None)
        doc_properties["manager"] = getattr(ext, "manager", None)
except Exception:
    pass

# Also try custom properties
custom_props_list = []
try:
    if hasattr(wb_formula, 'custom_doc_props'):
        for cp in wb_formula.custom_doc_props:
            custom_props_list.append({"name": cp.name, "value": str(cp.value)})
except Exception:
    pass

doc_properties["custom_properties"] = custom_props_list

# ── global defined names ─────────────────────────────────────────────────────

global_defined_names = []
sheet_scoped_names = defaultdict(list)

# openpyxl >= 3.1 uses a DefinedNameDict; iterate with .items()
for name, defn in wb_formula.defined_names.items():
    entry = {
        "name": name,
        "value": defn.value,
        "hidden": getattr(defn, 'hidden', False),
        "comment": getattr(defn, 'comment', None),
    }
    if defn.localSheetId is not None:
        try:
            sn = wb_formula.sheetnames[defn.localSheetId]
        except IndexError:
            sn = f"__unknown_sheet_id_{defn.localSheetId}__"
        sheet_scoped_names[sn].append(entry)
    else:
        global_defined_names.append(entry)

# Also gather sheet-level defined names
for sn in wb_formula.sheetnames:
    ws_tmp = wb_formula[sn]
    if hasattr(ws_tmp, 'defined_names'):
        for dn_name, dn_defn in ws_tmp.defined_names.items():
            entry = {
                "name": dn_name,
                "value": dn_defn.value,
                "hidden": getattr(dn_defn, 'hidden', False),
                "comment": getattr(dn_defn, 'comment', None),
            }
            sheet_scoped_names[sn].append(entry)

# ── per-sheet analysis ───────────────────────────────────────────────────────

sheets_data = []

for sheet_name in wb_formula.sheetnames:
    ws_f = wb_formula[sheet_name]
    ws_c = wb_cached[sheet_name]

    # Visibility
    vis = visibility_label(ws_f.sheet_state)

    # Dimensions
    min_row = ws_f.min_row
    max_row = ws_f.max_row
    min_col = ws_f.min_column
    max_col = ws_f.max_column
    dims = {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "min_col_letter": get_column_letter(min_col) if min_col else None,
        "max_col_letter": get_column_letter(max_col) if max_col else None,
        "total_rows_used": (max_row - min_row + 1) if max_row and min_row else 0,
        "total_cols_used": (max_col - min_col + 1) if max_col and min_col else 0,
    }

    # Protection
    protection_info = {}
    sp = ws_f.protection
    protection_info["sheet_protected"] = sp.sheet
    protection_info["password_set"] = bool(sp.password) if hasattr(sp, 'password') and sp.password else False
    protection_info["format_cells"] = sp.formatCells
    protection_info["format_columns"] = sp.formatColumns
    protection_info["format_rows"] = sp.formatRows
    protection_info["insert_columns"] = sp.insertColumns
    protection_info["insert_rows"] = sp.insertRows
    protection_info["delete_columns"] = sp.deleteColumns
    protection_info["delete_rows"] = sp.deleteRows
    protection_info["sort"] = sp.sort
    protection_info["auto_filter"] = sp.autoFilter

    # Count formulas vs values, collect formula samples, errors, external links
    formula_count = 0
    value_count = 0
    empty_count = 0
    formula_samples = []
    errors_found = []
    external_links = []

    # Iterate all cells in the used range
    if max_row and max_col:
        for row in ws_f.iter_rows(min_row=min_row, max_row=max_row,
                                   min_col=min_col, max_col=max_col):
            for cell in row:
                v = cell.value
                if v is None:
                    empty_count += 1
                    continue
                if isinstance(v, str) and v.startswith("="):
                    formula_count += 1
                    # Sample formulas (up to 10)
                    if len(formula_samples) < 10:
                        formula_samples.append({
                            "cell": cell.coordinate,
                            "formula": v,
                        })
                    # Check for external links in formula
                    if EXTERNAL_LINK_RE.search(v):
                        external_links.append({
                            "cell": cell.coordinate,
                            "formula": v,
                        })
                else:
                    value_count += 1

        # Check cached values for errors
        for row in ws_c.iter_rows(min_row=min_row, max_row=max_row,
                                   min_col=min_col, max_col=max_col):
            for cell in row:
                v = cell.value
                if v is not None and isinstance(v, str) and ERROR_PATTERNS.search(v):
                    errors_found.append({
                        "cell": cell.coordinate,
                        "cached_value": v,
                    })

    # Defined names scoped to this sheet
    scoped_names = sheet_scoped_names.get(sheet_name, [])

    sheet_info = {
        "name": sheet_name,
        "visibility": vis,
        "dimensions": dims,
        "protection": protection_info,
        "cell_counts": {
            "formulas": formula_count,
            "values": value_count,
            "empty": empty_count,
            "total_non_empty": formula_count + value_count,
        },
        "formula_samples": formula_samples,
        "errors_in_cached_values": errors_found,
        "external_links_in_formulas": external_links,
        "defined_names_scoped_here": scoped_names,
    }
    sheets_data.append(sheet_info)

# ── also check for external links at workbook level ──────────────────────────

wb_external_links = []
try:
    if hasattr(wb_formula, '_external_links') and wb_formula._external_links:
        for el in wb_formula._external_links:
            wb_external_links.append(str(el))
except Exception:
    pass

# ── assemble full result ─────────────────────────────────────────────────────

result = {
    "file": str(FILE),
    "file_size_bytes": FILE.stat().st_size,
    "sheet_count": len(wb_formula.sheetnames),
    "sheet_names": wb_formula.sheetnames,
    "document_properties": doc_properties,
    "global_defined_names": global_defined_names,
    "workbook_external_links": wb_external_links,
    "sheets": sheets_data,
}

# ── output JSON ──────────────────────────────────────────────────────────────

json_output = json.dumps(result, indent=2, default=str)
print("\n" + "=" * 80)
print("FULL JSON STRUCTURE")
print("=" * 80)
print(json_output)

# ── human-readable summary ───────────────────────────────────────────────────

print("\n" + "=" * 80)
print("HUMAN-READABLE AUDIT SUMMARY")
print("=" * 80)

print(f"\nFile: {FILE}")
print(f"Size: {FILE.stat().st_size:,} bytes")
print(f"Sheets: {len(wb_formula.sheetnames)}")

print("\n--- Document Properties ---")
for k, v in doc_properties.items():
    if k == "custom_properties":
        if v:
            print(f"  custom_properties:")
            for cp in v:
                print(f"    {cp['name']}: {cp['value']}")
        else:
            print(f"  custom_properties: (none)")
    else:
        print(f"  {k}: {v}")

print("\n--- Global Defined Names ---")
if global_defined_names:
    for dn in global_defined_names:
        hidden_flag = " [HIDDEN]" if dn.get("hidden") else ""
        print(f"  {dn['name']}{hidden_flag} = {dn['value']}")
else:
    print("  (none)")

print("\n--- Workbook-Level External Links ---")
if wb_external_links:
    for el in wb_external_links:
        print(f"  {el}")
else:
    print("  (none detected at workbook level)")

print("\n--- Per-Sheet Details ---")
for s in sheets_data:
    print(f"\n  Sheet: '{s['name']}' [{s['visibility']}]")
    d = s["dimensions"]
    print(f"    Range: {d['min_col_letter']}{d['min_row']}:{d['max_col_letter']}{d['max_row']}"
          f"  ({d['total_rows_used']} rows x {d['total_cols_used']} cols)")
    cc = s["cell_counts"]
    print(f"    Formulas: {cc['formulas']}  |  Values: {cc['values']}  |  Empty: {cc['empty']}"
          f"  |  Total non-empty: {cc['total_non_empty']}")

    prot = s["protection"]
    if prot["sheet_protected"]:
        print(f"    PROTECTION: ENABLED (password_set={prot['password_set']})")
    else:
        print(f"    Protection: not enabled")

    if s["defined_names_scoped_here"]:
        print(f"    Scoped Defined Names:")
        for dn in s["defined_names_scoped_here"]:
            hidden_flag = " [HIDDEN]" if dn.get("hidden") else ""
            print(f"      {dn['name']}{hidden_flag} = {dn['value']}")

    if s["formula_samples"]:
        print(f"    Formula Samples (up to 10):")
        for fs in s["formula_samples"]:
            print(f"      {fs['cell']}: {fs['formula']}")

    if s["errors_in_cached_values"]:
        print(f"    *** ERRORS in cached values ({len(s['errors_in_cached_values'])} found) ***")
        for err in s["errors_in_cached_values"]:
            print(f"      {err['cell']}: {err['cached_value']}")
    else:
        print(f"    Errors in cached values: none")

    if s["external_links_in_formulas"]:
        print(f"    *** EXTERNAL LINKS in formulas ({len(s['external_links_in_formulas'])} found) ***")
        for el in s["external_links_in_formulas"]:
            print(f"      {el['cell']}: {el['formula']}")
    else:
        print(f"    External links in formulas: none")

# ── overall totals ───────────────────────────────────────────────────────────

total_formulas = sum(s["cell_counts"]["formulas"] for s in sheets_data)
total_values = sum(s["cell_counts"]["values"] for s in sheets_data)
total_errors = sum(len(s["errors_in_cached_values"]) for s in sheets_data)
total_ext_links = sum(len(s["external_links_in_formulas"]) for s in sheets_data)
hidden_sheets = [s["name"] for s in sheets_data if s["visibility"] != "visible"]
protected_sheets = [s["name"] for s in sheets_data if s["protection"]["sheet_protected"]]

print("\n" + "=" * 80)
print("AUDIT TOTALS")
print("=" * 80)
print(f"  Total formulas across all sheets: {total_formulas}")
print(f"  Total value cells across all sheets: {total_values}")
print(f"  Total cached-value errors: {total_errors}")
print(f"  Total external-link formulas: {total_ext_links}")
print(f"  Hidden sheets: {hidden_sheets if hidden_sheets else '(none)'}")
print(f"  Protected sheets: {protected_sheets if protected_sheets else '(none)'}")

# ── metadata sensitivity flags ───────────────────────────────────────────────

print("\n--- Metadata Sensitivity Check ---")
sensitive_flags = []
if doc_properties.get("creator") and doc_properties["creator"].strip():
    sensitive_flags.append(f"Creator field set: '{doc_properties['creator']}'")
if doc_properties.get("last_modified_by") and doc_properties["last_modified_by"].strip():
    sensitive_flags.append(f"Last Modified By field set: '{doc_properties['last_modified_by']}'")
if doc_properties.get("company") and str(doc_properties["company"]).strip():
    sensitive_flags.append(f"Company field set: '{doc_properties['company']}'")
if doc_properties.get("manager") and str(doc_properties["manager"]).strip():
    sensitive_flags.append(f"Manager field set: '{doc_properties['manager']}'")
if doc_properties.get("keywords") and doc_properties["keywords"].strip():
    sensitive_flags.append(f"Keywords field set: '{doc_properties['keywords']}'")
if doc_properties.get("description") and doc_properties["description"].strip():
    sensitive_flags.append(f"Description field set: '{doc_properties['description']}'")
if custom_props_list:
    sensitive_flags.append(f"Custom document properties present: {len(custom_props_list)}")

if sensitive_flags:
    print("  *** POTENTIAL SENSITIVE METADATA FOUND ***")
    for flag in sensitive_flags:
        print(f"    - {flag}")
else:
    print("  No sensitive metadata detected.")

print("\n" + "=" * 80)
print("AUDIT COMPLETE")
print("=" * 80)

# Save JSON to file as well
json_path = FILE.parent / "audit_step0_structure.json"
with open(json_path, "w") as f:
    f.write(json_output)
print(f"\nJSON also saved to: {json_path}")

wb_formula.close()
wb_cached.close()
