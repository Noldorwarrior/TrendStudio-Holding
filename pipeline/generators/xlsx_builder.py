"""
generators/xlsx_builder.py — сборка artifacts/model.xlsx на 21 листе.

Вход: ValidatedInputs, RunAllResult + sensitivity/stress/mc/provenance/manifest
Выход: artifacts/model.xlsx

21 лист:
  00_Cover, 01_Summary, 02_Anchor_check, 03_Revenue_segments, 04_Slate_films,
  05_Costs, 06_PnL_cons, 07_PnL_base, 08_PnL_opt, 09_CashFlow_3scen,
  10_Quarterly_CF_base, 11_Valuation_DCF, 12_Sensitivity, 13_MonteCarlo,
  14_StressTests, 15_Investment_round, 16_InvestorReturns, 17_Macro,
  18_NWC_CapEx, 19_Provenance, 20_Manifest

Форматирование: header bold + fill #0070C0 white, money format #,##0, %-format.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from generators.core import RunAllResult
from generators.hash_manifest import build_manifest
from generators.monte_carlo import MonteCarloResult
from generators.provenance import ProvenanceRegistry
from generators.sensitivity import SensitivityTable
from generators.stress_tests import StressResults
from schemas.inputs import ValidatedInputs

YEARS = (2026, 2027, 2028)
SCENARIOS_ORDER = ("cons", "base", "opt")

HEADER_FILL = PatternFill("solid", fgColor="0070C0")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="0070C0")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="0070C0")
BORDER = Border(
    left=Side(border_style="thin", color="B0B0B0"),
    right=Side(border_style="thin", color="B0B0B0"),
    top=Side(border_style="thin", color="B0B0B0"),
    bottom=Side(border_style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


# ─────────────── helpers ───────────────

def _header_row(ws, row: int, headers: List[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _set_column_widths(ws, widths: List[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_rows(
    ws, start_row: int, rows: List[List[Any]], num_fmt: str = "#,##0"
) -> int:
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float)) and c_idx > 1:
                cell.number_format = num_fmt
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
    return start_row + len(rows)


def _title(ws, text: str) -> None:
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT


# ─────────────── sheets ───────────────

def _sheet_00_cover(ws, inputs: ValidatedInputs, run: RunAllResult) -> None:
    _title(ws, "ТрендСтудио — Финансовая модель 2026–2028")
    _set_column_widths(ws, [35, 22, 22, 22])
    ws.cell(row=3, column=1, value="Сценарий").font = SECTION_FONT
    ws.cell(row=3, column=2, value="Revenue, млн ₽").font = SECTION_FONT
    ws.cell(row=3, column=3, value="EBITDA, млн ₽").font = SECTION_FONT
    ws.cell(row=3, column=4, value="Net Income, млн ₽").font = SECTION_FONT
    r = 4
    for s in SCENARIOS_ORDER:
        m = run.get(s)
        rev = sum(m.revenue.total_by_year(y) for y in YEARS)
        eb = m.cumulative_ebitda
        ni = sum(m.pnl.net_income.values())
        _write_rows(ws, r, [[s.upper(), round(rev, 0), round(eb, 0), round(ni, 0)]])
        r += 1
    r += 2
    ws.cell(row=r, column=1, value="Якорный инвариант (Base EBITDA 3y)").font = SECTION_FONT
    r += 1
    status = "PASS ✓" if run.anchor_passed else "FAIL ✗"
    _write_rows(ws, r, [
        ["Target, млн ₽", round(run.anchor_value, 1)],
        ["Actual Base, млн ₽", round(run.anchor_actual, 1)],
        ["Отклонение, %", f"{run.anchor_deviation_pct:+.3f}%"],
        ["Статус", status],
    ])


def _sheet_01_summary(ws, run: RunAllResult) -> None:
    _title(ws, "Summary — ключевые метрики по сценариям")
    _set_column_widths(ws, [28, 14, 14, 14, 14])
    _header_row(ws, 3, ["Метрика", "2026", "2027", "2028", "Σ 3 года"])
    r = 4
    for s in SCENARIOS_ORDER:
        ws.cell(row=r, column=1, value=f"── {s.upper()} ──").font = SECTION_FONT
        r += 1
        m = run.get(s)
        rows: List[List[Any]] = []
        for label, d in [
            ("Revenue", {y: m.revenue.total_by_year(y) for y in YEARS}),
            ("COGS", m.costs.cogs),
            ("Gross Profit", m.pnl.gross_profit),
            ("OPEX", m.costs.opex),
            ("P&A", m.costs.pa),
            ("Contingency", m.costs.contingency),
            ("EBITDA", m.pnl.ebitda),
            ("D&A", m.costs.depreciation),
            ("EBIT", m.pnl.ebit),
            ("Taxes", m.pnl.taxes),
            ("Net Income", m.pnl.net_income),
            ("Operating CF", m.cashflow.operating_cf),
            ("CapEx total", m.cashflow.capex),
            ("Free Cash Flow", m.cashflow.free_cash_flow),
        ]:
            total = sum(d[y] for y in YEARS)
            rows.append([label, round(d[2026], 0), round(d[2027], 0), round(d[2028], 0), round(total, 0)])
        r = _write_rows(ws, r, rows)
        r += 1


def _sheet_02_anchor(ws, inputs: ValidatedInputs, run: RunAllResult) -> None:
    _title(ws, "Anchor Check — проверка якорного инварианта")
    _set_column_widths(ws, [40, 22])
    a = inputs.scenarios.anchor
    _write_rows(ws, 3, [
        ["Anchor metric", a.metric],
        ["Scenario", a.scenario],
        ["Target value, млн ₽", round(a.value_mln_rub, 1)],
        ["Tolerance, %", f"±{a.tolerance_pct:.1f}%"],
        ["Actual (Base) EBITDA cum, млн ₽", round(run.anchor_actual, 1)],
        ["Deviation, %", f"{run.anchor_deviation_pct:+.3f}%"],
        ["Status", "PASS ✓" if run.anchor_passed else "FAIL ✗"],
        ["Rationale", a.rationale or ""],
    ])


def _sheet_03_revenue(ws, run: RunAllResult) -> None:
    _title(ws, "Revenue — 5 сегментов × 3 сценария × 3 года")
    _set_column_widths(ws, [22, 14, 14, 14, 14])
    _header_row(ws, 3, ["Сегмент × Сценарий", "2026", "2027", "2028", "Σ"])
    r = 4
    segments = ["cinema", "advertising", "festivals", "education", "license_library"]
    for seg in segments:
        for s in SCENARIOS_ORDER:
            m = run.get(s)
            d = getattr(m.revenue, seg)
            total = sum(d.values())
            _write_rows(ws, r, [[
                f"{seg} [{s}]",
                round(d[2026], 0), round(d[2027], 0), round(d[2028], 0),
                round(total, 0),
            ]])
            r += 1
        r += 1


def _sheet_04_slate(ws, inputs: ValidatedInputs) -> None:
    _title(ws, "Slate — 12 фильмов релизного слата")
    _set_column_widths(ws, [20, 28, 8, 10, 12, 12, 12, 10])
    _header_row(ws, 3, [
        "film_id", "title", "year", "quarter",
        "BO cons", "BO base", "BO opt", "hit_rate base",
    ])
    r = 4
    for f in inputs.slate.films:
        _write_rows(ws, r, [[
            f.id, f.title, f.release_year, f.release_quarter,
            round(f.box_office_mln_rub.cons, 0),
            round(f.box_office_mln_rub.base, 0),
            round(f.box_office_mln_rub.opt, 0),
            round(f.hit_rate.base, 2),
        ]])
        r += 1


def _sheet_05_costs(ws, run: RunAllResult) -> None:
    _title(ws, "Costs — 7 категорий по сценариям")
    _set_column_widths(ws, [22, 14, 14, 14, 14])
    _header_row(ws, 3, ["Категория × Сценарий", "2026", "2027", "2028", "Σ"])
    r = 4
    categories = ["cogs", "pa", "opex", "contingency", "depreciation", "nwc_change", "taxes"]
    for cat in categories:
        for s in SCENARIOS_ORDER:
            m = run.get(s)
            d = getattr(m.costs, cat) if cat != "taxes" else m.pnl.taxes
            total = sum(d.values())
            _write_rows(ws, r, [[
                f"{cat} [{s}]",
                round(d[2026], 0), round(d[2027], 0), round(d[2028], 0),
                round(total, 0),
            ]])
            r += 1
        r += 1


def _sheet_pnl(ws, run: RunAllResult, scenario: str) -> None:
    m = run.get(scenario)
    p = m.pnl
    _title(ws, f"P&L — сценарий {scenario.upper()}")
    _set_column_widths(ws, [28, 14, 14, 14, 14])
    _header_row(ws, 3, ["Строка", "2026", "2027", "2028", "Σ"])
    rows: List[List[Any]] = []
    for label, d in [
        ("Revenue Total", p.revenue_total),
        ("(-) COGS", p.cogs),
        ("= Gross Profit", p.gross_profit),
        ("(-) P&A", p.pa),
        ("(-) OPEX", p.opex),
        ("(-) Contingency", p.contingency),
        ("= EBITDA", p.ebitda),
        ("(-) D&A", p.depreciation),
        ("= EBIT", p.ebit),
        ("(-) Taxes", p.taxes),
        ("= Net Income", p.net_income),
    ]:
        total = sum(d[y] for y in YEARS)
        rows.append([label, round(d[2026], 0), round(d[2027], 0), round(d[2028], 0), round(total, 0)])
    _write_rows(ws, 4, rows)


def _sheet_09_cashflow(ws, run: RunAllResult) -> None:
    _title(ws, "Cash Flow — 3 сценария")
    _set_column_widths(ws, [28, 14, 14, 14, 14])
    _header_row(ws, 3, ["Строка", "2026", "2027", "2028", "Σ"])
    r = 4
    for s in SCENARIOS_ORDER:
        ws.cell(row=r, column=1, value=f"── {s.upper()} ──").font = SECTION_FONT
        r += 1
        cf = run.get(s).cashflow
        rows: List[List[Any]] = []
        for label, d in [
            ("Net Income", cf.net_income),
            ("+ D&A", cf.depreciation_add),
            ("− Δ NWC", cf.nwc_change),
            ("= OCF", cf.operating_cf),
            ("− CapEx", cf.capex),
            ("= FCF", cf.free_cash_flow),
        ]:
            total = sum(d[y] for y in YEARS)
            rows.append([label, round(d[2026], 0), round(d[2027], 0), round(d[2028], 0), round(total, 0)])
        r = _write_rows(ws, r, rows)
        r += 1


def _sheet_10_quarterly(ws, run: RunAllResult) -> None:
    _title(ws, "Quarterly Cash Flow (Base) — Q1 2026 … Q4 2028")
    _set_column_widths(ws, [14, 14, 14, 14, 14])
    _header_row(ws, 3, ["Quarter", "Income", "Outcome", "Net", "Cumulative"])
    r = 4
    for qk, d in run.quarterly_cashflow_base.items():
        _write_rows(ws, r, [[qk, d["income"], d["outcome"], d["net"], d["cumulative"]]])
        r += 1


def _sheet_11_valuation(ws, run: RunAllResult) -> None:
    _title(ws, "Valuation — DCF по 3 WACC × 2 TV")
    _set_column_widths(ws, [28, 14, 14, 14])
    _header_row(ws, 3, ["Метрика", "Cons", "Base", "Opt"])
    r = 4
    metrics = [
        ("WACC (CAPM)", "wacc_capm"),
        ("WACC (Build-up)", "wacc_buildup"),
        ("WACC (Comparable)", "wacc_switcher"),
        ("NPV (CAPM)", "npv_capm"),
        ("NPV (Build-up)", "npv_buildup"),
        ("NPV (Comparable)", "npv_switcher"),
        ("IRR", "irr"),
        ("MOIC", "moic"),
        ("Payback, лет", "payback_years"),
        ("TV Gordon", "terminal_value_gordon"),
        ("TV Exit Multiple", "terminal_value_multiple"),
    ]
    for label, attr in metrics:
        row: List[Any] = [label]
        for s in SCENARIOS_ORDER:
            v = getattr(run.get(s).valuation, attr)
            row.append(round(v, 3) if attr in ("wacc_capm", "wacc_buildup", "wacc_switcher", "irr") else round(v, 1))
        _write_rows(ws, r, [row])
        r += 1


def _sheet_12_sensitivity(ws, sens: SensitivityTable) -> None:
    _title(ws, "Sensitivity: NPV (WACC × growth)")
    _set_column_widths(ws, [14] * (len(sens.growth_values) + 1))
    rows = sens.to_rows()
    _header_row(ws, 3, rows[0])
    for i, row in enumerate(rows[1:], start=4):
        _write_rows(ws, i, [row])


def _sheet_13_montecarlo(ws, mc: MonteCarloResult) -> None:
    _title(ws, f"Monte Carlo — {mc.n_sims} симуляций (seed {mc.seed})")
    _set_column_widths(ws, [36, 18])
    _write_rows(ws, 3, [
        ["EBITDA mean", mc.ebitda_mean],
        ["EBITDA median", mc.ebitda_median],
        ["EBITDA p5 (5% worst)", mc.ebitda_p5],
        ["EBITDA p95 (5% best)", mc.ebitda_p95],
        ["EBITDA std", mc.ebitda_std],
        ["FCF mean", mc.fcf_mean],
        ["FCF median", mc.fcf_median],
        ["FCF p5", mc.fcf_p5],
        ["FCF p95", mc.fcf_p95],
        ["P(EBITDA ≥ anchor)", f"{mc.prob_ebitda_above_anchor*100:.1f}%"],
        ["P(FCF > 0)", f"{mc.prob_fcf_positive*100:.1f}%"],
    ])


def _sheet_14_stress(ws, stress: StressResults) -> None:
    _title(ws, "Stress Tests — 6 сценариев + breakeven")
    _set_column_widths(ws, [26, 44, 14, 14, 18, 10])
    _header_row(ws, 3, ["Name", "Description", "Δ EBITDA %", "Δ FCF %", "New cum EBITDA", "Anchor"])
    r = 4
    for s in stress.scenarios:
        _write_rows(ws, r, [[
            s.name, s.description,
            f"{s.delta_ebitda_pct:+.1f}%", f"{s.delta_fcf_pct:+.1f}%",
            s.new_cumulative_ebitda,
            "PASS" if s.passes_anchor else "FAIL",
        ]])
        r += 1
    r += 2
    ws.cell(row=r, column=1, value="Breakeven Revenue Shock").font = SECTION_FONT
    r += 1
    _write_rows(ws, r, [
        ["% падения выручки для нуля EBITDA", f"{stress.breakeven_revenue_shock_pct:.1f}%"],
        ["Обоснование", stress.breakeven_rationale],
    ])


def _sheet_15_investment(ws, inputs: ValidatedInputs) -> None:
    inv = inputs.investment
    _title(ws, "Investment Round — структура раунда")
    _set_column_widths(ws, [32, 16, 20, 30])
    _write_rows(ws, 3, [
        ["Round type", inv.round_type, "", ""],
        ["Round stage", inv.round_stage, "", ""],
        ["Ask (cons)", round(inv.ask_mln_rub.cons, 0), "", ""],
        ["Ask (base)", round(inv.ask_mln_rub.base, 0), "", ""],
        ["Ask (opt)", round(inv.ask_mln_rub.opt, 0), "", ""],
        ["Headline ask", round(inv.headline_ask_mln_rub, 0), "", ""],
    ])
    r = 10
    ws.cell(row=r, column=1, value="Tranche structure").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Tranche", "Amount", "Instrument", "Title"])
    r += 1
    for t in inv.tranche_structure:
        _write_rows(ws, r, [[t.tranche_id, round(t.amount_mln_rub, 0), t.instrument, t.title]])
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Use of funds (млн ₽)").font = SECTION_FONT
    r += 1
    uof = inv.use_of_funds_mln_rub
    _write_rows(ws, r, [
        ["Production slate 2026-2027", round(uof.production_slate_2026_2027, 0), "", ""],
        ["Opening working capital", round(uof.opening_working_capital, 0), "", ""],
        ["Infrastructure CapEx", round(uof.infrastructure_capex, 0), "", ""],
        ["P&A marketing reserve", round(uof.pa_marketing_reserve, 0), "", ""],
        ["General contingency", round(uof.general_contingency, 0), "", ""],
        ["TOTAL", round(uof.total(), 0), "", ""],
    ])


def _sheet_16_returns(ws, inputs: ValidatedInputs) -> None:
    inv = inputs.investment
    _title(ws, "Investor Returns — MOIC / IRR / exit")
    _set_column_widths(ws, [28, 14, 14, 14])
    _header_row(ws, 3, ["Метрика", "Cons", "Base", "Opt"])
    r = 4
    _write_rows(ws, r, [[
        "MOIC",
        round(inv.investor_returns.expected_exit_multiple_moic.cons, 2),
        round(inv.investor_returns.expected_exit_multiple_moic.base, 2),
        round(inv.investor_returns.expected_exit_multiple_moic.opt, 2),
    ]])
    r += 1
    _write_rows(ws, r, [[
        "IRR",
        f"{inv.investor_returns.expected_irr_pct.cons*100:.1f}%",
        f"{inv.investor_returns.expected_irr_pct.base*100:.1f}%",
        f"{inv.investor_returns.expected_irr_pct.opt*100:.1f}%",
    ]])
    r += 2
    ws.cell(row=r, column=1, value=f"Exit strategy: {inv.investor_returns.exit_strategy}")
    r += 1
    ws.cell(row=r, column=1, value=f"Expected exit year: {inv.investor_returns.expected_exit_year}")


def _sheet_17_macro(ws, inputs: ValidatedInputs) -> None:
    m = inputs.macro
    _title(ws, "Macro assumptions")
    _set_column_widths(ws, [24, 14, 14, 14, 14])
    _header_row(ws, 3, ["Метрика / Год", "2026 cons", "2026 base", "2026 opt", "..."])
    r = 4
    for label, lst in [
        ("Inflation CPI", m.inflation_cpi),
        ("USD/RUB", m.usd_rub),
        ("Key rate CBR", m.key_rate_cbr),
    ]:
        ws.cell(row=r, column=1, value=f"── {label} ──").font = SECTION_FONT
        r += 1
        for rate in lst:
            _write_rows(ws, r, [[rate.year, rate.cons, rate.base, rate.opt, ""]])
            r += 1
        r += 1
    _write_rows(ws, r, [
        ["Profit tax rate", m.profit_tax_rate.rate, "", "", ""],
        ["VAT rate", m.vat_rate.rate, "cinema_exempt", m.vat_rate.cinema_exempt, ""],
    ])


def _sheet_18_nwc_capex(ws, inputs: ValidatedInputs) -> None:
    _title(ws, "NWC + CapEx")
    _set_column_widths(ws, [28, 14, 14, 14])
    nwc = inputs.nwc
    _header_row(ws, 3, ["Turnover days", "cons", "base", "opt"])
    r = 4
    for name in ("accounts_receivable", "accounts_payable", "inventory_wip", "advances_received"):
        td = getattr(nwc.turnover_days, name)
        _write_rows(ws, r, [[name, td.cons, td.base, td.opt]])
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Δ NWC change (млн ₽)").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Year", "cons", "base", "opt"])
    r += 1
    for nc in nwc.nwc_change_mln_rub:
        _write_rows(ws, r, [[nc.year, nc.cons, nc.base, nc.opt]])
        r += 1
    r += 2
    cap = inputs.capex
    ws.cell(row=r, column=1, value="Production CapEx (млн ₽)").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Year", "cons", "base", "opt"])
    r += 1
    for pc in cap.production_capex_mln_rub:
        _write_rows(ws, r, [[pc.year, pc.cons, pc.base, pc.opt]])
        r += 1


def _sheet_19_provenance(ws, registry: ProvenanceRegistry) -> None:
    _title(ws, f"Provenance Registry — {len(registry.entries)} источников")
    _set_column_widths(ws, [36, 42, 12, 12])
    _header_row(ws, 3, ["source_id", "used_in / title", "confidence", "last_updated"])
    r = 4
    for sid, entry in sorted(registry.entries.items()):
        used_in_str = ", ".join(entry.used_in_files) + (f" | {entry.title}" if entry.title else "")
        _write_rows(ws, r, [[
            entry.source_id, used_in_str,
            entry.confidence or "", entry.last_updated or "",
        ]])
        r += 1


def _sheet_20_manifest(ws, manifest: Dict) -> None:
    _title(ws, "Hash Manifest — воспроизводимость")
    _set_column_widths(ws, [50, 70])
    _write_rows(ws, 3, [
        ["generated_at", manifest.get("generated_at", "")],
        ["python_version", manifest.get("python_version", "")],
        ["combined_hash", manifest.get("combined_hash", "")],
    ])
    r = 7
    ws.cell(row=r, column=1, value="── Inputs hashes ──").font = SECTION_FONT
    r += 1
    for k, v in manifest.get("inputs_hashes", {}).items():
        _write_rows(ws, r, [[k, v[:16] + "..."]])
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="── Schemas hashes ──").font = SECTION_FONT
    r += 1
    for k, v in manifest.get("schemas_hashes", {}).items():
        _write_rows(ws, r, [[k, v[:16] + "..."]])
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="── Generators hashes ──").font = SECTION_FONT
    r += 1
    for k, v in manifest.get("generators_hashes", {}).items():
        _write_rows(ws, r, [[k, v[:16] + "..."]])
        r += 1


# ─────────────── orchestrator ───────────────

def build_xlsx(
    dst: Path,
    inputs: ValidatedInputs,
    run: RunAllResult,
    sensitivity: SensitivityTable,
    stress: StressResults,
    mc: MonteCarloResult,
    provenance: ProvenanceRegistry,
    manifest: Dict,
) -> None:
    wb = Workbook()
    # remove default
    default_ws = wb.active
    wb.remove(default_ws)

    sheets: List[tuple[str, Any]] = [
        ("00_Cover", lambda ws: _sheet_00_cover(ws, inputs, run)),
        ("01_Summary", lambda ws: _sheet_01_summary(ws, run)),
        ("02_Anchor_check", lambda ws: _sheet_02_anchor(ws, inputs, run)),
        ("03_Revenue_segments", lambda ws: _sheet_03_revenue(ws, run)),
        ("04_Slate_films", lambda ws: _sheet_04_slate(ws, inputs)),
        ("05_Costs", lambda ws: _sheet_05_costs(ws, run)),
        ("06_PnL_cons", lambda ws: _sheet_pnl(ws, run, "cons")),
        ("07_PnL_base", lambda ws: _sheet_pnl(ws, run, "base")),
        ("08_PnL_opt", lambda ws: _sheet_pnl(ws, run, "opt")),
        ("09_CashFlow_3scen", lambda ws: _sheet_09_cashflow(ws, run)),
        ("10_Quarterly_CF_base", lambda ws: _sheet_10_quarterly(ws, run)),
        ("11_Valuation_DCF", lambda ws: _sheet_11_valuation(ws, run)),
        ("12_Sensitivity", lambda ws: _sheet_12_sensitivity(ws, sensitivity)),
        ("13_MonteCarlo", lambda ws: _sheet_13_montecarlo(ws, mc)),
        ("14_StressTests", lambda ws: _sheet_14_stress(ws, stress)),
        ("15_Investment_round", lambda ws: _sheet_15_investment(ws, inputs)),
        ("16_InvestorReturns", lambda ws: _sheet_16_returns(ws, inputs)),
        ("17_Macro", lambda ws: _sheet_17_macro(ws, inputs)),
        ("18_NWC_CapEx", lambda ws: _sheet_18_nwc_capex(ws, inputs)),
        ("19_Provenance", lambda ws: _sheet_19_provenance(ws, provenance)),
        ("20_Manifest", lambda ws: _sheet_20_manifest(ws, manifest)),
    ]

    for name, builder in sheets:
        ws = wb.create_sheet(title=name)
        builder(ws)

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst))
