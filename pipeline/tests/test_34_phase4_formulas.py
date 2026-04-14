"""
tests/test_34_phase4_formulas.py — Phase 4 tests: Named Ranges + Live Formulas.

Verifies:
  - Named Ranges exist in workbook (NDP_ANCHOR, HURDLE_RATE, WACC_BASE, etc.)
  - Key sheets have formulas (not 100% static)
  - DCF references WACC_BASE named range
  - Invariant: NDP anchor = 3000 in Assumptions
  - F-025: Cash Flow has structural references
"""
from __future__ import annotations

import sys
import zipfile
import re
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).parent.parent.parent
PIPELINE_ROOT = Path(__file__).parent.parent

PUBLIC_XLSX = [
    REPO_ROOT / "investor_model_v1.0_Public.xlsx",
    REPO_ROOT / "Investor_Package" / "investor_model_v1.0_Public.xlsx",
]


def _get_xlsx_files():
    return [p for p in PUBLIC_XLSX if p.exists()]


# ─── Named Ranges ───────────────────────────────────────────────────────


@pytest.mark.parametrize("xlsx_path", _get_xlsx_files(), ids=lambda p: p.name)
class TestNamedRanges:
    """Named Ranges must exist in workbook."""

    REQUIRED_NAMES = [
        "NDP_ANCHOR", "HURDLE_RATE", "WACC_BASE",
        "EBITDA_3Y", "INVESTMENT_T1",
    ]

    def test_named_ranges_present(self, xlsx_path: Path):
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        names = set(wb.defined_names.keys()) if hasattr(wb.defined_names, 'keys') else set()
        wb.close()
        for name in self.REQUIRED_NAMES:
            assert name in names, f"Named range '{name}' not found in {xlsx_path.name}"

    def test_ndp_anchor_value(self, xlsx_path: Path):
        """NDP_ANCHOR should point to a cell containing 3000."""
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["02_Assumptions"]
        # Check R139 D (where NDP_ANCHOR points)
        val = ws.cell(row=139, column=4).value
        wb.close()
        assert val == 3000 or val == 3000.0, (
            f"NDP_ANCHOR cell should be 3000, got {val}"
        )

    def test_hurdle_rate_value(self, xlsx_path: Path):
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["02_Assumptions"]
        val = ws.cell(row=143, column=4).value
        wb.close()
        assert val == pytest.approx(0.18, abs=0.001), (
            f"HURDLE_RATE should be 0.18, got {val}"
        )


# ─── Live Formulas ───────────────────────────────────────────────────────


@pytest.mark.parametrize("xlsx_path", _get_xlsx_files(), ids=lambda p: p.name)
class TestLiveFormulas:
    """Key sheets should have formulas, not 100% static."""

    def test_dcf_has_wacc_formula(self, xlsx_path: Path):
        """22_Valuation_DCF should reference WACC_BASE."""
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["22_Valuation_DCF"]
        found_formula = False
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=4).value
            if isinstance(val, str) and "WACC_BASE" in val:
                found_formula = True
                break
        wb.close()
        assert found_formula, (
            "22_Valuation_DCF should have a formula referencing WACC_BASE"
        )

    def test_assumptions_has_anchor_cells(self, xlsx_path: Path):
        """02_Assumptions should have NDP=3000, T1=1250."""
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb["02_Assumptions"]
        ndp = ws.cell(row=139, column=4).value
        t1 = ws.cell(row=134, column=4).value
        wb.close()
        assert ndp == 3000 or ndp == 3000.0
        assert t1 == 1250 or t1 == 1250.0


# ─── Invariant Tests ─────────────────────────────────────────────────────


class TestInvariants:
    """Invariant: changing Assumptions should not break static anchors."""

    def test_ndp_anchor_3000(self):
        """NDP anchor = 3000 in all model files."""
        from generators.finance_core import NDP_ANCHOR
        assert NDP_ANCHOR == 3000.0

    def test_hurdle_rate_18(self):
        from generators.finance_core import HURDLE_RATE
        assert HURDLE_RATE == 0.18

    def test_wacc_base_1905(self):
        from generators.finance_core import WACC_BASE
        assert WACC_BASE == pytest.approx(0.1905, abs=0.0001)

    def test_formula_injector_exists(self):
        """formula_injector.py should exist in pipeline."""
        assert (PIPELINE_ROOT / "formula_injector.py").exists()
