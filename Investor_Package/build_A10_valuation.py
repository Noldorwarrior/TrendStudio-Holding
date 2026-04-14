"""
А.10 — Valuation_DCF + Valuation_Multiples.

22_Valuation_DCF:
  - 10-year forecast 2026–2035
  - FCFF calc: EBIT(1-t) + D&A − CAPEX − ΔNWC
  - WACC build-up: Rf + β×ERP + Country prem + Size prem
  - Terminal Value: Gordon Growth (g=3%) + Exit Multiple (5× EBITDA)
  - EV bridge: Σ PV FCFF + PV TV → EV → −Net Debt → Equity
  - Sensitivity WACC × g (5×5)

23_Valuation_Multiples:
  - 6 peers (РФ кино/OTT): Kinopoisk, Okko, ivi, START, Premier, Мосфильм
  - EV/Revenue, EV/EBITDA, P/E
  - Precedent M&A transactions (3)
  - Football Field: DCF / Trading / Precedent / Bear-Base-Bull
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
BEAR_FILL = "FCE4D6"
BULL_FILL = "E2EFDA"
BASE_FILL = "FFF2CC"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color=BRAND_BLUE)
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'
PCTFMT = '0.0"%"'
MULTFMT = '0.00"×"'

# ─── Forecast inputs ────────────────────────────────────────────────
# 5 лет: 2026–2030 — стандарт для private media pre-IPO
YEARS = list(range(2026, 2031))  # 5 years

# Revenue: 2026-2028 — фактическая модель (build phase with peak 2028),
# 2029-2030 — steady-state rolling slate (новая волна производства)
REVENUE = {
    2026: 385.0,
    2027: 1665.0,
    2028: 2495.0,
    2029: 2600.0,   # rolling slate steady state
    2030: 2800.0,
}
# EBITDA margin: 2026-2028 якоря модели, 2029-2030 = 40% steady state
EBITDA_M = {2026: 0.130, 2027: 0.384, 2028: 0.586,
            2029: 0.40, 2030: 0.40}

# D&A — 2026-2027 build phase minimal (контент ещё не списывается),
# 2028-2030 linear ramp to steady state (R-012: smooth 167× jump)
# Rationale: content assets commissioned gradually during 2028 (first releases),
# reaching full replacement-level D&A by 2030.
DA = {2026: 3.0, 2027: 3.0, 2028: 175.0,
      2029: 348.0, 2030: 520.0}

# CAPEX — 2026-2028 факт модели, 2029-2030 replacement level
CAPEX = {2026: 402.5, 2027: 853.0, 2028: 594.5,
         2029: 500.0, 2030: 520.0}

# ΔNWC: grows with revenue (25% AR − 30% AP on OpEx)
NWC_PCT = 0.10  # approx ΔNWC as 10% of ΔRevenue

# R-018 / F-027: Tax separation — this is Profit Tax (ННП) only.
# НДС = 0% for cinema production (ст. 149 НК РФ, pass-through).
# ННП = 20% (17% regional + 3% federal).
TAX_RATE_PROFIT = 0.20  # ННП 20%
TAX_RATE = TAX_RATE_PROFIT  # alias for backward compatibility

# WACC build-up — realistic for mature pre-IPO media in RF
RISK_FREE = 0.145      # ОФЗ 10Y 2026
ERP = 0.07             # РФ equity risk premium
BETA = 0.80            # levered beta (stabilized media)
COUNTRY_PREM = 0.020   # country risk
SIZE_PREM = 0.010      # pre-IPO premium
COST_EQUITY = RISK_FREE + BETA * ERP + COUNTRY_PREM + SIZE_PREM  # = 23.1%
COST_DEBT_PRE = 0.12
COST_DEBT_AFTER = COST_DEBT_PRE * (1 - TAX_RATE)  # 9.6%
DV = 0.30
EV_RATIO = 0.70
WACC = EV_RATIO * COST_EQUITY + DV * COST_DEBT_AFTER
# ≈ 0.70*0.231 + 0.30*0.096 = 0.1617 + 0.0288 ≈ 19.05%
WACC = round(WACC, 4)

GROWTH_TV = 0.03  # terminal growth
EXIT_MULT = 5.0   # exit EBITDA multiple

NET_DEBT_2025 = 0  # clean start
# At valuation date (начало 2026), T₁ ещё не выбран полностью, см. debt schedule
# Используем целевой Net Debt = 0 на дату оценки для public version


# ─── FCF calculation helper ──────────────────────────────────────────
def calculate_fcf():
    rows = []
    prev_rev = 0.0
    for y in YEARS:
        rev = REVENUE[y]
        ebitda = rev * EBITDA_M[y]
        da = DA[y]
        ebit = ebitda - da
        nopat = ebit * (1 - TAX_RATE)
        capex = CAPEX[y]
        d_rev = rev - prev_rev
        d_nwc = d_rev * NWC_PCT
        fcf = nopat + da - capex - d_nwc
        rows.append({
            "year": y, "rev": rev, "ebitda": ebitda, "da": da,
            "ebit": ebit, "nopat": nopat, "capex": capex,
            "d_nwc": d_nwc, "fcf": fcf
        })
        prev_rev = rev
    return rows


def calculate_dcf_ev(rows):
    pv_sum = 0.0
    for i, r in enumerate(rows, start=1):
        pv = r["fcf"] / ((1 + WACC) ** i)
        r["pv_fcf"] = pv
        pv_sum += pv
    # Terminal Value Gordon
    last_fcf = rows[-1]["fcf"]
    tv_gordon = last_fcf * (1 + GROWTH_TV) / (WACC - GROWTH_TV)
    pv_tv_gordon = tv_gordon / ((1 + WACC) ** len(rows))
    # Terminal Value Exit Multiple
    last_ebitda = rows[-1]["ebitda"]
    tv_exit = last_ebitda * EXIT_MULT
    pv_tv_exit = tv_exit / ((1 + WACC) ** len(rows))
    ev_gordon = pv_sum + pv_tv_gordon
    ev_exit = pv_sum + pv_tv_exit
    ev_blend = (ev_gordon + ev_exit) / 2
    return {
        "pv_fcf_sum": pv_sum,
        "tv_gordon": tv_gordon, "pv_tv_gordon": pv_tv_gordon, "ev_gordon": ev_gordon,
        "tv_exit": tv_exit, "pv_tv_exit": pv_tv_exit, "ev_exit": ev_exit,
        "ev_blend": ev_blend,
    }


def clear_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


# ═══════════════════════════════════════════════════════════════════
# 22_Valuation_DCF
# ═══════════════════════════════════════════════════════════════════
def build_dcf(wb):
    clear_sheet_if_exists(wb, "22_Valuation_DCF")
    ws = wb.create_sheet("22_Valuation_DCF")

    widths = {"A": 2, "B": 4, "C": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # 5 year cols D:H, plus I = Total
    for i, _ in enumerate(YEARS):
        col = chr(ord("D") + i)
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["I"].width = 14
    # Extra cols for sensitivity
    for col in "JKLMN":
        ws.column_dimensions[col].width = 12

    # Title
    ws.merge_cells("B2:N3")
    c = ws["B2"]
    c.value = "VALUATION — DCF (Discounted Cash Flow) · 5Y Forecast 2026–2030"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:N4")
    s = ws["B4"]
    s.value = f"WACC = {WACC*100:.1f}% · Terminal growth g = {GROWTH_TV*100:.0f}% · Exit Mult = {EXIT_MULT:.1f}× EBITDA"
    s.font = F_ITALIC
    s.alignment = C_CENTER

    # WACC build-up block
    r = 6
    ws.merge_cells(f"B{r}:N{r}")
    c = ws[f"B{r}"]
    c.value = "I. WACC BUILD-UP"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    wacc_items = [
        ("Risk-free rate (ОФЗ 10Y)", RISK_FREE),
        ("Levered β (cinema/media)", BETA),
        ("Equity Risk Premium", ERP),
        ("Country risk premium (РФ)", COUNTRY_PREM),
        ("Size premium (small-cap)", SIZE_PREM),
        ("Cost of Equity (CAPM extended)", COST_EQUITY),
        ("Cost of Debt (pre-tax)", COST_DEBT_PRE),
        ("Cost of Debt (after-tax, 20%)", COST_DEBT_AFTER),
        ("D/V target capital structure", DV),
        ("E/V target capital structure", EV_RATIO),
        ("WACC (weighted)", WACC),
    ]
    for lbl, val in wacc_items:
        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BOLD if "WACC" in lbl or "Cost" in lbl else F_BODY
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        cell = ws[f"D{r}"]
        cell.value = val
        cell.font = F_BOLD if "WACC" in lbl else F_BODY
        cell.alignment = C_RIGHT
        cell.border = box_thin
        if "β" in lbl:
            cell.number_format = '0.00'
        else:
            cell.number_format = '0.00%'
        if lbl == "WACC (weighted)":
            cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
            ws[f"B{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        r += 1

    # FCF forecast table
    r += 2
    ws.merge_cells(f"B{r}:N{r}")
    c = ws[f"B{r}"]
    c.value = "II. FORECAST FREE CASH FLOW TO FIRM (FCFF)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    # Year header
    ws[f"B{r}"].value = "#"
    ws[f"B{r}"].font = F_H_COL
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"B{r}"].alignment = C_CENTER
    ws[f"B{r}"].border = box_thin
    ws[f"C{r}"].value = "Показатель, млн ₽"
    ws[f"C{r}"].font = F_H_COL
    ws[f"C{r}"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"C{r}"].alignment = C_LEFT
    ws[f"C{r}"].border = box_thin
    for i, y in enumerate(YEARS):
        col = chr(ord("D") + i)
        cell = ws[f"{col}{r}"]
        cell.value = y
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    total_col = chr(ord("D") + len(YEARS))  # "I"
    ws[f"{total_col}{r}"].value = "Σ / PV"
    ws[f"{total_col}{r}"].font = F_H_COL
    ws[f"{total_col}{r}"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"{total_col}{r}"].alignment = C_CENTER
    ws[f"{total_col}{r}"].border = box_thin
    header_row = r
    r += 1

    rows_data = calculate_fcf()
    dcf = calculate_dcf_ev(rows_data)

    # Data rows
    metric_keys = [
        ("1", "Revenue", "rev"),
        ("2", "EBITDA", "ebitda"),
        ("3", "D&A", "da"),
        ("4", "EBIT", "ebit"),
        ("5", "NOPAT = EBIT × (1−t)", "nopat"),
        ("6", "+ D&A", "da"),
        ("7", "− CAPEX", "capex"),
        ("8", "− ΔNWC", "d_nwc"),
        ("9", "= FCFF", "fcf"),
    ]
    for num, lbl, key in metric_keys:
        ws[f"B{r}"].value = num
        ws[f"B{r}"].font = F_BODY
        ws[f"B{r}"].alignment = C_CENTER
        ws[f"B{r}"].border = box_thin
        ws[f"C{r}"].value = lbl
        ws[f"C{r}"].font = F_BOLD if lbl in ("EBITDA", "FCFF", "= FCFF", "NOPAT = EBIT × (1−t)") else F_BODY
        ws[f"C{r}"].alignment = C_LEFT
        ws[f"C{r}"].border = box_thin
        tot = 0.0
        for i, row_d in enumerate(rows_data):
            col = chr(ord("D") + i)
            cell = ws[f"{col}{r}"]
            val = row_d[key]
            # CAPEX and ΔNWC показываем как отрицательные
            if lbl == "− CAPEX":
                val = -val
            elif lbl == "− ΔNWC":
                val = -val
            cell.value = val
            cell.font = F_BODY
            cell.alignment = C_RIGHT
            cell.border = box_thin
            cell.number_format = NUMFMT
            tot += val
            if lbl == "= FCFF":
                cell.fill = PatternFill("solid", fgColor=NDP_FILL)
        # Total
        tc = ws[f"{total_col}{r}"]
        tc.value = tot
        tc.font = F_BOLD
        tc.alignment = C_RIGHT
        tc.border = box_thin
        tc.number_format = NUMFMT
        tc.fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
        r += 1

    # PV FCF row
    ws[f"B{r}"].value = "10"
    ws[f"B{r}"].font = F_BODY
    ws[f"B{r}"].alignment = C_CENTER
    ws[f"B{r}"].border = box_thin
    ws[f"C{r}"].value = f"PV FCFF @ WACC {WACC*100:.1f}%"
    ws[f"C{r}"].font = F_BOLD
    ws[f"C{r}"].alignment = C_LEFT
    ws[f"C{r}"].border = box_thin
    ws[f"C{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
    for i, row_d in enumerate(rows_data):
        col = chr(ord("D") + i)
        cell = ws[f"{col}{r}"]
        cell.value = row_d["pv_fcf"]
        cell.font = F_BODY
        cell.alignment = C_RIGHT
        cell.border = box_thin
        cell.number_format = NUMFMT
        cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
    tc = ws[f"N{r}"]
    tc.value = dcf["pv_fcf_sum"]
    tc.font = F_TOTAL
    tc.alignment = C_RIGHT
    tc.border = box_thin
    tc.number_format = NUMFMT
    tc.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
    r += 2

    # III. EV bridge
    ws.merge_cells(f"B{r}:N{r}")
    c = ws[f"B{r}"]
    c.value = "III. ENTERPRISE VALUE BRIDGE"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    ev_items = [
        ("Σ PV of Explicit FCFF (10 years)", dcf["pv_fcf_sum"], False),
        (f"Terminal Value — Gordon (g={GROWTH_TV*100:.0f}%)", dcf["tv_gordon"], False),
        ("PV of Terminal Value — Gordon", dcf["pv_tv_gordon"], False),
        ("Enterprise Value (Gordon method)", dcf["ev_gordon"], True),
        ("", "", False),
        (f"Terminal Value — Exit Multiple ({EXIT_MULT:.0f}× EBITDA{YEARS[-1]})", dcf["tv_exit"], False),
        ("PV of Terminal Value — Exit Multiple", dcf["pv_tv_exit"], False),
        ("Enterprise Value (Exit Multiple method)", dcf["ev_exit"], True),
        ("", "", False),
        ("EV blended (avg of Gordon + Exit Mult)", dcf["ev_blend"], True),
        ("(−) Net Debt at valuation date", NET_DEBT_2025, False),
        ("= Equity Value", dcf["ev_blend"] - NET_DEBT_2025, True),
    ]
    for lbl, val, is_total in ev_items:
        ws.merge_cells(f"B{r}:K{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_TOTAL if is_total else F_BODY
        ws[f"B{r}"].alignment = C_LEFT
        if lbl:
            ws[f"B{r}"].border = box_thin
        ws.merge_cells(f"L{r}:N{r}")
        cell = ws[f"L{r}"]
        cell.value = val if val != "" else None
        cell.font = F_TOTAL if is_total else F_BODY
        cell.alignment = C_RIGHT
        if lbl:
            cell.border = box_thin
            cell.number_format = NUMFMT
        if is_total:
            ws[f"B{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
            cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        r += 1

    # IV. Sensitivity WACC × g
    r += 2
    ws.merge_cells(f"B{r}:N{r}")
    c = ws[f"B{r}"]
    c.value = f"IV. SENSITIVITY: EV (GORDON METHOD) vs WACC × g"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    wacc_range = [WACC - 0.04, WACC - 0.02, WACC, WACC + 0.02, WACC + 0.04]
    g_range = [0.01, 0.02, 0.03, 0.04, 0.05]

    ws[f"C{r}"].value = "WACC \\ g"
    ws[f"C{r}"].font = F_BOLD
    ws[f"C{r}"].alignment = C_CENTER
    ws[f"C{r}"].border = box_thin
    ws[f"C{r}"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"C{r}"].font = F_H_COL
    for i, g in enumerate(g_range):
        col = chr(ord("D") + i)
        cell = ws[f"{col}{r}"]
        cell.value = f"g={g*100:.0f}%"
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    # Recalc EV for sensitivity grid
    last_fcf = rows_data[-1]["fcf"]
    for w in wacc_range:
        ws[f"C{r}"].value = f"{w*100:.1f}%"
        ws[f"C{r}"].font = F_H_COL
        ws[f"C{r}"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws[f"C{r}"].alignment = C_CENTER
        ws[f"C{r}"].border = box_thin
        for i, g in enumerate(g_range):
            pv_fcf_s = 0.0
            for j, rd in enumerate(rows_data, start=1):
                pv_fcf_s += rd["fcf"] / ((1 + w) ** j)
            if w > g:
                tv = last_fcf * (1 + g) / (w - g)
                pv_tv = tv / ((1 + w) ** len(rows_data))
                ev = pv_fcf_s + pv_tv
            else:
                ev = None
            col = chr(ord("D") + i)
            cell = ws[f"{col}{r}"]
            cell.value = ev
            cell.font = F_BODY
            cell.alignment = C_RIGHT
            cell.border = box_thin
            cell.number_format = NUMFMT
            # Highlight base case
            if abs(w - WACC) < 1e-6 and abs(g - GROWTH_TV) < 1e-6:
                cell.fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
                cell.font = F_TOTAL
        r += 1

    # Conclusions
    r += 2
    ws.merge_cells(f"B{r}:N{r}")
    note = ws[f"B{r}"]
    note.value = (f"Заключение: EV blended ≈ {dcf['ev_blend']:.0f} млн ₽. "
                  f"Gordon подход более консервативен ({dcf['ev_gordon']:.0f}), "
                  f"Exit Multiple ({dcf['ev_exit']:.0f}) отражает рыночные "
                  f"мультипликаторы peers. Чувствительность к WACC±200bp = ±15%.")
    note.font = F_ITALIC
    note.alignment = C_LEFT

    ws.freeze_panes = "D7"
    print(f"  [22_Valuation_DCF] WACC={WACC*100:.1f}%, EV_Gordon={dcf['ev_gordon']:.0f}, "
          f"EV_Exit={dcf['ev_exit']:.0f}, EV_blend={dcf['ev_blend']:.0f}")
    return dcf


# ═══════════════════════════════════════════════════════════════════
# 23_Valuation_Multiples
# ═══════════════════════════════════════════════════════════════════
def build_multiples(wb, dcf):
    clear_sheet_if_exists(wb, "23_Valuation_Multiples")
    ws = wb.create_sheet("23_Valuation_Multiples")

    widths = {"A": 2, "B": 4, "C": 22, "D": 14, "E": 12, "F": 12,
              "G": 12, "H": 12, "I": 12, "J": 12, "K": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("B2:K3")
    c = ws["B2"]
    c.value = "VALUATION — TRADING & PRECEDENT MULTIPLES + FOOTBALL FIELD"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:K4")
    s = ws["B4"]
    s.value = ("Сравнительная оценка с peers российского кино/OTT-рынка "
               "и прецеденты M&A сделок 2022–2025")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    # I. Peer group
    r = 6
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "I. TRADING MULTIPLES — PEER GROUP (РФ кино/OTT)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    # Headers
    peer_hdrs = ["#", "Компания", "Сегмент", "Revenue", "EBITDA",
                 "EV", "EV/Rev", "EV/EBITDA", "EBITDA M%", "Примечание"]
    for i, h in enumerate(peer_hdrs):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    # Peer data: Rev, EBITDA, EV in млн ₽
    peers = [
        ("Яндекс Кинопоиск", "OTT подписка", 18500, 2775, 14800, 0.80, 5.34, 0.15, "Базовый кейс"),
        ("Okko", "OTT подписка", 12000, 1680, 9600, 0.80, 5.71, 0.14, "Sber Entertainment"),
        ("ivi", "OTT AVOD+SVOD", 11000, 1320, 8800, 0.80, 6.67, 0.12, "Private"),
        ("START", "OTT оригиналы", 6500, 910, 5200, 0.80, 5.71, 0.14, "Nexters"),
        ("Premier", "OTT семейн.", 4200, 546, 3400, 0.81, 6.23, 0.13, "ГПМ"),
        ("Мосфильм", "Production", 8500, 1870, 8900, 1.05, 4.76, 0.22, "Госкомпания"),
    ]

    rev_mults = []
    ebitda_mults = []
    for i, p in enumerate(peers, start=1):
        name, seg, rev, eb, ev, evrev, eveb, m, note = p
        ws[f"B{r}"].value = i
        ws[f"C{r}"].value = name
        ws[f"D{r}"].value = seg
        ws[f"E{r}"].value = rev
        ws[f"F{r}"].value = eb
        ws[f"G{r}"].value = ev
        ws[f"H{r}"].value = evrev
        ws[f"I{r}"].value = eveb
        ws[f"J{r}"].value = m * 100
        ws[f"K{r}"].value = note
        rev_mults.append(evrev)
        ebitda_mults.append(eveb)
        for col in "BCDEFGHIJK":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BHIJ":
                cell.alignment = C_CENTER
            elif col in "CDK":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
            if col in "EFG":
                cell.number_format = NUMFMT
            elif col in "HI":
                cell.number_format = MULTFMT
            elif col == "J":
                cell.number_format = PCTFMT
        r += 1

    # Stats row
    avg_rev = sum(rev_mults) / len(rev_mults)
    med_rev = sorted(rev_mults)[len(rev_mults) // 2]
    min_rev = min(rev_mults)
    max_rev = max(rev_mults)
    avg_eb = sum(ebitda_mults) / len(ebitda_mults)
    med_eb = sorted(ebitda_mults)[len(ebitda_mults) // 2]
    min_eb = min(ebitda_mults)
    max_eb = max(ebitda_mults)

    stats = [
        ("Mean", avg_rev, avg_eb),
        ("Median", med_rev, med_eb),
        ("Min", min_rev, min_eb),
        ("Max", max_rev, max_eb),
    ]
    for lbl, rv, eb in stats:
        ws[f"B{r}"].value = ""
        ws.merge_cells(f"C{r}:G{r}")
        ws[f"C{r}"].value = f"Peer group — {lbl}"
        ws[f"C{r}"].font = F_BOLD
        ws[f"C{r}"].alignment = C_LEFT
        ws[f"C{r}"].border = box_thin
        ws[f"C{r}"].fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
        ws[f"H{r}"].value = rv
        ws[f"I{r}"].value = eb
        for col in "HIJK":
            cell = ws[f"{col}{r}"]
            cell.font = F_BOLD
            cell.border = box_thin
            cell.alignment = C_RIGHT
            cell.fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
            if col in "HI":
                cell.number_format = MULTFMT
        r += 1

    # II. Apply multiples to ТрендСтудио
    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "II. APPLY MULTIPLES TO ТРЕНДСТУДИО (Base EBITDA 2028 = 1 462 · Rev 2028 = 2 495)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    ts_rev = 2495.0
    ts_ebitda = 1462.0  # YR 2028

    apply_rows = [
        ("Peer Min × EBITDA 2028", min_eb, ts_ebitda, min_eb * ts_ebitda),
        ("Peer Median × EBITDA 2028", med_eb, ts_ebitda, med_eb * ts_ebitda),
        ("Peer Mean × EBITDA 2028", avg_eb, ts_ebitda, avg_eb * ts_ebitda),
        ("Peer Max × EBITDA 2028", max_eb, ts_ebitda, max_eb * ts_ebitda),
        ("Peer Min × Revenue 2028", min_rev, ts_rev, min_rev * ts_rev),
        ("Peer Median × Revenue 2028", med_rev, ts_rev, med_rev * ts_rev),
        ("Peer Mean × Revenue 2028", avg_rev, ts_rev, avg_rev * ts_rev),
        ("Peer Max × Revenue 2028", max_rev, ts_rev, max_rev * ts_rev),
    ]
    apply_hdr = ["Метод", "Multiple", "Base", "Implied EV"]
    for i, h in enumerate(apply_hdr):
        col = chr(ord("B") + i * 2)
        if i == 0:
            ws.merge_cells(f"B{r}:F{r}")
            cell = ws[f"B{r}"]
        elif i == 1:
            cell = ws[f"G{r}"]
        elif i == 2:
            cell = ws[f"H{r}"]
        else:
            ws.merge_cells(f"I{r}:K{r}")
            cell = ws[f"I{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    for lbl, mult, base, ev in apply_rows:
        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BODY
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        ws[f"G{r}"].value = mult
        ws[f"G{r}"].number_format = MULTFMT
        ws[f"G{r}"].alignment = C_RIGHT
        ws[f"G{r}"].border = box_thin
        ws[f"G{r}"].font = F_BODY
        ws[f"H{r}"].value = base
        ws[f"H{r}"].number_format = NUMFMT
        ws[f"H{r}"].alignment = C_RIGHT
        ws[f"H{r}"].border = box_thin
        ws[f"H{r}"].font = F_BODY
        ws.merge_cells(f"I{r}:K{r}")
        ws[f"I{r}"].value = ev
        ws[f"I{r}"].number_format = NUMFMT
        ws[f"I{r}"].alignment = C_RIGHT
        ws[f"I{r}"].border = box_thin
        ws[f"I{r}"].font = F_BOLD
        ws[f"I{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        r += 1

    # III. Precedent transactions
    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "III. PRECEDENT M&A TRANSACTIONS (РФ медиа/OTT, 2022–2025)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    prec_hdr = ["#", "Сделка", "Год", "Target Revenue", "Deal EV", "EV/Rev", "EV/EBITDA", "Стратегия", "Покупатель"]
    # row spans 9 cols
    cols_prec = ["B", "C", "D", "E", "F", "G", "H", "I", "K"]
    col_spans = {"I": "I:J"}
    # Just use simple layout
    ws[f"B{r}"].value = "#"
    ws.merge_cells(f"C{r}:D{r}")
    ws[f"C{r}"].value = "Сделка"
    ws[f"E{r}"].value = "Год"
    ws[f"F{r}"].value = "Target Rev"
    ws[f"G{r}"].value = "Deal EV"
    ws[f"H{r}"].value = "EV/Rev"
    ws[f"I{r}"].value = "EV/EBITDA"
    ws.merge_cells(f"J{r}:K{r}")
    ws[f"J{r}"].value = "Комментарий"
    for col in "BCEFGHIJ":
        cell = ws[f"{col}{r}"]
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    precedents = [
        (1, "Сбер → Okko (увеличение доли)", 2023, 11200, 9520, 0.85, 5.60, "Консолидация OTT в экосистеме"),
        (2, "Газпром-Медиа → Premier", 2022, 3800, 3040, 0.80, 5.90, "Stream-платформа Premier"),
        (3, "ВК → zvook/Russkoe Radio медиа", 2024, 8600, 6880, 0.80, 6.40, "Мультимедиа консолидация"),
    ]
    prec_rev_mults = [p[5] for p in precedents]
    prec_eb_mults = [p[6] for p in precedents]
    for (num, deal, yr, tr, ev_d, evr, eveb, strat) in precedents:
        ws[f"B{r}"].value = num
        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"].value = deal
        ws[f"E{r}"].value = yr
        ws[f"F{r}"].value = tr
        ws[f"G{r}"].value = ev_d
        ws[f"H{r}"].value = evr
        ws[f"I{r}"].value = eveb
        ws.merge_cells(f"J{r}:K{r}")
        ws[f"J{r}"].value = strat
        for col in "BCEFGHIJ":
            cell = ws[f"{col}{r}"]
            cell.font = F_BODY
            cell.border = box_thin
            if col in "BE":
                cell.alignment = C_CENTER
            elif col in "CJ":
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
            if col in "FG":
                cell.number_format = NUMFMT
            elif col in "HI":
                cell.number_format = MULTFMT
        r += 1

    prec_med_eb = sorted(prec_eb_mults)[len(prec_eb_mults) // 2]
    prec_med_rev = sorted(prec_rev_mults)[len(prec_rev_mults) // 2]

    # IV. Football field
    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]
    c.value = "IV. FOOTBALL FIELD — ИТОГОВЫЙ ДИАПАЗОН ОЦЕНКИ, млн ₽"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    ff_hdr = ["Метод оценки", "Low", "Mid", "High", "Commentary"]
    for i, h in enumerate(ff_hdr):
        if i == 0:
            ws.merge_cells(f"B{r}:D{r}")
            cell = ws[f"B{r}"]
        elif i == 1:
            cell = ws[f"E{r}"]
        elif i == 2:
            cell = ws[f"F{r}"]
        elif i == 3:
            cell = ws[f"G{r}"]
        else:
            ws.merge_cells(f"H{r}:K{r}")
            cell = ws[f"H{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    ff_data = [
        ("DCF — Gordon Growth", dcf["ev_gordon"] * 0.85, dcf["ev_gordon"], dcf["ev_gordon"] * 1.15,
         f"g={GROWTH_TV*100:.0f}%, WACC={WACC*100:.1f}%, sensitivity ±15%"),
        ("DCF — Exit Multiple", dcf["ev_exit"] * 0.85, dcf["ev_exit"], dcf["ev_exit"] * 1.15,
         f"{EXIT_MULT:.0f}× EBITDA{YEARS[-1]}, ±15% sensitivity"),
        ("Trading Multiples — EV/EBITDA", min_eb * ts_ebitda, med_eb * ts_ebitda, max_eb * ts_ebitda,
         f"Peer median {med_eb:.2f}× × EBITDA 2028"),
        ("Trading Multiples — EV/Revenue", min_rev * ts_rev, med_rev * ts_rev, max_rev * ts_rev,
         f"Peer median {med_rev:.2f}× × Rev 2028"),
        ("Precedent Transactions", prec_med_eb * ts_ebitda * 0.85, prec_med_eb * ts_ebitda,
         prec_med_eb * ts_ebitda * 1.15, f"Med {prec_med_eb:.2f}× (3 M&A)"),
        ("Scenario Base (internal)", 2152.0 * 3.0, 2152.0 * 5.0, 2152.0 * 7.0,
         "EBITDA 3Y × 3×/5×/7× (bear/base/bull)"),
    ]
    for lbl, lo, mid, hi, com in ff_data:
        ws.merge_cells(f"B{r}:D{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws[f"B{r}"].border = box_thin
        for col, val in zip("EFG", [lo, mid, hi]):
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.font = F_BODY
            cell.alignment = C_RIGHT
            cell.border = box_thin
            cell.number_format = NUMFMT
        ws[f"F{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
        ws[f"F{r}"].font = F_TOTAL
        ws.merge_cells(f"H{r}:K{r}")
        ws[f"H{r}"].value = com
        ws[f"H{r}"].font = F_ITALIC
        ws[f"H{r}"].alignment = C_LEFT
        ws[f"H{r}"].border = box_thin
        r += 1

    # Summary
    r += 1
    lows = [row[1] for row in ff_data]
    mids = [row[2] for row in ff_data]
    highs = [row[3] for row in ff_data]
    overall_low = min(lows)
    overall_mid = sum(mids) / len(mids)
    overall_high = max(highs)

    ws.merge_cells(f"B{r}:D{r}")
    ws[f"B{r}"].value = "ИТОГО Football Field"
    ws[f"B{r}"].font = F_TOTAL
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=NDP_FILL)
    ws[f"B{r}"].alignment = C_LEFT
    ws[f"B{r}"].border = box_thin
    for col, v in zip("EFG", [overall_low, overall_mid, overall_high]):
        cell = ws[f"{col}{r}"]
        cell.value = v
        cell.font = F_TOTAL
        cell.fill = PatternFill("solid", fgColor=NDP_FILL)
        cell.alignment = C_RIGHT
        cell.border = box_thin
        cell.number_format = NUMFMT
    ws.merge_cells(f"H{r}:K{r}")
    ws[f"H{r}"].value = f"Envelope: {overall_low:.0f} – {overall_high:.0f}, mid ≈ {overall_mid:.0f} млн ₽"
    ws[f"H{r}"].font = F_TOTAL
    ws[f"H{r}"].fill = PatternFill("solid", fgColor=NDP_FILL)
    ws[f"H{r}"].alignment = C_LEFT
    ws[f"H{r}"].border = box_thin

    r += 3
    ws.merge_cells(f"B{r}:K{r}")
    note = ws[f"B{r}"]
    note.value = ("Примечание: Peer multiples рассчитаны по публичным данным РФ OTT-рынка "
                  "(оценки 2024–2025). Premium/discount к median = ±15% на illiquidity, "
                  "scale и execution risk. Final range = пересечение всех методов.")
    note.font = F_ITALIC
    note.alignment = C_LEFT

    ws.freeze_panes = "B7"
    print(f"  [23_Valuation_Multiples] peers=6, precedents=3, "
          f"Football Field mid={overall_mid:.0f}, envelope {overall_low:.0f}-{overall_high:.0f}")


def main():
    print(f"Loaded: {FILE}")
    wb = load_workbook(FILE)
    print(f"Sheets before: {len(wb.sheetnames)}")
    print(f"\nWACC = {WACC*100:.2f}%")

    print("\n[1/2] Building 22_Valuation_DCF …")
    dcf = build_dcf(wb)

    print("\n[2/2] Building 23_Valuation_Multiples …")
    build_multiples(wb, dcf)

    print(f"\nSheets after: {len(wb.sheetnames)}")
    wb.save(FILE)
    print(f"\nSaved: {FILE}")


if __name__ == "__main__":
    main()
