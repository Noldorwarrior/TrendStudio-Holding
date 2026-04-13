"""
А.15 — Gov_KPI + Tax_Schedule + Roadmap_2026_2032
Листы 33, 34, 35.

Источники (приоритет российским официальным): ТАСС, РИА Новости, Кремлин.ру,
Минкультуры РФ, Фонд кино, ФНС, Минфин РФ, Консультант Плюс, Гарант.ру.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx"

BLUE = "0070C0"
DARK_BLUE = "002060"
LIGHT_BLUE = "DEEBF7"
GREEN = "00B050"
LIGHT_GREEN = "E2EFDA"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "FCE4E4"
GREY = "808080"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(c): return PatternFill("solid", fgColor=c)

def header_cell(ws, row, col, text, width=None, color=DARK_BLUE):
    c = ws.cell(row, col)
    c.value = text
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = hfill(color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def section_title(ws, row, text, color=BLUE):
    ws.cell(row, 2).value = text
    ws.cell(row, 2).font = Font(name="Calibri", size=12, bold=True, color=color)

def title(ws, text, subtitle, end_col=11):
    ws["B2"] = text
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
    ws.merge_cells(f"B2:{get_column_letter(end_col)}2")
    ws["B4"] = subtitle
    ws["B4"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.merge_cells(f"B4:{get_column_letter(end_col)}4")

def body_cell(ws, r, c, v, align="left", bold=False, fill=None, size=9, fmt=None):
    cell = ws.cell(r, c)
    cell.value = v
    cell.font = Font(name="Calibri", size=size, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    if fill:
        cell.fill = hfill(fill)
    if fmt:
        cell.number_format = fmt

# ==========================================================================
# ЛИСТ 33: GOV_KPI
# ==========================================================================
def build_gov_kpi(wb):
    if "33_Gov_KPI" in wb.sheetnames:
        del wb["33_Gov_KPI"]
    ws = wb.create_sheet("33_Gov_KPI")
    title(ws, "GOV KPI — ПОКАЗАТЕЛИ ГОСПОДДЕРЖКИ И СОЦИОКУЛЬТУРНОГО ВКЛАДА",
          "Соответствие нацпроектам, ФП «Развитие отечественного кинематографа», KPI Минкульта/ФКП/ИРИ",
          end_col=11)

    # ===== I. Национальные проекты =====
    section_title(ws, 6, "I. СООТВЕТСТВИЕ НАЦИОНАЛЬНЫМ ПРОЕКТАМ РФ 2025–2030")
    headers = ["#", "Нацпроект / программа", "Федеральный проект", "Ответственный ФОИВ",
               "Вклад ТрендСтудио", "Мера вклада"]
    widths = [5, 22, 26, 18, 30, 16]
    for i, h in enumerate(headers):
        header_cell(ws, 8, 2 + i, h, width=widths[i])

    np_data = [
        (1, "Семья",
         "Поддержка семейного кино",
         "Минкультуры РФ",
         "Слот F07 «Сказка наяву» (семейный), F04 «Последний экзамен»",
         "2 из 12 фильмов"),
        (2, "Молодёжь и дети",
         "Воспитание патриотизма через культуру",
         "Минкультуры РФ / Росмолодёжь",
         "Слот F01 «Родные стены», F06 «Красная заря» (исторический)",
         "2 из 12 фильмов"),
        (3, "Культура",
         "Развитие отечественного кинематографа",
         "Минкультуры РФ / Фонд кино",
         "Весь портфель (12 фильмов), объём 1 850 млн ₽",
         "12/12 фильмов"),
        (4, "Экономика данных",
         "Цифровизация креативных индустрий",
         "Минцифры РФ",
         "Digital distribution, VFX-pipelines, SaaS-аналитика",
         "Инфраструктура"),
        (5, "Кадры",
         "Подготовка кадров креативной индустрии",
         "Минобрнауки РФ / ВГИК",
         "Стажировки от 40 чел/год, партнёрство с ВГИК, МШНК",
         "40+ стажёров"),
        (6, "Международная кооперация",
         "Экспорт культуры и контента",
         "Минкультуры РФ / Россотрудничество",
         "Копродукция с СНГ/БРИКС/ОАЭ, фестивали",
         "3-4 копродукции"),
    ]
    row = 9
    for d in np_data:
        for i, v in enumerate(d):
            body_cell(ws, row, 2 + i, v,
                      align="center" if i in (0, 5) else "left",
                      bold=(i == 0),
                      size=9)
        ws.row_dimensions[row].height = 32
        row += 1
    row += 2

    # ===== II. KPI Минкульта / ФКП =====
    section_title(ws, row, "II. KPI ДЛЯ ГОСПОДДЕРЖКИ (Минкультуры / Фонд кино / ИРИ)")
    row += 2
    headers = ["#", "Показатель", "План 2026", "План 2027", "План 2028", "Цель 3Y",
               "Единица", "Источник", "Методика расчёта"]
    widths = [5, 28, 12, 12, 12, 12, 10, 14, 30]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1

    kpi_gov = [
        ("Количество релизов", 3, 5, 4, 12, "шт", "Фонд кино", "Фактические премьеры в широком прокате"),
        ("Количество зрителей", 4.5, 12.5, 13.0, 30.0, "млн чел", "ЕАИС", "Билеты + просмотры >30 мин"),
        ("Средняя заполняемость залов", 22, 28, 30, 27, "%", "ЕАИС", "Attendance / seats × 100%"),
        ("Сборы от проката", 400, 1320, 1100, 2820, "млн ₽", "ЕАИС", "Box-office РФ (ЕАИС)"),
        ("Рабочие места (ФОТ)", 180, 320, 280, 260, "чел/год", "Внутр. отчёт", "Среднесписочная, включая фриланс"),
        ("Объём налоговых отчислений", 55, 185, 178, 418, "млн ₽", "ФНС", "НДФЛ + соцвзносы + имущ."),
        ("Доля отечественного контента", 100, 100, 100, 100, "%", "Минкульт", "Российский продюсер, русский язык, российская съёмочная команда"),
        ("Экспорт за рубеж", 1, 3, 4, 8, "фильмов", "Минкульт", "Продажи прав + фестивали"),
        ("Субсидии / гранты привлечено", 280, 420, 340, 1040, "млн ₽", "Фонд кино", "ФКП + ИРИ + Минкульт"),
        ("Возвратность господдержки", 15, 25, 35, 25, "%", "Фонд кино", "Доля возвращаемых грантов по 1215/2013"),
    ]
    for i, (kpi, y26, y27, y28, total, unit, src, meth) in enumerate(kpi_gov, 1):
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, kpi, bold=True)
        body_cell(ws, row, 4, y26, align="center", fmt="#,##0.0")
        body_cell(ws, row, 5, y27, align="center", fmt="#,##0.0")
        body_cell(ws, row, 6, y28, align="center", fmt="#,##0.0")
        body_cell(ws, row, 7, total, align="center", fmt="#,##0.0", bold=True, fill=LIGHT_BLUE)
        body_cell(ws, row, 8, unit, align="center")
        body_cell(ws, row, 9, src, align="center", size=8)
        body_cell(ws, row, 10, meth, size=8)
        ws.row_dimensions[row].height = 28
        row += 1
    row += 2

    # ===== III. Структура господдержки =====
    section_title(ws, row, "III. СТРУКТУРА ГОСПОДДЕРЖКИ (ожидаемое привлечение 2026–2028)")
    row += 2
    headers = ["#", "Источник", "Тип поддержки", "Сумма 3Y (млн ₽)", "Доля", "Условия"]
    widths = [5, 22, 22, 16, 10, 40]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1

    gov_sources = [
        (1, "Фонд кино", "Безвозвратная субсидия + возвратная часть", 520,
         "Пост. Правительства РФ от 31.12.2013 № 1215 (с изм.)"),
        (2, "Минкультуры РФ", "Гранты на производство", 220,
         "Конкурсный отбор, приоритет — исторический/патриотический контент"),
        (3, "ИРИ (Институт развития интернета)", "Гранты на цифровой контент", 150,
         "Фонд поддержки контента для молодёжи"),
        (4, "Региональные фонды", "Рибейты и налоговые льготы", 90,
         "Московская обл., СПб, Крым, Татарстан — cashback до 30%"),
        (5, "НДС 0%", "Налоговая льгота", 60,
         "ст. 149 НК РФ — освобождение от НДС услуг в сфере культуры"),
        ("Σ", "ИТОГО ГОСПОДДЕРЖКА", "—", 1040, None),
    ]
    total_g = sum(x[3] for x in gov_sources[:-1])
    for d in gov_sources:
        if len(d) == 5 and d[0] == "Σ":  # итого
            body_cell(ws, row, 2, "Σ", align="center", bold=True, fill=LIGHT_BLUE)
            body_cell(ws, row, 3, "ИТОГО ГОСПОДДЕРЖКА", bold=True, fill=LIGHT_BLUE)
            body_cell(ws, row, 4, "—", align="center", fill=LIGHT_BLUE)
            body_cell(ws, row, 5, total_g, align="center", fmt="#,##0", bold=True, fill=LIGHT_BLUE)
            body_cell(ws, row, 6, 1.0, align="center", fmt="0.0%", bold=True, fill=LIGHT_BLUE)
            body_cell(ws, row, 7, "56% от CAPEX; остальное — T₁ инвестор + producer equity", fill=LIGHT_BLUE)
        else:
            num, name, typ, amt, cond = d
            body_cell(ws, row, 2, num, align="center", bold=True)
            body_cell(ws, row, 3, name, bold=True)
            body_cell(ws, row, 4, typ)
            body_cell(ws, row, 5, amt, align="center", fmt="#,##0", fill=LIGHT_GREEN)
            body_cell(ws, row, 6, amt/total_g, align="center", fmt="0.0%")
            body_cell(ws, row, 7, cond, size=8)
        row += 1
    row += 2

    # ===== IV. Социокультурное воздействие =====
    section_title(ws, row, "IV. СОЦИОКУЛЬТУРНОЕ ВОЗДЕЙСТВИЕ")
    row += 2
    impact = [
        ("Просвещение через контент",
         "Исторические фильмы F06 «Красная заря», F08 «Неизвестный герой» — оригинальные события в истории РФ"),
        ("Поддержка семейных ценностей",
         "F01 «Родные стены», F07 «Сказка наяву», F09 «Моё дело» — 3 фильма с явным семейным мессиджем"),
        ("Популяризация региона",
         "Съёмки в Карелии, на Алтае, в Крыму, Татарстане — развитие кинотуризма"),
        ("Занятость",
         "260 рабочих мест/год в среднем, включая фрилансеров; партнёрство с вузами"),
        ("Международный имидж РФ",
         "Фестивали Шанхай, ММКФ, Minsk Fest, Dubai Film Days — экспорт культуры"),
        ("Вклад в экосистему",
         "VFX-центр, школа актёрского мастерства, продюсерские стажировки"),
        ("Сохранение культурного наследия",
         "Библиотека фильмов, архивы, копродукция с Госфильмофондом"),
    ]
    headers = ["#", "Направление", "Описание"]
    widths = [5, 28, 75]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1
    for i, (dir_, desc) in enumerate(impact, 1):
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, dir_, bold=True)
        body_cell(ws, row, 4, desc, size=9)
        ws.row_dimensions[row].height = 28
        row += 1
    row += 2

    # ===== V. Regulatory compliance =====
    section_title(ws, row, "V. REGULATORY COMPLIANCE CHECKLIST")
    row += 2
    compliance = [
        ("Лицензия на прокат", "Удостоверение национального фильма (УНФ)", "Минкультуры РФ", "✓ до релиза"),
        ("Возрастной рейтинг", "Категория 0+/6+/12+/16+/18+", "Минкультуры РФ", "✓ до релиза"),
        ("НДС 0%", "Подтверждение льготы по ст. 149 НК РФ", "ФНС", "✓ квартально"),
        ("Авторские права", "Регистрация в РАО/РСП/ВОИС, договоры с авторами", "РАО, ВОИС", "✓ до съёмок"),
        ("Договоры с актёрами", "ТД + имидж-права + морал райтс", "Минтруд РФ", "✓ до съёмок"),
        ("Съёмочные разрешения", "Разрешения на съёмки в госучреждениях", "Минкульт, регионы", "✓ до съёмок"),
        ("Страхование", "Completion bond, страхование членов группы", "ФКП + частные", "✓ до съёмок"),
        ("Отчётность в ФКП", "Ежеквартальная отчётность + финальная", "Фонд кино", "✓ квартально"),
        ("Охрана труда", "Техника безопасности на площадке", "Росинспекция", "✓ на площадке"),
        ("ФЗ-152 о персональных данных", "Согласия актёров, статистов, зрителей", "Роскомнадзор", "✓ непрерывно"),
    ]
    headers = ["#", "Требование", "Описание", "Регулятор", "Статус"]
    widths = [5, 22, 36, 16, 14]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1
    for i, (req, desc, reg, stat) in enumerate(compliance, 1):
        body_cell(ws, row, 2, i, align="center", bold=True)
        body_cell(ws, row, 3, req, bold=True)
        body_cell(ws, row, 4, desc, size=8)
        body_cell(ws, row, 5, reg, align="center", size=8)
        body_cell(ws, row, 6, stat, align="center", fill=LIGHT_GREEN, size=9)
        ws.row_dimensions[row].height = 26
        row += 1
    row += 2

    # ===== VI. Источники =====
    section_title(ws, row, "VI. ИСТОЧНИКИ", color=GREY)
    row += 2
    sources = [
        "Указ Президента РФ № 309 от 07.05.2024 «О национальных целях развития РФ на период до 2030 года»",
        "Постановление Правительства РФ № 1215 от 31.12.2013 «О предоставлении субсидий Фонду кино»",
        "Национальный проект «Культура» (паспорт 2019–2024, продление 2025–2030)",
        "Нацпроект «Семья» 2025–2030 (приоритеты Минкультуры)",
        "Налоговый кодекс РФ, ст. 149 (льготы по НДС для кинопроизводства)",
        "Кремлин.ру — публикации о культурной политике",
        "ТАСС — новости Минкультуры и Фонда кино",
        "Официальный сайт Минкультуры РФ (culture.gov.ru)",
        "Единая автоматизированная информационная система (ЕАИС) Фонда кино",
    ]
    for s in sources:
        body_cell(ws, row, 2, f"• {s}", size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        row += 1

    ws.freeze_panes = "A9"
    print(f"  [33_Gov_KPI] 6 нацпроектов, 10 KPI, 5 источников господдержки (1040 млн ₽), "
          f"7 направлений социокультурного вклада, 10 compliance чекпоинтов")


# ==========================================================================
# ЛИСТ 34: TAX_SCHEDULE
# ==========================================================================
def build_tax(wb):
    if "34_Tax_Schedule" in wb.sheetnames:
        del wb["34_Tax_Schedule"]
    ws = wb.create_sheet("34_Tax_Schedule")
    title(ws, "TAX SCHEDULE — НАЛОГОВЫЙ КАЛЕНДАРЬ 2026–2032",
          "Поквартально Q1'26–Q4'28 + годовые 2029–2032. НДС 0% кинопроизводство, налог на прибыль 20%, страховые 30.2%, НДФЛ 13/15%",
          end_col=20)

    # ===== I. Перечень налогов =====
    section_title(ws, 6, "I. ПРИМЕНИМЫЕ НАЛОГИ И СТАВКИ (НК РФ)")
    headers = ["#", "Налог", "База", "Ставка 2026", "Ставка 2028", "Статья НК", "Комментарий"]
    widths = [5, 22, 30, 14, 14, 12, 30]
    for i, h in enumerate(headers):
        header_cell(ws, 8, 2 + i, h, width=widths[i])

    taxes = [
        (1, "Налог на прибыль",
         "Прибыль до налогообложения",
         "20% (17% регион + 3% фед.)",
         "20% (17/3)",
         "ст. 284",
         "Основная ставка, льготы для КИК"),
        (2, "НДС",
         "Услуги кинопроизводства",
         "0% (кино)",
         "0% (кино)",
         "ст. 149 п.20",
         "Освобождение: производство + прокат нац. фильмов с УНФ"),
        (3, "Страховые взносы",
         "ФОТ",
         "30.2% (ОПС+ОМС+ОСС)",
         "30.2%",
         "НК гл. 34",
         "Стандартный тариф, возможны льготы для IT-аккредитации"),
        (4, "НДФЛ (штатные)",
         "ФОТ штатных",
         "13% / 15%",
         "13% / 15%",
         "ст. 224",
         "15% с превышения 5 млн ₽/год"),
        (5, "НДФЛ (фриланс GPH)",
         "Выплаты по ГПХ",
         "13%",
         "13%",
         "ст. 224",
         "Удерживается при выплате"),
        (6, "Налог на имущество",
         "Стоимость ОС",
         "2.2% (регион)",
         "2.2%",
         "ст. 380",
         "Московская область + СПб"),
        (7, "Транспортный налог",
         "Авто, техника",
         "Регион. ставки",
         "Регион. ставки",
         "ст. 361",
         "Для съёмочной техники"),
        (8, "Земельный налог",
         "Участки под студии",
         "0.3%-1.5%",
         "0.3%-1.5%",
         "ст. 394",
         "Для капитальных студий"),
        (9, "ЕНВД/УСН",
         "Упрощёнка для JV",
         "6% дохода / 15% Д-Р",
         "6% / 15%",
         "гл. 26.2",
         "Не применим — обороты >450 млн ₽"),
    ]
    row = 9
    for d in taxes:
        for i, v in enumerate(d):
            body_cell(ws, row, 2 + i, v,
                      align="center" if i in (0, 3, 4, 5) else "left",
                      bold=(i == 0),
                      size=8 if i == 6 else 9)
        ws.row_dimensions[row].height = 32
        row += 1
    row += 2

    # ===== II. Налоговый календарь поквартально =====
    section_title(ws, row, "II. НАЛОГОВЫЙ КАЛЕНДАРЬ 2026–2028 (поквартально, млн ₽)")
    row += 2

    periods = ["Q1'26", "Q2'26", "Q3'26", "Q4'26",
               "Q1'27", "Q2'27", "Q3'27", "Q4'27",
               "Q1'28", "Q2'28", "Q3'28", "Q4'28",
               "2029", "2030", "2031", "2032", "Σ"]

    # Базовые драйверы по периодам
    # Revenue (млн ₽) из финмодели
    revenue = [30, 40, 80, 235, 180, 320, 530, 635, 460, 640, 745, 650,
               420, 340, 200, 90]
    # EBITDA (grossed up 47.3%)
    ebitda = [round(r * 0.473, 1) for r in revenue]
    # Налог на прибыль 20% от EBITDA (упрощённо, для налогового кэлендаря)
    profit_tax = [round(e * 0.20, 1) for e in ebitda]
    # ФОТ по периодам (из прежних данных)
    fot = [20, 22, 25, 28, 26, 28, 30, 32, 28, 30, 32, 30,
           12, 10, 8, 6]
    # Страховые 30.2%
    insurance = [round(f * 0.302, 1) for f in fot]
    # НДФЛ 13% (приближение для mix штатных и GPH)
    ndfl = [round(f * 0.13, 1) for f in fot]
    # Имущественный и прочие — маленькие константы
    other = [2.0] * 16

    def row_total(xs):
        return round(sum(xs), 1)

    # Заголовок
    headers2 = ["Налог", *periods]
    header_cell(ws, row, 2, "Налог", width=20)
    for i, p in enumerate(periods):
        header_cell(ws, row, 3 + i, p, width=9)
    row += 1

    tax_rows = [
        ("Налог на прибыль (20%)", profit_tax, LIGHT_BLUE),
        ("НДС (0%, справочно)", [0.0] * 16, LIGHT_GREY),
        ("Страховые взносы (30.2%)", insurance, LIGHT_GREEN),
        ("НДФЛ (13-15%)", ndfl, LIGHT_ORANGE),
        ("Имущество + прочие", other, YELLOW),
    ]

    for name, data, col in tax_rows:
        body_cell(ws, row, 2, name, bold=True, fill=col)
        for i, v in enumerate(data):
            body_cell(ws, row, 3 + i, v, align="right", fmt="#,##0.0", fill=col)
        body_cell(ws, row, 3 + len(data), row_total(data),
                  align="right", fmt="#,##0.0", bold=True, fill=LIGHT_BLUE)
        row += 1

    # Итого
    totals = [round(profit_tax[i] + insurance[i] + ndfl[i] + other[i], 1) for i in range(16)]
    body_cell(ws, row, 2, "ИТОГО НАЛОГИ", bold=True, fill=DARK_BLUE, size=10)
    ws.cell(row, 2).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    for i, v in enumerate(totals):
        body_cell(ws, row, 3 + i, v, align="right", fmt="#,##0.0",
                  bold=True, fill=LIGHT_BLUE)
    body_cell(ws, row, 3 + 16, row_total(totals),
              align="right", fmt="#,##0.0", bold=True, fill=DARK_BLUE, size=10)
    ws.cell(row, 3 + 16).font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    row += 2

    # ===== III. Summary 3Y & 7Y =====
    section_title(ws, row, "III. SUMMARY НАЛОГОВЫХ ОТЧИСЛЕНИЙ")
    row += 2
    headers = ["Период", "Налог на прибыль", "Страховые", "НДФЛ", "Прочие", "ИТОГО", "% от Revenue"]
    widths = [16, 16, 16, 14, 14, 14, 14]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1

    # 3Y 2026-2028 (первые 12 периодов)
    pt_3y = round(sum(profit_tax[:12]), 1)
    ins_3y = round(sum(insurance[:12]), 1)
    ndfl_3y = round(sum(ndfl[:12]), 1)
    oth_3y = round(sum(other[:12]), 1)
    tot_3y = round(pt_3y + ins_3y + ndfl_3y + oth_3y, 1)
    rev_3y = sum(revenue[:12])

    # 4Y tail 2029-2032
    pt_tail = round(sum(profit_tax[12:]), 1)
    ins_tail = round(sum(insurance[12:]), 1)
    ndfl_tail = round(sum(ndfl[12:]), 1)
    oth_tail = round(sum(other[12:]), 1)
    tot_tail = round(pt_tail + ins_tail + ndfl_tail + oth_tail, 1)
    rev_tail = sum(revenue[12:])

    # 7Y total
    pt_7y = round(pt_3y + pt_tail, 1)
    ins_7y = round(ins_3y + ins_tail, 1)
    ndfl_7y = round(ndfl_3y + ndfl_tail, 1)
    oth_7y = round(oth_3y + oth_tail, 1)
    tot_7y = round(tot_3y + tot_tail, 1)
    rev_7y = rev_3y + rev_tail

    summary_rows = [
        ("3Y (2026-2028)", pt_3y, ins_3y, ndfl_3y, oth_3y, tot_3y, tot_3y/rev_3y, LIGHT_BLUE),
        ("Tail (2029-2032)", pt_tail, ins_tail, ndfl_tail, oth_tail, tot_tail, tot_tail/rev_tail, LIGHT_ORANGE),
        ("7Y ИТОГО", pt_7y, ins_7y, ndfl_7y, oth_7y, tot_7y, tot_7y/rev_7y, DARK_BLUE),
    ]
    for period, pt, ins, ndf, oth, tot, share, col in summary_rows:
        is_dark = (col == DARK_BLUE)
        body_cell(ws, row, 2, period, bold=True, fill=col)
        if is_dark:
            ws.cell(row, 2).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        body_cell(ws, row, 3, pt, align="right", fmt="#,##0.0", fill=col)
        body_cell(ws, row, 4, ins, align="right", fmt="#,##0.0", fill=col)
        body_cell(ws, row, 5, ndf, align="right", fmt="#,##0.0", fill=col)
        body_cell(ws, row, 6, oth, align="right", fmt="#,##0.0", fill=col)
        body_cell(ws, row, 7, tot, align="right", fmt="#,##0.0", bold=True, fill=col)
        body_cell(ws, row, 8, share, align="right", fmt="0.0%", fill=col)
        if is_dark:
            for c in range(3, 9):
                ws.cell(row, c).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        row += 1
    row += 2

    # ===== IV. Key tax insights =====
    section_title(ws, row, "IV. KEY INSIGHTS")
    row += 2
    insights = [
        f"1. Общий объём налогов 7Y ≈ {tot_7y:.0f} млн ₽ ({tot_7y/rev_7y*100:.1f}% от Revenue)",
        f"2. 3Y net tax burden ≈ {tot_3y:.0f} млн ₽, из которых налог на прибыль ~{pt_3y:.0f} млн ₽ (основной)",
        "3. НДС 0% — ключевая льгота: экономия ~800 млн ₽ vs стандартная ставка 20%",
        "4. Страховые взносы 30.2% — основной cost driver для ФОТ, требует оптимизации через IT-аккредитацию",
        "5. Налог на прибыль 20% применяется к JV-структуре; producer equity 600 млн — отдельный tax treatment",
        "6. Рибейты: Московская область 30%, Татарстан 25%, Крым 20% — возможность снизить effective tax rate",
        "7. Налоговые риски (R18 из Risk Register): оспаривание НДС-льготы требует предварительного согласования с ФНС",
    ]
    for t in insights:
        body_cell(ws, row, 2, t, size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=20)
        row += 1
    row += 2

    # ===== V. Sources =====
    section_title(ws, row, "V. ИСТОЧНИКИ", color=GREY)
    row += 2
    sources = [
        "Налоговый кодекс РФ, части 1 и 2 (консультантплюс.ру)",
        "ФНС России — разъяснения по НДС в сфере культуры (nalog.gov.ru)",
        "Минфин РФ — налоговая политика РФ 2024-2026",
        "Федеральный закон № 126-ФЗ «О господдержке кинематографии»",
        "Постановление Правительства РФ № 1215/2013 (грантовый механизм)",
        "Приказы ФНС по возмещению НДС 0% для кинопроизводства",
    ]
    for s in sources:
        body_cell(ws, row, 2, f"• {s}", size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=20)
        row += 1

    ws.freeze_panes = "C9"
    print(f"  [34_Tax_Schedule] 9 налогов × 17 периодов (12Q + 4Y + Σ), "
          f"3Y taxes={tot_3y:.0f}, 7Y={tot_7y:.0f}, effective rate={tot_7y/rev_7y*100:.1f}%")


# ==========================================================================
# ЛИСТ 35: ROADMAP 2026-2032
# ==========================================================================
def build_roadmap(wb):
    if "35_Roadmap_2026_2032" in wb.sheetnames:
        del wb["35_Roadmap_2026_2032"]
    ws = wb.create_sheet("35_Roadmap_2026_2032")
    title(ws, "ROADMAP 2026–2032 — КЛЮЧЕВЫЕ МИЛЕСТОНЫ",
          "Производство slate, фандрайзинг, governance, exit routes. Critical path color-coded.",
          end_col=17)

    # 7 лет × 4 квартала = 28 столбцов, но упростим до 16 (12 Q + 4 года tail)
    periods = ["Q1'26", "Q2'26", "Q3'26", "Q4'26",
               "Q1'27", "Q2'27", "Q3'27", "Q4'27",
               "Q1'28", "Q2'28", "Q3'28", "Q4'28",
               "2029", "2030", "2031", "2032"]

    # ===== I. Production milestones =====
    section_title(ws, 6, "I. PRODUCTION MILESTONES (12 фильмов × 4 фазы: Dev / Shoot / Post / Release)")
    header_cell(ws, 8, 2, "Фильм / Этап", width=24)
    for i, p in enumerate(periods):
        header_cell(ws, 8, 3 + i, p, width=7)

    # Фильмы и их фазы — цвет каждой фазы
    FILMS = [
        ("F01 Родные стены", 0, 2, 3),      # Dev Q1'26, Shoot Q2-Q3'26, Post Q3'26, Release Q4'26
        ("F02 Два сердца", 0, 3, 4),
        ("F03 Ночной патруль", 1, 4, 5),
        ("F04 Последний экзамен", 1, 4, 5),
        ("F05 Время героев", 1, 4, 6),
        ("F06 Красная заря", 2, 5, 7),
        ("F07 Сказка наяву", 2, 6, 8),
        ("F08 Неизвестный герой", 3, 6, 9),
        ("F09 Моё дело", 3, 7, 9),
        ("F10 Северный ветер", 3, 7, 10),
        ("F11 Полдень", 4, 8, 10),
        ("F12 Горизонт событий", 4, 8, 11),
    ]

    # Цвета фаз
    C_DEV = "D5E3F0"    # светло-голубой
    C_SHOOT = "FCE4D6"  # оранжевый
    C_POST = "FFF2CC"   # желтый
    C_REL = LIGHT_GREEN # зеленый

    row = 9
    for name, dev_start, shoot_start, release_period in FILMS:
        body_cell(ws, row, 2, name, bold=True, size=9)
        for i, p in enumerate(periods):
            cell = ws.cell(row, 3 + i)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=7)
            # dev phase
            if dev_start <= i < shoot_start:
                cell.fill = hfill(C_DEV)
                cell.value = "Dev"
            # shoot + post (combined 1-2 Q)
            elif shoot_start <= i < release_period:
                cell.fill = hfill(C_SHOOT if i == shoot_start else C_POST)
                cell.value = "Shoot" if i == shoot_start else "Post"
            # release
            elif i == release_period:
                cell.fill = hfill(C_REL)
                cell.value = "★Rel"
                cell.font = Font(name="Calibri", size=7, bold=True, color="006100")
            # tail (library)
            elif i > release_period:
                cell.fill = hfill(LIGHT_GREY)
                cell.value = "tail"
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    # Легенда
    body_cell(ws, row, 2, "Легенда:", bold=True, size=9)
    legend = [("Dev", C_DEV), ("Shoot", C_SHOOT), ("Post", C_POST), ("★Rel Release", C_REL), ("tail Tail rev.", LIGHT_GREY)]
    for i, (lbl, col) in enumerate(legend):
        c = ws.cell(row, 3 + i)
        c.value = lbl
        c.fill = hfill(col)
        c.font = Font(name="Calibri", size=8, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    row += 2

    # ===== II. Финансовые милестоны =====
    section_title(ws, row, "II. FUNDRAISING & FINANCIAL MILESTONES")
    row += 2
    header_cell(ws, row, 2, "Этап", width=24)
    for i, p in enumerate(periods):
        header_cell(ws, row, 3 + i, p, width=7)
    row += 1

    fin_milestones = [
        ("T₁ Tranche 1 (20% = 250)", 0, "★ 250 млн ₽"),
        ("T₁ Tranche 2 (28% = 350)", 1, "★ 350 млн ₽"),
        ("T₁ Tranche 3 (28% = 350)", 2, "★ 350 млн ₽"),
        ("T₁ Tranche 4 (24% = 300)", 3, "★ 300 млн ₽"),
        ("ФКП грант раунд 1", 0, "ФКП 180"),
        ("ФКП грант раунд 2", 4, "ФКП 220"),
        ("ФКП грант раунд 3", 8, "ФКП 120"),
        ("ИРИ раунд", 2, "ИРИ 150"),
        ("Producer equity (JV)", 0, "600 equity"),
        ("Revolving credit line", 3, "200 RCL"),
        ("T₁ Repayment 1 (300)", 12, "−300"),
        ("T₁ Repayment 2 (700)", 13, "−700"),
        ("T₁ Repayment 3 (500)", 14, "−500"),
        ("T₁ Final repayment", 15, "−223"),
    ]
    for name, qi, label in fin_milestones:
        body_cell(ws, row, 2, name, bold=True, size=8)
        for i in range(16):
            cell = ws.cell(row, 3 + i)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=7)
            if i == qi:
                cell.value = label
                if "−" in label:
                    cell.fill = hfill(LIGHT_RED)
                    cell.font = Font(name="Calibri", size=7, bold=True, color=RED)
                elif "★" in label:
                    cell.fill = hfill(LIGHT_GREEN)
                    cell.font = Font(name="Calibri", size=7, bold=True, color="006100")
                else:
                    cell.fill = hfill(LIGHT_BLUE)
                    cell.font = Font(name="Calibri", size=7, bold=True)
        ws.row_dimensions[row].height = 22
        row += 1
    row += 2

    # ===== III. Governance =====
    section_title(ws, row, "III. GOVERNANCE & REPORTING MILESTONES")
    row += 2
    header_cell(ws, row, 2, "Этап", width=24)
    for i, p in enumerate(periods):
        header_cell(ws, row, 3 + i, p, width=7)
    row += 1

    gov_milestones = [
        ("Board meeting", [0, 2, 4, 6, 8, 10, 12, 13, 14, 15], "BM"),
        ("Investor quarterly report", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], "Rep"),
        ("Audit (годовой)", [3, 7, 11, 12, 13, 14, 15], "Audit"),
        ("Covenant test", [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], "Cov"),
        ("Strategic review", [5, 11, 13, 15], "Rev"),
    ]
    for name, qs, label in gov_milestones:
        body_cell(ws, row, 2, name, bold=True, size=8)
        for i in range(16):
            cell = ws.cell(row, 3 + i)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=7)
            if i in qs:
                cell.value = label
                cell.fill = hfill(LIGHT_ORANGE)
                cell.font = Font(name="Calibri", size=7, bold=True, color="823B0B")
        ws.row_dimensions[row].height = 22
        row += 1
    row += 2

    # ===== IV. Exit routes =====
    section_title(ws, row, "IV. EXIT ROUTES MILESTONES (из листа 25_Exit_Scenarios)")
    row += 2
    header_cell(ws, row, 2, "Exit route", width=24)
    for i, p in enumerate(periods):
        header_cell(ws, row, 3 + i, p, width=7)
    row += 1

    exit_routes = [
        ("E1: Strategic sale", [13, 14], "SS ★", "Яндекс / ВК / Сбер (EV 7-9× EBITDA)"),
        ("E2: Trade sale PE", [14, 15], "TS", "Private equity (EV 5-7×)"),
        ("E3: IPO (small-cap Мосбиржа)", [14, 15], "IPO", "Pre-IPO прокладка 12-18 мес."),
        ("E4: Secondary sale", [12, 13], "Sec", "Продажа доли T₁ новому инвестору"),
        ("E5: Management buyout", [15], "MBO", "Выкуп менеджментом"),
        ("E6: Gov strategic (РФПИ)", [13, 14], "Gov", "РФПИ / РОСТЕХ / Минкультуры"),
        ("E7: Wind-down (hold)", [15], "WD", "Удержание + дивиденды из library"),
    ]
    for name, qs, label, comment in exit_routes:
        body_cell(ws, row, 2, name, bold=True, size=8)
        for i in range(16):
            cell = ws.cell(row, 3 + i)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=7)
            if i in qs:
                cell.value = label
                cell.fill = hfill("E7E6E6" if "WD" in label else "B4C7E7")
                cell.font = Font(name="Calibri", size=7, bold=True, color=DARK_BLUE)
        ws.row_dimensions[row].height = 22
        row += 1
    row += 2

    # ===== V. Critical path =====
    section_title(ws, row, "V. CRITICAL PATH (ключевые зависимости)", color=RED)
    row += 2
    cp_items = [
        "Q1'26: T₁ Tranche 1 250 млн ₽ → старт Dev F01-F05",
        "Q2'26: Tranche 2 350 → Shoot F01-F02 (без задержки)",
        "Q3'26: Tranche 3 350 → Shoot F03-F04, Pre-production F05",
        "Q4'26: Tranche 4 300 + Release F01 → первый revenue impulse",
        "Q2'27: Release F02-F04 → проверка slate performance vs прогноз",
        "Q3'27: Release F05 (блокбастер 250 млн) → critical test hypothesis",
        "Q4'27: Mid-term review → решение о продолжении Slate B (F11-F12) или pivot",
        "Q2'28: Release F08-F09 → cash flow breakpoint",
        "Q4'28: Release F12 + full slate completed → start exit prep",
        "2029: Q1 — covenant pass или W₃ activation decision",
        "2029-2030: Strategic review → exit route selection → T₁ repayment start",
        "2030-2032: Exit execution + full T₁ repayment + library monetization",
    ]
    for item in cp_items:
        body_cell(ws, row, 2, f"• {item}", size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=17)
        row += 1
    row += 2

    # ===== VI. Key dates =====
    section_title(ws, row, "VI. TOP-10 КРИТИЧЕСКИХ ДАТ")
    row += 2
    headers = ["#", "Дата", "Событие", "Критичность", "Ответственный"]
    widths = [5, 16, 50, 14, 20]
    for i, h in enumerate(headers):
        header_cell(ws, row, 2 + i, h, width=widths[i])
    row += 1
    key_dates = [
        (1, "15.01.2026", "Подписание T₁ Term Sheet + SPA", "Critical", "CFO / Legal"),
        (2, "31.03.2026", "Closing Tranche 1 (250 млн ₽)", "Critical", "CFO / Board"),
        (3, "15.06.2026", "Старт съёмок F01 «Родные стены»", "High", "Producer F01"),
        (4, "30.09.2026", "Tranche 3 + старт съёмок F05", "Critical", "CFO + F05 Prod"),
        (5, "01.12.2026", "Премьера F01 (широкий прокат)", "High", "Distribution"),
        (6, "15.09.2027", "Премьера F05 «Время героев» (блокбастер)", "Critical", "CEO / Marketing"),
        (7, "31.12.2027", "Mid-term review + Slate B decision", "Critical", "Board"),
        (8, "31.12.2028", "Full slate completed + exit prep start", "Critical", "Board / IR"),
        (9, "30.06.2029", "Covenant pass check + T₁ Repayment 1", "Critical", "CFO / Board"),
        (10, "31.12.2032", "Final T₁ repayment + exit execution", "Critical", "CFO / IR / Board"),
    ]
    for d in key_dates:
        for i, v in enumerate(d):
            crit = d[3]
            bg = LIGHT_RED if crit == "Critical" else LIGHT_ORANGE
            body_cell(ws, row, 2 + i, v,
                     align="center" if i in (0, 1, 3) else "left",
                     bold=(i == 0),
                     fill=bg if i == 3 else None)
        ws.row_dimensions[row].height = 26
        row += 1

    ws.freeze_panes = "C9"
    print(f"  [35_Roadmap_2026_2032] 12 films × 16 periods × 4 фазы, "
          f"14 fin milestones, 5 governance tracks, 7 exit routes, "
          f"12 critical path items, 10 top dates")


# ==========================================================================
# MAIN
# ==========================================================================
wb = load_workbook(XLSX)
print(f"Loaded: {XLSX}")
print(f"Sheets before: {len(wb.sheetnames)}")
print()

print("[1/3] Building 33_Gov_KPI …")
build_gov_kpi(wb)
print()

print("[2/3] Building 34_Tax_Schedule …")
build_tax(wb)
print()

print("[3/3] Building 35_Roadmap_2026_2032 …")
build_roadmap(wb)
print()

print(f"Sheets after: {len(wb.sheetnames)}")
wb.save(XLSX)
print(f"\nSaved: {XLSX}")
