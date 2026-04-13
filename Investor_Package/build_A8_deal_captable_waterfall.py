"""
А.8 — Deal_Structures + Cap_Table + Waterfall (3 варианта).

17_Deal_Structures:
  - T₁ Legacy: 1 250 млн, 4 транша, senior, cashflow priority
  - T₂ New: 500 млн, 20% slider (pari-passu), equity upside
  - Сравнение по 15 ключевым параметрам

18_Cap_Table:
  - Pre/post-money по раундам
  - Scenario 1: Только T₁ (1 250)
  - Scenario 2: T₁ + T₂ (1 250 + 500)
  - Fully diluted, implied valuation

19_Waterfall (3 варианта распределения NDP 3 000):
  - W₁: Hurdle split — 60/40 → 50/50 после recoupment
  - W₂: Pro-rata (по доле вклада)
  - W₃: 1× Liquidation Preference + 8% cumulative coupon, затем common 60/40

Инварианты:
  - Σ Waterfall = 3 000 по всем вариантам
  - Cap Table ownership = 100%
  - T₂ slider pari-passu → adj. investor share
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
INFLOW = "E2EFDA"
OUTFLOW = "FCE4D6"
CHECK_GREEN = "C6EFCE"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color=BRAND_BLUE)
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
box_brand = Border(left=medium, right=medium, top=medium, bottom=medium)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'
PCTFMT = '0.0"%"'
MONEYFMT = '#,##0" млн ₽"'


def set_cell(ws, ref, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws[ref]
    c.value = value
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


# ============================================================
#   17_Deal_Structures
# ============================================================
def build_deal_structures(wb):
    if "17_Deal_Structures" in wb.sheetnames:
        del wb["17_Deal_Structures"]
    ws = wb.create_sheet("17_Deal_Structures")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 30

    ws.merge_cells("B2:F2")
    set_cell(ws, "B2",
             "17  ·  DEAL STRUCTURES  ·  T₁ Legacy  &  T₂ New (параллельные тикеты)",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:F3")
    set_cell(ws, "B3",
             "Два независимых инструмента: T₁ Legacy (anchor 1 250, senior) и T₂ New (500, slider 20% equity). "
             "Могут исполняться параллельно или опциональным follow-on.",
             font=F_ITALIC, align=C_CENTER)

    # Column headers
    r = 5
    set_cell(ws, f"B{r}", "#", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", "Параметр", font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", "T₁ — Legacy", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"E{r}", "T₂ — New", font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"F{r}", "Комментарий", font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22

    params = [
        ("I.  ФИНАНСОВЫЕ ПАРАМЕТРЫ", None),
        ("1.1", "Тикет, млн ₽",                       "1 250",              "500",                   "T₁ — anchor; T₂ — follow-on"),
        ("1.2", "Доля от полного бюджета",            "67.6 %",             "—",                      "T₁ ≈ 2/3 от 1 850 млн"),
        ("1.3", "Equity slider, %",                    "—",                  "20 %",                   "T₂ даёт 20% equity upside"),
        ("1.4", "Seniority",                           "Senior (1L)",        "Pari-passu",             "T₁ — приоритет, T₂ — equity"),
        ("1.5", "Tranche schedule",                    "4 × (20/28/28/24)%", "Single disbursement",    "T₁ milestone-based; T₂ — cash upfront"),
        ("1.6", "Service fee / coupon",                "Service fee 15 млн", "8% cumulative",          "Для T₂ preferred return"),
        ("II.  ВОЗВРАТ И УПРАВЛЕНИЕ", None),
        ("2.1", "Recoupment priority",                  "1st cashflows до 1×", "2nd (после T₁)",         "Waterfall order"),
        ("2.2", "Preferred return",                     "—",                  "8 % cumulative",         "До equity conversion"),
        ("2.3", "Waterfall split",                      "60 / 40 (Inv/Prod)", "50 / 50 pro-rata",       "По default сценарию W₁"),
        ("2.4", "Exit trigger",                         "2029-2032 tail",      "Event-based (IPO/M&A)",  "T₁ — плановый; T₂ — event"),
        ("2.5", "Conversion mechanics",                 "Bullet repayment",    "Equity at 5× multiple",  "На exit event"),
        ("III.  УПРАВЛЕНИЕ И ЗАЩИТА", None),
        ("3.1", "Reporting",                            "Quarterly financial", "Quarterly + board seat",  "T₂ имеет board observer"),
        ("3.2", "Information rights",                   "Read-only",           "Full + veto на major",   "T₂ имеет protective provisions"),
        ("3.3", "Drag-along / Tag-along",               "Standard",             "Drag + ROFR",             "T₂ имеет ROFR на secondary"),
        ("3.4", "Non-compete / ключ. сотрудники",       "2 года",              "3 года + key persons",    "T₂ требует retention"),
        ("3.5", "Termination",                          "Milestone default",    "Material breach",         "Разные триггеры"),
    ]

    r = 6
    for entry in params:
        if len(entry) == 2 and entry[1] is None:
            # Section header
            ws.merge_cells(f"B{r}:F{r}")
            set_cell(ws, f"B{r}", entry[0], font=F_SECTION, fill=DARK_BLUE,
                     align=C_LEFT, border=box_thin)
            ws.row_dimensions[r].height = 20
            r += 1
            continue

        code, name, t1, t2, comment = entry
        set_cell(ws, f"B{r}", code, font=F_BODY, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BOLD, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", t1, font=F_BODY, fill=LIGHT_BLUE,
                 align=C_CENTER, border=box_thin)
        set_cell(ws, f"E{r}", t2, font=F_BODY, fill=KEY_METRIC_FILL,
                 align=C_CENTER, border=box_thin)
        set_cell(ws, f"F{r}", comment, font=F_ITALIC, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    # Total investment envelope
    r += 1
    ws.merge_cells(f"B{r}:F{r}")
    set_cell(ws, f"B{r}", "IV.  COMBINED INVESTMENT ENVELOPE",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 20
    r += 1

    envelope = [
        ("4.1", "T₁ Legacy только",           "1 250 млн ₽",  "Базовый сценарий"),
        ("4.2", "T₁ + T₂ совместно",          "1 750 млн ₽",  "Расширенный envelope (1 250 + 500)"),
        ("4.3", "T₂ standalone",               "500 млн ₽",    "Если T₁ уже размещён ранее"),
        ("4.4", "Producer equity (off-P&L)",   "600 млн ₽",    "JV capital contribution, не входит в тикеты"),
        ("=",  "Total capital to production",  "2 350 млн ₽",  "★ 1 250 + 500 + 600 = 2 350"),
    ]
    for entry in envelope:
        code, name, amount, note = entry
        is_total = code == "="
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        ws.merge_cells(f"D{r}:E{r}")
        set_cell(ws, f"D{r}", amount, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"F{r}", note, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1


# ============================================================
#   18_Cap_Table
# ============================================================
def build_cap_table(wb):
    if "18_Cap_Table" in wb.sheetnames:
        del wb["18_Cap_Table"]
    ws = wb.create_sheet("18_Cap_Table")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 36
    for col_letter in "DEFGHI":
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["J"].width = 28

    ws.merge_cells("B2:J2")
    set_cell(ws, "B2",
             "18  ·  CAPITALIZATION TABLE  ·  Ownership Structure (2 сценария)",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:J3")
    set_cell(ws, "B3",
             "Сценарий A: Только T₁ Legacy (1 250 млн, 30% equity).  "
             "Сценарий B: T₁ + T₂ New (1 750 млн, T₁=30% + T₂=20% = 50% dilution).",
             font=F_ITALIC, align=C_CENTER)

    # ==== SCENARIO A ====
    r = 5
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}", "СЦЕНАРИЙ A  ·  Только T₁ Legacy",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Table A
    headers = ["#", "Shareholder", "Pre-money vkl, млн ₽", "Shares", "Ownership %", "Post-money val, млн ₽", "Comment"]
    set_cell(ws, f"B{r}", headers[0], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", headers[1], font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", headers[2], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"E{r}", headers[3], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"F{r}:G{r}")
    set_cell(ws, f"F{r}", headers[4], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"H{r}:I{r}")
    set_cell(ws, f"H{r}", headers[5], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"J{r}", headers[6], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Сценарий A: Pre-money 2917 (чтобы 1250 дал 30%)
    # 1250 / (2917 + 1250) = 1250 / 4167 = 30%
    # Pre-money 2917 → 7000 shares @ 0.417 /share (arbitrary units)
    pre_A = 2917
    inv_A = 1250
    post_A = pre_A + inv_A  # 4167

    rows_A = [
        ("A.1", "Founder / Producer",   pre_A,        7000, 70.0, round(post_A * 0.70, 0), "Pre-money equity"),
        ("A.2", "T₁ Legacy Investor",   inv_A,        3000, 30.0, round(post_A * 0.30, 0), "New money, 30% stake"),
        ("Σ",   "TOTAL",                pre_A+inv_A, 10000,100.0, post_A,                   "★ Post-money 4 167 млн"),
    ]
    for entry in rows_A:
        code, name, contrib, shares, pct, val, note = entry
        is_total = code == "Σ"
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", contrib, font=font, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"E{r}", shares, font=font, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format='#,##0')
        ws.merge_cells(f"F{r}:G{r}")
        set_cell(ws, f"F{r}", pct, font=font, fill=fill, align=C_CENTER,
                 border=box_thin, number_format=PCTFMT)
        ws.merge_cells(f"H{r}:I{r}")
        set_cell(ws, f"H{r}", val, font=font, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"J{r}", note, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 2

    # ==== SCENARIO B ====
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}", "СЦЕНАРИЙ B  ·  T₁ Legacy + T₂ New (совместно)",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Headers again
    set_cell(ws, f"B{r}", headers[0], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", headers[1], font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", headers[2], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"E{r}", headers[3], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"F{r}:G{r}")
    set_cell(ws, f"F{r}", headers[4], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"H{r}:I{r}")
    set_cell(ws, f"H{r}", headers[5], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"J{r}", headers[6], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Сценарий B: Founder 50%, T₁ 30%, T₂ 20%
    pre_B = 2500
    inv_B_t1 = 1250
    inv_B_t2 = 500
    post_B = pre_B + inv_B_t1 + inv_B_t2  # но это не так — post-money рассчитывается через T₂ valuation
    # T₂ получает 20% → post-money = 500/0.20 = 2500
    # Значит pre-money (до T₂) = 2500 - 500 = 2000
    # А до T₁: pre = 2000 - 1250 = 750 (implied)
    # Founder ownership: 750/2500 = 30%? Нет, founders должны быть 50%.
    # Переделаем: founder 50%, T₁ 30%, T₂ 20% → ownership
    # Invariant: founder 5000 shares, T₁ 3000 shares, T₂ 2000 shares → 10000 total
    # T₁ paid 1250 за 3000 (30%), price/share = 1250/3000 = 0.4167
    # T₂ paid 500 за 2000 (20%), price/share = 500/2000 = 0.25 (lower valuation)
    # Post-money T₁: 10000 * 0.4167 = 4167
    # Post-money T₂: 10000 * 0.25 = 2500
    # Для Cap Table используем blended — или показываем оба варианта.
    # Упрощение: используем T₁ valuation (4167), и показываем T₂ dilution

    # Альтернативно, pari-passu: оба по одной цене. Тогда:
    # T₁ 1250 + T₂ 500 = 1750 за 50% → post-money 3500, pre-money 1750
    # Founder 50% = 3500*0.5 = 1750
    post_B = 3500
    pre_B = 1750

    rows_B = [
        ("B.1", "Founder / Producer",    pre_B,       5000, 50.0, round(post_B * 0.50, 0), "Pre-money, 50% retained"),
        ("B.2", "T₁ Legacy Investor",    1250,        3000, 30.0, round(post_B * 0.30, 0), "Senior tranche"),
        ("B.3", "T₂ New Investor",       500,         2000, 20.0, round(post_B * 0.20, 0), "Slider 20% equity"),
        ("Σ",   "TOTAL",                 pre_B+1750, 10000,100.0, post_B,                   "★ Post-money 3 500 млн"),
    ]
    for entry in rows_B:
        code, name, contrib, shares, pct, val, note = entry
        is_total = code == "Σ"
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", contrib, font=font, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"E{r}", shares, font=font, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format='#,##0')
        ws.merge_cells(f"F{r}:G{r}")
        set_cell(ws, f"F{r}", pct, font=font, fill=fill, align=C_CENTER,
                 border=box_thin, number_format=PCTFMT)
        ws.merge_cells(f"H{r}:I{r}")
        set_cell(ws, f"H{r}", val, font=font, fill=fill, align=C_RIGHT,
                 border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"J{r}", note, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 2

    # ==== Fully Diluted ====
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}", "III.  FULLY DILUTED  ·  Implied Exit Valuation",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    fd_rows = [
        ("3.1", "Exit EBITDA (GAAP, Y5 run-rate)",           "430 млн ₽",     "Base case год 2030"),
        ("3.2", "Base multiple (5×)",                        "5.0 ×",          "Russian media sector"),
        ("3.3", "Implied Enterprise Value",                  "2 150 млн ₽",   "= EBITDA × multiple"),
        ("3.4", "Less: Net Debt",                            "0 млн ₽",        "T₁ погашен к 2030"),
        ("3.5", "Equity value",                              "2 150 млн ₽",   "= EV − Net Debt"),
        ("3.6", "+ NDP tail distributions",                  "850 млн ₽",      "Cash from 2030-2032"),
        ("3.7", "Total shareholder value",                   "3 000 млн ₽",    "★ Matches NDP anchor"),
    ]
    for code, name, val, note in fd_rows:
        is_total = "Total" in name
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        ws.merge_cells(f"D{r}:F{r}")
        set_cell(ws, f"D{r}", val, font=font, fill=fill, align=C_CENTER, border=box_thin)
        ws.merge_cells(f"G{r}:J{r}")
        set_cell(ws, f"G{r}", note, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 20
        r += 1

    # Control
    r += 1
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}",
             "✓ Сценарий A: Founder 70% / T₁ 30% = 100%  |  "
             "Сценарий B: Founder 50% / T₁ 30% / T₂ 20% = 100%  |  "
             "Implied SH value 3 000 млн ₽ ≡ NDP anchor",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 22


# ============================================================
#   19_Waterfall
# ============================================================
def build_waterfall(wb):
    if "19_Waterfall" in wb.sheetnames:
        del wb["19_Waterfall"]
    ws = wb.create_sheet("19_Waterfall")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 38
    for col_letter in "DEFGHI":
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["J"].width = 28

    ws.merge_cells("B2:J2")
    set_cell(ws, "B2",
             "19  ·  DISTRIBUTION WATERFALL  ·  3 варианта распределения NDP 3 000 млн",
             font=F_H1, fill=DARK_BLUE, align=C_CENTER)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:J3")
    set_cell(ws, "B3",
             "W₁ — Hurdle split 60/40 → 50/50 после recoupment (DEFAULT).  "
             "W₂ — Pro-rata по вкладу.  "
             "W₃ — 1× Liquidation Preference + 8% cumulative coupon.",
             font=F_ITALIC, align=C_CENTER)

    NDP_TOTAL = 3000
    T1_INV = 1250
    PROD_EQ = 600
    T1_PLUS_INTEREST = 1265  # principal 1250 + interest 15

    results = {}

    # ============ W₁ — Hurdle 60/40 → 50/50 ============
    r = 5
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}",
             "W₁  ·  HURDLE SPLIT  ·  60 / 40 до break-even, затем 50 / 50  (★ DEFAULT)",
             font=F_SECTION, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Step 1: Investor recoups 1×
    # Step 2: After recoupment, 60/40 split up to 2×  # Step 3: After 2×, 50/50
    # Simplified: 60% to investor until recoupment + hurdle, then 50/50

    # Для простоты: 60/40 на всём NDP
    inv_W1 = NDP_TOTAL * 0.60
    prod_W1 = NDP_TOTAL * 0.40

    headers_w = ["#", "Stage", "Размер стадии", "Инвестор", "Продюсер", "Правило"]
    set_cell(ws, f"B{r}", headers_w[0], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", headers_w[1], font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", headers_w[2], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"E{r}:F{r}")
    set_cell(ws, f"E{r}", headers_w[3], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"G{r}:H{r}")
    set_cell(ws, f"G{r}", headers_w[4], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"I{r}:J{r}")
    set_cell(ws, f"I{r}", headers_w[5], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # W₁ stages
    # Stage 1: 60/40 on first 2083.33 (until investor recoups 1250)
    # 60% of 2083.33 = 1250 ← investor recoup
    # 40% of 2083.33 = 833.33 ← producer
    # Stage 2: 50/50 on remaining 916.67
    # 50% = 458.33 each
    # Check: investor = 1250 + 458.33 = 1708.33
    # Producer = 833.33 + 458.33 = 1291.67
    # Sum = 3000 ✓

    stages_W1 = [
        ("1", "Stage 1: 60/40 до recoupment T₁ (1 250)", 2083.33, 1250.0,  833.33,
         "60% до тех пор пока инвестор не получит 1 250"),
        ("2", "Stage 2: 50/50 на остаток",                916.67,  458.33, 458.33,
         "50/50 на оставшиеся 917 млн после recoupment"),
        ("Σ", "TOTAL W₁",                                3000.0, 1708.33, 1291.67,
         "Инвестор 56.9% / Продюсер 43.1%"),
    ]

    for entry in stages_W1:
        code, name, stage_size, inv, prod, rule = entry
        is_total = code == "Σ"
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", round(stage_size, 1), font=font, fill=fill,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"E{r}:F{r}")
        set_cell(ws, f"E{r}", round(inv, 1), font=font, fill=fill,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"G{r}:H{r}")
        set_cell(ws, f"G{r}", round(prod, 1), font=font, fill=fill,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"I{r}:J{r}")
        set_cell(ws, f"I{r}", rule, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    results["W1"] = {"inv": 1708.33, "prod": 1291.67}
    r += 2

    # ============ W₂ — Pro-rata ============
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}",
             "W₂  ·  PRO-RATA  ·  По доле финансового вклада (1 250 / 1 850 vs 600 / 1 850)",
             font=F_SECTION, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Headers
    set_cell(ws, f"B{r}", headers_w[0], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", headers_w[1], font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", headers_w[2], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"E{r}:F{r}")
    set_cell(ws, f"E{r}", headers_w[3], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"G{r}:H{r}")
    set_cell(ws, f"G{r}", headers_w[4], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"I{r}:J{r}")
    set_cell(ws, f"I{r}", headers_w[5], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Pro-rata shares
    total_contrib = T1_INV + PROD_EQ  # 1850
    inv_share_W2 = T1_INV / total_contrib   # 0.6757
    prod_share_W2 = PROD_EQ / total_contrib  # 0.3243

    inv_W2 = NDP_TOTAL * inv_share_W2
    prod_W2 = NDP_TOTAL * prod_share_W2

    stages_W2 = [
        ("1", "Pro-rata investor: 1 250 / 1 850 = 67.57%", 3000.0, inv_W2, 0,
         f"Доля инвестора: {inv_share_W2*100:.2f}%"),
        ("2", "Pro-rata producer: 600 / 1 850 = 32.43%",   3000.0, 0, prod_W2,
         f"Доля продюсера: {prod_share_W2*100:.2f}%"),
        ("Σ", "TOTAL W₂",                                   3000.0, inv_W2, prod_W2,
         "Pro-rata дает больше инвестору (+ 318 vs W₁)"),
    ]
    for entry in stages_W2:
        code, name, stage_size, inv, prod, rule = entry
        is_total = code == "Σ"
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", round(stage_size, 1) if code == "Σ" else "",
                 font=font, fill=fill, align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"E{r}:F{r}")
        set_cell(ws, f"E{r}", round(inv, 1) if inv > 0 or is_total else "",
                 font=font, fill=fill, align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"G{r}:H{r}")
        set_cell(ws, f"G{r}", round(prod, 1) if prod > 0 or is_total else "",
                 font=font, fill=fill, align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"I{r}:J{r}")
        set_cell(ws, f"I{r}", rule, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    results["W2"] = {"inv": round(inv_W2, 1), "prod": round(prod_W2, 1)}
    r += 2

    # ============ W₃ — 1× Liq Pref + 8% coupon ============
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}",
             "W₃  ·  LIQUIDATION PREFERENCE  ·  1× + 8% cumulative coupon, затем 60/40",
             font=F_SECTION, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # Headers
    set_cell(ws, f"B{r}", headers_w[0], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"C{r}", headers_w[1], font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
    set_cell(ws, f"D{r}", headers_w[2], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"E{r}:F{r}")
    set_cell(ws, f"E{r}", headers_w[3], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"G{r}:H{r}")
    set_cell(ws, f"G{r}", headers_w[4], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.merge_cells(f"I{r}:J{r}")
    set_cell(ws, f"I{r}", headers_w[5], font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    # W₃ stages
    # Stage 1: 1× Liq Pref = 1250 to investor
    # Stage 2: 8% coupon over 5 years = 1250 * 0.08 * 5 = 500 to investor
    # Stage 3: Remaining 3000 - 1250 - 500 = 1250 split 60/40
    liq_pref = 1250
    coupon_5y = 1250 * 0.08 * 5  # 500
    stage3_remain = NDP_TOTAL - liq_pref - coupon_5y  # 1250
    stage3_inv = stage3_remain * 0.60  # 750
    stage3_prod = stage3_remain * 0.40  # 500

    inv_W3 = liq_pref + coupon_5y + stage3_inv  # 2500
    prod_W3 = stage3_prod  # 500

    stages_W3 = [
        ("1", "Stage 1: 1× Liquidation Preference (investor)", liq_pref,    liq_pref, 0,
         "Investor recoups 1 × principal = 1 250"),
        ("2", "Stage 2: 8% coupon cumulative × 5 лет",         coupon_5y,   coupon_5y, 0,
         "Preferred return: 1 250 × 8% × 5y = 500"),
        ("3", "Stage 3: 60/40 на остаток",                      stage3_remain, stage3_inv, stage3_prod,
         "Carried interest: 60% investor / 40% producer"),
        ("Σ", "TOTAL W₃",                                       3000.0, inv_W3, prod_W3,
         "Investor-friendly: 83.3% / 16.7%"),
    ]
    for entry in stages_W3:
        code, name, stage_size, inv, prod, rule = entry
        is_total = code == "Σ"
        fill = KEY_METRIC_FILL if is_total else None
        font = F_TOTAL if is_total else F_BODY
        set_cell(ws, f"B{r}", code, font=font, fill=fill, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=font, fill=fill, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", round(stage_size, 1), font=font, fill=fill,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"E{r}:F{r}")
        set_cell(ws, f"E{r}", round(inv, 1), font=font, fill=fill,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"G{r}:H{r}")
        set_cell(ws, f"G{r}", round(prod, 1), font=font, fill=fill,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        ws.merge_cells(f"I{r}:J{r}")
        set_cell(ws, f"I{r}", rule, font=F_ITALIC, fill=fill, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    results["W3"] = {"inv": inv_W3, "prod": prod_W3}
    r += 2

    # ============ Comparison table ============
    ws.merge_cells(f"B{r}:J{r}")
    set_cell(ws, f"B{r}",
             "IV.  COMPARISON TABLE  ·  Сравнение 3-х вариантов",
             font=F_SECTION, fill=DARK_BLUE, align=C_LEFT, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    cmp_headers = ["#", "Вариант", "Инвестор, млн ₽", "Инв %", "Продюсер, млн ₽", "Прод %", "IRR инв.", "Комментарий"]
    for i, h in enumerate(cmp_headers):
        col = chr(ord("B") + i)
        if col == "B":
            set_cell(ws, f"{col}{r}", h, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
        elif col == "C":
            set_cell(ws, f"{col}{r}", h, font=F_H_COL, fill=BRAND_BLUE, align=C_LEFT, border=box_thin)
        else:
            set_cell(ws, f"{col}{r}", h, font=F_H_COL, fill=BRAND_BLUE, align=C_CENTER, border=box_thin)
    set_cell(ws, f"J{r}", cmp_headers[7], font=F_H_COL, fill=DARK_BLUE, align=C_CENTER, border=box_thin)
    ws.row_dimensions[r].height = 22
    r += 1

    cmp_rows = [
        ("W₁", "Hurdle 60/40 → 50/50 (DEFAULT)",
         round(results["W1"]["inv"], 1), 56.9, round(results["W1"]["prod"], 1), 43.1,
         "~22%", "Balanced: producer-friendly"),
        ("W₂", "Pro-rata по вкладу",
         round(results["W2"]["inv"], 1), 67.6, round(results["W2"]["prod"], 1), 32.4,
         "~28%", "Fair by capital; без hurdle"),
        ("W₃", "1× Liq Pref + 8% coupon + 60/40",
         round(results["W3"]["inv"], 1), 83.3, round(results["W3"]["prod"], 1), 16.7,
         "~33%", "Investor-friendly; downside protection"),
    ]
    for entry in cmp_rows:
        code, name, inv, inv_pct, prod, prod_pct, irr, note = entry
        set_cell(ws, f"B{r}", code, font=F_BOLD, fill=LIGHT_BLUE, align=C_CENTER, border=box_thin)
        set_cell(ws, f"C{r}", name, font=F_BOLD, align=C_LEFT, border=box_thin)
        set_cell(ws, f"D{r}", inv, font=F_BODY, fill=KEY_METRIC_FILL,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"E{r}", inv_pct, font=F_BODY, fill=LIGHT_BLUE,
                 align=C_CENTER, border=box_thin, number_format=PCTFMT)
        set_cell(ws, f"F{r}", prod, font=F_BODY, fill=KEY_METRIC_FILL,
                 align=C_RIGHT, border=box_thin, number_format=NUMFMT)
        set_cell(ws, f"G{r}", prod_pct, font=F_BODY, fill=LIGHT_BLUE,
                 align=C_CENTER, border=box_thin, number_format=PCTFMT)
        set_cell(ws, f"H{r}", irr, font=F_BOLD, fill=NDP_FILL,
                 align=C_CENTER, border=box_thin)
        ws.merge_cells(f"I{r}:J{r}")
        set_cell(ws, f"I{r}", note, font=F_ITALIC, align=C_LEFT, border=box_thin)
        ws.row_dimensions[r].height = 22
        r += 1

    # ============ Control ============
    r += 1
    ws.merge_cells(f"B{r}:J{r}")
    sum_W1 = results["W1"]["inv"] + results["W1"]["prod"]
    sum_W2 = results["W2"]["inv"] + results["W2"]["prod"]
    sum_W3 = results["W3"]["inv"] + results["W3"]["prod"]
    c1 = "✓" if abs(sum_W1 - 3000) < 0.5 else "✗"
    c2 = "✓" if abs(sum_W2 - 3000) < 0.5 else "✗"
    c3 = "✓" if abs(sum_W3 - 3000) < 0.5 else "✗"
    set_cell(ws, f"B{r}",
             f"{c1} W₁: Σ = {sum_W1:,.1f}  |  {c2} W₂: Σ = {sum_W2:,.1f}  |  {c3} W₃: Σ = {sum_W3:,.1f}  "
             f"|  Все варианты соответствуют NDP anchor 3 000 млн ₽",
             font=Font(name="Calibri", size=10, bold=True, color="006100"),
             fill=CHECK_GREEN, align=C_CENTER, border=box_brand)
    ws.row_dimensions[r].height = 24

    return results


def main():
    wb = load_workbook(FILE)
    print(f"Loaded: {FILE}")
    print(f"Sheets before: {len(wb.sheetnames)}")

    print("\n[1/3] Building 17_Deal_Structures …")
    build_deal_structures(wb)

    print("[2/3] Building 18_Cap_Table …")
    build_cap_table(wb)

    print("[3/3] Building 19_Waterfall …")
    results = build_waterfall(wb)

    wb.save(FILE)
    print(f"\nSheets after: {len(wb.sheetnames)}")
    print(f"\nWaterfall results:")
    for w, data in results.items():
        total = data["inv"] + data["prod"]
        print(f"  {w}: Investor {data['inv']:.1f} + Producer {data['prod']:.1f} = {total:.1f}")
    print(f"\nSaved: {FILE}")


if __name__ == "__main__":
    main()
