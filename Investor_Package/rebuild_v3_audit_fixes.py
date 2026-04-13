#!/usr/bin/env python3
"""
Master rebuild script — applies audit fixes (FIX-01..FIX-06) to investor_model_v3.

Rebuilds sheets:
  09_P&L_Statement  (FIX-01 revenue floor, FIX-02 opex cap, FIX-03 renewal scenario)
  10_Cash_Flow      (FIX-01, FIX-02)
  11_Balance_Sheet   (FIX-04 asset floors, FIX-05 loan floor)
  22_Valuation_DCF   (FIX-03 reconciliation note)

Then applies post-processing fixes:
  FIX-06: OOXML orphan rels
  FIX-07: Localization duplicates
  FIX-08: Cover Letter date/version sync
"""
import os
import sys
import importlib.util

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)


def load_module(name, path):
    """Dynamically load a Python module from path."""
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def rebuild_pnl_cf_bs(xlsx_path):
    """Rebuild P&L, Cash Flow, Balance Sheet sheets with audit fixes."""
    from openpyxl import load_workbook

    print(f"\n{'='*60}")
    print(f"  Rebuilding: {os.path.basename(xlsx_path)}")
    print(f"{'='*60}")

    wb = load_workbook(xlsx_path)
    print(f"  Loaded {len(wb.sheetnames)} sheets")

    # ── Import fixed build modules ─────────────────────────────
    # We need to monkey-patch the FILE variable in each module

    # 1. Rebuild 09_P&L_Statement via build_A4
    print("\n  [1/4] Rebuilding 09_P&L_Statement (FIX-01, FIX-02, FIX-03)...")
    build_a4 = load_module("build_A4", os.path.join(BASE_DIR, "build_A4_variantC_pnl.py"))

    # Patch cover
    try:
        build_a4.patch_cover(wb)
        print("    ✓ 01_Cover patched (dual metric)")
    except Exception as e:
        print(f"    ⚠ 01_Cover patch skipped: {e}")

    # Patch assumptions
    try:
        build_a4.patch_assumptions(wb)
        print("    ✓ 02_Assumptions patched (NDP bridge)")
    except Exception as e:
        print(f"    ⚠ 02_Assumptions patch skipped: {e}")

    # Build P&L
    try:
        if "09_P&L_Statement" in wb.sheetnames:
            del wb["09_P&L_Statement"]
        build_a4.build_pnl(wb)
        print("    ✓ 09_P&L_Statement rebuilt with revenue floor + opex cap + renewal scenario")
    except Exception as e:
        print(f"    ✗ 09_P&L_Statement FAILED: {e}")
        import traceback; traceback.print_exc()

    # 2. Rebuild 10_Cash_Flow and 11_Balance_Sheet via build_A5
    print("\n  [2/4] Rebuilding 10_Cash_Flow (FIX-01, FIX-02)...")
    build_a5 = load_module("build_A5", os.path.join(BASE_DIR, "build_A5_cf_bs.py"))

    try:
        cf_data = build_a5.build_cash_flow(wb)
        print("    ✓ 10_Cash_Flow rebuilt with revenue floor + opex cap")
    except Exception as e:
        print(f"    ✗ 10_Cash_Flow FAILED: {e}")
        import traceback; traceback.print_exc()
        cf_data = None

    print("\n  [3/4] Rebuilding 11_Balance_Sheet (FIX-04, FIX-05)...")
    if cf_data:
        try:
            bs_data = build_a5.build_balance_sheet(wb, cf_data)
            print(f"    ✓ 11_Balance_Sheet rebuilt with asset floors + loan floor")
            print(f"    ✓ Balance check max diff: {bs_data['max_diff']:.4f}")
        except Exception as e:
            print(f"    ✗ 11_Balance_Sheet FAILED: {e}")
            import traceback; traceback.print_exc()

    # 3. Rebuild 22_Valuation_DCF via build_A10
    print("\n  [4/4] Rebuilding 22_Valuation_DCF (FIX-03 reconciliation note)...")
    build_a10 = load_module("build_A10", os.path.join(BASE_DIR, "build_A10_valuation.py"))

    try:
        if "22_Valuation_DCF" in wb.sheetnames:
            del wb["22_Valuation_DCF"]
        dcf = build_a10.build_dcf(wb)
        print(f"    ✓ 22_Valuation_DCF rebuilt with reconciliation note")
    except Exception as e:
        print(f"    ✗ 22_Valuation_DCF FAILED: {e}")
        import traceback; traceback.print_exc()

    # Save
    wb.save(xlsx_path)
    print(f"\n  ✓ Saved: {xlsx_path}")
    print(f"  Total sheets: {len(wb.sheetnames)}")


def run_post_fixes(xlsx_path):
    """Run post-processing fix scripts on xlsx."""
    print(f"\n  Post-processing: {os.path.basename(xlsx_path)}")

    # FIX-07: Localization duplicates
    fix_loc = load_module("fix_loc", os.path.join(BASE_DIR, "fix_localization_duplicates.py"))
    n = fix_loc.fix_duplicates(xlsx_path)
    print(f"    FIX-07: {n} localization duplicates removed")

    # FIX-08: Cover Letter sync
    fix_cover = load_module("fix_cover", os.path.join(BASE_DIR, "fix_cover_letter_sync.py"))
    n = fix_cover.fix_dates_and_versions(xlsx_path)
    print(f"    FIX-08: {n} date/version cells synced")

    # FIX-06: OOXML orphan rels
    fix_ooxml = load_module("fix_ooxml", os.path.join(BASE_DIR, "fix_ooxml_orphan_rels.py"))
    fix_ooxml.fix_orphan_rels(xlsx_path)
    print(f"    FIX-06: OOXML orphan rels check complete")


def verify_fixes(xlsx_path):
    """Quick verification of applied fixes."""
    from openpyxl import load_workbook
    import re

    wb = load_workbook(xlsx_path, data_only=True)
    issues = []

    # Check 1: Revenue floor — tail years should be >= 380
    if "09_P&L_Statement" in wb.sheetnames:
        ws = wb["09_P&L_Statement"]
        # Look for Total Revenue row in tail columns
        for row in ws.iter_rows(max_row=80, max_col=22):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "Total Revenue" in str(cell.value):
                    # Check columns P, Q, R, S (tail years)
                    r = cell.row
                    for col_letter in ["P", "Q", "R", "S"]:
                        val = ws[f"{col_letter}{r}"].value
                        if val is not None and isinstance(val, (int, float)) and val < 380:
                            issues.append(f"Revenue in {col_letter}{r} = {val} < floor 380")
                    break

    # Check 2: No negative assets in BS
    if "11_Balance_Sheet" in wb.sheetnames:
        ws = wb["11_Balance_Sheet"]
        for row in ws.iter_rows(max_row=30, max_col=22):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    v = str(cell.value)
                    if "Content library" in v or "PP&E" in v or "Property" in v:
                        r = cell.row
                        for col_letter in ["P", "Q", "R", "S"]:
                            val = ws[f"{col_letter}{r}"].value
                            if val is not None and isinstance(val, (int, float)) and val < -0.01:
                                issues.append(f"Negative asset in {col_letter}{r} = {val}")

    # Check 3: No localization duplicates
    dupe_count = 0
    pattern = re.compile(r'(\b[А-ЯЁа-яё]+(?:\s+[А-ЯЁа-яё]+)*\b)\s*\(\1\)')
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=100, max_col=30):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if pattern.search(cell.value):
                        dupe_count += 1

    if dupe_count > 0:
        issues.append(f"Still {dupe_count} localization duplicates remaining")

    # Check 4: Reconciliation note in DCF
    if "22_Valuation_DCF" in wb.sheetnames:
        ws = wb["22_Valuation_DCF"]
        found_recon = False
        for row in ws.iter_rows(max_row=150, max_col=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "RECONCILIATION" in cell.value:
                    found_recon = True
                    break
        if not found_recon:
            issues.append("No reconciliation note found in 22_Valuation_DCF")

    return issues


def main():
    targets = [
        os.path.join(BASE_DIR, "investor_model_v3_Internal.xlsx"),
        os.path.join(BASE_DIR, "investor_model_v3_Public.xlsx"),
    ]

    # Also rebuild root-level copies
    parent = os.path.dirname(BASE_DIR)
    targets.extend([
        os.path.join(parent, "investor_model_v3_Internal.xlsx"),
        os.path.join(parent, "investor_model_v3_Public.xlsx"),
    ])

    for path in targets:
        if not os.path.exists(path):
            print(f"SKIP (not found): {path}")
            continue

        # Phase 1: Rebuild sheets
        rebuild_pnl_cf_bs(path)

        # Phase 2: Post-processing
        run_post_fixes(path)

        # Phase 3: Verify
        print(f"\n  Verifying fixes...")
        issues = verify_fixes(path)
        if issues:
            print(f"  ⚠ ISSUES FOUND:")
            for issue in issues:
                print(f"    - {issue}")
        else:
            print(f"  ✓ ALL FIXES VERIFIED")

    print(f"\n{'='*60}")
    print("  REBUILD COMPLETE")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
