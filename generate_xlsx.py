#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side

# ЧИТАЕМ СПЕЦИФИКАЦИЮ
with open('/sessions/friendly-gracious-dirac/mnt/Холдинг/structure.json', 'r', encoding='utf-8') as f:
    spec = json.load(f)

# СОЗДАЁМ WORKBOOK
wb = Workbook()
wb.remove(wb.active)

# §12.2: WORKAROUND для default font
wb._fonts[0] = Font(
    name="Calibri",
    size=12,
    family=2,
    scheme="minor",
    charset=204,
    color=Color(theme=1)
)
wb._named_styles['Normal'].font = Font(
    name="Calibri",
    size=12,
    family=2,
    scheme="minor",
    charset=204,
    color=Color(theme=1)
)

def apply_font(cell, font_spec):
    if not font_spec:
        return
    font_kwargs = {}
    if "name" in font_spec:
        font_kwargs["name"] = font_spec["name"]
    if "size" in font_spec:
        font_kwargs["size"] = font_spec["size"]
    if "b" in font_spec:
        font_kwargs["bold"] = font_spec["b"]
    if "i" in font_spec:
        font_kwargs["italic"] = font_spec["i"]
    if "color" in font_spec:
        color_spec = font_spec["color"]
        if isinstance(color_spec, dict):
            if "theme" in color_spec:
                color_kwargs = {"theme": color_spec["theme"]}
                if "tint" in color_spec:
                    color_kwargs["tint"] = color_spec["tint"]
                font_kwargs["color"] = Color(**color_kwargs)
                if "family" in font_spec:
                    font_kwargs["family"] = font_spec["family"]
                if "scheme" in font_spec:
                    font_kwargs["scheme"] = font_spec["scheme"]
                if "charset" in font_spec:
                    font_kwargs["charset"] = font_spec["charset"]
            elif "rgb" in color_spec:
                rgb = color_spec["rgb"]
                if not rgb.startswith("FF"):
                    rgb = "FF" + rgb
                font_kwargs["color"] = Color(rgb=rgb)
            elif "indexed" in color_spec:
                font_kwargs["color"] = Color(indexed=color_spec["indexed"])
    if font_kwargs:
        cell.font = Font(**font_kwargs)

def apply_fill(cell, fill_spec):
    if not fill_spec:
        return
    if isinstance(fill_spec, str):
        color = fill_spec if fill_spec.startswith("FF") else "FF" + fill_spec
        cell.fill = PatternFill(fill_type="solid", fgColor=Color(rgb=color))

def apply_alignment(cell, al_spec):
    if not al_spec:
        return
    al_kwargs = {}
    if "h" in al_spec:
        al_kwargs["horizontal"] = al_spec["h"]
    if "v" in al_spec:
        al_kwargs["vertical"] = al_spec["v"]
    if "wrap" in al_spec:
        al_kwargs["wrap_text"] = al_spec["wrap"]
    if al_kwargs:
        cell.alignment = Alignment(**al_kwargs)

def parse_side(side_spec):
    """Парсит спецификацию стороны границы и возвращает Side"""
    if not side_spec:
        return None
    if isinstance(side_spec, dict):
        style = side_spec.get("style")
        if not style:
            return None
        color_spec = side_spec.get("color")
        color = None
        if color_spec:
            if isinstance(color_spec, dict):
                if "theme" in color_spec:
                    color_kwargs = {"theme": color_spec["theme"]}
                    if "tint" in color_spec:
                        color_kwargs["tint"] = color_spec["tint"]
                    color = Color(**color_kwargs)
                elif "rgb" in color_spec:
                    rgb = color_spec["rgb"]
                    if not rgb.startswith("FF"):
                        rgb = "FF" + rgb
                    color = Color(rgb=rgb)
                elif "auto" in color_spec:
                    color = Color(auto=True)
        return Side(style=style, color=color)
    return None

def apply_borders(cell, bd_spec):
    if not bd_spec:
        return
    
    thin_side = Side(style="thin")
    
    if bd_spec == "box":
        cell.border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
    elif isinstance(bd_spec, dict):
        border_kwargs = {}
        for s in ("top", "bottom", "left", "right"):
            if s in bd_spec:
                side = parse_side(bd_spec[s])
                if side:
                    border_kwargs[s] = side
        if border_kwargs:
            cell.border = Border(**border_kwargs)

def apply_number_format(cell, nf_spec):
    if nf_spec:
        cell.number_format = nf_spec

# СОЗДАНИЕ ЛИСТОВ И ЯЧЕЕК
sheet_order = spec.get("sheet_order", [])
sheet_counts = {}

for sheet_name in sheet_order:
    sheet_data = spec["sheets"][sheet_name]
    ws = wb.create_sheet(sheet_name)
    sheet_counts[sheet_name] = len(sheet_data.get("cells", {}))

    # Ширины колонок
    for col, width in sheet_data.get("column_widths", {}).items():
        ws.column_dimensions[col].width = width

    # Высоты строк
    for row_str, height in sheet_data.get("row_heights", {}).items():
        row_num = int(row_str)
        ws.row_dimensions[row_num].height = height

    # Zoom
    zoom = sheet_data.get("zoom")
    if zoom:
        ws.sheet_view.zoomScale = zoom

    # Page margins
    page_margins = spec["meta"].get("page_margins_all_sheets", {})
    if page_margins:
        ws.page_margins.left = page_margins.get("left", 0.7)
        ws.page_margins.right = page_margins.get("right", 0.7)
        ws.page_margins.top = page_margins.get("top", 0.75)
        ws.page_margins.bottom = page_margins.get("bottom", 0.75)

    # Merged cells
    for merge_range in sheet_data.get("merged", []):
        ws.merge_cells(merge_range)

    # Ячейки (§12.11: ВСЕ ячейки, включая пустые с стилями)
    # ВАЖНО: надо явно обратиться к ячейке через ws[addr], чтобы она была создана в _cells
    for addr, cell_spec in sheet_data.get("cells", {}).items():
        cell = ws[addr]
        
        # Значение (§12.10: Формулы имеют ключ "formula")
        if "formula" in cell_spec:
            cell.value = cell_spec["formula"]
        elif "v" in cell_spec and cell_spec["v"] is not None:
            cell.value = cell_spec["v"]
        # Для пустых ячеек (v=null или отсутствует) — значение не устанавливаем
        
        # Стили — применяем ВСЕГДА, даже если нет значения
        apply_font(cell, cell_spec.get("font"))
        apply_fill(cell, cell_spec.get("fill"))
        apply_alignment(cell, cell_spec.get("al"))
        apply_borders(cell, cell_spec.get("bd"))
        apply_number_format(cell, cell_spec.get("nf"))

# СОХРАНИТЬ ФАЙЛ
output_path = '/sessions/friendly-gracious-dirac/mnt/Холдинг/subagent_v22_output.xlsx'
wb.save(output_path)

print(f"File saved: {output_path}")
print(f"\nCell counts processed from spec:")
for sn in sheet_order:
    print(f"  {sn}: {sheet_counts[sn]}")
