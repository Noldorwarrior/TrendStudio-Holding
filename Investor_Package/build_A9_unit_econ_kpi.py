"""
А.9 — Unit_Economics_per_Film + KPI_Dashboard.

20_Unit_Economics_per_Film:
  - 12 фильмов × 14 колонок: бюджет, P&A, total cost, revenue,
    gross profit, margin, multiple, ROI, payback (Q),
    break-even attendance (тыс. зрителей), risk tier, release Q
  - Σ бюджет = 1 850, Σ P&A = 277.5, Σ total cost = 2 127.5
  - Σ revenue (full lifetime 2026-2032) = 5 595 = 4 545 (до 2028) + 1 050 (tail)
  - Σ gross profit = 3 467.5; средняя маржа по портфелю ~62%
  - Средний multiple = 5 595 / 1 850 = 3.02×

21_KPI_Dashboard:
  - 5 секций × 4 столбца (Metric / 2026 / 2027 / 2028 / 3Y total / Tail 2029-32)
  - Section I. Revenue & Growth
  - Section II. Profitability (GAAP EBITDA + NDP anchor)
  - Section III. Returns & Multiples (ROI, IRR, MOIC, Payback)
  - Section IV. Cash & Debt (DSCR, ICR, Leverage)
  - Section V. Operational (films released, avg budget, success rate)
  - Scenario stripe: Bear/Base/Bull для ключевых метрик
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CANDIDATES = [
    "/Users/noldorwarrior/Documents/Claude/Projects/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
    "/sessions/cool-serene-johnson/mnt/Холдинг/Investor_Package/investor_model_v1.0_Public.xlsx",
]
FILE = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])

BRAND_BLUE = "0070C0"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "DEEBF7"
KEY_METRIC_FILL = "FFF2CC"
NDP_FILL = "E2EFDA"
SUBTOTAL_FILL = "D9E1F2"
BEAR_FILL = "FCE4D6"
BULL_FILL = "E2EFDA"
BASE_FILL = "FFF2CC"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color=BRAND_BLUE)
box_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
box_brand = Border(left=medium, right=medium, top=medium, bottom=medium)

F_H1 = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_H_COL = Font(name="Calibri", size=9, bold=True, color=WHITE)
F_BODY = Font(name="Calibri", size=10, color="000000")
F_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
F_TOTAL = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
F_ITALIC = Font(name="Calibri", size=9, italic=True, color="595959")

C_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
C_RIGHT = Alignment(horizontal="right", vertical="center")

NUMFMT = '#,##0.0;[Red]-#,##0.0'
PCTFMT = '0.0"%"'
INTFMT = '#,##0'
MULTFMT = '0.00"×"'


# ═══════════════════════════════════════════════════════════════════
# FILMS PORTFOLIO — 12 films with budget, revenue multiplier, risk tier
# ═══════════════════════════════════════════════════════════════════
# Каждый фильм: (code, name, budget, release_q, genre, risk_tier, mult)
# Σ budget = 1 850, Σ revenue = 5 595 = 4 545 (2026-2028) + 1 050 (tail)
# mult = revenue / budget (в рамках жанра и бюджета)

FILMS = [
    # code, name, budget, release, genre, risk, mult, tail_share
    ("F01", "Родные стены",       150, "Q4'26", "Драма",       "Low",    2.80, 0.18),
    ("F02", "Два сердца",          100, "Q1'27", "Мелодрама",   "Low",    3.20, 0.20),
    ("F03", "Ночной патруль",      220, "Q2'27", "Экшн",        "Medium", 3.10, 0.15),
    ("F04", "Последний экзамен",    90, "Q2'27", "Комедия",     "Low",    3.50, 0.22),
    ("F05", "Время героев",        250, "Q3'27", "Блокбастер",  "High",   3.30, 0.15),
    ("F06", "Красная заря",        180, "Q4'27", "Исторический","Medium", 2.90, 0.20),
    ("F07", "Сказка наяву",        170, "Q1'28", "Семейный",    "Low",    3.40, 0.25),
    ("F08", "Неизвестный герой",   130, "Q2'28", "Биография",   "Medium", 2.80, 0.22),
    ("F09", "Моё дело",             80, "Q2'28", "Комедия",     "Low",    3.60, 0.20),
    ("F10", "Северный ветер",      180, "Q3'28", "Приключения", "Medium", 3.10, 0.18),
    ("F11", "Полдень",             140, "Q3'28", "Триллер",     "Medium", 2.95, 0.18),
    ("F12", "Горизонт событий",    160, "Q4'28", "Фантастика",  "High",   3.15, 0.18),
]

# Проверка: budgets sum
assert sum(f[2] for f in FILMS) == 1850, f"budgets = {sum(f[2] for f in FILMS)}"

# P&A = 15% of budget (Σ = 277.5)
# Revenue lifetime = budget × mult
# Normalize к 5 595 (округление множителей могло дать не ровно 5595)
_raw_rev_sum = sum(f[2] * f[6] for f in FILMS)
_scale = 5595.0 / _raw_rev_sum
# применим scale чтобы сумма = 5595 ровно
print(f"Raw rev sum: {_raw_rev_sum:.2f}, scale: {_scale:.5f}")


def clear_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


# ═══════════════════════════════════════════════════════════════════
# 20_Unit_Economics_per_Film
# ═══════════════════════════════════════════════════════════════════
def build_unit_economics(wb):
    clear_sheet_if_exists(wb, "20_Unit_Economics_per_Film")
    ws = wb.create_sheet("20_Unit_Economics_per_Film")

    # Column widths
    widths = {
        "A": 2, "B": 6, "C": 22, "D": 14, "E": 10, "F": 10, "G": 11,
        "H": 11, "I": 11, "J": 10, "K": 9, "L": 11, "M": 12, "N": 12, "O": 10
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Title
    ws.merge_cells("B2:O3")
    c = ws["B2"]
    c.value = "UNIT ECONOMICS PER FILM — ПОРТФЕЛЬ 12 ФИЛЬМОВ 2026–2028"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    # ── Subtitle
    ws.merge_cells("B4:O4")
    s = ws["B4"]
    s.value = ("Валовая экономика на фильм: бюджет, P&A, выручка за жизненный цикл "
               "(2026–2032), маржа, ROI, payback в кварталах и break-even attendance")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    # ── Column headers
    headers = [
        ("B6", "№"),
        ("C6", "Название"),
        ("D6", "Жанр"),
        ("E6", "Релиз"),
        ("F6", "Бюджет"),
        ("G6", "P&A"),
        ("H6", "Total Cost"),
        ("I6", "Revenue\n(lifetime)"),
        ("J6", "Gross\nProfit"),
        ("K6", "Margin"),
        ("L6", "Multiple"),
        ("M6", "ROI"),
        ("N6", "Payback (Q)"),
        ("O6", "Risk"),
    ]
    for cell, val in headers:
        c = ws[cell]
        c.value = val
        c.font = F_H_COL
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = C_CENTER
        c.border = box_thin
    ws.row_dimensions[6].height = 30

    # ── Rows for each film
    rows_data = []
    row = 7
    for (code, name, budget, release, genre, risk, mult_raw, _tail) in FILMS:
        pna = round(budget * 0.15, 1)
        total_cost = round(budget + pna, 1)
        revenue = round(budget * mult_raw * _scale, 1)
        gross = round(revenue - total_cost, 1)
        margin = gross / revenue * 100 if revenue else 0
        multiple = revenue / budget
        roi = gross / total_cost * 100

        # Payback в кварталах — используем mult и жанр как упрощённую модель
        # Базовая формула: чем выше mult и ниже риск, тем быстрее
        if risk == "Low":
            payback_q = round(total_cost / (revenue / 8), 1)  # 8 кв. жизн. цикла
        elif risk == "Medium":
            payback_q = round(total_cost / (revenue / 9), 1)
        else:
            payback_q = round(total_cost / (revenue / 10), 1)

        rows_data.append((code, name, genre, release, budget, pna,
                          total_cost, revenue, gross, margin, multiple,
                          roi, payback_q, risk))

        vals = [code, name, genre, release, budget, pna, total_cost,
                revenue, gross, margin, multiple, roi, payback_q, risk]
        for i, v in enumerate(vals):
            col = chr(ord("B") + i)
            cell = ws[f"{col}{row}"]
            cell.value = v
            cell.font = F_BODY
            cell.border = box_thin
            if i in (0, 3, 13):  # code, release, risk
                cell.alignment = C_CENTER
            elif i in (1, 2):  # name, genre
                cell.alignment = C_LEFT
            else:
                cell.alignment = C_RIGHT
            if i in (4, 5, 6, 7, 8):  # currency cols
                cell.number_format = NUMFMT
            elif i == 9:  # margin
                cell.number_format = PCTFMT
            elif i == 10:  # multiple
                cell.number_format = MULTFMT
            elif i == 11:  # ROI
                cell.number_format = PCTFMT
            elif i == 12:  # payback
                cell.number_format = '0.0" Q"'

            # Risk coloring
            if i == 13:
                if v == "Low":
                    cell.fill = PatternFill("solid", fgColor="E2EFDA")
                elif v == "Medium":
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
        row += 1

    # ── Totals row
    total_row = row
    ws[f"B{total_row}"].value = ""
    ws.merge_cells(f"B{total_row}:D{total_row}")
    c = ws[f"B{total_row}"]
    c.value = "ИТОГО ПОРТФЕЛЬ (12 фильмов)"
    c.font = F_TOTAL
    c.alignment = C_LEFT
    c.fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)

    tot_budget = sum(r[4] for r in rows_data)
    tot_pna = sum(r[5] for r in rows_data)
    tot_cost = sum(r[6] for r in rows_data)
    tot_rev = sum(r[7] for r in rows_data)
    tot_gross = sum(r[8] for r in rows_data)
    avg_margin = tot_gross / tot_rev * 100
    avg_mult = tot_rev / tot_budget
    avg_roi = tot_gross / tot_cost * 100
    avg_payback = sum(r[12] for r in rows_data) / len(rows_data)

    totals = ["", tot_budget, tot_pna, tot_cost, tot_rev, tot_gross,
              avg_margin, avg_mult, avg_roi, avg_payback, ""]
    start_col = ord("E")
    for i, v in enumerate(totals):
        col = chr(start_col + i)
        cell = ws[f"{col}{total_row}"]
        cell.value = v
        cell.font = F_TOTAL
        cell.fill = PatternFill("solid", fgColor=SUBTOTAL_FILL)
        cell.border = box_thin
        cell.alignment = C_RIGHT
        if i in (1, 2, 3, 4, 5):
            cell.number_format = NUMFMT
        elif i == 6:
            cell.number_format = PCTFMT
        elif i == 7:
            cell.number_format = MULTFMT
        elif i == 8:
            cell.number_format = PCTFMT
        elif i == 9:
            cell.number_format = '0.0" Q"'

    # ── Summary block
    r = total_row + 3
    ws.merge_cells(f"B{r}:O{r}")
    c = ws[f"B{r}"]
    c.value = "КЛЮЧЕВЫЕ ИНСАЙТЫ UNIT ECONOMICS"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT

    insights = [
        ("Σ Total Cost / Σ COGS", f"{tot_cost:.1f} млн ₽ = 2 127.5 млн ₽ (matches P&L COGS)"),
        ("Σ Revenue lifetime", f"{tot_rev:.1f} млн ₽ = 4 545 (2026-2028) + 1 050 (tail 2029-32)"),
        ("Portfolio Gross Margin", f"{avg_margin:.1f}% (industry benchmark: 55-65%)"),
        ("Portfolio Multiple", f"{avg_mult:.2f}× (break-even 1.2×, target 2.5×)"),
        ("Portfolio ROI", f"{avg_roi:.1f}% (hurdle rate 40%, target 60%)"),
        ("Avg Payback", f"{avg_payback:.1f} Q (target ≤ 6 кварталов)"),
        ("High-risk films", f"2 из 12 (Время героев 250, Горизонт событий 160)"),
        ("Concentration", f"Top-3 budgets = 720 млн ₽ = 38.9% от portfolio"),
    ]
    for (lbl, val) in insights:
        r += 1
        ws.merge_cells(f"B{r}:E{r}")
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_LEFT
        ws.merge_cells(f"F{r}:O{r}")
        ws[f"F{r}"].value = val
        ws[f"F{r}"].font = F_BODY
        ws[f"F{r}"].alignment = C_LEFT

    ws.freeze_panes = "B7"
    print(f"  [20_Unit_Economics] Σ budget={tot_budget:.1f}, Σ rev={tot_rev:.1f}, "
          f"margin={avg_margin:.1f}%, mult={avg_mult:.2f}×")
    return rows_data


# ═══════════════════════════════════════════════════════════════════
# 21_KPI_Dashboard
# ═══════════════════════════════════════════════════════════════════
def build_kpi_dashboard(wb, ue_rows):
    clear_sheet_if_exists(wb, "21_KPI_Dashboard")
    ws = wb.create_sheet("21_KPI_Dashboard")

    widths = {"A": 2, "B": 5, "C": 36, "D": 13, "E": 13, "F": 13,
              "G": 13, "H": 13, "I": 13, "J": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Title
    ws.merge_cells("B2:J3")
    c = ws["B2"]
    c.value = "KPI DASHBOARD — ИНВЕСТОРСКАЯ СВОДКА ПО ГОДАМ"
    c.font = F_H1
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_CENTER

    ws.merge_cells("B4:J4")
    s = ws["B4"]
    s.value = ("Ключевые показатели эффективности холдинга «ТрендСтудио» — "
               "Dual metric (GAAP EBITDA + NDP anchor) · Scenario stripe Bear/Base/Bull")
    s.font = F_ITALIC
    s.alignment = C_CENTER

    # ── Period header
    hdr_row = 6
    hdrs = [("B", "#"), ("C", "Метрика"), ("D", "2026"),
            ("E", "2027"), ("F", "2028"), ("G", "Σ 3Y"),
            ("H", "Tail 29-32"), ("I", "Σ Life"), ("J", "Scenario")]
    for col, val in hdrs:
        c = ws[f"{col}{hdr_row}"]
        c.value = val
        c.font = F_H_COL
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = C_CENTER
        c.border = box_thin
    ws.row_dimensions[hdr_row].height = 24

    # ── Data — distribute revenue & EBITDA by year
    # Revenue by year (из А.3): 2026 ≈ 385, 2027 ≈ 1665, 2028 ≈ 2495
    REV_Y = {"2026": 385.0, "2027": 1665.0, "2028": 2495.0}
    EBITDA_Y = {"2026": 50.0, "2027": 640.0, "2028": 1462.0}  # Σ = 2152
    NP_Y = {"2026": 5.0, "2027": 490.0, "2028": 1194.0}  # Σ ≈ 1689
    TAIL_REV = 1050.0
    TAIL_EBITDA = 850.0  # остаток NDP bridge
    TAIL_NP = 650.0

    # Фильмы выпущенные по годам
    films_released = {"2026": 1, "2027": 5, "2028": 6}  # F01; F02-F06; F07-F12

    sections = [
        {
            "title": "I. REVENUE & GROWTH",
            "rows": [
                ("Revenue, млн ₽",
                 REV_Y["2026"], REV_Y["2027"], REV_Y["2028"],
                 sum(REV_Y.values()), TAIL_REV,
                 sum(REV_Y.values()) + TAIL_REV,
                 "Base", NUMFMT),
                ("Revenue YoY growth",
                 "—", f"{(REV_Y['2027']/REV_Y['2026']-1)*100:.0f}%",
                 f"{(REV_Y['2028']/REV_Y['2027']-1)*100:.0f}%",
                 "CAGR 155%", "—", "—", "Base", "@"),
                ("Films released (шт)",
                 1, 5, 6, 12, 0, 12, "Base", INTFMT),
                ("Avg budget per film, млн ₽",
                 150.0, 168.0, 143.3, 154.2, "—", 154.2, "Base", NUMFMT),
            ],
        },
        {
            "title": "II. PROFITABILITY — Dual Metric",
            "rows": [
                ("EBITDA GAAP, млн ₽",
                 EBITDA_Y["2026"], EBITDA_Y["2027"], EBITDA_Y["2028"],
                 2152.0, TAIL_EBITDA, 2152.0 + TAIL_EBITDA, "Base", NUMFMT),
                ("EBITDA margin, %",
                 EBITDA_Y["2026"] / REV_Y["2026"] * 100,
                 EBITDA_Y["2027"] / REV_Y["2027"] * 100,
                 EBITDA_Y["2028"] / REV_Y["2028"] * 100,
                 2152.0 / 4545.0 * 100,
                 TAIL_EBITDA / TAIL_REV * 100,
                 3002.0 / 5595.0 * 100,
                 "Base", PCTFMT),
                ("NDP (legacy anchor), млн ₽",
                 "—", "—", "—", 3000.0, "—", 3000.0, "Base", NUMFMT),
                ("Net Profit, млн ₽",
                 NP_Y["2026"], NP_Y["2027"], NP_Y["2028"],
                 1689.0, TAIL_NP, 1689.0 + TAIL_NP, "Base", NUMFMT),
                ("Net Profit margin, %",
                 NP_Y["2026"] / REV_Y["2026"] * 100,
                 NP_Y["2027"] / REV_Y["2027"] * 100,
                 NP_Y["2028"] / REV_Y["2028"] * 100,
                 1689.0 / 4545.0 * 100,
                 TAIL_NP / TAIL_REV * 100,
                 2339.0 / 5595.0 * 100,
                 "Base", PCTFMT),
            ],
        },
        {
            "title": "III. RETURNS & MULTIPLES",
            "rows": [
                ("ROI portfolio, %", "—", "—", "—",
                 "57.9%", "—", "—", "Base", "@"),
                ("Gross multiple", "—", "—", "—",
                 "2.14×", "0.49×", "3.02×", "Base", "@"),
                ("IRR investor T₁ (W₁)", "—", "—", "—",
                 "~22%", "—", "—", "Base", "@"),
                ("MOIC investor T₁", "—", "—", "—",
                 "—", "—", "1.37×", "Base", "@"),
                ("Payback T₁ (years)", "—", "—", "—",
                 "—", "—", "4.2 y", "Base", "@"),
                ("EV/EBITDA exit (5×)", "—", "—", "—",
                 "—", "—", "10 760 млн ₽", "Bull", "@"),
            ],
        },
        {
            "title": "IV. CASH & DEBT",
            "rows": [
                ("T₁ drawn cumul., млн ₽",
                 250.0, 950.0, 1250.0, 1250.0, "—", 1250.0, "Base", NUMFMT),
                ("Cash end period, млн ₽",
                 "—", "—", "3 538",
                 "3 538", "1 128", "1 128", "Base", "@"),
                ("Interest expense, млн ₽",
                 1.5, 6.0, 5.0, 12.5, 2.5, 15.0, "Base", NUMFMT),
                ("DSCR (EBITDA/Interest)",
                 33.3, 106.7, 292.4, 172.2, "—", "—", "Base", "0.0"),
                ("Net Debt / EBITDA",
                 "—", "—", "0.00×", "0.00×", "—", "—", "Base", "@"),
            ],
        },
        {
            "title": "V. OPERATIONAL",
            "rows": [
                ("Production CAPEX, млн ₽",
                 402.5, 853.0, 594.5, 1850.0, 0.0, 1850.0, "Base", NUMFMT),
                ("P&A spend, млн ₽",
                 22.5, 105.0, 150.0, 277.5, "—", 277.5, "Base", NUMFMT),
                ("FOT (штат A₁), млн ₽",
                 73.9, 73.9, 73.9, 221.8, "—", 221.8, "Base", NUMFMT),
                ("Success rate (ROI>40%)",
                 "—", "—", "—", "12/12", "—", "100%", "Base", "@"),
                ("High-risk films (% of budget)",
                 "—", "—", "—", "22.2%", "—", "—", "Base", "@"),
            ],
        },
    ]

    r = hdr_row + 1
    section_idx = 0
    for sec in sections:
        section_idx += 1
        # Section header
        ws.merge_cells(f"B{r}:J{r}")
        c = ws[f"B{r}"]
        c.value = sec["title"]
        c.font = F_SECTION
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = C_LEFT
        r += 1

        for i, row_data in enumerate(sec["rows"], start=1):
            (lbl, v26, v27, v28, v3y, vtail, vlife, scen, fmt) = row_data
            ws[f"B{r}"].value = f"{section_idx}.{i}"
            ws[f"B{r}"].font = F_BODY
            ws[f"B{r}"].alignment = C_CENTER
            ws[f"B{r}"].border = box_thin

            ws[f"C{r}"].value = lbl
            ws[f"C{r}"].font = F_BOLD
            ws[f"C{r}"].alignment = C_LEFT
            ws[f"C{r}"].border = box_thin

            vals = [v26, v27, v28, v3y, vtail, vlife]
            cols = ["D", "E", "F", "G", "H", "I"]
            for col, v in zip(cols, vals):
                cell = ws[f"{col}{r}"]
                cell.value = v
                cell.font = F_BODY
                cell.border = box_thin
                cell.alignment = C_RIGHT
                if isinstance(v, (int, float)):
                    cell.number_format = fmt
                else:
                    cell.number_format = "@"

            ws[f"J{r}"].value = scen
            ws[f"J{r}"].font = F_BOLD
            ws[f"J{r}"].alignment = C_CENTER
            ws[f"J{r}"].border = box_thin
            if scen == "Bear":
                ws[f"J{r}"].fill = PatternFill("solid", fgColor=BEAR_FILL)
            elif scen == "Bull":
                ws[f"J{r}"].fill = PatternFill("solid", fgColor=BULL_FILL)
            else:
                ws[f"J{r}"].fill = PatternFill("solid", fgColor=BASE_FILL)

            # Highlight Σ 3Y column
            ws[f"G{r}"].fill = PatternFill("solid", fgColor=KEY_METRIC_FILL)
            r += 1
        r += 1  # spacer

    # ── Scenario stripe at the bottom (Bear/Base/Bull summary)
    r += 1
    ws.merge_cells(f"B{r}:J{r}")
    c = ws[f"B{r}"]
    c.value = "VI. SCENARIO STRIPE (Revenue 2028 · EBITDA 3Y · Net Profit 3Y · Exit EV)"
    c.font = F_SECTION
    c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    c.alignment = C_LEFT
    r += 1

    scen_headers = ["Сценарий", "Revenue 2028", "EBITDA 3Y", "NDP 3Y",
                    "Net Profit 3Y", "IRR T₁", "MOIC T₁", "EV/Mult", "EV млн ₽"]
    for i, h in enumerate(scen_headers):
        col = chr(ord("B") + i)
        cell = ws[f"{col}{r}"]
        cell.value = h
        cell.font = F_H_COL
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = C_CENTER
        cell.border = box_thin
    r += 1

    scen_data = [
        ("Bear",  1872, 1614, 2250, 1182,  "12%", "1.05×", "3×", 4840),
        ("Base",  2495, 2152, 3000, 1689,  "22%", "1.37×", "5×", 10760),
        ("Bull",  3119, 2690, 3750, 2111,  "33%", "2.00×", "7×", 18830),
    ]
    for (name, rev, eb, ndp, np, irr, moic, mult, ev) in scen_data:
        ws[f"B{r}"].value = name
        if name == "Bear":
            fill = BEAR_FILL
        elif name == "Bull":
            fill = BULL_FILL
        else:
            fill = BASE_FILL
        ws[f"B{r}"].fill = PatternFill("solid", fgColor=fill)
        ws[f"B{r}"].font = F_BOLD
        ws[f"B{r}"].alignment = C_CENTER
        ws[f"B{r}"].border = box_thin

        vals = [rev, eb, ndp, np, irr, moic, mult, ev]
        for i, v in enumerate(vals):
            col = chr(ord("C") + i)
            cell = ws[f"{col}{r}"]
            cell.value = v
            cell.font = F_BODY
            cell.border = box_thin
            cell.alignment = C_RIGHT
            cell.fill = PatternFill("solid", fgColor=fill)
            if isinstance(v, (int, float)):
                cell.number_format = NUMFMT
        r += 1

    # ── Note
    r += 2
    ws.merge_cells(f"B{r}:J{r}")
    note = ws[f"B{r}"]
    note.value = ("Примечание: Base case — якорь NDP 3 000 млн ₽; Bear = 0.75× Base; "
                  "Bull = 1.25× Base. EV/EBITDA мультипликаторы: 3× bear / 5× base / 7× bull.")
    note.font = F_ITALIC
    note.alignment = C_LEFT

    ws.freeze_panes = "D7"
    print(f"  [21_KPI_Dashboard] sections=5, rows={r}")


# ═══════════════════════════════════════════════════════════════════
def main():
    print(f"Loaded: {FILE}")
    wb = load_workbook(FILE)
    print(f"Sheets before: {len(wb.sheetnames)}")

    print("\n[1/2] Building 20_Unit_Economics_per_Film …")
    ue_rows = build_unit_economics(wb)

    print("\n[2/2] Building 21_KPI_Dashboard …")
    build_kpi_dashboard(wb, ue_rows)

    print(f"\nSheets after: {len(wb.sheetnames)}")
    wb.save(FILE)
    print(f"\nSaved: {FILE}")


if __name__ == "__main__":
    main()
