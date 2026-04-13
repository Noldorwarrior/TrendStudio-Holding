#!/usr/bin/env python3
"""
Generate the TrendStudio investor model audit remediation roadmap workbook.

Output: audit_public_v1_REMEDIATION_ROADMAP.xlsx
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

HEADERS = [
    "ID",
    "Category",
    "Description",
    "Priority",
    "Severity",
    "Responsible",
    "Effort (h)",
    "Deadline",
    "Dependencies",
    "Platform",
    "Acceptance Criteria",
    "Status",
]

ROWS = [
    ("R-001", "Leakage", "Scrub Internal W5 V-D reference from 24_Investor_Returns!B49", "P0", "CRITICAL", "Model Owner", 1, "Immediate", "None", "Excel", 'Cell contains no "Internal" or "W5" text', "OPEN"),
    ("R-002", "Leakage", 'Remove all "Internal" keyword from sharedStrings (8 occurrences)', "P0", "CRITICAL", "Model Owner", 2, "Immediate", "None", "Excel/Python", 'grep -i "internal" on unzipped xlsx returns 0 hits (excluding legitimate "Internal Rate of Return")', "OPEN"),
    ("R-003", "Leakage", "Strip absPath from xl/workbook.xml", "P0", "CRITICAL", "Model Owner", 1, "Immediate", "None", "Python/XML", "No x15ac:absPath element in workbook.xml", "OPEN"),
    ("R-004", "Metadata", 'Clean docProps (description, keywords "L3", lastModifiedBy)', "P0", "CRITICAL", "Model Owner", 1, "Immediate", "R-003", "Excel/Python", 'No anchor values, no "L3", proper author name', "OPEN"),
    ("R-005", "Leakage", 'Fix Internal xlsx title metadata (says "Public")', "P1", "HIGH", "Model Owner", 0.5, "Week 1", "None", "Python", 'Title = "investor_model_v1.0_Internal"', "OPEN"),
    ("R-006", "Docs", "Remove template placeholders from 42_Cover_Letter", "P1", "HIGH", "Model Owner", 1, "Week 1", "None", "Excel", "No [Name], [Date], TODO in cover letter", "OPEN"),
    ("R-007", "Consistency", "Reconcile scenario probabilities (5/15/50/20/10 vs 10/20/40/20/10)", "P1", "HIGH", "Model Owner", 2, "Week 1", "None", "Python/Excel", "Single probability set used everywhere", "OPEN"),
    ("R-008", "Math", "Standardize IRR calculation method", "P1", "HIGH", "Model Owner", 8, "Week 1", "None", "Python", "One method (numpy_financial.irr) used across all scripts", "OPEN"),
    ("R-009", "Math", "Fix MC revenue blend +6% upward bias", "P1", "HIGH", "Model Owner", 4, "Week 1", "R-008", "Python", "Blend formula centers at 1.0 at mode inputs", "OPEN"),
    ("R-010", "Valuation", "Add source citations to peer comps", "P1", "HIGH", "Model Owner", 8, "Week 2", "None", "Excel", "Every comp has date, source (Pitchbook/Mergermarket/press), link", "OPEN"),
    ("R-011", "Valuation", "Reconcile valuation spread to \u22642\u00d7", "P1", "HIGH", "Model Owner", 16, "Week 2", "R-010", "Excel/Python", "DCF/Comps/MC gap \u2264 2\u00d7 with documented explanation", "OPEN"),
    ("R-012", "Math", "Fix D&A jump 167\u00d7 in DCF (3M\u2192500M)", "P1", "HIGH", "Model Owner", 4, "Week 2", "None", "Excel", "D&A transition justified with asset base schedule", "OPEN"),
    ("R-013", "Math", "Clarify MoIC metrics (aggregate vs T1)", "P1", "HIGH", "Model Owner", 4, "Week 2", "None", "Excel", "Both MoIC shown with labels and reconciliation bridge", "OPEN"),
    ("R-014", "Architecture", "Convert P&L sheet to live formulas", "P2", "MEDIUM", "Model Owner", 24, "Week 3", "R-007,R-008", "Excel/Python", "P&L cells reference Assumptions sheet, changes propagate", "OPEN"),
    ("R-015", "Architecture", "Convert Cash Flow sheet to live formulas", "P2", "MEDIUM", "Model Owner", 16, "Week 3", "R-014", "Excel/Python", "CF cells reference P&L, changes propagate", "OPEN"),
    ("R-016", "Architecture", "Convert Waterfall sheet to live formulas", "P2", "MEDIUM", "Model Owner", 16, "Week 3", "R-015", "Excel/Python", "Waterfall recalculates when NDP changes", "OPEN"),
    ("R-017", "Architecture", "Convert Valuation DCF to live formulas", "P2", "MEDIUM", "Model Owner", 16, "Week 4", "R-015", "Excel/Python", "DCF recalculates when WACC/revenue changes", "OPEN"),
    ("R-018", "Tax", "Separate VAT and profit tax in P&L", "P2", "MEDIUM", "Model Owner", 8, "Week 3", "R-014", "Excel", "Distinct rows for \u041d\u0414\u0421 (pass-through) and \u041d\u041d\u041f 20%", "OPEN"),
    ("R-019", "Valuation", "Add CAPM decomposition to WACC", "P2", "MEDIUM", "Model Owner", 4, "Week 2", "None", "Excel", "Rf, Beta, ERP, Country premium, Size premium all sourced", "OPEN"),
    ("R-020", "Risk", "Add risk scoring rubric", "P2", "MEDIUM", "Model Owner", 4, "Week 2", "None", "Excel", "Rubric tab with probability/impact ranges", "OPEN"),
    ("R-021", "Structure", "Create Capital Stack sheet for Producer equity", "P2", "MEDIUM", "Model Owner", 8, "Week 3", "None", "Excel", "Clear hierarchy: Senior LP, Producer equity, T2 alt", "OPEN"),
    ("R-022", "Pipeline", "Pin requirements.txt with ==", "P2", "MEDIUM", "Pipeline Owner", 2, "Week 1", "None", "Python", "pip freeze > requirements.txt", "OPEN"),
    ("R-023", "Pipeline", "Add missing deps (plotly, python-pptx)", "P2", "MEDIUM", "Pipeline Owner", 1, "Week 1", "R-022", "requirements.txt", "All imports resolve on clean install", "OPEN"),
    ("R-024", "Pipeline", "Fix verify.py hardcoded count (14\u219218)", "P2", "MEDIUM", "Pipeline Owner", 0.5, "Week 1", "None", "Python", "len(INPUT_FILES) used dynamically", "OPEN"),
    ("R-025", "Pipeline", "Remove hardcoded macOS paths from 3 scripts", "P2", "MEDIUM", "Pipeline Owner", 2, "Week 1", "None", "Python", "All paths via CLI args or env vars", "OPEN"),
]

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

PRIORITY_FILLS = {
    "P0": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "P1": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "P2": PatternFill(start_color="4DA6FF", end_color="4DA6FF", fill_type="solid"),
}
PRIORITY_FONTS = {
    "P0": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
    "P1": Font(name="Calibri", bold=True, color="000000", size=11),
    "P2": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
}

SEVERITY_FILLS = {
    "CRITICAL": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "HIGH": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="4DA6FF", end_color="4DA6FF", fill_type="solid"),
}
SEVERITY_FONTS = {
    "CRITICAL": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
    "HIGH": Font(name="Calibri", bold=True, color="000000", size=11),
    "MEDIUM": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
}

STATUS_FILL = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
STATUS_FONT = Font(name="Calibri", bold=True, color="000000", size=11)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DEFAULT_FONT = Font(name="Calibri", size=11)

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remediation Roadmap"

    # Freeze the header row
    ws.freeze_panes = "A2"

    # ---- Header row ----
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # ---- Data rows ----
    for row_idx, row_data in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DEFAULT_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Priority column (col 4)
        priority_cell = ws.cell(row=row_idx, column=4)
        prio = str(priority_cell.value)
        if prio in PRIORITY_FILLS:
            priority_cell.fill = PRIORITY_FILLS[prio]
            priority_cell.font = PRIORITY_FONTS[prio]
            priority_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Severity column (col 5)
        severity_cell = ws.cell(row=row_idx, column=5)
        sev = str(severity_cell.value)
        if sev in SEVERITY_FILLS:
            severity_cell.fill = SEVERITY_FILLS[sev]
            severity_cell.font = SEVERITY_FONTS[sev]
            severity_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Status column (col 12)
        status_cell = ws.cell(row=row_idx, column=12)
        if str(status_cell.value) == "OPEN":
            status_cell.fill = STATUS_FILL
            status_cell.font = STATUS_FONT
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Center-align specific narrow columns: ID(1), Priority(4), Severity(5), Effort(7), Deadline(8), Status(12)
        for c in (1, 7, 8):
            ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # ---- Summary row ----
    summary_row = len(ROWS) + 3
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=summary_row, column=1).border = THIN_BORDER

    total_effort = sum(r[6] for r in ROWS)
    effort_cell = ws.cell(row=summary_row, column=7, value=total_effort)
    effort_cell.font = Font(name="Calibri", bold=True, size=11)
    effort_cell.alignment = Alignment(horizontal="center", vertical="center")
    effort_cell.border = THIN_BORDER

    weeks_cell = ws.cell(row=summary_row, column=8, value=f"~{total_effort / 40:.0f} weeks @ 40h/wk")
    weeks_cell.font = Font(name="Calibri", bold=True, size=11)
    weeks_cell.border = THIN_BORDER

    # ---- Auto-width columns ----
    for col_idx in range(1, len(HEADERS) + 1):
        max_len = len(str(HEADERS[col_idx - 1]))
        for row_idx in range(2, len(ROWS) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        # Cap width to avoid excessively wide columns
        adjusted = min(max_len + 3, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # ---- Auto-filter ----
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    out_dir = Path(__file__).resolve().parent
    out_path = out_dir / "audit_public_v1_REMEDIATION_ROADMAP.xlsx"

    wb = build_workbook()
    wb.save(str(out_path))
    print(f"Written: {out_path}")
    print(f"Rows:    {len(ROWS)} remediation items")
    total = sum(r[6] for r in ROWS)
    print(f"Effort:  ~{total}h (~{total / 40:.0f} weeks at 40h/week)")


if __name__ == "__main__":
    main()
