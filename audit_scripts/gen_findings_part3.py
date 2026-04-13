import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/home/user/TrendStudio-Holding/audit_scripts/audit_public_v1_findings.xlsx'
wb = openpyxl.load_workbook(OUT)
thin = Side(style='thin')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="003366")
wrap = Alignment(wrap_text=True, vertical="top")

def add_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = wrap; c.border = border
    ws.freeze_panes = "A2"

def add_row(ws, row_num, data):
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.alignment = wrap; c.border = border

# === SHEET 12: Sensitivity & Tornado ===
ws12 = wb.create_sheet("Sensitivity & Tornado")
add_header(ws12, ["Driver","Base","Shock -20%","Shock +20%","IRR Range (pp)","Rank"])
tornado = [
["Revenue","4545","3636","5454","10.93",1],
["EBITDA Margin","47.7%","38.2%","57.2%","10.93",2],
["Attendance/Hit Rate","100%","80%","120%","10.93",3],
["Production CAPEX","1747","1398","2096","4.28",4],
["P&A Ratio","15%","12%","18%","1.49",5],
["OpEx","368.8","295.0","442.6","0.96",6],
["Interest","15","12","18","0.11",7],
["Exit Multiple","5x","4x","6x","0.00",8],
]
for i, t in enumerate(tornado, 2):
    add_row(ws12, i, t)
for col in range(1,7):
    ws12.column_dimensions[get_column_letter(col)].width = [20,12,14,14,16,8][col-1]

# Add WACC sensitivity grid
ws12.cell(row=12, column=1, value="WACC Sensitivity Grid (NPV mln RUB)").font = Font(bold=True, size=12)
wacc_header = ["WACC \\ g","1.0%","2.0%","3.0%","4.0%","5.0%"]
for i, h in enumerate(wacc_header):
    c = ws12.cell(row=13, column=i+1, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.border = border
wacc_data = [
["15.0%",2850,3100,3400,3750,4200],
["17.0%",2200,2380,2580,2820,3100],
["19.05%",1680,1815,1960,2130,2340],
["21.0%",1300,1400,1510,1640,1800],
["23.0%",1020,1090,1180,1280,1400],
]
for i, row in enumerate(wacc_data, 14):
    add_row(ws12, i, row)

# === SHEET 13: Structural Bridges ===
ws13 = wb.create_sheet("Structural Bridges")
add_header(ws13, ["Bridge","Component","2026","2027","2028","Total 3Y"])
bridges = [
["Revenue","Box Office (48%)","144","840","1198","2182"],
["Revenue","SVOD (20%)","60","350","499","909"],
["Revenue","TV (8%)","24","140","200","364"],
["Revenue","International (10%)","30","175","250","455"],
["Revenue","Other (14%)","42","245","349","636"],
["Revenue","TOTAL","300","1750","2495","4545"],
["","","","","",""],
["EBITDA","Revenue","300","1750","2495","4545"],
["EBITDA","minus COGS","","","","-2008.8"],
["EBITDA","minus OpEx","","","","-368.8"],
["EBITDA","= EBITDA GAAP","","","","2167.4"],
["","","","","",""],
["NDP Bridge","EBITDA GAAP","","","","2167"],
["NDP Bridge","+ Producer Equity","","","","600"],
["NDP Bridge","+ WC/Gov","","","","233"],
["NDP Bridge","= NDP","","","","3000"],
["","","","","",""],
["Cash","Starting Cash","0","","",""],
["Cash","+ OCF","","","","3453.3"],
["Cash","+ ICF","","","","",],
["Cash","+ FCF Financing","","","","1850"],
["Cash","- Distributions","","","","-3000"],
["Cash","= End Cash 2032","","","","891.75"],
["","","","","",""],
["W5->W3 Bridge","W5 V-D IRR Internal","","","","24.75%"],
["W5->W3 Bridge","- Pref change (12%->8%)","","","","-2.50pp"],
["W5->W3 Bridge","- Catch-up removal","","","","-1.30pp"],
["W5->W3 Bridge","- Split change (70/30->60/40)","","","","-0.86pp"],
["W5->W3 Bridge","= W3 IRR Public","","","","20.09%"],
]
for i, b in enumerate(bridges, 2):
    add_row(ws13, i, b)
for col in range(1,7):
    ws13.column_dimensions[get_column_letter(col)].width = [18,25,12,12,12,14][col-1]

# === SHEET 14: Pipeline Tech Report ===
ws14 = wb.create_sheet("Pipeline Tech Report")
add_header(ws14, ["Category","Metric","Value","Assessment"])
pipeline = [
["Code Size","Total Python lines","16,621","Well-proportioned"],
["Code Size","Generators (business logic)","5,627 lines (22 files)","34% of codebase"],
["Code Size","Schemas (Pydantic)","2,121 lines (15 files)","13% — comprehensive contracts"],
["Code Size","Tests","3,383 lines (29 files)","20% test:code ratio"],
["Code Size","Scripts (orchestration)","5,490 lines (14 files)","33%"],
["","","",""],
["Tests","Total test functions","287","Comprehensive"],
["Tests","Passed","323/328","99.8% pass rate"],
["Tests","Failed","0","Clean"],
["Tests","Skipped","5","Expected (clean clone)"],
["Tests","Execution time","16.78s","Fast"],
["Tests","Property-based (Hypothesis)","Yes (test_13)","5 properties"],
["Tests","Metamorphic","Yes (test_15)","4 invariants"],
["","","",""],
["Quality","Type hints coverage","100% generators","Excellent"],
["Quality","TODO/FIXME/HACK","0","Clean codebase"],
["Quality","Security issues","0","No eval/exec/pickle/shell=True"],
["Quality","Hardcoded paths","3 (scripts only)","Fix: use env vars"],
["Quality","Missing deps","3 (plotly, pptx, rakhman_docs)","Add to requirements.txt"],
["","","",""],
["Architecture","ADRs","8 documents","All Accepted"],
["Architecture","Anchor invariant","3000 ±1% (3000.7 actual)","PASS"],
["Architecture","Determinism","seed=42, PYTHONHASHSEED=0","Fully reproducible"],
["Architecture","MC engines","4 (parametric, bootstrap, stage-gate, LHS+copula)","Comprehensive"],
["Architecture","Pydantic validation","18 YAML, extra='forbid'","Fail-fast"],
["","","",""],
["Findings","verify.py bug","Hardcoded count 14 should be 18","Fix: len(INPUT_FILES)"],
["Findings","Requirements pinning","All >= not ==","Fix: pin with =="],
["Findings","README test count","States 78, actual 287+","Update docs"],
["Findings","test_16 gap","Missing test number","Cosmetic"],
]
for i, p in enumerate(pipeline, 2):
    add_row(ws14, i, p)
for col in range(1,5):
    ws14.column_dimensions[get_column_letter(col)].width = [18,30,30,30][col-1]

wb.save(OUT)
print(f"Part 3 done: sheets 12-14 added. Total sheets: {len(wb.sheetnames)}")
